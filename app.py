"""
Movi Motors ERP — Streamlit + Supabase (USD base, multimoneda BS/USDT).

Conexión: cliente oficial `supabase` con secretos en [connections.supabase].
(Streamlit también documenta `st_supabase_connection.SupabaseConnection`; si lo
instalas, puedes sustituir `get_supabase()` por `st.connection(...)`.)

Requisitos: ejecutar supabase/schema_erp_multimoneda.sql (o patch_004 si ya tenías
BD) y configurar `.streamlit/secrets.toml` con Supabase (no subir a Git).

Acceso: cada persona entra con su **usuario** y **contraseña** (guardada hasheada
en `erp_users`). El **superusuario** crea usuarios y asigna rol: administrador,
vendedor o almacén. Tras entrar, la sesión se guarda en una **cookie firmada**
(por usuario y por cada login); al refrescar la página sigues identificado hasta
que pulses **Cerrar sesión** o venza la vigencia (p. ej. 90 días).
"""

from __future__ import annotations

import base64
from collections import Counter, defaultdict
import gzip
import hashlib
import hmac
import html
import json
import math
import re
import secrets
import unicodedata
import time
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote
from zoneinfo import ZoneInfo

import bcrypt
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from supabase import Client


from movi.paths import APP_DIR as _APP_DIR, BRAND_LOGO_PATH
from movi.brand import brand_logo_file, brand_logo_data_uri, render_brand_logo

_brand_logo_data_uri = brand_logo_data_uri  # alias usado en PDF/HTML

from movi.theme import render_movi_ui_theme_styles, render_movi_theme_picker
from movi.rbac import MOVI_MOD_ICONS, movi_nav_options_for_role, role_can
from movi.nav import render_movi_main_module_nav
from movi.services.dashboard_kpis import compute_dashboard_kpis_periodo
from movi.modules.tasas import TasasModuleDeps, render_module_tasas

_PAGE_ICON: str = brand_logo_file() or "⚙️"
st.set_page_config(
    page_title="Movi Motor's Importadora · ERP",
    page_icon=_PAGE_ICON,
    layout="wide",
    initial_sidebar_state="expanded",
)

render_movi_ui_theme_styles()


def _movi_ss_pop_keys(*keys: str) -> None:
    for k in keys:
        st.session_state.pop(k, None)


def _movi_ss_pop_key_prefixes(*prefixes: str) -> None:
    for k in list(st.session_state.keys()):
        if any(str(k).startswith(p) for p in prefixes):
            st.session_state.pop(k, None)


def _movi_bump_form_nonce(name: str) -> None:
    st.session_state[name] = int(st.session_state.get(name, 0)) + 1


def _movi_reset_venta_form_fields() -> None:
    """Quita el estado de widgets del formulario de venta (incl. cobros y líneas de producto)."""
    _movi_ss_pop_key_prefixes("vp_", "vq_", "vpu_", "vcb_", "vca_", "vsrl_")
    _movi_ss_pop_keys(
        "venta_doc_tasa_bs",
        "venta_tasa_bs_override",
        "venta_abono_credito",
        "venta_cli",
        "venta_forma",
        "venta_fv",
        "venta_notas",
    )


def _rpc_resp_uuid(resp: Any) -> str | None:
    """Extrae UUID devuelto por RPC `RETURNS UUID` (PostgREST / supabase-py)."""
    data = getattr(resp, "data", None)
    if data is None:
        return None
    if isinstance(data, str) and len(data) >= 32:
        return data.strip()
    if isinstance(data, list) and data:
        x = data[0]
        if isinstance(x, str) and len(x) >= 32:
            return x.strip()
    return None


def _rpc_resp_int(resp: Any) -> int | None:
    """Extrae entero devuelto por RPC `RETURNS INTEGER` (PostgREST / supabase-py)."""
    data = getattr(resp, "data", None)
    if data is None:
        return None
    if isinstance(data, bool):
        return int(data)
    if isinstance(data, int):
        return data
    if isinstance(data, float):
        return int(data)
    if isinstance(data, str) and data.strip().lstrip("-").isdigit():
        return int(data.strip())
    if isinstance(data, list) and data:
        x = data[0]
        if isinstance(x, bool):
            return int(x)
        if isinstance(x, int):
            return x
        if isinstance(x, float):
            return int(x)
    return None


def _venta_firma_registro(cliente: str, est_total: float, new_lines: list[dict[str, Any]]) -> str:
    """Firma del intento de venta para detectar doble envío inmediato."""
    try:
        payload = json.dumps(
            {"c": (cliente or "").strip(), "t": round(float(est_total), 2), "l": new_lines},
            sort_keys=True,
            default=str,
        )
    except Exception:
        payload = str(cliente) + str(est_total) + str(new_lines)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _venta_set_ok_banner_desde_id(
    sb: Client,
    *,
    venta_id: str,
    cliente: str,
    est_total: float,
    forma: str,
    extra_caption: str | None = None,
) -> None:
    try:
        rr = (
            sb.table("ventas")
            .select("numero,total_usd,forma_pago,cliente")
            .eq("id", venta_id)
            .limit(1)
            .execute()
        )
        row = (rr.data or [{}])[0] if rr.data else {}
    except Exception:
        row = {}
    st.session_state["venta_ok_banner"] = {
        "id": venta_id,
        "numero": row.get("numero"),
        "total_usd": float(row.get("total_usd") if row.get("total_usd") is not None else est_total),
        "cliente": (row.get("cliente") if row.get("cliente") is not None else cliente) or "",
        "forma_pago": str(row.get("forma_pago") or forma or ""),
        "extra": extra_caption,
    }


def _movi_reset_venta_session_nueva(plist: list[dict[str, Any]], id_to_price: dict[str, float]) -> None:
    """Una línea de producto por defecto, un cobro, sin datos del cliente; nuevo `form` vía nonce."""
    if not plist:
        return
    pid0 = str(plist[0]["id"])
    st.session_state["venta_lines"] = [
        {"producto_id": pid0, "cantidad": 1, "precio_unitario_usd": float(id_to_price.get(pid0, 0))}
    ]
    st.session_state["venta_n_cobros"] = 1
    _movi_reset_venta_form_fields()
    _movi_bump_form_nonce("venta_form_nonce")


def _movi_reset_compra_form_fields() -> None:
    _movi_ss_pop_key_prefixes("cp_", "cq_", "ccu_")
    _movi_ss_pop_keys(
        "forma_compra",
        "caja_compra",
        "fv_compra",
        "compra_doc_tasa_bs",
        "compra_prov",
        "compra_notas",
    )


def _movi_reset_producto_alta_fields() -> None:
    _movi_ss_pop_keys(
        "inv_alta_prod_desc",
        "inv_alta_prod_cat",
        "inv_alta_marca_prod",
        "inv_alta_cond",
        "inv_alta_cod_auto",
        "inv_alta_prod_codigo",
        "inv_alta_prod_sku_oem",
        "inv_alta_marcas_pick",
        "inv_alta_marcas_veh",
        "inv_alta_anos",
        "inv_alta_ubic",
        "inv_alta_img",
        "inv_alta_img_file",
        "inv_alta_stock",
        "inv_alta_seriales",
        "inv_alta_smin",
        "inv_alta_costo",
        "inv_alta_pv",
    )


def _movi_api_response_first_row_data(data: Any) -> dict[str, Any] | None:
    """Normaliza `data` de APIResponse: lista de filas, un solo dict, o vacío."""
    if data is None:
        return None
    if isinstance(data, dict):
        if data.get("message") and (
            "code" in data or "hint" in data or "details" in data
        ):
            return None
        return data
    if isinstance(data, list):
        if not data:
            return None
        r0 = data[0]
        return r0 if isinstance(r0, dict) else None
    return None


def _inv_resolve_producto_id_after_insert(
    sb: Client, *, ins_data: Any, codigo: str | None
) -> str:
    """Obtiene `id` del cuerpo del insert o, si falta, con SELECT por código único."""
    row = _movi_api_response_first_row_data(ins_data)
    rid = str((row or {}).get("id") or "").strip()
    if rid:
        return rid
    c = (codigo or "").strip()
    if not c:
        return ""
    try:
        chk = sb.table("productos").select("id").eq("codigo", c).limit(1).execute()
        row2 = _movi_api_response_first_row_data(chk.data)
        return str((row2 or {}).get("id") or "").strip()
    except Exception:
        return ""


def _inv_alta_producto_id_missing_help() -> str:
    return (
        "El alta no devolvió el ID del producto y no apareció al buscar por código. "
        "En `.streamlit/secrets.toml` usá **SUPABASE_KEY** = clave **service_role** "
        "(Supabase → Project Settings → API), no la clave **anon**. "
        "Con anon, RLS suele bloquear lecturas aunque el insert parezca OK."
    )


def _movi_productos_insert_execute(sb: Client, row: dict[str, Any]) -> Any:
    """
    Insert en `productos`. No encadenar `.select()` tras `.insert()`: en postgrest-py
    reciente el builder es SyncQueryRequestBuilder y **no tiene** `.select()` (AttributeError).
    El insert ya usa `return=representation` por defecto; si el cuerpo viniera vacío,
    `_inv_resolve_producto_id_after_insert` busca por `codigo`.
    """
    return sb.table("productos").insert(row).execute()


def _erp_user_uuid_or_none(erp_uid: str) -> str | None:
    s = str(erp_uid or "").strip()
    if re.fullmatch(
        r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}",
        s,
    ):
        return s
    return None


def _movi_reset_inv_ficha_product_keys(product_id: str) -> None:
    pid = str(product_id).strip()
    if not pid:
        return
    suf = f"_{pid}"
    for k in list(st.session_state.keys()):
        sk = str(k)
        if sk.startswith("inv_ficha_") and sk.endswith(suf):
            st.session_state.pop(k, None)


def _secrets_ready() -> bool:
    try:
        _ = st.secrets["connections"]["supabase"]["SUPABASE_URL"]
        _ = st.secrets["connections"]["supabase"]["SUPABASE_KEY"]
        return True
    except Exception:
        return False


ERP_SESSION_COOKIE = "movi_erp_session"
SESSION_MAX_DAYS = 90


def _session_signing_key() -> bytes:
    try:
        auth = st.secrets.get("auth")
        if isinstance(auth, dict):
            sk = auth.get("SESSION_SIGNING_KEY")
            if sk:
                return str(sk).encode("utf-8")
    except Exception:
        pass
    k = st.secrets["connections"]["supabase"]["SUPABASE_KEY"]
    return hashlib.sha256((str(k) + "|movi_erp_session_v1").encode("utf-8")).digest()


def _encode_session_token(uid: str, sid: str, exp_unix: int) -> str:
    payload = {"v": 1, "uid": str(uid), "sid": str(sid), "exp": int(exp_unix)}
    body = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    sig = hmac.new(_session_signing_key(), body, hashlib.sha256).hexdigest()
    wrapped = json.dumps({"p": payload, "sig": sig}, separators=(",", ":"))
    return base64.urlsafe_b64encode(wrapped.encode("utf-8")).decode("ascii").rstrip("=")


def _decode_session_token(token: str) -> dict[str, Any] | None:
    try:
        pad = "=" * (-len(token) % 4)
        raw = base64.urlsafe_b64decode(token + pad)
        o = json.loads(raw.decode("utf-8"))
        p = o.get("p")
        sig = o.get("sig")
        if not isinstance(p, dict) or not isinstance(sig, str):
            return None
        body = json.dumps(p, sort_keys=True, separators=(",", ":")).encode("utf-8")
        expect = hmac.new(_session_signing_key(), body, hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expect, sig):
            return None
        if int(p.get("exp", 0)) < int(datetime.now(timezone.utc).timestamp()):
            return None
        return p
    except Exception:
        return None


def _erp_cookie_manager():
    from extra_streamlit_components import CookieManager

    return CookieManager(key="movi_erp_cookie_mgr")


def _persist_new_session_cookie(cm: Any | None, row: dict[str, Any]) -> None:
    """Un token nuevo en cada login (sid distinto); la cookie queda ligada a ese usuario (uid)."""
    if cm is None:
        return
    uid = str(row["id"])
    sid = secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    exp_unix = int(now.timestamp()) + SESSION_MAX_DAYS * 86400
    tok = _encode_session_token(uid, sid, exp_unix)
    expires_at = now + timedelta(days=SESSION_MAX_DAYS)
    cm.set(
        ERP_SESSION_COOKIE,
        tok,
        key="erp_cookie_set_login",
        path="/",
        expires_at=expires_at,
        same_site="lax",
    )


def _try_restore_session_from_cookie(sb: Client, cm: Any | None) -> None:
    if cm is None or st.session_state.get("erp_uid"):
        return
    tok = cm.get(ERP_SESSION_COOKIE)
    if not tok or not isinstance(tok, str):
        return
    payload = _decode_session_token(tok)
    if not payload:
        try:
            cm.delete(ERP_SESSION_COOKIE, key="erp_cookie_del_bad")
        except Exception:
            pass
        return
    uid = str(payload.get("uid", "")).strip()
    if not uid:
        try:
            cm.delete(ERP_SESSION_COOKIE, key="erp_cookie_del_bad2")
        except Exception:
            pass
        return
    r = (
        sb.table("erp_users")
        .select("id,username,nombre,rol,activo")
        .eq("id", uid)
        .limit(1)
        .execute()
    )
    row = (r.data or [None])[0]
    if not row or row.get("activo") is False:
        try:
            cm.delete(ERP_SESSION_COOKIE, key="erp_cookie_del_inactive")
        except Exception:
            pass
        return
    st.session_state["erp_uid"] = str(row["id"])
    st.session_state["erp_rol"] = str(row["rol"])
    st.session_state["erp_nombre"] = str(row.get("nombre") or row["username"])
    st.session_state["erp_username"] = str(row["username"])
    st.rerun()


def _logout() -> None:
    for k in ("erp_uid", "erp_rol", "erp_nombre", "erp_username"):
        st.session_state.pop(k, None)
    try:
        cm = _erp_cookie_manager()
        cm.delete(ERP_SESSION_COOKIE, key="erp_cookie_del_logout")
    except Exception:
        pass


def _cookie_support() -> bool:
    try:
        import extra_streamlit_components  # noqa: F401

        return True
    except ImportError:
        return False


def _hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")


def _password_ok(plain: str, stored_hash: str) -> bool:
    h = (stored_hash or "").strip()
    if not h or not h.startswith("$2"):
        return False
    b = plain.encode("utf-8")
    try:
        if bcrypt.checkpw(b, h.encode("utf-8")):
            return True
    except Exception:
        pass
    # PostgreSQL pgcrypto suele guardar $2a$; la librería bcrypt en Python a veces falla: probamos $2b$
    if h.startswith("$2a$"):
        try:
            h2 = "$2b$" + h[4:]
            return bool(bcrypt.checkpw(b, h2.encode("utf-8")))
        except Exception:
            return False
    return False


def _fetch_erp_user_by_login(sb: Client, username_normalized: str) -> dict[str, Any] | None:
    """Coincide usuario ignorando mayúsculas (p. ej. Admin = admin)."""
    resp = sb.table("erp_users").select("id,username,nombre,rol,password_hash,activo").execute()
    u = username_normalized.strip().lower()
    for row in resp.data or []:
        if str(row.get("username", "")).strip().lower() == u:
            return row
    return None


def gate_user_login(sb: Client, cm: Any | None) -> dict[str, Any] | None:
    if st.session_state.get("erp_uid"):
        return {
            "id": st.session_state["erp_uid"],
            "rol": st.session_state["erp_rol"],
            "nombre": st.session_state["erp_nombre"],
            "username": st.session_state["erp_username"],
        }

    _, c_logo, _ = st.columns([1, 2.2, 1])
    with c_logo:
        render_brand_logo()
    st.markdown("---")
    st.subheader("Iniciar sesión")
    _persist_hint = (
        f"En este equipo la sesión se mantiene al refrescar (hasta **{SESSION_MAX_DAYS} días** o **Cerrar sesión**)."
        if _cookie_support()
        else "Para recordar la sesión al refrescar, instala: `python -m pip install extra-streamlit-components`."
    )
    st.caption(
        "Usuario y contraseña los asigna el superusuario en **Mantenimiento → Usuarios del sistema**. "
        "El nombre de usuario no distingue mayúsculas (admin = Admin). "
        + _persist_hint
    )
    with st.expander("Tema visual (sin entrar)", expanded=False):
        st.caption("Elegí colores más alegres o elegantes; el contraste del texto está cuidado.")
        render_movi_theme_picker(key_suffix="login")
    user = st.text_input("Usuario", autocomplete="username")
    pwd = st.text_input("Contraseña", type="password", autocomplete="current-password")
    if st.button("Entrar"):
        u = (user or "").strip().lower()
        if not u or not pwd:
            st.error("Complete usuario y contraseña.")
            return None
        row = _fetch_erp_user_by_login(sb, u)
        if not row:
            st.error("Usuario o contraseña incorrectos.")
            return None
        if row.get("activo") is False:
            st.error("Este usuario está desactivado.")
            return None
        ph = (row.get("password_hash") or "").strip()
        if not ph:
            st.error("Tu cuenta aún no tiene contraseña. Pide al superusuario que te la asigne.")
            return None
        if not _password_ok(pwd, ph):
            st.error("Usuario o contraseña incorrectos.")
            return None
        st.session_state["erp_uid"] = str(row["id"])
        st.session_state["erp_rol"] = str(row["rol"])
        st.session_state["erp_nombre"] = str(row.get("nombre") or row["username"])
        st.session_state["erp_username"] = str(row["username"])
        _persist_new_session_cookie(cm, row)
        st.rerun()
    return None


@st.cache_resource
def get_supabase() -> Client:
    from supabase import create_client

    u = st.secrets["connections"]["supabase"]["SUPABASE_URL"]
    k = st.secrets["connections"]["supabase"]["SUPABASE_KEY"]
    return create_client(str(u), str(k))


def latest_tasas(sb: Client) -> dict[str, Any] | None:
    r = (
        sb.table("tasas_dia")
        .select("*")
        .order("fecha", desc=True)
        .limit(1)
        .execute()
    )
    rows = r.data or []
    return rows[0] if rows else None


def _today_caracas() -> date:
    return datetime.now(ZoneInfo("America/Caracas")).date()


def _tasas_para_fecha(sb: Client, d: date) -> dict[str, Any] | None:
    r = sb.table("tasas_dia").select("*").eq("fecha", str(d)).limit(1).execute()
    row = (r.data or [None])[0]
    return row if isinstance(row, dict) else None


# Sincronización automática tasas web → `tasas_dia` + recálculo Bs en productos (RPC).
# Si REL = 0, solo cuenta la diferencia absoluta en Bs/USD (más “tiempo real”).
AUTO_TASA_SYNC_REL_MIN = 0.0
# Mínimo cambio en Bs por 1 USD para escribir (evita ruido de redondeo de la API).
AUTO_TASA_ABS_MIN_BS = 0.02
# Alineado con caché de `get_live_exchange_rates` (~120 s): cada refresco web puede persistirse.
AUTO_TASA_SYNC_MIN_SECONDS = 120.0


def _refresh_productos_bs_equiv_note(sb: Client, t_oper: float) -> str:
    """Tras actualizar `tasa_bs`, recalcula precio_v_bs_ref y costo_bs_ref en productos."""
    try:
        sb.rpc("refresh_productos_bs_equiv", {"p_tasa_bs": float(t_oper)}).execute()
        return " Productos: equivalentes Bs recalculados en BD."
    except Exception:
        return " (Ejecuta `supabase/patch_007_productos_bs_ref.sql` para recalcular Bs en productos.)"


def maybe_auto_sync_tasas_from_web(sb: Client) -> dict[str, Any] | None:
    """
    Si el USD×Bs web difiere del guardado (abs ≥ AUTO_TASA_ABS_MIN_BS y/o rel ≥ AUTO_TASA_SYNC_REL_MIN),
    hace upsert del día (Caracas): actualiza la **ref. Bs/USD mercado (API / alineada a P2P)**; **no pisa el BCV** que tengas guardado.
    `tasa_bs` sigue **BCV** si en el último guardado estabas operando con BCV; si operabas con mercado P2P, sigue esa cotización web.
    Tras guardar, llama `refresh_productos_bs_equiv` con esa `tasa_bs`.
    """
    t_latest = latest_tasas(sb)
    if not t_latest:
        return None

    now = time.monotonic()
    last = float(st.session_state.get("_auto_tasas_sync_mono", 0.0))
    if now - last < AUTO_TASA_SYNC_MIN_SECONDS:
        return latest_tasas(sb)

    live = get_live_exchange_rates()
    if not live.get("ok"):
        return latest_tasas(sb)

    ves = _nf(live.get("ves_bs_por_usd"))
    if ves is None or ves <= 0:
        return latest_tasas(sb)

    hoy = _today_caracas()
    t_hoy = _tasas_para_fecha(sb, hoy)
    base = t_hoy or t_latest
    saved_par = _nf(base.get("paralelo_bs_por_usd")) or _nf(base.get("tasa_bs"))
    if saved_par is None or saved_par <= 0:
        return latest_tasas(sb)

    abs_diff = abs(float(ves) - float(saved_par))
    rel_diff = abs_diff / float(saved_par) if saved_par > 0 else 0.0
    rel_ok = AUTO_TASA_SYNC_REL_MIN > 0 and rel_diff >= AUTO_TASA_SYNC_REL_MIN
    abs_ok = abs_diff >= AUTO_TASA_ABS_MIN_BS
    if not rel_ok and not abs_ok:
        return latest_tasas(sb)

    bcv = _nf(base.get("bcv_bs_por_usd"))
    if bcv is None or bcv <= 0:
        bcv = _nf(t_latest.get("bcv_bs_por_usd")) or _nf(t_latest.get("tasa_bs")) or float(saved_par)

    tusdt = _nf(base.get("tasa_usdt")) or _nf(t_latest.get("tasa_usdt")) or 1.0
    if tusdt <= 0:
        tusdt = 1.0

    usd_eur = _nf(live.get("usd_por_eur")) or _nf(base.get("usd_por_eur")) or _nf(t_latest.get("usd_por_eur")) or 1.08
    if usd_eur <= 0:
        usd_eur = 1.08

    p2p = _nf(live.get("usdt_x_ves_p2p")) or _nf(live.get("p2p_bs_por_usdt_aprox"))
    if p2p is None or p2p <= 0:
        p2p = _nf(base.get("p2p_bs_por_usdt")) or _nf(t_latest.get("p2p_bs_por_usdt")) or float(ves)

    par = float(ves)
    prev_tb = _nf(base.get("tasa_bs"))
    prev_bcv = _nf(base.get("bcv_bs_por_usd")) or prev_tb
    prev_par = _nf(base.get("paralelo_bs_por_usd")) or prev_tb
    if (
        prev_tb is not None
        and prev_bcv is not None
        and prev_par is not None
        and float(prev_bcv) > 0
        and float(prev_par) > 0
    ):
        d_b = abs(float(prev_tb) - float(prev_bcv))
        d_p = abs(float(prev_tb) - float(prev_par))
        t_oper = float(bcv) if d_b <= d_p else par
    else:
        t_oper = par
    row = {
        "fecha": str(hoy),
        "tasa_bs": float(t_oper),
        "tasa_usdt": float(tusdt),
        "bcv_bs_por_usd": float(bcv),
        "paralelo_bs_por_usd": float(par),
        "usd_por_eur": float(usd_eur),
        "p2p_bs_por_usdt": float(p2p),
    }
    try:
        sb.table("tasas_dia").upsert(row, on_conflict="fecha").execute()
        st.session_state["_auto_tasas_sync_mono"] = now
        prod_note = _refresh_productos_bs_equiv_note(sb, float(t_oper))
        modo = (
            "tasa_bs = BCV guardado"
            if abs(float(t_oper) - float(bcv)) < 1e-6
            else "tasa_bs = mercado P2P / ref. web"
        )
        st.session_state["_tasas_auto_sync_msg"] = (
            f"Tasas auto: mercado/ref. web {par:,.2f} Bs/USD (antes {saved_par:,.2f}). {modo}.{prod_note}"
        )
    except Exception:
        pass
    return latest_tasas(sb)


_INV_PRODUCTO_COLS = [
    "id",
    "codigo",
    "sku_oem",
    "descripcion",
    "marca_producto",
    "condicion",
    "ubicacion",
    "compatibilidad",
    "imagen_url",
    "stock_actual",
    "stock_minimo",
    "costo_usd",
    "precio_v_usd",
    "precio_v_bs_ref",
    "costo_bs_ref",
    "activo",
    "categoria_id",
    "es_compuesto",
]


def _inv_compat_as_dict(raw: Any) -> dict[str, Any]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return {}
    if isinstance(raw, dict):
        return dict(raw)
    if isinstance(raw, str):
        s = raw.strip()
        if not s:
            return {}
        try:
            o = json.loads(s)
            return o if isinstance(o, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _inv_compat_marcas_str(d: dict[str, Any]) -> str:
    m = d.get("marcas_vehiculo") if "marcas_vehiculo" in d else d.get("marcas")
    if isinstance(m, list):
        return ", ".join(str(x).strip() for x in m if str(x).strip())
    if isinstance(m, str) and m.strip():
        return m.strip()
    return ""


def _inv_compat_anos_str(d: dict[str, Any]) -> str:
    a = d.get("años") if "años" in d else d.get("anos")
    if a is None or (isinstance(a, float) and pd.isna(a)):
        return ""
    return str(a).strip()


def _inv_build_compat_dict(marcas_csv: str, anos: str) -> dict[str, Any]:
    raw = (marcas_csv or "").replace(";", ",")
    marcas = [x.strip() for x in raw.split(",") if x.strip()]
    out: dict[str, Any] = {}
    if marcas:
        out["marcas_vehiculo"] = marcas
    a = (anos or "").strip()
    if a:
        out["años"] = a
    return out


def _inv_categoria_sugiere_seriales_motor(nombre_categoria: str) -> bool:
    """True si el nombre de categoría indica motores (p. ej. «Motores», «Motor usado»)."""
    n = (nombre_categoria or "").strip().lower()
    return "motor" in n


def _inv_compat_seriales_motor_list(d: dict[str, Any]) -> list[str]:
    raw = d.get("seriales_motor")
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    return []


def _venta_serial_en_pool_motor(serial_vendido: str, pool: list[str]) -> bool:
    """Igual que en `crear_venta_erp` (patch_023): compara con trim."""
    a = (serial_vendido or "").strip()
    if not a:
        return False
    for x in pool:
        if (x or "").strip() == a:
            return True
    return False


def _inv_parse_seriales_motor_texto(s: str) -> list[str]:
    out: list[str] = []
    for line in (s or "").splitlines():
        for part in line.replace(";", ",").split(","):
            t = part.strip()
            if t:
                out.append(t)
    return out


def _inv_compat_merge_seriales(base: dict[str, Any], seriales: list[str]) -> dict[str, Any]:
    out = dict(base)
    if seriales:
        out["seriales_motor"] = seriales
    else:
        out.pop("seriales_motor", None)
    return out


def _inv_compat_seriales_motor_resumen(d: dict[str, Any], *, max_vis: int = 4) -> str:
    lst = _inv_compat_seriales_motor_list(d)
    if not lst:
        return ""
    head = lst[:max_vis]
    suf = f" (+{len(lst) - max_vis} más)" if len(lst) > max_vis else ""
    return ", ".join(head) + suf


def _inv_merge_marcas_catalogo_texto(seleccion: list[str], texto_extra: str) -> str:
    """Une marcas elegidas en multiselect + las que escribís a mano (coma o punto y coma)."""
    nombres: set[str] = set()
    for x in seleccion or []:
        s = str(x).strip()
        if s:
            nombres.add(s)
    raw = (texto_extra or "").replace(";", ",")
    for x in raw.split(","):
        s = x.strip()
        if s:
            nombres.add(s)
    return ", ".join(sorted(nombres, key=str.casefold))


def _codigo_interno_slug(s: str, *, max_len: int = 3) -> str:
    """Fragmento corto en mayúsculas (solo letras/números) para armar códigos tipo FIL-BOS-0001."""
    if not s or not str(s).strip():
        return "GEN"[:max_len]
    t = unicodedata.normalize("NFKD", str(s).strip())
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = "".join(ch for ch in t.upper() if ch.isalnum())
    if not t:
        return "GEN"[:max_len]
    return t[:max_len]


def _siguiente_codigo_interno_producto(sb: Client, nombre_categoria: str, marca_repuesto: str) -> str:
    """
    Código automático: CATEGORÍA(3)-MARCA_REP(3)-NNNN.
    Ej.: categoría «Filtros» + marca «Bosch» → FIL-BOS-0007
    """
    cat = _codigo_interno_slug(nombre_categoria or "CAT", max_len=3)
    mar = _codigo_interno_slug(marca_repuesto or "GEN", max_len=3)
    prefix = f"{cat}-{mar}-"
    try:
        r = sb.table("productos").select("codigo").like("codigo", f"{prefix}%").execute()
    except Exception:
        r = type("R", (), {"data": []})()
    max_n = 0
    rx = re.compile(re.escape(prefix) + r"(\d{1,6})$", re.IGNORECASE)
    for row in r.data or []:
        c = (row.get("codigo") or "").strip()
        m = rx.match(c)
        if m:
            try:
                max_n = max(max_n, int(m.group(1)))
            except ValueError:
                pass
    return f"{prefix}{max_n + 1:04d}"


def _fetch_marcas_vehiculo_catalogo(sb: Client) -> list[str]:
    try:
        mr = (
            sb.table("marcas_vehiculo")
            .select("nombre")
            .eq("activo", True)
            .order("orden")
            .execute()
        )
        return [str(r["nombre"]) for r in (mr.data or []) if r.get("nombre")]
    except Exception:
        return []


def _inv_row_matches_query(row: pd.Series, q: str) -> bool:
    ql = q.lower()
    for k in ("descripcion", "codigo", "sku_oem", "marca_producto"):
        if ql in str(row.get(k) or "").lower():
            return True
    d = _inv_compat_as_dict(row.get("compatibilidad"))
    for m in d.get("marcas_vehiculo") or d.get("marcas") or []:
        if ql in str(m).lower():
            return True
    if ql in _inv_compat_anos_str(d).lower():
        return True
    return False


def _inv_stock_int(x: Any) -> int:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return 0
        return int(round(float(x)))
    except (TypeError, ValueError):
        return 0


def _inv_fetch_productos_para_dropdown_eliminar(sb: Client) -> tuple[list[dict[str, Any]], str]:
    """
    Todos los productos para el desplegable de eliminar. Los kits también viven en `productos`
    (`es_compuesto`); `productos_kit_items` solo guarda componentes. Pagina resultados para no
    cortar en el límite típico de PostgREST (~1000 filas).
    """
    page = 1000
    out: list[dict[str, Any]] = []
    use_created = True
    order_hint = "Listado: **últimos cargados primero** (por `created_at`)."
    offset = 0
    while True:
        try:
            if use_created:
                r = (
                    sb.table("productos")
                    .select("id,codigo,descripcion,stock_actual,es_compuesto,created_at")
                    .order("created_at", desc=True)
                    .range(offset, offset + page - 1)
                    .execute()
                )
            else:
                r = (
                    sb.table("productos")
                    .select("id,codigo,descripcion,stock_actual,es_compuesto")
                    .order("descripcion")
                    .range(offset, offset + page - 1)
                    .execute()
                )
        except Exception:
            if use_created and offset == 0:
                use_created = False
                order_hint = "Listado: por **descripción** (A–Z); si falla `created_at`, sin orden de fecha."
                continue
            break
        batch = r.data or []
        out.extend(batch)
        if len(batch) < page:
            break
        offset += page
        if offset > 200_000:
            break
    return out, order_hint


def _line_qty_int(x: Any, *, default: int = 1) -> int:
    """Cantidad en líneas de venta/compra: entero ≥ 1."""
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return max(1, default)
        v = int(round(float(x)))
        return max(1, v)
    except (TypeError, ValueError):
        return max(1, default)


def _inv_eliminar_producto_stock_cero(sb: Client, producto_id: str, confirmacion: str) -> tuple[bool, str]:
    if (confirmacion or "").strip().upper() != "ELIMINAR":
        return False, "Escribí **ELIMINAR** en mayúsculas para confirmar."
    r = sb.table("productos").select("id,stock_actual,codigo,descripcion").eq("id", producto_id).limit(1).execute()
    rows = r.data or []
    if not rows:
        return False, "Producto no encontrado."
    row = rows[0]
    if _inv_stock_int(row.get("stock_actual")) != 0:
        return False, "Solo se puede eliminar con **stock en cero**."
    try:
        sb.table("productos").delete().eq("id", producto_id).execute()
    except Exception as e:
        es = str(e).lower()
        if "foreign key" in es or "23503" in es or "violates" in es:
            return (
                False,
                "No se puede borrar: el producto tiene **ventas o compras** registradas. "
                "Desactivá el ítem (**Activo** = no) en la tabla de abajo para dejar de usarlo.",
            )
        return False, str(e)
    cd = _export_cell_txt(row.get("codigo")) or "—"
    ds = _export_cell_txt(row.get("descripcion"))[:80]
    return True, f"Eliminado **{cd}** · {ds}"


def _inv_aplicar_movimiento_stock(
    sb: Client,
    erp_uid: str,
    producto_id: str,
    tipo: str,
    cantidad: int,
    motivo: str,
) -> tuple[bool, str]:
    if tipo not in ("Entrada", "Salida"):
        return False, "Tipo inválido."
    if cantidad < 1:
        return False, "La cantidad debe ser al menos **1**."
    r = sb.table("productos").select("id,stock_actual,descripcion,codigo").eq("id", producto_id).limit(1).execute()
    rows = r.data or []
    if not rows:
        return False, "Producto no encontrado."
    row = rows[0]
    st_antes = _inv_stock_int(row.get("stock_actual"))
    if tipo == "Salida":
        if st_antes < cantidad:
            return False, f"Stock insuficiente para descargar (hay **{st_antes}**)."
        st_desp = st_antes - cantidad
    else:
        st_desp = st_antes + cantidad
    mot = (motivo or "").strip()
    if not mot:
        return False, "Indicá un **motivo** (ej. inventario físico, merma, hallazgo, traslado)."
    try:
        sb.table("productos").update({"stock_actual": st_desp}).eq("id", producto_id).execute()
    except Exception as e:
        return False, str(e)
    try:
        sb.table("movimientos_inventario").insert(
            {
                "producto_id": producto_id,
                "tipo": tipo,
                "cantidad": int(cantidad),
                "motivo": mot[:2000],
                "stock_antes": st_antes,
                "stock_despues": st_desp,
                "usuario_id": erp_uid,
            }
        ).execute()
    except Exception as e:
        try:
            sb.table("productos").update({"stock_actual": st_antes}).eq("id", producto_id).execute()
        except Exception:
            pass
        return (
            False,
            f"{e} · Stock revertido. Ejecutá **supabase/patch_013_movimientos_inventario.sql** si falta la tabla.",
        )
    cd = _export_cell_txt(row.get("codigo")) or "—"
    return True, f"**{cd}** · Stock **{st_antes}** → **{st_desp}** ({tipo} **{cantidad}** unidades)."


def _fetch_kit_items_by_kit(sb: Client) -> dict[str, list[dict[str, Any]]]:
    try:
        r = (
            sb.table("productos_kit_items")
            .select("kit_producto_id,componente_producto_id,cantidad")
            .execute()
        )
    except Exception:
        return {}
    out: dict[str, list[dict[str, Any]]] = {}
    for row in r.data or []:
        kid = str(row.get("kit_producto_id") or "").strip()
        cid = str(row.get("componente_producto_id") or "").strip()
        if not kid or not cid:
            continue
        try:
            q = float(row.get("cantidad") or 0)
        except (TypeError, ValueError):
            continue
        if q <= 0:
            continue
        out.setdefault(kid, []).append({"componente_producto_id": cid, "cantidad": q})
    return out


def _kit_cantidad_armable(stock_by_id: dict[str, int], items: list[dict[str, Any]]) -> int:
    """Cantidad máxima de kits armables según stock de cada componente."""
    mins: list[int] = []
    for it in items:
        cid = it["componente_producto_id"]
        need = float(it["cantidad"])
        st = int(stock_by_id.get(cid, 0))
        if need <= 0:
            continue
        mins.append(max(0, int(st / need)))
    return min(mins) if mins else 0


def _inv_enrich_compat_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    if "compatibilidad" not in df.columns:
        df["compatibilidad"] = pd.NA
        df["vehiculos_compat"] = ""
        df["años_compat"] = ""
        return df
    vc: list[str] = []
    va: list[str] = []
    vs: list[str] = []
    for _, row in df.iterrows():
        d = _inv_compat_as_dict(row.get("compatibilidad"))
        vc.append(_inv_compat_marcas_str(d))
        va.append(_inv_compat_anos_str(d))
        vs.append(_inv_compat_seriales_motor_resumen(d))
    df["vehiculos_compat"] = vc
    df["años_compat"] = va
    df["seriales_motor"] = vs
    return df


def _fetch_productos_inventario_df(sb: Client) -> pd.DataFrame:
    cols_full = (
        "id,codigo,sku_oem,descripcion,marca_producto,condicion,ubicacion,compatibilidad,imagen_url,"
        "stock_actual,stock_minimo,costo_usd,precio_v_usd,precio_v_bs_ref,costo_bs_ref,activo,categoria_id,es_compuesto"
    )
    cols_base = (
        "id,codigo,descripcion,stock_actual,stock_minimo,costo_usd,precio_v_usd,"
        "precio_v_bs_ref,costo_bs_ref,activo,categoria_id"
    )
    cols_min = "id,codigo,descripcion,stock_actual,stock_minimo,costo_usd,precio_v_usd,activo,categoria_id"
    page = 1000

    def _paged(cols: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        offset = 0
        while True:
            r = (
                sb.table("productos")
                .select(cols)
                .order("descripcion")
                .range(offset, offset + page - 1)
                .execute()
            )
            batch = r.data or []
            rows.extend(batch)
            if len(batch) < page:
                break
            offset += page
            if offset > 200_000:
                break
        return rows

    try:
        data = _paged(cols_full)
    except Exception:
        try:
            data = _paged(cols_base)
        except Exception:
            data = _paged(cols_min)
    return pd.DataFrame(data)


def _normalize_productos_inventario_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Si no hay filas, Supabase devuelve [] y pandas crea un DataFrame **sin columnas**.
    Normalizamos al esquema esperado por el editor y los reportes.
    """
    if df.empty and len(df.columns) == 0:
        return pd.DataFrame(columns=_INV_PRODUCTO_COLS)
    out = df.copy()
    for c in _INV_PRODUCTO_COLS:
        if c not in out.columns:
            out[c] = pd.NA
    if "condicion" in out.columns:
        out["condicion"] = out["condicion"].apply(
            lambda x: (
                "Nuevo"
                if x is None or (isinstance(x, float) and pd.isna(x)) or str(x).strip() not in ("Nuevo", "Usado")
                else str(x).strip()
            )
        )
    return out


def _categoria_maps_from_rows(
    cat_rows: list[dict[str, Any]],
) -> tuple[dict[str, str], dict[str, str], list[str]]:
    """id->nombre, nombre->id, opciones selectbox ('' = sin categoría)."""
    id_to_n: dict[str, str] = {}
    n_to_id: dict[str, str] = {}
    for c in cat_rows:
        cid = str(c.get("id") or "").strip()
        nom = str(c.get("nombre") or "").strip()
        if cid and nom:
            id_to_n[cid] = nom
            n_to_id[nom] = cid
    opts = [""] + sorted(n_to_id.keys(), key=str.casefold)
    return id_to_n, n_to_id, opts


def _resolve_categoria_id_por_nombre(nombre: str, n_to_id: dict[str, str]) -> str | None:
    s = nombre.strip()
    if not s:
        return None
    if s in n_to_id:
        return n_to_id[s]
    sl = s.lower()
    for k, vid in n_to_id.items():
        if k.lower() == sl:
            return vid
    return None


def _inv_cat_display(celda: Any) -> str:
    if celda is None or (isinstance(celda, float) and pd.isna(celda)):
        return "(Sin categoría)"
    s = str(celda).strip()
    return s if s else "(Sin categoría)"


def _df_inventario_filtrado_impresion(
    df: pd.DataFrame,
    *,
    categorias_sel: list[str],
    costo_min: float,
    costo_max: float,
    precio_min: float,
    precio_max: float,
    solo_activos: bool,
) -> pd.DataFrame:
    out = df.copy()
    if solo_activos and "activo" in out.columns:
        out = out[out["activo"] == True]  # noqa: E712
    if categorias_sel:
        disp = out["categoria"].map(_inv_cat_display)
        out = out[disp.isin(categorias_sel)]
    c_usd = pd.to_numeric(out["costo_usd"], errors="coerce").fillna(0.0)
    p_usd = pd.to_numeric(out["precio_v_usd"], errors="coerce").fillna(0.0)
    m = pd.Series(True, index=out.index)
    if costo_min > 0:
        m &= c_usd >= float(costo_min)
    if costo_max > 0:
        m &= c_usd <= float(costo_max)
    if precio_min > 0:
        m &= p_usd >= float(precio_min)
    if precio_max > 0:
        m &= p_usd <= float(precio_max)
    return out.loc[m]


def _df_inventario_orden_impresion(df: pd.DataFrame, orden_key: str, *, agrupar_categoria: bool) -> pd.DataFrame:
    keys: list[tuple[str, bool]] = []
    if orden_key == "codigo":
        keys = [("codigo", True)]
    elif orden_key == "costo_asc":
        keys = [("costo_usd", True)]
    elif orden_key == "costo_desc":
        keys = [("costo_usd", False)]
    elif orden_key == "precio_asc":
        keys = [("precio_v_usd", True)]
    elif orden_key == "precio_desc":
        keys = [("precio_v_usd", False)]
    else:
        keys = [("descripcion", True)]
    if agrupar_categoria:
        keys = [("categoria", True)] + keys
    work = df.copy()
    for col, _asc in keys:
        if col not in work.columns:
            continue
        if work[col].dtype == object or str(work[col].dtype) == "object":
            work[col] = work[col].fillna("").astype(str)
    col_names = [k[0] for k in keys if k[0] in work.columns]
    ascending = [k[1] for k in keys if k[0] in work.columns]
    if not col_names:
        return work
    return work.sort_values(by=col_names, ascending=ascending, kind="mergesort")


# --- Reporte inventario: columnas (misma lógica HTML / PDF / Excel) ---
INV_REP_COL_META: list[tuple[str, str]] = [
    ("codigo", "Código"),
    ("sku_oem", "OEM"),
    ("descripcion", "Descripción"),
    ("marca_producto", "Marca rep."),
    ("condicion", "Cond."),
    ("_veh_rep", "Marcas carro"),
    ("_anos_rep", "Años"),
    ("categoria_display", "Categoría"),
    ("stock_actual", "Stock"),
    ("stock_minimo", "Stock mín."),
    ("costo_usd", "Costo USD"),
    ("precio_v_usd", "Precio venta USD"),
    ("precio_v_bs_ref", "Precio venta Bs (ref.)"),
    ("costo_bs_ref", "Costo Bs (ref.)"),
    ("_pv_usdt_ref", "Precio venta USDT (ref.)"),
    ("_cu_usdt_ref", "Costo USDT (ref.)"),
    ("ubicacion", "Ubicación"),
    ("activo", "Activo"),
]

# Multiselect “personalizado”: las USDT se activan con el checkbox de moneda, no hace falta listarlas acá.
INV_REP_COL_META_DICT: dict[str, str] = dict(INV_REP_COL_META)
INV_REP_COL_KEYS_PERSONALIZADO: tuple[str, ...] = tuple(
    k for k, _ in INV_REP_COL_META if k not in ("_pv_usdt_ref", "_cu_usdt_ref")
)

INV_REP_PDF_ABBR: dict[str, str] = {
    "codigo": "Cód.",
    "sku_oem": "OEM",
    "descripcion": "Descripción",
    "marca_producto": "M. rep.",
    "condicion": "Cond.",
    "_veh_rep": "Vehíc.",
    "_anos_rep": "Años",
    "categoria_display": "Cat.",
    "stock_actual": "St",
    "stock_minimo": "Mín",
    "costo_usd": "C.U.",
    "precio_v_usd": "P.V.",
    "precio_v_bs_ref": "P.Bs",
    "costo_bs_ref": "C.Bs",
    "_pv_usdt_ref": "P.USDT",
    "_cu_usdt_ref": "C.USDT",
    "ubicacion": "Ubic.",
    "activo": "Act.",
}

INV_REP_KEY_TO_EXCEL: dict[str, str] = {
    "codigo": "Código",
    "sku_oem": "OEM",
    "descripcion": "Descripción",
    "marca_producto": "Marca repuesto",
    "condicion": "Condición",
    "_veh_rep": "Marcas carro",
    "_anos_rep": "Años",
    "categoria_display": "Categoría",
    "stock_actual": "Stock",
    "stock_minimo": "Stock mín.",
    "costo_usd": "Costo USD",
    "precio_v_usd": "Precio venta USD",
    "precio_v_bs_ref": "Precio venta Bs (ref.)",
    "costo_bs_ref": "Costo Bs (ref.)",
    "_pv_usdt_ref": "Precio venta USDT (ref.)",
    "_cu_usdt_ref": "Costo USDT (ref.)",
    "ubicacion": "Ubicación",
    "activo": "Activo",
}

INV_REP_NUMERIC_KEYS: frozenset[str] = frozenset(
    {
        "stock_actual",
        "stock_minimo",
        "costo_usd",
        "precio_v_usd",
        "precio_v_bs_ref",
        "costo_bs_ref",
        "_pv_usdt_ref",
        "_cu_usdt_ref",
    }
)

INV_REP_SYNTH_USDT_KEYS: frozenset[str] = frozenset({"_pv_usdt_ref", "_cu_usdt_ref"})

INV_REP_META_KEYS_SIN_SYNTH: frozenset[str] = frozenset(k for k, _ in INV_REP_COL_META if k not in INV_REP_SYNTH_USDT_KEYS)

# Columnas “de detalle”: por defecto fuera del reporte; el usuario las activa con checkboxes (presets).
INV_REP_DETAIL_OPT_KEYS: frozenset[str] = frozenset(
    {"marca_producto", "condicion", "_veh_rep", "_anos_rep", "stock_minimo", "ubicacion"}
)

# Pesos relativos para repartir ancho (HTML colgroup + PDF). Más alto = más ancho.
_INV_REP_COL_W: dict[str, float] = {
    "codigo": 1.7,
    "sku_oem": 1.7,
    "descripcion": 10.0,
    "marca_producto": 0.9,
    "condicion": 0.55,
    "_veh_rep": 1.05,
    "_anos_rep": 0.75,
    "categoria_display": 0.85,
    "stock_actual": 0.62,
    "stock_minimo": 0.55,
    "costo_usd": 0.92,
    "precio_v_usd": 0.92,
    "precio_v_bs_ref": 0.88,
    "costo_bs_ref": 0.88,
    "_pv_usdt_ref": 0.92,
    "_cu_usdt_ref": 0.92,
    "ubicacion": 0.85,
    "activo": 0.48,
}


def _inv_rep_col_width_fracs(keys: list[str]) -> list[float]:
    wts = [_INV_REP_COL_W.get(k, 1.0) for k in keys]
    s = sum(wts) or 1.0
    fr = [w / s for w in wts]
    if "descripcion" in keys:
        i = keys.index("descripcion")
        # 30% menos ancho para Descripcion (vs min_f=0.30 anterior).
        min_f = 0.21
        if fr[i] < min_f:
            rest = 1.0 - min_f
            sum_other = sum(fr[j] for j in range(len(fr)) if j != i)
            if sum_other <= 1e-9:
                n = len(fr)
                return [1.0 / n] * n
            for j in range(len(fr)):
                if j == i:
                    fr[j] = min_f
                else:
                    fr[j] = (fr[j] / sum_other) * rest
    return fr


INV_REP_PRESET_INTERNO_CORE: frozenset[str] = frozenset(
    {
        "codigo",
        "sku_oem",
        "descripcion",
        "categoria_display",
        "stock_actual",
        "costo_usd",
        "precio_v_usd",
        "precio_v_bs_ref",
        "costo_bs_ref",
        "activo",
    }
)

INV_REP_PRESET_COLS: dict[str, frozenset[str] | None] = {
    "interno": INV_REP_PRESET_INTERNO_CORE,
    "lista_cliente": frozenset({"codigo", "descripcion", "categoria_display", "precio_v_usd"}),
    "analisis_precios": frozenset(
        {
            "codigo",
            "sku_oem",
            "descripcion",
            "categoria_display",
            "stock_actual",
            "costo_usd",
            "precio_v_usd",
        }
    ),
}


def _inv_rep_merge_template_keys(column_keys: frozenset[str] | None) -> frozenset[str]:
    """`None` = personalizado vacío → todas las columnas físicas (sin USDT sintéticas)."""
    if column_keys is None:
        return frozenset(INV_REP_META_KEYS_SIN_SYNTH)
    return column_keys


def _inv_rep_extend_currency_columns(
    keys: frozenset[str],
    *,
    show_bs: bool,
    show_usdt: bool,
) -> frozenset[str]:
    """Añade columnas de Bs ref. / USDT ref. solo si el usuario las pidió (p. ej. lista cliente sin Bs por defecto)."""
    k = set(keys)
    if show_bs:
        k.add("precio_v_bs_ref")
        if "costo_usd" in k:
            k.add("costo_bs_ref")
    if show_usdt:
        k.update(INV_REP_SYNTH_USDT_KEYS)
    return frozenset(k)


def _inv_rep_apply_currency_prefs(
    keys: frozenset[str],
    *,
    show_usd: bool,
    show_bs: bool,
    show_usdt: bool,
) -> frozenset[str]:
    drop: set[str] = set()
    if not show_usd:
        drop.update({"costo_usd", "precio_v_usd"})
    if not show_bs:
        drop.update({"precio_v_bs_ref", "costo_bs_ref"})
    if not show_usdt:
        drop.update(INV_REP_SYNTH_USDT_KEYS)
    return frozenset(k for k in keys if k not in drop)


def _inv_rep_extend_detail_columns(
    keys: frozenset[str],
    *,
    marca: bool,
    cond: bool,
    veh: bool,
    anos: bool,
    stock_min: bool,
    ubi: bool,
) -> frozenset[str]:
    """Añade columnas de detalle solo si el usuario las pidió (presets interno / lista / análisis)."""
    k = set(keys)
    if marca:
        k.add("marca_producto")
    if cond:
        k.add("condicion")
    if veh:
        k.add("_veh_rep")
    if anos:
        k.add("_anos_rep")
    if stock_min:
        k.add("stock_minimo")
    if ubi:
        k.add("ubicacion")
    return frozenset(k)


def _inv_format_usdt_ref_cell(val_usd: Any, tasa_usdt: float | None) -> str:
    """USDT ref. = USD × `tasa_usdt` (USDT por 1 USD en sistema), igual que en el resto de la app."""
    if tasa_usdt is None or tasa_usdt <= 0 or not _inv_is_finite_num(val_usd):
        return ""
    return _rep_fmt_precio_entero(float(val_usd) * float(tasa_usdt))


def _inv_rep_cols_for_export(work: pd.DataFrame, column_keys: frozenset[str] | None) -> list[tuple[str, str]]:
    """Orden fijo de metadatos; respeta columnas que existan en `work`."""
    out: list[tuple[str, str]] = []
    for key, lab in INV_REP_COL_META:
        if column_keys is not None and key not in column_keys:
            continue
        if key in INV_REP_SYNTH_USDT_KEYS:
            out.append((key, lab))
            continue
        if key in ("precio_v_bs_ref", "costo_bs_ref"):
            if key not in work.columns:
                continue
        elif key in ("_veh_rep", "_anos_rep"):
            if "compatibilidad" not in work.columns:
                continue
        elif key == "categoria_display":
            if "categoria" not in work.columns:
                continue
        elif key not in work.columns:
            continue
        out.append((key, lab))
    return out


def _inv_rep_prepare_work_df(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    if "categoria" in work.columns:
        work["categoria_display"] = work["categoria"].map(_inv_cat_display)
    else:
        work["categoria_display"] = ""
    if "compatibilidad" in work.columns:
        work["_veh_rep"] = work["compatibilidad"].map(
            lambda x: _inv_compat_marcas_str(_inv_compat_as_dict(x))
        )
        work["_anos_rep"] = work["compatibilidad"].map(
            lambda x: _inv_compat_anos_str(_inv_compat_as_dict(x))
        )
    else:
        work["_veh_rep"] = ""
        work["_anos_rep"] = ""
    return work


def _html_inventario_listado(
    df: pd.DataFrame,
    t: dict[str, Any] | None,
    *,
    agrupar_categoria: bool,
    subtitulo_filtros: str,
    column_keys: frozenset[str] | None = None,
) -> str:
    tz = ZoneInfo("America/Caracas")
    fecha = datetime.now(tz).strftime("%d/%m/%Y %H:%M")
    work = _inv_rep_prepare_work_df(df)
    cols_print = _inv_rep_cols_for_export(work, column_keys)
    if not cols_print:
        return (
            "<!DOCTYPE html><html lang=\"es\"><head><meta charset=\"utf-8\"/><title>Inventario</title></head>"
            "<body><p>No hay columnas para mostrar con la plantilla elegida (revisá datos o elegí otras columnas).</p></body></html>"
        )

    _t_bs_rep = float(t["tasa_bs"]) if (t and _nf(t.get("tasa_bs")) is not None) else None
    _t_usdt_rep = float(t["tasa_usdt"]) if (t and _nf(t.get("tasa_usdt")) is not None) else None

    _k_list = [k for k, _ in cols_print]
    _fracs = _inv_rep_col_width_fracs(_k_list)
    _code_chars = 0
    _oem_chars = 0
    if "codigo" in work.columns:
        _code_chars = int(work["codigo"].fillna("").astype(str).map(len).max() or 0)
    if "sku_oem" in work.columns:
        _oem_chars = int(work["sku_oem"].fillna("").astype(str).map(len).max() or 0)
    # Reserva ancho por caracteres para mostrar Codigo/OEM completos en horizontal.
    _code_oem_ch = min(52, max(14, max(_code_chars, _oem_chars) + 4))
    _col_parts: list[str] = []
    for _fk, _f in zip(_k_list, _fracs):
        _cls_parts: list[str] = []
        if _fk == "descripcion":
            _cls_parts.append("col-desc")
        if _fk == "codigo":
            _cls_parts.append("col-code")
        if _fk == "sku_oem":
            _cls_parts.append("col-oem")
        if _fk == "categoria_display":
            _cls_parts.append("col-cat")
        _cls = f' class="{" ".join(_cls_parts)}"' if _cls_parts else ""
        _col_parts.append(f'<col{_cls} style="width:{100 * _f:.2f}%" />')
    _colgroup = "".join(_col_parts)
    ths_parts: list[str] = []
    for _k, _lab in cols_print:
        _th_classes: list[str] = []
        if _k == "codigo":
            _th_classes.append("code")
        if _k == "sku_oem":
            _th_classes.append("oem")
        if _k == "categoria_display":
            _th_classes.append("cat")
        if _k in INV_REP_NUMERIC_KEYS:
            _th_classes.append("num")
        _cls = f' class="{" ".join(_th_classes)}"' if _th_classes else ""
        _lab_short = INV_REP_PDF_ABBR.get(_k, _lab)
        ths_parts.append(f"<th{_cls} title=\"{html.escape(_lab)}\">{html.escape(_lab_short)}</th>")
    ths = "".join(ths_parts)
    body_rows: list[str] = []
    current_cat: str | None = None
    for _, row in work.iterrows():
        if agrupar_categoria:
            cd = str(row["categoria_display"])
            if cd != current_cat:
                current_cat = cd
                body_rows.append(
                    f'<tr class="catgrp"><td colspan="{len(cols_print)}">{html.escape(cd)}</td></tr>'
                )
        tds: list[str] = []
        for key, _lab in cols_print:
            val = row.get(key)
            td_cls = ""
            if key in ("stock_actual", "stock_minimo"):
                try:
                    cell = f"{_inv_stock_int(val):,d}"
                    td_cls = "num"
                except (TypeError, ValueError):
                    cell = html.escape("" if val is None else str(val))
            elif key in ("costo_usd", "precio_v_usd"):
                cell = _rep_fmt_precio_entero(val)
                td_cls = "num"
            elif key == "precio_v_bs_ref":
                cell = _inv_format_bs_ref_cell(val, row.get("precio_v_usd"), _t_bs_rep)
                td_cls = "num"
            elif key == "costo_bs_ref":
                cell = _inv_format_bs_ref_cell(val, row.get("costo_usd"), _t_bs_rep)
                td_cls = "num"
            elif key == "_pv_usdt_ref":
                cell = _inv_format_usdt_ref_cell(row.get("precio_v_usd"), _t_usdt_rep)
                td_cls = "num"
            elif key == "_cu_usdt_ref":
                cell = _inv_format_usdt_ref_cell(row.get("costo_usd"), _t_usdt_rep)
                td_cls = "num"
            elif key == "activo":
                cell = "Sí" if bool(val) else "No"
            else:
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    cell = ""
                else:
                    cell = html.escape(str(val))
                if key == "descripcion":
                    td_cls = "desc"
                elif key == "codigo":
                    td_cls = "code"
                elif key == "sku_oem":
                    td_cls = "oem"
                elif key == "categoria_display":
                    td_cls = "cat"
            _cls_attr = f' class="{td_cls}"' if td_cls else ""
            _title_attr = ""
            if key in {"codigo", "sku_oem"}:
                _title_attr = f' title="{html.escape("" if val is None else str(val))}"'
            tds.append(f"<td{_cls_attr}{_title_attr}>{cell}</td>")
        body_rows.append("<tr>" + "".join(tds) + "</tr>")

    n = len(work)
    tasa_note = _inv_rep_tasas_footer_html(cols_print, t)

    filt_html = f"<p class=\"sub\">{html.escape(subtitulo_filtros)}</p>" if subtitulo_filtros.strip() else ""

    _logo_uri = _brand_logo_data_uri()
    _logo_block = (
        f'<div class="logo-wrap"><img class="logo" src="{_logo_uri}" alt="Movi Motors"/></div>'
        if _logo_uri
        else '<div class="logo-wrap logo-missing">Movi Motor\'s Importadora</div>'
    )

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8"/>
<title>Inventario — Movi Motors</title>
<style>
  @page {{
    size: A4 portrait;
    margin: 14mm 16mm 16mm 16mm;
  }}
  * {{ box-sizing: border-box; }}
  body {{
    font-family: Segoe UI, Roboto, Arial, sans-serif;
    margin: 0;
    padding: 0.75rem 0.5rem;
    width: 100%;
    max-width: none;
    color: #111;
  }}
  .logo-wrap {{
    text-align: center;
    margin: 0 0 0.6rem 0;
  }}
  .logo {{
    max-height: 22mm;
    max-width: 52mm;
    width: auto;
    height: auto;
    object-fit: contain;
  }}
  .logo-missing {{
    font-weight: 800;
    font-size: 1rem;
    color: #5c2d91;
    letter-spacing: 0.02em;
  }}
  h1 {{ font-size: 1.1rem; margin: 0 0 0.35rem 0; text-align: center; color: #2a1f45; }}
  .meta {{ color: #444; font-size: 0.82rem; margin-bottom: 0.65rem; text-align: center; }}
  .sub {{ font-size: 0.78rem; color: #333; margin: 0.3rem 0; text-align: center; }}
  table.inv-grid {{ border-collapse: collapse; width: 100%; min-width: 100%; font-size: 0.72rem; table-layout: fixed; }}
  col.col-desc {{ min-width: 8.4rem; }}
  col.col-code, col.col-oem {{ min-width: {_code_oem_ch}ch; }}
  col.col-cat {{ min-width: 6.8rem; }}
  th, td {{ border: 1px solid #bbb; padding: 0.35rem 0.45rem; text-align: left; vertical-align: top;
    word-wrap: break-word; overflow-wrap: break-word; hyphens: auto; }}
  th {{ background: #2a1f45; color: #fff; font-weight: 600; font-size: 0.68rem; line-height: 1.1;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis; word-wrap: normal; overflow-wrap: normal; }}
  th.num {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  td.desc {{ white-space: normal; font-size: 0.8rem; line-height: 1.4; }}
  th.code, td.code {{ white-space: nowrap; word-break: keep-all; overflow-wrap: normal; }}
  th.oem, td.oem {{ white-space: nowrap; word-break: keep-all; overflow-wrap: normal; }}
  th.cat, td.cat {{ white-space: nowrap; word-break: keep-all; overflow-wrap: normal; }}
  tr.catgrp td {{ background: #fff3e0; font-weight: 700; color: #e65100; border-color: #ffcc80;
    font-family: Segoe UI, Roboto, Arial, sans-serif; font-style: normal; }}
  tr:nth-child(even) td {{ background: #fafafa; }}
  .foot {{ margin-top: 0.85rem; font-size: 0.75rem; color: #555; text-align: center; }}
  .print-actions {{ margin-top: 0.75rem; text-align: center; }}
  @media print {{
    body {{ padding: 0; max-width: 210mm; margin-left: auto; margin-right: auto; }}
    .print-actions {{ display: none !important; }}
    tr.catgrp td {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    th {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    table {{ font-size: 0.68rem; }}
  }}
</style>
</head>
<body>
  {_logo_block}
  <h1>Listado de inventario</h1>
  <div class="meta">Generado: {html.escape(fecha)} (America/Caracas) · <strong>{n}</strong> ítem(s)</div>
  {tasa_note}
  {filt_html}
  <table class="inv-grid">
    <colgroup>{_colgroup}</colgroup>
    <thead><tr>{ths}</tr></thead>
    <tbody>
      {''.join(body_rows)}
    </tbody>
  </table>
  <p class="foot">Precios y costos en USD son los maestros del sistema. Columnas Bs (ref.) dependen de la tasa guardada al momento del reporte.</p>
  <div class="print-actions">
    <script>function imprimir(){{ window.print(); }}</script>
    <button type="button" onclick="imprimir()" style="padding:0.5rem 1.2rem;font-size:1rem;cursor:pointer;background:#2a1f45;color:#fff;border:none;border-radius:6px;">Imprimir</button>
  </div>
</body>
</html>"""


def _export_cell_txt(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def _inv_is_finite_num(val: Any) -> bool:
    if val is None:
        return False
    try:
        if isinstance(val, float) and pd.isna(val):
            return False
        v = float(val)
        return math.isfinite(v)
    except (TypeError, ValueError):
        return False


def _rep_fmt_precio_entero(val: Any) -> str:
    """Montos / precios en reportes: entero con separador de miles, sin decimales."""
    if not _inv_is_finite_num(val):
        return ""
    return f"{int(round(float(val))):,d}"


def _rep_series_montos_enteros(s: pd.Series) -> pd.Series:
    """Serie numérica → enteros redondeados (nullable) para tablas Excel/CSV de reportes."""
    num = pd.to_numeric(s, errors="coerce")
    return num.round(0).astype("Int64")


def _inv_rep_tasas_footer_html(cols_print: list[tuple[str, str]], t: dict[str, Any] | None) -> str:
    bits: list[str] = []
    keys_in = {k for k, _ in cols_print}
    if keys_in & {"precio_v_bs_ref", "costo_bs_ref"} and t and _nf(t.get("tasa_bs")) is not None:
        bits.append(f"Ref. Bs/USD (precios en Bs): <strong>{float(t['tasa_bs']):,.2f}</strong> Bs por 1 USD")
    if keys_in & INV_REP_SYNTH_USDT_KEYS and t and _nf(t.get("tasa_usdt")) is not None:
        bits.append(f"Ref. USDT/USD: <strong>{float(t['tasa_usdt']):,.6f}</strong> USDT por 1 USD")
    if not bits:
        return ""
    return '<p class="sub">' + " · ".join(bits) + "</p>"


def _inv_format_bs_ref_cell(val_bs: Any, val_usd: Any, tasa_bs: float | None) -> str:
    """Bs ref.: usa columna BD si es número; si no, USD × tasa del reporte (evita nan en PDF/HTML)."""
    if _inv_is_finite_num(val_bs):
        return _rep_fmt_precio_entero(val_bs)
    if tasa_bs is not None and tasa_bs > 0 and _inv_is_finite_num(val_usd):
        return _rep_fmt_precio_entero(float(val_usd) * float(tasa_bs))
    return ""


def _df_inventario_export_flat(
    df: pd.DataFrame,
    t: dict[str, Any] | None = None,
    column_keys: frozenset[str] | None = None,
) -> pd.DataFrame:
    work = _inv_rep_prepare_work_df(df.copy())
    _t_bs_x = _nf(t.get("tasa_bs")) if t else None
    base: dict[str, Any] = {
        "Código": work["codigo"].map(_export_cell_txt),
        "Descripción": work["descripcion"].map(_export_cell_txt),
        "Categoría": work["categoria"].map(_inv_cat_display),
        "Stock": work["stock_actual"].map(_inv_stock_int),
        "Stock mín.": work["stock_minimo"].map(_inv_stock_int),
        "Costo USD": _rep_series_montos_enteros(work["costo_usd"]),
        "Precio venta USD": _rep_series_montos_enteros(work["precio_v_usd"]),
    }
    if "sku_oem" in work.columns:
        base["OEM"] = work["sku_oem"].map(_export_cell_txt)
    if "marca_producto" in work.columns:
        base["Marca repuesto"] = work["marca_producto"].map(_export_cell_txt)
    if "condicion" in work.columns:
        base["Condición"] = work["condicion"].map(_export_cell_txt)
    if "compatibilidad" in work.columns:
        base["Marcas carro"] = work["compatibilidad"].map(
            lambda x: _inv_compat_marcas_str(_inv_compat_as_dict(x))
        )
        base["Años"] = work["compatibilidad"].map(
            lambda x: _inv_compat_anos_str(_inv_compat_as_dict(x))
        )
    if "ubicacion" in work.columns:
        base["Ubicación"] = work["ubicacion"].map(_export_cell_txt)
    out = pd.DataFrame(base)
    if "precio_v_bs_ref" in work.columns:
        _pv = pd.to_numeric(work["precio_v_usd"], errors="coerce")
        _pbs = pd.to_numeric(work["precio_v_bs_ref"], errors="coerce")
        if _t_bs_x is not None and _t_bs_x > 0:
            _pbs = _pbs.fillna(_pv * float(_t_bs_x))
        out["Precio venta Bs (ref.)"] = _rep_series_montos_enteros(_pbs)
    if "costo_bs_ref" in work.columns:
        _cv = pd.to_numeric(work["costo_usd"], errors="coerce")
        _cbs = pd.to_numeric(work["costo_bs_ref"], errors="coerce")
        if _t_bs_x is not None and _t_bs_x > 0:
            _cbs = _cbs.fillna(_cv * float(_t_bs_x))
        out["Costo Bs (ref.)"] = _rep_series_montos_enteros(_cbs)
    _t_ut_x = _nf(t.get("tasa_usdt")) if t else None
    _pvu_num = pd.to_numeric(work["precio_v_usd"], errors="coerce")
    _cvu_num = pd.to_numeric(work["costo_usd"], errors="coerce")
    if _t_ut_x is not None and _t_ut_x > 0:
        out["Precio venta USDT (ref.)"] = _rep_series_montos_enteros(_pvu_num * float(_t_ut_x))
        out["Costo USDT (ref.)"] = _rep_series_montos_enteros(_cvu_num * float(_t_ut_x))
    else:
        out["Precio venta USDT (ref.)"] = pd.Series([pd.NA] * len(work), index=work.index, dtype="Float64")
        out["Costo USDT (ref.)"] = pd.Series([pd.NA] * len(work), index=work.index, dtype="Float64")
    if "activo" in work.columns:
        out["Activo"] = work["activo"].astype(bool)
    keys_order = [k for k, _ in _inv_rep_cols_for_export(work, column_keys)]
    excel_pick = [INV_REP_KEY_TO_EXCEL[k] for k in keys_order if k in INV_REP_KEY_TO_EXCEL]
    cols_ok = [c for c in excel_pick if c in out.columns]
    if cols_ok:
        out = out[cols_ok]
    return out


def _xlsx_inventario_bytes(df_flat: pd.DataFrame) -> bytes:
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_flat.to_excel(writer, index=False, sheet_name="Inventario")
        ws = writer.sheets["Inventario"]
        _code_oem_w: float | None = None
        _cols_l = [str(c).lower() for c in df_flat.columns]
        if any(c in {"código", "codigo"} for c in _cols_l) or "oem" in _cols_l:
            _max_len = 0
            for c in df_flat.columns:
                lc = str(c).lower()
                if lc not in {"código", "codigo", "oem"}:
                    continue
                _lens = df_flat[c].fillna("").astype(str).map(len)
                _m = max(int(_lens.max()) if len(_lens) > 0 else 0, len(str(c)))
                _max_len = max(_max_len, _m)
            _code_oem_w = float(min(50, max(12, _max_len + 2)))
        for i, col in enumerate(df_flat.columns, start=1):
            lens = df_flat[col].astype(str).map(len)
            m = max(int(lens.max()) if len(lens) > 0 else 0, len(str(col)))
            base = max(10, m + 2)
            lc = str(col).lower()
            if lc in {"código", "codigo", "oem"} and _code_oem_w is not None:
                base = max(base, _code_oem_w)
            if lc == "categoría" or lc == "categoria":
                base = max(base, 22)
            if "escripci" in str(col).lower() or str(col).lower() == "descripción":
                # Descripcion mas compacta (vs min ancho 44 anterior).
                base = max(base, 32)
            ws.column_dimensions[get_column_letter(i)].width = float(min(50, base))
    return buf.getvalue()


def _reporte_tabla_a_excel(df: pd.DataFrame, *, nombre_hoja: str = "Datos") -> bytes:
    """Excel genérico para reportes (hoja única). Requiere openpyxl."""
    hn = (nombre_hoja or "Datos")[:31]
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=hn)
    return buf.getvalue()


def _reporte_tabla_a_csv(df: pd.DataFrame) -> bytes:
    """CSV con BOM para que Excel en Windows abra bien las tildes."""
    return df.to_csv(index=False).encode("utf-8-sig")


def _md_celda_ia(x: Any, *, max_len: int = 320) -> str:
    """Texto seguro para tablas Markdown (evita `|` y saltos de línea que rompen la tabla)."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        s = ""
    else:
        s = str(x).replace("\r\n", " ").replace("\n", " ").replace("|", "·").strip()
    if len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s or "—"


def inventario_activos_markdown_ia(sb: Client, _t: dict[str, Any] | None) -> str:
    """
    Documento Markdown: solo productos **activos**, con stock y precio de venta **solo en USD**.
    Pensado para subirlo a un asistente / IA que consulte disponibilidad.
    """
    try:
        cats_list = (sb.table("categorias").select("id,nombre").order("nombre").execute().data or [])
    except Exception:
        cats_list = []
    id_to_n, _, _ = _categoria_maps_from_rows(cats_list)
    df = _inv_enrich_compat_columns(_normalize_productos_inventario_df(_fetch_productos_inventario_df(sb)))
    if df.empty:
        now_vz = datetime.now(ZoneInfo("America/Caracas"))
        return "\n".join(
            [
                "# Inventario Movi Motors — consulta de disponibilidad y precios",
                "",
                f"- **Generado (Caracas):** {now_vz.strftime('%Y-%m-%d %H:%M')} ({now_vz.tzname()})",
                "",
                "## Productos activos",
                "",
                "*No hay productos en la base.*",
                "",
            ]
        )
    if "categoria_id" in df.columns:
        df["categoria"] = df["categoria_id"].apply(
            lambda x: id_to_n.get(str(x).strip(), "")
            if x is not None and not (isinstance(x, float) and pd.isna(x)) and str(x).strip()
            else ""
        )
    else:
        df["categoria"] = ""

    if "activo" not in df.columns:
        activos = df.iloc[0:0]
    else:
        activos = df.loc[df["activo"] == True].copy()  # noqa: E712

    now_vz = datetime.now(ZoneInfo("America/Caracas"))

    lines: list[str] = [
        "# Inventario Movi Motors — consulta de disponibilidad y precios",
        "",
        f"- **Generado (Caracas):** {now_vz.strftime('%Y-%m-%d %H:%M')} ({now_vz.tzname()})",
        "- **Alcance:** solo productos marcados como **activos** en el ERP.",
        "- **Moneda de precios:** solo **USD** (dólares), según `precio_v_usd` en el maestro.",
        "",
        "## Cómo usar este archivo",
        "",
        "Cada fila es un producto. **Stock** = unidades disponibles según el sistema (ventas bajan, compras suben). "
        "**Precio venta (USD)** es el valor de venta cargado en el ERP.",
        "",
    ]

    n = len(activos)
    if n == 0:
        lines.extend(["## Productos activos", "", "*No hay productos activos en la base.*", ""])
        return "\n".join(lines)

    con_stock = sum(1 for _, r in activos.iterrows() if _inv_stock_int(r.get("stock_actual")) > 0)
    sin_stock = n - con_stock
    lines.extend(
        [
            "## Resumen",
            "",
            f"- **Productos activos:** {n}",
            f"- **Con stock > 0:** {con_stock}",
            f"- **Sin unidades (stock 0):** {sin_stock}",
            "",
            "## Detalle (tabla)",
            "",
            "| Código | OEM | Descripción | Categoría | Marca rep. | Cond. | Stock | Mín. | ¿Bajo mín.? | Precio venta USD | Ubicación | Vehículos compat. | Años | Kit | Disponibilidad |",
            "| --- | --- | --- | --- | --- | --- | ---: | ---: | --- | ---: | --- | --- | --- | --- | --- |",
        ]
    )

    activos = activos.assign(
        _cs=activos["categoria"].fillna("").map(lambda x: str(x).casefold()),
        _ds=activos["descripcion"].fillna("").map(lambda x: str(x).casefold()),
    )
    activos = activos.sort_values(["_cs", "_ds"]).drop(columns=["_cs", "_ds"])

    for _, r in activos.iterrows():
        st_i = _inv_stock_int(r.get("stock_actual"))
        smin = _inv_stock_int(r.get("stock_minimo"))
        bajo = "sí" if st_i <= smin and smin > 0 else "no"
        if st_i <= 0:
            disp = "Sin unidades"
        elif st_i <= smin and smin > 0:
            disp = f"En stock ({st_i}) — atención: en o bajo mínimo ({smin})"
        else:
            disp = f"En stock ({st_i} uds.)"
        try:
            pv_usd = float(r.get("precio_v_usd") or 0)
        except (TypeError, ValueError):
            pv_usd = 0.0
        es_kit = r.get("es_compuesto")
        kit_txt = "sí" if es_kit else "no"

        row_cells = [
            _md_celda_ia(r.get("codigo"), max_len=48),
            _md_celda_ia(r.get("sku_oem"), max_len=64),
            _md_celda_ia(r.get("descripcion"), max_len=280),
            _md_celda_ia(r.get("categoria"), max_len=64),
            _md_celda_ia(r.get("marca_producto"), max_len=48),
            _md_celda_ia(r.get("condicion"), max_len=12),
            str(st_i),
            str(smin),
            bajo,
            f"{pv_usd:,.2f}",
            _md_celda_ia(r.get("ubicacion"), max_len=120),
            _md_celda_ia(r.get("vehiculos_compat"), max_len=200),
            _md_celda_ia(r.get("años_compat"), max_len=80),
            kit_txt,
            _md_celda_ia(disp, max_len=120),
        ]
        lines.append("| " + " | ".join(row_cells) + " |")

    lines.append("")
    return "\n".join(lines)


def _pdf_inventario_col_widths_for_keys(keys: list[str], total_w: float) -> list[float]:
    """Ancho PDF por columnas; evita celdas tan angostas que ReportLab falle (availWidth negativo con Paragraph)."""
    if not keys or total_w <= 0:
        return []

    n = len(keys)
    fr = _inv_rep_col_width_fracs(keys)
    raw_ws = [total_w * float(f) for f in fr]

    # ReportLab usa paddings internos por celda. Si una columna queda demasiado angosta,
    # Paragraph puede fallar con `negative availWidth`. Mantenemos mínimos por columna,
    # dando prioridad a `codigo` para que se lea horizontal.
    base_min = max(16.0, total_w * 0.03)
    min_w_by_key: list[float] = []
    for k in keys:
        if k == "codigo":
            min_w_by_key.append(max(42.0, base_min * 2.35))
        elif k == "sku_oem":
            min_w_by_key.append(max(42.0, base_min * 2.35))
        elif k == "categoria_display":
            min_w_by_key.append(max(20.0, base_min * 1.2))
        else:
            min_w_by_key.append(base_min)

    mins_sum = sum(min_w_by_key)
    if mins_sum <= 0:
        return [total_w / n] * n

    if mins_sum >= total_w:
        # Espacio insuficiente: escalamos los mínimos proporcionalmente.
        scale = total_w / mins_sum
        ws = [m * scale for m in min_w_by_key]
        drift = total_w - sum(ws)
        ws[-1] += drift
        return [max(0.0, float(w)) for w in ws]

    # 1) Base: aplicar mínimo a cada columna.
    ws = [max(w, min_w_by_key[i]) for i, w in enumerate(raw_ws)]
    base_sum = sum(ws)
    remaining = total_w - base_sum

    # 2) Resto: se reparte solo entre columnas que estaban por encima de su mínimo en el reparto "raw".
    if remaining > 1e-6:
        extras = [max(0.0, raw_ws[i] - min_w_by_key[i]) for i in range(n)]
        extras_sum = sum(extras)
        if extras_sum > 0:
            for i in range(n):
                ws[i] += remaining * (extras[i] / extras_sum)
        else:
            idx = keys.index("descripcion") if "descripcion" in keys else n - 1
            ws[idx] += remaining

    # Ajuste final por redondeo.
    drift = total_w - sum(ws)
    if abs(drift) > 1e-6:
        idx = keys.index("descripcion") if "descripcion" in keys else n - 1
        ws[idx] += drift

    return [max(0.0, float(w)) for w in ws]


def _pdf_inventario_bytes(
    df: pd.DataFrame,
    t: dict[str, Any] | None,
    *,
    agrupar_categoria: bool,
    subtitulo_filtros: str,
    column_keys: frozenset[str] | None = None,
) -> bytes:
    from copy import deepcopy

    from xml.sax.saxutils import escape as xml_esc

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    page = A4
    lm = 18 * mm
    rm = 18 * mm
    tm = 15 * mm
    bm = 18 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=lm,
        rightMargin=rm,
        topMargin=tm,
        bottomMargin=bm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []
    tw = float(page[0] - lm - rm)

    if BRAND_LOGO_PATH.is_file():
        try:
            ir = ImageReader(str(BRAND_LOGO_PATH))
            iw, ih = ir.getSize()
            if iw > 0 and ih > 0:
                target_w = 44 * mm
                target_h = target_w * (float(ih) / float(iw))
                max_h = 22 * mm
                if target_h > max_h:
                    target_h = max_h
                    target_w = target_h * (float(iw) / float(ih))
                lg = RLImage(str(BRAND_LOGO_PATH), width=target_w, height=target_h)
                lg.hAlign = "CENTER"
                story.append(lg)
                story.append(Spacer(1, 2.5 * mm))
        except Exception:
            pass

    tz = ZoneInfo("America/Caracas")
    fecha = datetime.now(tz).strftime("%d/%m/%Y %H:%M")
    tit = deepcopy(styles["Title"])
    tit.fontSize = 13
    tit.leading = 15
    tit.textColor = colors.HexColor("#2a1f45")
    tit.alignment = TA_CENTER
    story.append(Paragraph(xml_esc("Listado de inventario"), tit))
    meta = deepcopy(styles["Normal"])
    meta.fontSize = 8.5
    meta.alignment = TA_CENTER
    story.append(
        Paragraph(
            xml_esc(f"Movi Motor's Importadora · Generado: {fecha} (America/Caracas) · {len(df)} ítem(s)"),
            meta,
        )
    )
    work_pdf = _inv_rep_prepare_work_df(df)
    col_defs_pdf = _inv_rep_cols_for_export(work_pdf, column_keys)
    pdf_key_list = [k for k, _ in col_defs_pdf]
    _keys_pdf_set = set(pdf_key_list)
    tbs = _nf(t.get("tasa_bs")) if t else None
    if tbs is not None and _keys_pdf_set & {"precio_v_bs_ref", "costo_bs_ref"}:
        story.append(Paragraph(xml_esc(f"Tasa Bs/USD ref.: {tbs:,.2f}"), meta))
    tusd = _nf(t.get("tasa_usdt")) if t else None
    if tusd is not None and _keys_pdf_set & INV_REP_SYNTH_USDT_KEYS:
        story.append(Paragraph(xml_esc(f"Tasa USDT/USD ref.: {tusd:,.6f}"), meta))
    if subtitulo_filtros.strip():
        sf = deepcopy(styles["BodyText"])
        sf.fontSize = 8
        sf.alignment = TA_CENTER
        story.append(Paragraph(xml_esc(subtitulo_filtros), sf))
    story.append(Spacer(1, 2.5 * mm))

    if not pdf_key_list:
        story.append(Paragraph(xml_esc("Sin columnas para mostrar con la plantilla elegida."), meta))
        doc.build(story)
        return buf.getvalue()
    headers = [INV_REP_PDF_ABBR.get(k, k) for k in pdf_key_list]
    n_h = len(headers)
    col_ws = _pdf_inventario_col_widths_for_keys(pdf_key_list, tw)
    t_bs_pdf = _nf(t.get("tasa_bs")) if t else None

    cell_l = ParagraphStyle(
        name="invCellL",
        fontName="Helvetica",
        fontSize=6.5,
        leading=7.8,
        alignment=TA_LEFT,
        spaceAfter=0,
        spaceBefore=0,
    )
    cell_r = ParagraphStyle(
        name="invCellR",
        fontName="Helvetica",
        fontSize=6.5,
        leading=7.8,
        alignment=TA_RIGHT,
        spaceAfter=0,
        spaceBefore=0,
    )
    hdr_ps = ParagraphStyle(
        name="invHdr",
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=8,
        alignment=TA_CENTER,
        textColor=colors.whitesmoke,
        spaceAfter=0,
        spaceBefore=0,
    )

    def Pcell(txt: str | None, ps: ParagraphStyle) -> Paragraph:
        return Paragraph(xml_esc("" if txt is None else str(txt)), ps)

    def fmt_int_st(x: Any) -> str:
        try:
            return f"{_inv_stock_int(x):d}"
        except (TypeError, ValueError):
            return ""

    def cell_txt_for_key(r: pd.Series, key: str) -> str:
        if key == "categoria_display":
            return _inv_cat_display(r.get("categoria")) or ""
        if key == "_veh_rep":
            return _inv_compat_marcas_str(_inv_compat_as_dict(r.get("compatibilidad"))) or ""
        if key == "_anos_rep":
            return _inv_compat_anos_str(_inv_compat_as_dict(r.get("compatibilidad"))) or ""
        if key == "activo":
            return "Sí" if bool(r.get("activo")) else "No"
        if key in ("stock_actual", "stock_minimo"):
            return fmt_int_st(r.get(key))
        if key in ("costo_usd", "precio_v_usd"):
            return _rep_fmt_precio_entero(r.get(key))
        if key == "precio_v_bs_ref":
            return _inv_format_bs_ref_cell(r.get("precio_v_bs_ref"), r.get("precio_v_usd"), t_bs_pdf)
        if key == "costo_bs_ref":
            return _inv_format_bs_ref_cell(r.get("costo_bs_ref"), r.get("costo_usd"), t_bs_pdf)
        if key == "_pv_usdt_ref":
            return _inv_format_usdt_ref_cell(r.get("precio_v_usd"), tusd)
        if key == "_cu_usdt_ref":
            return _inv_format_usdt_ref_cell(r.get("costo_usd"), tusd)
        return _export_cell_txt(r.get(key))

    def row_cells(r: pd.Series) -> list[Paragraph]:
        texts = [cell_txt_for_key(r, k) for k in pdf_key_list]
        out: list[Paragraph] = []
        for k, tx in zip(pdf_key_list, texts):
            out.append(Pcell(tx, cell_r if k in INV_REP_NUMERIC_KEYS else cell_l))
        return out

    hdr_row = [Pcell(h, hdr_ps) for h in headers]

    tbl_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2a1f45")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("LEADING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
    )

    work = work_pdf
    work["_cdisp"] = work["categoria_display"]

    if agrupar_categoria:
        h4_base = ParagraphStyle(
            name="invCatHdr",
            fontName="Helvetica-Bold",
            fontSize=9.5,
            leading=11,
            textColor=colors.HexColor("#e65100"),
            alignment=TA_LEFT,
            spaceAfter=2,
            spaceBefore=2,
        )
        for cat_name, grp in work.groupby("_cdisp", sort=False):
            story.append(Paragraph(f"<b>{xml_esc(str(cat_name))}</b>", h4_base))
            data: list[list[Any]] = [hdr_row]
            for _, r in grp.iterrows():
                data.append(row_cells(r))
            tbl = Table(data, colWidths=col_ws, repeatRows=1)
            tbl.setStyle(tbl_style)
            story.append(tbl)
            story.append(Spacer(1, 2 * mm))
    else:
        data = [hdr_row]
        for _, r in work.iterrows():
            data.append(row_cells(r))
        tbl = Table(data, colWidths=col_ws, repeatRows=1)
        tbl.setStyle(tbl_style)
        story.append(tbl)

    doc.build(story)
    return buf.getvalue()


def _pdf_toma_inventario_fisico_bytes(df: pd.DataFrame, *, subtitulo_filtros: str) -> bytes:
    """PDF para conteo físico: código, categoría, descripción, stock sistema y columna vacía para anotar."""
    from copy import deepcopy
    from xml.sax.saxutils import escape as xml_esc

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    page = A4
    lm = 14 * mm
    rm = 14 * mm
    tm = 12 * mm
    bm = 14 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=lm,
        rightMargin=rm,
        topMargin=tm,
        bottomMargin=bm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []
    tw = float(page[0] - lm - rm)

    if BRAND_LOGO_PATH.is_file():
        try:
            ir = ImageReader(str(BRAND_LOGO_PATH))
            iw, ih = ir.getSize()
            if iw > 0 and ih > 0:
                target_w = 40 * mm
                target_h = target_w * (float(ih) / float(iw))
                max_h = 20 * mm
                if target_h > max_h:
                    target_h = max_h
                    target_w = target_h * (float(iw) / float(ih))
                lg = RLImage(str(BRAND_LOGO_PATH), width=target_w, height=target_h)
                lg.hAlign = "CENTER"
                story.append(lg)
                story.append(Spacer(1, 2 * mm))
        except Exception:
            pass

    tz = ZoneInfo("America/Caracas")
    fecha = datetime.now(tz).strftime("%d/%m/%Y %H:%M")
    tit = deepcopy(styles["Title"])
    tit.fontSize = 14
    tit.leading = 17
    tit.textColor = colors.HexColor("#2a1f45")
    tit.alignment = TA_CENTER
    story.append(Paragraph(xml_esc("Toma de inventario físico"), tit))
    meta = deepcopy(styles["Normal"])
    meta.fontSize = 8.5
    meta.alignment = TA_CENTER
    story.append(
        Paragraph(
            xml_esc(f"Movi Motor's Importadora · Generado: {fecha} (America/Caracas)"),
            meta,
        )
    )
    work = _inv_rep_prepare_work_df(df.copy())
    n_items = len(work)
    story.append(
        Paragraph(
            xml_esc(f"Ítems: {n_items} — Anotar el conteo real en la última columna."),
            meta,
        )
    )
    if subtitulo_filtros.strip():
        sf = deepcopy(styles["BodyText"])
        sf.fontSize = 8
        sf.alignment = TA_CENTER
        story.append(Paragraph(xml_esc(subtitulo_filtros), sf))
    story.append(Spacer(1, 3 * mm))

    cell_l = ParagraphStyle(
        name="tfCellL",
        fontName="Helvetica",
        fontSize=7,
        leading=8.5,
        alignment=TA_LEFT,
        spaceAfter=0,
        spaceBefore=0,
    )
    cell_r = ParagraphStyle(
        name="tfCellR",
        fontName="Helvetica",
        fontSize=7,
        leading=8.5,
        alignment=TA_RIGHT,
        spaceAfter=0,
        spaceBefore=0,
    )
    hdr_ps = ParagraphStyle(
        name="tfHdr",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.whitesmoke,
        spaceAfter=0,
        spaceBefore=0,
    )

    def Pcell(txt: str | None, ps: ParagraphStyle) -> Paragraph:
        return Paragraph(xml_esc("" if txt is None else str(txt)), ps)

    hdr_labels = ["Código", "Categoría", "Descripción", "Stock sist.", "Conteo físico"]
    hdr_row = [Pcell(h, hdr_ps) for h in hdr_labels]

    # Anchos: descripción flexible
    w_code = 22 * mm
    w_cat = 26 * mm
    w_stk = 16 * mm
    w_fis = 28 * mm
    w_desc = max(35 * mm, tw - w_code - w_cat - w_stk - w_fis)
    col_ws = [w_code, w_cat, w_desc, w_stk, w_fis]

    data: list[list[Any]] = [hdr_row]
    if work.empty:
        story.append(Paragraph(xml_esc("No hay productos con los filtros actuales."), meta))
        doc.build(story)
        return buf.getvalue()

    for _, r in work.iterrows():
        cod = _export_cell_txt(r.get("codigo"))
        cat = _inv_cat_display(r.get("categoria"))
        des = _export_cell_txt(r.get("descripcion"))
        try:
            stv = f"{_inv_stock_int(r.get('stock_actual')):d}"
        except (TypeError, ValueError):
            stv = ""
        data.append(
            [
                Pcell(cod, cell_l),
                Pcell(cat, cell_l),
                Pcell(des, cell_l),
                Pcell(stv, cell_r),
                Pcell("\u00a0", cell_l),
            ]
        )

    tbl = Table(data, colWidths=col_ws, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2a1f45")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("ALIGN", (4, 1), (4, -1), "LEFT"),
                ("TOPPADDING", (4, 1), (4, -1), 6),
                ("BOTTOMPADDING", (4, 1), (4, -1), 10),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()


def _backup_file_timestamp() -> str:
    return datetime.now(ZoneInfo("America/Caracas")).strftime("%Y%m%d_%H%M%S")


def _json_backup_bytes(payload: dict[str, Any]) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def _json_backup_bytes_compact_gzip(payload: dict[str, Any]) -> bytes:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=str).encode("utf-8")
    return gzip.compress(raw, compresslevel=9)


def decode_backup_upload_bytes(raw: bytes) -> dict[str, Any]:
    if len(raw) >= 2 and raw[0] == 0x1F and raw[1] == 0x8B:
        raw = gzip.decompress(raw)
    return json.loads(raw.decode("utf-8"))


_ERP_KV_KEY_AUTO_DAY = "movi_auto_backup_day_v1"


def _auto_backup_config() -> dict[str, Any]:
    defaults: dict[str, Any] = {"enabled": True, "retain_days": 14, "storage_bucket": ""}
    try:
        ab = st.secrets.get("auto_backup")
    except Exception:
        return defaults
    if ab is None:
        return defaults
    try:
        rd = int(ab.get("retain_days", 14))
        rd = max(1, min(365, rd))
        return {
            "enabled": bool(ab.get("enabled", True)),
            "retain_days": rd,
            "storage_bucket": str(ab.get("storage_bucket") or "").strip(),
        }
    except Exception:
        return defaults


def _auto_backup_dir() -> Path:
    return _APP_DIR / "auto_backups"


def _local_auto_backup_day_path() -> Path:
    return _auto_backup_dir() / ".last_auto_day_v1"


def _read_local_auto_backup_day() -> str | None:
    try:
        s = _local_auto_backup_day_path().read_text(encoding="utf-8").strip()
        return s or None
    except Exception:
        return None


def _write_local_auto_backup_day(day: str) -> None:
    d = _auto_backup_dir()
    d.mkdir(parents=True, exist_ok=True)
    _local_auto_backup_day_path().write_text(day, encoding="utf-8")


def _erp_kv_get(sb: Client, key: str) -> str | None:
    try:
        r = sb.table("erp_kv").select("value").eq("key", key).limit(1).execute()
        row = (r.data or [{}])[0]
        s = str(row.get("value") or "").strip()
        return s or None
    except Exception:
        return None


def _erp_kv_set(sb: Client, key: str, value: str) -> bool:
    try:
        sb.table("erp_kv").upsert({"key": key, "value": value}).execute()
        return True
    except Exception:
        return False


def _prune_old_auto_backups(*, retain_days: int) -> None:
    dir_path = _auto_backup_dir()
    if not dir_path.is_dir():
        return
    today_c = _today_caracas()
    cutoff = today_c - timedelta(days=max(1, retain_days))
    for p in dir_path.glob("movi_erp_auto_*.json.gz"):
        if not p.is_file():
            continue
        name = p.name
        if not name.startswith("movi_erp_auto_") or not name.endswith(".json.gz"):
            continue
        mid = name[len("movi_erp_auto_") : -len(".json.gz")]
        try:
            d = date.fromisoformat(mid)
        except ValueError:
            continue
        if d < cutoff:
            try:
                p.unlink()
            except OSError:
                pass


def _storage_auto_backup_exists(sb: Client, bucket: str, day_str: str) -> bool:
    needle = f"movi_erp_auto_{day_str}.json.gz"
    try:
        items = sb.storage.from_(bucket).list("auto")
        for it in items or []:
            if str(it.get("name") or "") == needle:
                return True
    except Exception:
        pass
    return False


def _try_storage_auto_backup(sb: Client, bucket: str, day_str: str, data: bytes) -> bool:
    if not bucket:
        return False
    path = f"auto/movi_erp_auto_{day_str}.json.gz"
    try:
        sb.storage.from_(bucket).upload(
            path,
            data,
            file_options={"content-type": "application/gzip", "upsert": "true"},
        )
        return True
    except Exception:
        return False


def _catalogo_bucket_name() -> str:
    """Bucket de Supabase Storage para fotos de productos."""
    try:
        cfg = st.secrets.get("catalogo")
        if isinstance(cfg, dict):
            b = str(cfg.get("bucket") or "").strip()
            if b:
                return b
    except Exception:
        pass
    return "movi-productos"


def _catalogo_storage_portada_enabled() -> bool:
    """
    Subir archivos a Storage, galería y foto de portada en la nube.
    Si es False: NO se oculta el catálogo HTML en Reportes (listados/etiquetas siguen); solo se quitan subida y galería.
    Inventario: solo campo URL para imagen_url, sin selector de archivo.

    secrets.toml → [catalogo]:
      storage_fotos = false   (recomendado)
      enabled = false         (mismo efecto; nombre viejo por compatibilidad)
    Si definís ambos, manda `storage_fotos`.
    """
    try:
        cfg = st.secrets.get("catalogo")
        if not isinstance(cfg, dict):
            return True
        if "storage_fotos" in cfg:
            v = cfg.get("storage_fotos")
            if isinstance(v, str):
                return v.strip().lower() not in ("0", "false", "no", "off", "")
            return bool(v)
        if "enabled" in cfg:
            v = cfg.get("enabled")
            if isinstance(v, str):
                return v.strip().lower() not in ("0", "false", "no", "off", "")
            return bool(v)
        return True
    except Exception:
        return True


def _movi_foto_upload_bucket_hint(bucket: str, ex: BaseException) -> str:
    """Si Storage devuelve bucket inexistente, guía para crearlo en el panel de Supabase."""
    low = str(ex).lower()
    if "bucket not found" in low or ("404" in str(ex) and "bucket" in low):
        b = str(bucket or _catalogo_bucket_name()).strip() or "movi-productos"
        return (
            f"\n\n**Falta el bucket en Supabase Storage:** creá uno llamado exactamente **`{b}`** "
            "(menú **Storage** → **New bucket**). Para que las URLs públicas del catálogo funcionen, "
            "dejalo **público** o configurá políticas de lectura. "
            "Otro nombre: en `secrets.toml` → `[catalogo]` → `bucket = \"...\"`."
        )
    return ""


def _supabase_url_base() -> str:
    u = st.secrets["connections"]["supabase"]["SUPABASE_URL"]
    return str(u).rstrip("/")


def _storage_public_object_url(bucket: str, path: str) -> str:
    base = _supabase_url_base()
    b = quote(str(bucket).strip(), safe="")
    p = quote(str(path).lstrip("/"), safe="/")
    return f"{base}/storage/v1/object/public/{b}/{p}"


def _catalogo_upload_producto_foto(
    sb: Client,
    *,
    bucket: str,
    producto_id: str,
    filename: str,
    content_type: str,
    data: bytes,
) -> str:
    pid = str(producto_id).strip()
    if not pid:
        raise ValueError("producto_id vacío")
    name = (filename or "foto").strip().replace("\\", "_").replace("/", "_")
    if not name:
        name = "foto"
    ext = ""
    if "." in name and len(name.rsplit(".", 1)[-1]) <= 8:
        ext = "." + name.rsplit(".", 1)[-1].lower()
    suf = secrets.token_hex(8)
    obj_path = f"productos/{pid}/{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{suf}{ext}"
    sb.storage.from_(bucket).upload(
        obj_path,
        data,
        file_options={"content-type": str(content_type or "application/octet-stream"), "upsert": "true"},
    )
    return obj_path


def _catalogo_fetch_fotos(sb: Client, producto_id: str) -> list[dict[str, Any]]:
    pid = str(producto_id).strip()
    if not pid:
        return []
    r = (
        sb.table("producto_fotos")
        .select("id,producto_id,storage_path,is_primary,created_at,created_by")
        .eq("producto_id", pid)
        .order("is_primary", desc=True)
        .order("created_at", desc=True)
        .execute()
    )
    return list(r.data or [])


def _catalogo_set_primary(sb: Client, *, producto_id: str, foto_id: str) -> None:
    pid = str(producto_id).strip()
    fid = str(foto_id).strip()
    if not pid or not fid:
        return
    sb.table("producto_fotos").update({"is_primary": False}).eq("producto_id", pid).execute()
    sb.table("producto_fotos").update({"is_primary": True}).eq("id", fid).execute()


def _catalogo_delete_foto(sb: Client, *, bucket: str, foto_row: dict[str, Any]) -> None:
    fid = str(foto_row.get("id") or "").strip()
    path = str(foto_row.get("storage_path") or "").strip()
    if path:
        try:
            sb.storage.from_(bucket).remove([path])
        except Exception:
            pass
    if fid:
        sb.table("producto_fotos").delete().eq("id", fid).execute()


def _catalogo_row_is_primary(row: dict[str, Any]) -> bool:
    """PostgREST a veces devuelve booleanos raros; normalizamos."""
    v = row.get("is_primary")
    if v is True:
        return True
    if v is False or v is None:
        return False
    if isinstance(v, (int, float)):
        return v != 0
    s = str(v).strip().lower()
    return s in ("true", "t", "1", "yes", "si", "sí")


def _catalogo_sync_primary_foto(
    sb: Client, *, bucket: str, producto_id: str, storage_path: str
) -> None:
    """
    Tras insertar una fila en producto_fotos: marca esa foto como principal y actualiza productos.imagen_url.
    Evita el caso en que el insert llegue con is_primary false por defecto en BD y el catálogo no muestre la imagen.
    """
    pid = str(producto_id).strip()
    path = str(storage_path).strip()
    if not pid or not path:
        return
    fotos_now = _catalogo_fetch_fotos(sb, pid)
    match = next((r for r in fotos_now if str(r.get("storage_path") or "").strip() == path), None)
    if match and str(match.get("id") or "").strip():
        _catalogo_set_primary(sb, producto_id=pid, foto_id=str(match["id"]).strip())
    sb.table("productos").update({"imagen_url": _storage_public_object_url(bucket, path)}).eq("id", pid).execute()


def _catalogo_primary_path_for_producto(sb: Client, producto_id: str) -> str | None:
    fotos = _catalogo_fetch_fotos(sb, producto_id)
    if not fotos:
        return None
    prim = next((x for x in fotos if _catalogo_row_is_primary(x)), None) or fotos[0]
    sp = str(prim.get("storage_path") or "").strip()
    return sp or None


def _html_catalogo_imprimible(
    items: list[dict[str, Any]],
    *,
    titulo: str,
    subtitulo: str,
) -> str:
    tz = ZoneInfo("America/Caracas")
    fecha = datetime.now(tz).strftime("%d/%m/%Y %H:%M")
    _logo_uri = _brand_logo_data_uri()
    _logo_block = (
        f'<div class="logo-wrap"><img class="logo" src="{_logo_uri}" alt="Movi Motors"/></div>'
        if _logo_uri
        else '<div class="logo-wrap logo-missing">Movi Motor\'s Importadora</div>'
    )

    cards: list[str] = []
    for it in items:
        desc = html.escape(str(it.get("descripcion") or ""))
        cod = html.escape(str(it.get("codigo") or ""))
        oem = html.escape(str(it.get("sku_oem") or ""))
        precio = it.get("precio_v_usd")
        precio_txt = _rep_fmt_precio_entero(precio)
        img = str(it.get("imagen_url") or "").strip()
        img_tag = f'<img class="ph" src="{html.escape(img)}" alt="{desc}"/>' if img else '<div class="ph ph-missing">Sin foto</div>'
        oem_line = f'<div class="oem">OEM: {oem}</div>' if oem else ""
        cards.append(
            f"""
            <div class="card">
              <div class="photo">{img_tag}</div>
              <div class="meta">
                <div class="desc">{desc}</div>
                <div class="code">{cod}</div>
                {oem_line}
                <div class="price">US$ {html.escape(precio_txt) if precio_txt else "—"}</div>
              </div>
            </div>
            """
        )

    sub_html = f"<div class=\"sub\">{html.escape(subtitulo)}</div>" if subtitulo.strip() else ""
    return f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8"/>
<title>{html.escape(titulo)}</title>
<style>
  @page {{ size: A4 portrait; margin: 12mm 12mm 14mm 12mm; }}
  * {{ box-sizing: border-box; }}
  body {{
    font-family: Segoe UI, Roboto, Arial, sans-serif;
    margin: 0;
    padding: 0.75rem 0.5rem;
    color: #111;
  }}
  .logo-wrap {{ text-align:center; margin: 0 0 0.5rem 0; }}
  .logo {{ max-height: 18mm; max-width: 52mm; width:auto; height:auto; object-fit:contain; }}
  .logo-missing {{ font-weight:800; font-size:1rem; color:#5c2d91; letter-spacing:0.02em; }}
  h1 {{ font-size: 1.15rem; margin: 0 0 0.2rem 0; text-align:center; color:#2a1f45; }}
  .meta-top {{ text-align:center; color:#444; font-size:0.82rem; margin-bottom:0.45rem; }}
  .sub {{ text-align:center; color:#333; font-size:0.78rem; margin: 0.2rem 0 0.6rem 0; }}
  .grid {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 0.55rem;
    align-items: stretch;
  }}
  .card {{
    border: 1px solid #c9c9c9;
    border-radius: 10px;
    overflow: hidden;
    background: #fff;
    min-height: 200px;
  }}
  .photo {{
    width: 100%;
    height: 148px;
    min-height: 148px;
    background: #f3f3f3;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 8px;
    box-sizing: border-box;
  }}
  .ph {{
    max-width: 100%;
    max-height: 100%;
    width: auto;
    height: auto;
    object-fit: contain;
    object-position: center;
    display: block;
    vertical-align: middle;
  }}
  .ph-missing {{
    width: 100%;
    height: 100%;
    min-height: 120px;
    display: flex;
    align-items: center;
    justify-content: center;
    color: #666;
    font-size: 0.85rem;
  }}
  .meta {{ padding: 0.55rem 0.6rem; }}
  .desc {{ font-weight: 700; font-size: 0.86rem; line-height: 1.2; min-height: 2.1em; }}
  .code {{ color:#2a1f45; font-weight: 700; margin-top:0.2rem; font-size:0.8rem; }}
  .oem {{ color:#444; margin-top:0.15rem; font-size:0.74rem; }}
  .price {{ margin-top:0.35rem; font-weight: 800; font-size: 0.95rem; color:#111; }}
  .print-actions {{ margin-top: 0.75rem; text-align:center; }}
  @media print {{
    body {{ padding: 0; }}
    .print-actions {{ display:none !important; }}
  }}
</style>
</head>
<body>
  {_logo_block}
  <h1>{html.escape(titulo)}</h1>
  <div class="meta-top">Generado: {html.escape(fecha)} (America/Caracas) · <strong>{len(items)}</strong> producto(s)</div>
  {sub_html}
  <div class="grid">
    {''.join(cards)}
  </div>
  <div class="print-actions">
    <script>function imprimir(){{ window.print(); }}</script>
    <button type="button" onclick="imprimir()" style="padding:0.5rem 1.2rem;font-size:1rem;cursor:pointer;background:#2a1f45;color:#fff;border:none;border-radius:6px;">Imprimir</button>
  </div>
</body>
</html>"""


def maybe_run_daily_auto_backup(sb: Client, rol: str) -> None:
    """
    Una vez al día (America/Caracas), si entra un superusuario: genera el mismo JSON
    que el respaldo completo pero compacto + gzip. Guarda en `auto_backups/` y/o
    Storage (secrets). Coordina con `erp_kv` o archivo local para no repetir.
    """
    if rol != "superuser":
        return
    cfg = _auto_backup_config()
    if not cfg["enabled"]:
        return
    today_str = _today_caracas().isoformat()

    if _erp_kv_get(sb, _ERP_KV_KEY_AUTO_DAY) == today_str:
        st.session_state["_movi_auto_backup_session_day"] = today_str
        return
    if _read_local_auto_backup_day() == today_str:
        st.session_state["_movi_auto_backup_session_day"] = today_str
        return
    bucket = str(cfg.get("storage_bucket") or "")
    if bucket and _storage_auto_backup_exists(sb, bucket, today_str):
        st.session_state["_movi_auto_backup_session_day"] = today_str
        return
    if st.session_state.get("_movi_auto_backup_session_day") == today_str:
        return
    if st.session_state.get("_movi_auto_backup_busy"):
        return

    st.session_state["_movi_auto_backup_busy"] = True
    try:
        payload = build_backup_erp_completo(sb)
        gz = _json_backup_bytes_compact_gzip(payload)
        ok_any = False
        bdir = _auto_backup_dir()
        try:
            bdir.mkdir(parents=True, exist_ok=True)
            (bdir / f"movi_erp_auto_{today_str}.json.gz").write_bytes(gz)
            ok_any = True
        except OSError:
            pass

        if bucket and _try_storage_auto_backup(sb, bucket, today_str, gz):
            ok_any = True

        if not ok_any:
            return

        _prune_old_auto_backups(retain_days=int(cfg["retain_days"]))
        if not _erp_kv_set(sb, _ERP_KV_KEY_AUTO_DAY, today_str):
            _write_local_auto_backup_day(today_str)
        st.session_state["_movi_auto_backup_session_day"] = today_str
        kb = max(1, len(gz) // 1024)
        st.session_state["_movi_auto_backup_toast"] = f"Respaldo automático del día guardado (~{kb} KB gzip)."
    except Exception as ex:
        st.session_state["_movi_auto_backup_toast_err"] = str(ex)
    finally:
        st.session_state["_movi_auto_backup_busy"] = False


def build_backup_inventario(sb: Client) -> dict[str, Any]:
    """Snapshot lógico: categorías + productos (todo el maestro de inventario)."""
    return {
        "meta": {
            "tipo": "inventario",
            "version": 1,
            "exportado_en_utc": datetime.now(timezone.utc).isoformat(),
            "app": "Movi Motors ERP",
        },
        "categorias": (sb.table("categorias").select("*").order("nombre").execute().data or []),
        "productos": (sb.table("productos").select("*").order("descripcion").execute().data or []),
    }


def build_backup_erp_completo(sb: Client) -> dict[str, Any]:
    """
    Snapshot amplio para recuperación ante fallos o antes de depuración.
    No exporta password_hash (seguridad); reasignar claves tras restaurar usuarios.
    """
    payload: dict[str, Any] = {
        "meta": {
            "tipo": "erp_completo",
            "version": 1,
            "exportado_en_utc": datetime.now(timezone.utc).isoformat(),
            "app": "Movi Motors ERP",
            "nota_usuarios": "erp_users sin password_hash. Restauración: volver a definir contraseñas en el módulo Usuarios o en Supabase.",
        },
    }
    specs: list[tuple[str, str]] = [
        ("categorias", "*"),
        ("productos", "*"),
        ("productos_kit_items", "*"),
        ("tasas_dia", "*"),
        ("cajas_bancos", "*"),
        ("erp_users", "id,username,nombre,email,rol,activo,created_at"),
        ("ventas", "*"),
        ("ventas_detalles", "*"),
        ("compras", "*"),
        ("compras_detalles", "*"),
        ("cuentas_por_cobrar", "*"),
        ("cuentas_por_pagar", "*"),
        ("movimientos_caja", "*"),
        ("cambios_tesoreria", "*"),
    ]
    errs: list[dict[str, str]] = []
    for tbl, cols in specs:
        try:
            payload[tbl] = sb.table(tbl).select(cols).execute().data or []
        except Exception as ex:
            errs.append({"tabla": tbl, "error": str(ex)})
            payload[tbl] = []
    if errs:
        payload["meta"]["errores_al_exportar"] = errs
    return payload


_PHONY_UUID = "00000000-0000-0000-0000-000000000000"


def _delete_all_rows(sb: Client, table: str) -> None:
    sb.table(table).delete().neq("id", _PHONY_UUID).execute()


def _insert_rows_batched(sb: Client, table: str, rows: list[dict[str, Any]], *, batch: int = 75) -> None:
    if not rows:
        return
    for i in range(0, len(rows), batch):
        chunk = rows[i : i + batch]
        sb.table(table).insert(chunk).execute()


def _merge_erp_users_from_backup(sb: Client, rows: list[dict[str, Any]]) -> list[str]:
    """Actualiza usuarios existentes; crea faltantes con contraseña temporal."""
    notes: list[str] = []
    for row in rows:
        uid = str(row.get("id") or "").strip()
        un = (row.get("username") or "").strip()
        if not uid or not un:
            continue
        ex = sb.table("erp_users").select("id").eq("id", uid).limit(1).execute()
        base = {
            "username": un.lower(),
            "nombre": (row.get("nombre") or un).strip(),
            "email": (row.get("email") or "").strip() or None,
            "rol": row.get("rol") or "vendedor",
            "activo": bool(row.get("activo", True)),
        }
        if ex.data:
            sb.table("erp_users").update(base).eq("id", uid).execute()
        else:
            try:
                sb.table("erp_users").insert(
                    {**base, "id": uid, "password_hash": _hash_password("Restaurar2025!")}
                ).execute()
                notes.append(
                    f"Usuario **{un}** creado desde respaldo — contraseña temporal: **Restaurar2025!** (cambiar en Usuarios)."
                )
            except Exception as ex:
                notes.append(f"No se pudo crear usuario `{un}`: {ex}")
    return notes


def restore_erp_completo_desde_json(sb: Client, payload: dict[str, Any]) -> tuple[bool, str, list[str]]:
    """
    Reemplaza datos operativos desde un JSON generado por *Descargar respaldo completo*.
    Orden: borra hijos primero; inserta categorías → … → movimientos.
    """
    meta = payload.get("meta") or {}
    if meta.get("tipo") != "erp_completo":
        return False, "El archivo no es un respaldo **erp_completo** (usá el JSON de Mantenimiento).", []
    if int(meta.get("version") or 0) < 1:
        return False, "Versión de respaldo no soportada.", []

    warns: list[str] = []
    try:
        _delete_all_rows(sb, "movimientos_caja")
        _delete_all_rows(sb, "ventas")
        _delete_all_rows(sb, "compras")
        try:
            _delete_all_rows(sb, "productos_kit_items")
        except Exception:
            pass
        _delete_all_rows(sb, "productos")
        _delete_all_rows(sb, "categorias")
        _delete_all_rows(sb, "tasas_dia")
        try:
            _delete_all_rows(sb, "cambios_tesoreria")
        except Exception:
            pass
        _delete_all_rows(sb, "cajas_bancos")

        _insert_rows_batched(sb, "categorias", list(payload.get("categorias") or []))
        _insert_rows_batched(sb, "productos", list(payload.get("productos") or []))
        try:
            _insert_rows_batched(sb, "productos_kit_items", list(payload.get("productos_kit_items") or []))
        except Exception:
            warns.append(
                "No se pudieron restaurar filas de **productos_kit_items** (¿falta patch_014?). Los kits hay que redefinirlos en Inventario."
            )
        _insert_rows_batched(sb, "tasas_dia", list(payload.get("tasas_dia") or []))
        _insert_rows_batched(sb, "cajas_bancos", list(payload.get("cajas_bancos") or []))
        try:
            _insert_rows_batched(sb, "cambios_tesoreria", list(payload.get("cambios_tesoreria") or []))
        except Exception:
            warns.append(
                "No se pudieron restaurar filas de **cambios_tesoreria** (¿falta patch_018?). Podés reimportarlas manualmente o ignorar si no usabas el registro."
            )

        urows = list(payload.get("erp_users") or [])
        if urows:
            warns.extend(_merge_erp_users_from_backup(sb, urows))

        _insert_rows_batched(sb, "ventas", list(payload.get("ventas") or []))
        _insert_rows_batched(sb, "ventas_detalles", list(payload.get("ventas_detalles") or []))
        _insert_rows_batched(sb, "compras", list(payload.get("compras") or []))
        _insert_rows_batched(sb, "compras_detalles", list(payload.get("compras_detalles") or []))
        _insert_rows_batched(sb, "cuentas_por_cobrar", list(payload.get("cuentas_por_cobrar") or []))
        _insert_rows_batched(sb, "cuentas_por_pagar", list(payload.get("cuentas_por_pagar") or []))
        _insert_rows_batched(sb, "movimientos_caja", list(payload.get("movimientos_caja") or []))

        try:
            sb.rpc("sync_erp_sequences").execute()
        except Exception:
            warns.append(
                "Ejecutá en Supabase `supabase/patch_009_sync_sequences.sql` para alinear números de venta/compra."
            )

        return True, "Restauración completa aplicada. Revisá avisos abajo si los hay.", warns
    except Exception as e:
        return False, f"Error durante la restauración (la base puede quedar incompleta): {e}", warns


def restore_inventario_desde_json(sb: Client, payload: dict[str, Any]) -> tuple[bool, str]:
    meta = payload.get("meta") or {}
    if meta.get("tipo") != "inventario":
        return False, "El archivo no es un respaldo de **inventario**."
    try:
        _delete_all_rows(sb, "productos")
        _delete_all_rows(sb, "categorias")
        _insert_rows_batched(sb, "categorias", list(payload.get("categorias") or []))
        _insert_rows_batched(sb, "productos", list(payload.get("productos") or []))
        return True, "Inventario restaurado desde el JSON."
    except Exception as e:
        err = str(e).lower()
        if "foreign key" in err or "23503" in err:
            return (
                False,
                "No se puede borrar productos: hay **ventas o compras** que los referencian. "
                "Usá **Mantenimiento → Restaurar todo** con el JSON completo, o depurá operaciones primero.",
            )
        return False, str(e)


def fmt_tri(usd: float, t_bs: float, t_usdt: float) -> str:
    """Equivalentes para pies de reporte: montos en enteros (sin decimales)."""
    usd = float(usd)
    return (
        f"**USD** {int(round(usd)):,d} · **Bs** {int(round(usd * t_bs)):,d} · **USDT** {int(round(usd * t_usdt)):,d}"
    )


def _nf(x: Any, default: float | None = None) -> float | None:
    try:
        if x is None:
            return default
        v = float(x)
        return v
    except (TypeError, ValueError):
        return default


def _pct_vs_bcv(mercado: float | None, bcv: float | None) -> float | None:
    if mercado is None or bcv is None or bcv <= 0:
        return None
    return (mercado - bcv) / bcv * 100.0


def _p2p_bs_equiv_por_usd(t: dict[str, Any]) -> float | None:
    p2p = _nf(t.get("p2p_bs_por_usdt"))
    tusdt = _nf(t.get("tasa_usdt"))
    if p2p is None or tusdt is None or tusdt <= 0:
        return None
    return p2p * tusdt


def build_tasas_tabla_detalle(t: dict[str, Any]) -> pd.DataFrame:
    """Cruces: USD×Bs, EUR×VES, USDT×VES (P2P), más referencias."""
    bcv = _nf(t.get("bcv_bs_por_usd")) or _nf(t.get("tasa_bs"))
    par = _nf(t.get("paralelo_bs_por_usd")) or _nf(t.get("tasa_bs"))
    t_bs_oper = _nf(t.get("tasa_bs"))
    usd_eur = _nf(t.get("usd_por_eur"))
    p2p = _nf(t.get("p2p_bs_por_usdt"))
    tusdt = _nf(t.get("tasa_usdt"))
    p2p_usd = _p2p_bs_equiv_por_usd(t)

    eur_x_ves_par = (par * usd_eur) if (par is not None and usd_eur is not None and usd_eur > 0) else None
    eur_x_ves_bcv = (bcv * usd_eur) if (bcv is not None and usd_eur is not None and usd_eur > 0) else None

    rows: list[dict[str, Any]] = []

    def add_row(nombre: str, valor: float | None, unidad: str, vs_bcv: float | None = None) -> None:
        rows.append(
            {
                "Cruce": nombre,
                "Valor": float(valor) if valor is not None else None,
                "Unidad": unidad,
                "vs BCV (%)": float(vs_bcv) if vs_bcv is not None else None,
            }
        )

    if bcv is not None:
        add_row(
            "USD × Bs — oficial BCV",
            bcv,
            "Bolívares por 1 USD — Banco Central de Venezuela (no es paralelo)",
            None,
        )
    if par is not None:
        add_row(
            "USD × Bs — Binance P2P / mercado (ref.)",
            par,
            "Bolívares por 1 USD — referencia que cargás desde mercado (p. ej. alineada a Binance P2P); no es BCV",
            _pct_vs_bcv(par, bcv),
        )
    if t_bs_oper is not None:
        add_row(
            "USD × Bs — operativa (facturación)",
            t_bs_oper,
            "Valor guardado como `tasa_bs` (elegiste BCV o mercado al guardar)",
            _pct_vs_bcv(t_bs_oper, bcv),
        )
    if eur_x_ves_bcv is not None:
        add_row(
            "EUR × VES — referencia BCV",
            eur_x_ves_bcv,
            "Bs por 1 EUR si usaras solo BCV × (USD por 1 EUR)",
            None,
        )
    if eur_x_ves_par is not None and usd_eur is not None:
        pv_ev = (
            _pct_vs_bcv(eur_x_ves_par, eur_x_ves_bcv)
            if (eur_x_ves_bcv is not None and eur_x_ves_bcv > 0)
            else None
        )
        add_row(
            "EUR × VES — vía ref. mercado (P2P)",
            eur_x_ves_par,
            f"Bs por 1 EUR = (USD×Bs ref. mercado) × ({float(usd_eur):.6f} USD por 1 EUR)",
            pv_ev,
        )
    if p2p is not None:
        add_row(
            "USDT × VES (P2P)",
            p2p,
            "Bolívares por 1 USDT — mercado P2P / cripto",
            _pct_vs_bcv(p2p, bcv),
        )
    if p2p_usd is not None:
        add_row(
            "USD × Bs — vía USDT (equiv.)",
            p2p_usd,
            "(USDT×VES P2P) × (USDT por 1 USD del sistema)",
            _pct_vs_bcv(p2p_usd, bcv),
        )
    if tusdt is not None:
        add_row("Ref. USDT por 1 USD", tusdt, "Para pasar USD↔USDT en pantallas", None)
    if usd_eur is not None and usd_eur > 0:
        add_row("Ref. USD por 1 EUR", usd_eur, "1 EUR = X USD (dato para armar EUR×VES)", None)

    return pd.DataFrame(rows)


def _infer_tasa_bs_oper_index(lt: dict[str, Any]) -> int:
    """0 = operar con BCV; 1 = operar con ref. mercado P2P (campo 2) — según último `tasa_bs` guardado."""
    if not lt:
        return 0
    tb = _nf(lt.get("tasa_bs"))
    b0 = _nf(lt.get("bcv_bs_por_usd")) or tb
    p0 = _nf(lt.get("paralelo_bs_por_usd")) or tb
    if tb is None or b0 is None or p0 is None:
        return 0
    d_b = abs(float(tb) - float(b0))
    d_p = abs(float(tb) - float(p0))
    return 0 if d_b <= d_p else 1


DOC_TASA_BS_OPTS = ("BCV oficial", "P2P Binance (mercado)")

# Gastos operativos (no compra de inventario): categoría va a movimientos_caja.categoria_gasto (patch_025).
GASTO_OPERATIVO_CATEGORIAS: tuple[str, ...] = (
    "Alquiler y local",
    "Servicios (luz, agua, internet, teléfono)",
    "Nómina y cargas sociales",
    "Impuestos y tasas",
    "Marketing y publicidad",
    "Transporte y fletes",
    "Mantenimiento y reparaciones (local, equipos)",
    "Honorarios profesionales",
    "Gastos bancarios y comisiones",
    "Otros",
)
GASTO_OPERATIVO_OTRO = "Otro (escribir categoría)"

# Cobros en caja: ZELLE se contabiliza 1:1 como USD (mismo tratamiento en RPC).
COBRO_MONEDAS: tuple[str, ...] = ("USD", "ZELLE", "VES", "USDT")

_COBR_MON_LBL: dict[str, str] = {
    "USD": "USD (efectivo o transferencia local en dólares)",
    "ZELLE": "Zelle (USD en cuenta USA; 1:1 con USD)",
    "VES": "Bolívares (VES)",
    "USDT": "USDT (cripto)",
}


def _fmt_moneda_cobro(code: str) -> str:
    return _COBR_MON_LBL.get(code, code)


def _error_msg_from_supabase_exc(e: BaseException) -> str:
    """Intenta extraer el mensaje legible de errores de PostgREST/Supabase."""
    m = getattr(e, "message", None)
    if m:
        return str(m)
    s = str(e).strip()
    if s.startswith("{") and "message" in s:
        import ast

        try:
            d = ast.literal_eval(s)
            if isinstance(d, dict) and d.get("message"):
                return str(d["message"])
        except (SyntaxError, ValueError, TypeError):
            pass
    return s


def _monto_nativo_a_usd(mon: str, monto: float, t_bs: float, t_usdt: float) -> float:
    """Convierte monto cobrado en VES / USD / USDT / ZELLE a equivalente USD (tasas de la venta)."""
    u = (mon or "").strip().upper()
    if u in ("USD", "ZELLE"):
        return float(monto)
    if u == "USDT":
        return float(monto) / float(t_usdt) if t_usdt else 0.0
    if u in ("VES", "BS"):
        return float(monto) / float(t_bs) if t_bs else 0.0
    return 0.0


def _tasa_bs_para_documento(t: dict[str, Any], *, usar_bcv: bool) -> float:
    """Bs por 1 USD para esta venta/compra: BCV o ref. mercado (campo P2P). Los montos USD van 1:1."""
    tb = _nf(t.get("tasa_bs"))
    bcv = _nf(t.get("bcv_bs_por_usd")) or tb
    par = _nf(t.get("paralelo_bs_por_usd")) or tb
    if usar_bcv:
        v = bcv if bcv is not None and float(bcv) > 0 else tb
    else:
        v = par if par is not None and float(par) > 0 else tb
    if v is None or float(v) <= 0:
        raise ValueError("Sin tasa Bs/USD válida; cargá tasas en el Dashboard.")
    return float(v)


def render_tabla_tasas_ui(df: pd.DataFrame) -> None:
    if df.empty:
        st.caption("No hay filas para mostrar.")
        return
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Cruce": st.column_config.TextColumn("Cruce", width="large"),
            "Valor": st.column_config.NumberColumn("Valor", format="%.6f"),
            "Unidad": st.column_config.TextColumn("Unidad / nota", width="large"),
            "vs BCV (%)": st.column_config.NumberColumn(
                "vs BCV (%)",
                format="%.2f",
                help="Sobre Bs/USD o cruces derivados comparables con el BCV guardado.",
            ),
        },
    )


@st.cache_data(ttl=120, show_spinner="Consultando tasas en internet…")
def get_live_exchange_rates() -> dict[str, Any]:
    from tasas_live import fetch_live_rates

    return fetch_live_rates()


def render_sidebar_welcome(*, nombre: str, username: str, rol: str) -> None:
    safe_n = html.escape(str(nombre or username or "Usuario"))
    safe_u = html.escape(str(username or ""))
    safe_r = html.escape(str(rol or ""))
    line_user = safe_n if safe_n else safe_u
    st.markdown(
        f"""
<div class="sb-welcome">
  <div class="sb-welcome-title">Movi Motor's Importadora</div>
  <div class="sb-welcome-sub">ERP · Multimoneda</div>
  <div class="sb-welcome-user">{line_user} · <span class="sb-role">{safe_r}</span></div>
</div>
""",
        unsafe_allow_html=True,
    )


def render_sidebar_calculadora() -> None:
    """Calculadora aritmética en la barra lateral (A y B, operación, resultado)."""
    with st.expander("Calculadora", expanded=False, key="sidebar_exp_calculadora"):
        _ca, _cb = st.columns(2)
        with _ca:
            _va = st.number_input("A", value=0.0, format="%.6g", key="movi_calc_a")
        with _cb:
            _vb = st.number_input("B", value=0.0, format="%.6g", key="movi_calc_b")
        _op = st.selectbox(
            "Operación",
            ["Suma (+)", "Resta (−)", "Multiplicación (×)", "División (÷)"],
            key="movi_calc_op",
        )
        _res: float | None = None
        try:
            if _op.startswith("Suma"):
                _res = float(_va) + float(_vb)
            elif _op.startswith("Resta"):
                _res = float(_va) - float(_vb)
            elif _op.startswith("Multi"):
                _res = float(_va) * float(_vb)
            else:
                if abs(float(_vb)) < 1e-15:
                    st.caption("No se puede dividir por cero.")
                else:
                    _res = float(_va) / float(_vb)
        except (TypeError, ValueError, OverflowError):
            st.caption("Valores no válidos.")
        if _res is not None:
            st.markdown(f"**Resultado:** `{_res:,.8g}`")


def render_tasas_tiempo_real(*, key_suffix: str, t_guardado: dict[str, Any] | None) -> dict[str, Any]:
    """Muestra tasas públicas en vivo (caché ~2 min) y opción de forzar refresco."""
    st.markdown("##### Tasas en tiempo real (internet)")
    st.caption(
        "Cruces: **USD×Bs**, **EUR×VES**, **USDT×VES (Binance P2P)**. Caché **~2 min**; usa **Refrescar ahora** para forzar. "
        "El **USD×Bs** web **no** es BCV oficial. **EUR×VES** = (Bs/USD web) × (USD por 1 EUR, Frankfurter). "
        "**USDT×VES** = mediana de anuncios en **Binance P2P** (VES/USDT); si Binance falla, se usa el mismo Bs/USD de la API de mercado como respaldo."
    )
    _, br = st.columns([3, 1])
    with br:
        if st.button("Refrescar ahora", key=f"live_refresh_{key_suffix}", help="Ignora la caché y vuelve a pedir datos"):
            get_live_exchange_rates.clear()
            st.rerun()

    data = get_live_exchange_rates()
    for err in data.get("errors") or []:
        st.warning(str(err))

    if not data.get("ok"):
        st.info("No se obtuvieron tasas web. Comprueba tu conexión o inténtalo más tarde.")
        return data

    bcv_ref = None
    if t_guardado:
        bcv_ref = _nf(t_guardado.get("bcv_bs_por_usd")) or _nf(t_guardado.get("tasa_bs"))

    rows: list[dict[str, Any]] = []
    ves = data.get("ves_bs_por_usd")
    eur_ves = data.get("eur_x_ves")
    p2p = data.get("usdt_x_ves_p2p") or data.get("p2p_bs_por_usdt_aprox")
    p2p_src = data.get("usdt_x_ves_p2p_source")
    eur = data.get("usd_por_eur")
    ut = data.get("usdt_por_usd")

    if ves is not None:
        pv = _pct_vs_bcv(float(ves), bcv_ref)
        rows.append(
            {
                "Cruce": "USD × Bs (web)",
                "Valor": float(ves),
                "Unidad": "Bs por 1 USD — ref. mercado web (no BCV); contrastá con tu precio en Binance P2P",
                "vs BCV guardado (%)": float(pv) if pv is not None else None,
            }
        )
    if eur_ves is not None:
        bcv_eur = None
        if t_guardado and eur:
            b0 = _nf(t_guardado.get("bcv_bs_por_usd")) or _nf(t_guardado.get("tasa_bs"))
            if b0 is not None and eur:
                bcv_eur = float(b0) * float(eur)
        pv_e = _pct_vs_bcv(float(eur_ves), bcv_eur) if bcv_eur else None
        rows.append(
            {
                "Cruce": "EUR × VES (web)",
                "Valor": float(eur_ves),
                "Unidad": f"Bs por 1 EUR = (USD×Bs web) × ({float(eur):.6f} USD por 1 EUR)" if eur else "Bs por 1 EUR",
                "vs BCV guardado (%)": float(pv_e) if pv_e is not None else None,
            }
        )
    if p2p is not None:
        pv2 = _pct_vs_bcv(float(p2p), bcv_ref)
        if p2p_src == "binance_p2p_median_buy":
            p2p_label = "USDT × VES (Binance P2P)"
            p2p_unit = "Bs por 1 USDT — mediana hasta 20 anuncios (comprar USDT con VES en Binance)"
        else:
            p2p_label = "USDT × VES (respaldo API)"
            p2p_unit = "Bs por 1 USDT — respaldo: mismo valor que USD×Bs web (Binance no respondió)"
        rows.append(
            {
                "Cruce": p2p_label,
                "Valor": float(p2p),
                "Unidad": p2p_unit,
                "vs BCV guardado (%)": float(pv2) if pv2 is not None else None,
            }
        )
    if eur is not None and eur_ves is None:
        rows.append(
            {
                "Cruce": "Ref. USD por 1 EUR (web)",
                "Valor": float(eur),
                "Unidad": "Dato para calcular EUR×VES cuando cargue USD×Bs",
                "vs BCV guardado (%)": None,
            }
        )
    if ut is not None:
        rows.append(
            {
                "Cruce": "Ref. USDT por 1 USD",
                "Valor": float(ut),
                "Unidad": "Solo referencia en la app al guardar tasas",
                "vs BCV guardado (%)": None,
            }
        )

    st.dataframe(
        pd.DataFrame(rows),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Cruce": st.column_config.TextColumn("Cruce", width="large"),
            "Valor": st.column_config.NumberColumn("Valor", format="%.6f"),
            "Unidad": st.column_config.TextColumn("Unidad / nota", width="large"),
            "vs BCV guardado (%)": st.column_config.NumberColumn(
                "vs tu BCV en BD (%)",
                format="%.2f",
                help="Compara con el BCV guardado en Supabase (cuando aplica).",
            ),
        },
    )
    meta = " · ".join(data.get("sources") or [])
    if data.get("time_next_update_utc"):
        meta += f" · Próx. actualización (API VES): {data['time_next_update_utc']}"
    st.caption(meta)
    return data


def _plotly_apply_dash_theme(fig: Any, *, title: str | None = None) -> Any:
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(22,27,34,0.55)",
        font=dict(color="#c9d1d9", family="system-ui, sans-serif"),
        margin=dict(t=56, b=48, l=24, r=24),
        hovermode="x unified",
        title=dict(text=title, font=dict(size=15, color="#00e5ff")) if title else None,
    )
    fig.update_xaxes(gridcolor="rgba(255,255,255,0.06)", zeroline=False)
    fig.update_yaxes(gridcolor="rgba(255,255,255,0.06)", zeroline=False)
    return fig


def _dash_liquidity_bucket(*, tipo: str, nombre: str, entidad: str = "") -> str:
    n = f"{nombre or ''} {entidad or ''}".lower()
    if "binance" in n or "usdt" in n or "crypto" in n:
        return "Binance (crypto)"
    t = (tipo or "").strip()
    if t == "Efectivo":
        return "Caja fuerte"
    if t == "Banco":
        return "Bancos nacionales"
    if t == "Wallet":
        return "Binance (crypto)"
    return "Otros"


def _caja_etiqueta_lista(c: dict[str, Any]) -> str:
    """Texto para selectores: banco + alias + moneda de cuenta + Nº + titular."""
    nom = str(c.get("nombre") or "").strip()
    ent = str(c.get("entidad") or "").strip()
    mon = (str(c.get("moneda_cuenta") or "USD")).strip().upper() or "USD"
    num = str(c.get("numero_cuenta") or "").strip()
    tit = str(c.get("titular") or "").strip()
    tipo = str(c.get("tipo") or "").strip()
    bits: list[str] = []
    if ent:
        bits.append(ent)
    if nom:
        bits.append(nom)
    if not bits:
        bits.append("Caja")
    bits.append(mon)
    if num:
        bits.append(f"Nº {num}")
    if tit:
        bits.append(tit)
    s = " · ".join(bits)
    return f"{s} ({tipo})" if tipo else s


def _round_money_2(x: Any) -> float:
    """Montos en pantalla / tablas: máximo 2 decimales (evita colas de ceros por precisión float)."""
    try:
        return round(float(x), 2)
    except (TypeError, ValueError):
        return 0.0


def _caja_saldo_cuenta_y_equiv(
    moneda_cuenta: str | None,
    saldo_actual_usd: Any,
    t: dict[str, Any] | None,
) -> tuple[str, str]:
    """Texto de saldo en la moneda de la cuenta + texto de equiv. USD interno (`saldo_actual_usd`)."""
    mon = (moneda_cuenta or "USD").strip().upper() or "USD"
    su = _round_money_2(_nf(saldo_actual_usd) or 0)
    if mon == "VES":
        t_bs = float(_nf((t or {}).get("tasa_bs")) or 0) if t else 0.0
        if t_bs <= 0:
            return (
                "Bs — (cargá **tasa Bs** en Tasas)",
                f"equiv. USD {_round_money_2(su):,.2f}",
            )
        bs = su * t_bs
        return (
            f"Bs {_round_money_2(bs):,.2f}",
            f"US$ {_round_money_2(su):,.2f} equiv.",
        )
    if mon == "USDT":
        t_ut = float(_nf((t or {}).get("tasa_usdt")) or 0) if t else 0.0
        if t_ut <= 0:
            return (
                "USDT — (cargá **tasa USDT** en Tasas)",
                f"equiv. USD {_round_money_2(su):,.2f}",
            )
        ut = su * t_ut
        return (
            f"USDT {round(ut, 4):,.4f}",
            f"US$ {_round_money_2(su):,.2f} equiv.",
        )
    return (
        f"US$ {_round_money_2(su):,.2f}",
        f"US$ {_round_money_2(su):,.2f} (ref. sistema)",
    )


def _modulo_titulo_info(titulo: str, *, ayuda_md: str, key: str) -> None:
    """Título del módulo y texto de ayuda en un expander (menos ruido visual hasta que lo abras)."""
    _mt1, _mt2 = st.columns([5.2, 1.05], gap="small")
    with _mt1:
        st.subheader(titulo)
    with _mt2:
        with st.expander("Información", expanded=False, key=f"modinfo_exp_{key}"):
            st.markdown(ayuda_md)


def _cajas_fetch_rows(sb: Client, *, solo_activas: bool) -> list[dict[str, Any]]:
    q = (
        sb.table("cajas_bancos")
        .select("id,nombre,tipo,saldo_actual_usd,activo,entidad,numero_cuenta,titular,moneda_cuenta")
        .order("nombre")
    )
    if solo_activas:
        q = q.eq("activo", True)
    try:
        return q.execute().data or []
    except Exception:
        q2 = sb.table("cajas_bancos").select("id,nombre,tipo,saldo_actual_usd,activo").order("nombre")
        if solo_activas:
            q2 = q2.eq("activo", True)
        rows = q2.execute().data or []
        for r in rows:
            r.setdefault("entidad", None)
            r.setdefault("numero_cuenta", None)
            r.setdefault("titular", None)
            r.setdefault("moneda_cuenta", "USD")
        return rows


def _caja_select_options(rows: list[dict[str, Any]]) -> tuple[list[str], Any]:
    """(ids, format_func) para st.selectbox por id único."""
    by_id = {str(r["id"]): r for r in rows}
    ids = [str(r["id"]) for r in rows]
    bases = [_caja_etiqueta_lista(by_id[i]) for i in ids]
    dup = Counter(bases)

    def fmt(cid: str) -> str:
        b = _caja_etiqueta_lista(by_id[cid])
        if dup[b] > 1:
            return f"{b} · ref {cid[:8]}…"
        return b

    return ids, fmt


def _dash_trend_pct(curr: float, prev: float) -> float | None:
    if prev > 1e-9:
        return (curr - prev) / prev * 100.0
    if curr > 1e-9:
        return 100.0
    return None


def _dash_kpi_card(label: str, value: str, trend_pct: float | None = None, sub: str | None = None) -> None:
    if trend_pct is None:
        tr = '<span class="dash-kpi-trend-flat">— vs período anterior</span>'
    elif trend_pct > 0.05:
        tr = f'<span class="dash-kpi-trend-up">▲ {trend_pct:+.1f}%</span>'
    elif trend_pct < -0.05:
        tr = f'<span class="dash-kpi-trend-down">▼ {trend_pct:+.1f}%</span>'
    else:
        tr = '<span class="dash-kpi-trend-flat">≈ estable</span>'
    sub_html = f'<div class="dash-kpi-sub">{html.escape(sub)}</div>' if sub else ""
    st.markdown(
        f"""
<div class="dash-bento">
  <div class="dash-kpi-label">{html.escape(label)}</div>
  <div class="dash-kpi-value">{value}</div>
  <div class="dash-kpi-sub">{tr}</div>
  {sub_html}
</div>
""",
        unsafe_allow_html=True,
    )


def _dash_mercado_card(label: str, value: str, *, foot: str | None = None) -> None:
    """Tarjeta estilo dashboard para cotizaciones (sin tendencia vs período)."""
    ft = (
        f'<div class="dash-kpi-sub">{html.escape(foot)}</div>'
        if foot
        else '<div class="dash-kpi-sub"> </div>'
    )
    st.markdown(
        f"""
<div class="dash-bento">
  <div class="dash-kpi-label">{html.escape(label)}</div>
  <div class="dash-kpi-value">{html.escape(value)}</div>
  {ft}
</div>
""",
        unsafe_allow_html=True,
    )


def render_dashboard_mercado_live_tarjetas(t: dict[str, Any] | None) -> None:
    """Cotizaciones web (caché ~2 min) en tarjetas — referencia de mercado en Resumen ejecutivo."""
    st.markdown("##### Referencia de mercado (tiempo real)")
    st.caption(
        "Fuentes públicas en internet · caché **~2 min** · **no** es BCV oficial. "
        "Ventas y compras siguen usando **`tasa_bs` / `tasa_usdt` guardadas en BD** hasta que las actualices o corra el **auto-sync**."
    )

    _bref, _ = st.columns([1, 5])
    with _bref:
        if st.button(
            "Refrescar ahora",
            key="dash_mercado_refresh_live",
            help="Ignora la caché y vuelve a consultar las APIs",
        ):
            get_live_exchange_rates.clear()
            st.rerun()

    live = get_live_exchange_rates()
    ves = live.get("ves_bs_por_usd")
    p2p = live.get("usdt_x_ves_p2p") or live.get("p2p_bs_por_usdt_aprox")
    ut_ref = _nf(live.get("usdt_por_usd")) or 1.0
    t_bs_bd = _nf(t.get("tasa_bs")) if t else None

    if not live.get("ok") or ves is None:
        st.info("Sin datos web por ahora. Revisá la conexión o probá **Refrescar ahora**.")
        for err in (live.get("errors") or [])[:3]:
            st.caption(html.escape(str(err)[:160]))
        if t:
            st.markdown("**Operativo en documentos (BD)**")
            c1, c2 = st.columns(2)
            with c1:
                _dash_mercado_card("Bs por 1 USD (guardado)", f"{float(t['tasa_bs']):,.2f}", foot="Facturación / compras")
            with c2:
                _dash_mercado_card("USDT por 1 USD (guardado)", f"{float(t['tasa_usdt']):,.6f}", foot="Sistema")
        return

    fv = float(ves)
    delta_vs = None
    if t_bs_bd is not None and float(t_bs_bd) > 0:
        delta_vs = fv - float(t_bs_bd)

    m1, m2, m3 = st.columns(3)
    with m1:
        _foot = (
            f"{delta_vs:+,.2f} Bs vs tu Bs/USD guardado en BD" if delta_vs is not None else "Mercado web · no BCV"
        )
        _dash_mercado_card("Bs por 1 USD (web)", f"{fv:,.2f}", foot=_foot)
    with m2:
        _dash_mercado_card("USDT por 1 USD (ref. web)", f"{float(ut_ref):,.4f}", foot="Cruce USDT/USD")
    with m3:
        if p2p is not None:
            src = live.get("usdt_x_ves_p2p_source")
            p2p_lbl = "Binance P2P" if src == "binance_p2p_median_buy" else "API referencia"
            _dash_mercado_card("Bs por 1 USDT", f"{float(p2p):,.4f}", foot=f"Fuente: {p2p_lbl}")
        else:
            _dash_mercado_card("Bs por 1 USDT", "—", foot="Sin dato P2P (revisá conexión)")

    if t:
        tbusd = float(t["tasa_bs"])
        tusdt = float(t["tasa_usdt"])
        d_bd_web = tbusd - fv
        st.caption("**Comparación con lo guardado en base (lo que usan los documentos hoy)**")
        b1, b2 = st.columns(2)
        with b1:
            _dash_mercado_card(
                "Bs por 1 USD (guardado en BD)",
                f"{tbusd:,.2f}",
                foot=f"{d_bd_web:+,.2f} vs web ahora",
            )
        with b2:
            _dash_mercado_card("USDT por 1 USD (guardado en BD)", f"{tusdt:,.6f}", foot="Misma fila de tasas del día")


def _bitacora_filas_con_cajas(
    sb: Client, rows: list[dict[str, Any]], t: dict[str, Any] | None
) -> list[dict[str, Any]]:
    """Filas legibles: salida de Bs desde cuenta VES → entrada de valor en cuenta USD/USDT."""
    if not rows:
        return []
    cj = {str(c["id"]): c for c in _cajas_fetch_rows(sb, solo_activas=False)}
    t_ut = float(_nf(t.get("tasa_usdt")) or 0) if t else 0.0
    out: list[dict[str, Any]] = []
    for r in rows:
        oid = r.get("caja_origen_id")
        did = r.get("caja_destino_id")
        co = cj.get(str(oid)) if oid else None
        cd = cj.get(str(did)) if did else None
        origen = _caja_etiqueta_lista(co) if co else "—"
        destino = _caja_etiqueta_lista(cd) if cd else "—"
        mon_d = str((cd or {}).get("moneda_cuenta") or "").strip().upper() or "—"
        m_usd = float(r.get("monto_usd_obtenido") or 0)
        m_bs = float(r.get("monto_ves") or 0)
        if mon_d == "USDT" and t_ut > 0:
            detalle = f"≈ {_round_money_2(m_usd * t_ut):,.4f} USDT (tasa día {t_ut:,.6f} USDT/USD)"
        elif mon_d == "USDT":
            detalle = "USDT — cargá tasa USDT/USD en Tasas para estimar cantidad"
        elif mon_d == "USD":
            detalle = "Valor en cuenta USD (Zelle u otra si esa es la caja)"
        else:
            detalle = f"Cuenta destino ({mon_d})"
        out.append(
            {
                "Fecha": str(r.get("fecha") or "")[:19].replace("T", " "),
                "Salió de (bolívares)": origen,
                "Llegó a (destino)": destino,
                "Moneda destino": mon_d,
                "Bs usados": _round_money_2(m_bs),
                "USD equivalente": _round_money_2(m_usd),
                "Qué representa": detalle,
                "Nota": (str(r.get("nota") or ""))[:120],
            }
        )
    return out


def _caja_map_por_id(sb: Client) -> dict[str, dict[str, Any]]:
    return {str(c["id"]): c for c in _cajas_fetch_rows(sb, solo_activas=False)}


def _movimiento_monto_explicito_columnas(row: dict[str, Any]) -> tuple[str | None, float | None]:
    """Columnas nativas de patch_030: la primera con valor > 0 define bucket y monto."""
    for col, bk in (("monto_bs", "VES"), ("monto_usdt", "USDT"), ("monto_usd_caja", "USD")):
        v = row.get(col)
        if v is None or str(v).strip() == "":
            continue
        try:
            x = float(v)
        except (TypeError, ValueError):
            continue
        if x > 0:
            return bk, x
    return None, None


def _movimiento_caja_flow_bucket_amount(
    r: dict[str, Any], caja_by_id: dict[str, dict[str, Any]]
) -> tuple[str, float]:
    """Clave de visualización y monto a sumar.

    - Con **monto_bs** / **monto_usdt** / **monto_usd_caja** (patch_030) o **monto_moneda**:
      bucket **VES** / **USD** / **USDT** (monto nativo).
    - Sin nativo: el número sigue siendo **monto_usd** del movimiento (equiv. interno del ERP),
      pero el bucket distingue **desde qué tipo de cuenta** salió para no mostrarlo como
      «dólares en efectivo» cuando la caja es **VES** o **USDT**:
      **VES_MOTOR_USD** / **USDT_MOTOR_USD** / **USD_equiv** (cuenta USD o caja sin moneda).
    """
    bn, an = _gasto_op_bucket_solo_monto_cargado(r, caja_by_id)
    if bn is not None and an is not None:
        return bn, an
    cid = str(r.get("caja_id") or "")
    cj = caja_by_id.get(cid) or {}
    cmon_raw = str(cj.get("moneda_cuenta") or "").strip().upper()
    cmon = "VES" if cmon_raw in ("VES", "BS") else cmon_raw
    musd = float(r.get("monto_usd") or 0)
    if cmon == "VES":
        return "VES_MOTOR_USD", musd
    if cmon == "USDT":
        return "USDT_MOTOR_USD", musd
    return "USD_equiv", musd


def _flow_ingreso_egreso_por_moneda(
    sb: Client, dsl: str, r_fut: str, caja_by_id: dict[str, dict[str, Any]]
) -> tuple[dict[str, float], dict[str, float]]:
    ing: dict[str, float] = defaultdict(float)
    egr: dict[str, float] = defaultdict(float)
    sel_ext = "tipo,monto_usd,moneda,monto_moneda,monto_bs,monto_usdt,monto_usd_caja,caja_id"
    sel_leg = "tipo,monto_usd,moneda,monto_moneda,caja_id"
    try:
        try:
            m_all = (
                sb.table("movimientos_caja")
                .select(sel_ext)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .execute()
            )
        except Exception:
            m_all = (
                sb.table("movimientos_caja")
                .select(sel_leg)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .execute()
            )
        for r in m_all.data or []:
            k, amt = _movimiento_caja_flow_bucket_amount(r, caja_by_id)
            if r.get("tipo") == "Ingreso":
                ing[k] += amt
            elif r.get("tipo") == "Egreso":
                egr[k] += amt
    except Exception:
        pass
    return dict(ing), dict(egr)


def _gastos_op_por_categoria_multimoneda(
    sb: Client, dsl: str, r_fut: str, caja_by_id: dict[str, dict[str, Any]]
) -> dict[str, dict[str, float]]:
    out: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    sel_ext = "monto_usd,categoria_gasto,moneda,monto_moneda,monto_bs,monto_usdt,monto_usd_caja,caja_id"
    sel_leg = "monto_usd,categoria_gasto,moneda,monto_moneda,caja_id"
    try:
        try:
            mr = (
                sb.table("movimientos_caja")
                .select(sel_ext)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .eq("tipo", "Egreso")
                .execute()
            )
        except Exception:
            mr = (
                sb.table("movimientos_caja")
                .select(sel_leg)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .eq("tipo", "Egreso")
                .execute()
            )
        for row in mr.data or []:
            cg = row.get("categoria_gasto")
            if cg is None or not str(cg).strip():
                continue
            k, amt = _movimiento_caja_flow_bucket_amount(row, caja_by_id)
            out[str(cg).strip()][k] += amt
    except Exception:
        return {}
    return {c: dict(v) for c, v in out.items()}


def _gasto_op_bucket_solo_monto_cargado(
    row: dict[str, Any], caja_by_id: dict[str, dict[str, Any]]
) -> tuple[str | None, float | None]:
    """Monto en moneda de cuenta: columnas explícitas (patch_030) o monto_moneda; sin tasas ni monto_usd."""
    ex_bk, ex_amt = _movimiento_monto_explicito_columnas(row)
    if ex_bk is not None and ex_amt is not None:
        return ex_bk, ex_amt
    mm = row.get("monto_moneda")
    if mm is None or str(mm).strip() == "":
        return None, None
    try:
        mmf = float(mm)
    except (TypeError, ValueError):
        return None, None
    if mmf <= 0:
        return None, None
    cid = str(row.get("caja_id") or "")
    cj = caja_by_id.get(cid) or {}
    cmon_raw = str(cj.get("moneda_cuenta") or "").strip().upper()
    cmon = "VES" if cmon_raw in ("VES", "BS") else cmon_raw
    rmon = str(row.get("moneda") or "").strip().upper()
    bk = rmon or cmon or "USD"
    if bk not in ("VES", "USD", "USDT"):
        bk = "USD"
    return bk, mmf


def _gastos_op_por_categoria_solo_cargado(
    sb: Client, dsl: str, r_fut: str, caja_by_id: dict[str, dict[str, Any]]
) -> dict[str, dict[str, float]]:
    """Gastos con categoría: suma montos en moneda de cuenta (columnas patch_030 o monto_moneda)."""
    out: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    sel_ext = "categoria_gasto,moneda,monto_moneda,monto_bs,monto_usdt,monto_usd_caja,caja_id"
    sel_leg = "categoria_gasto,moneda,monto_moneda,caja_id"
    try:
        try:
            mr = (
                sb.table("movimientos_caja")
                .select(sel_ext)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .eq("tipo", "Egreso")
                .execute()
            )
        except Exception:
            mr = (
                sb.table("movimientos_caja")
                .select(sel_leg)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .eq("tipo", "Egreso")
                .execute()
            )
        for row in mr.data or []:
            cg = row.get("categoria_gasto")
            if cg is None or not str(cg).strip():
                continue
            bk, amt = _gasto_op_bucket_solo_monto_cargado(row, caja_by_id)
            if bk is None or amt is None:
                continue
            out[str(cg).strip()][bk] += amt
    except Exception:
        return {}
    return {c: dict(v) for c, v in out.items()}


def _gastos_op_totales_solo_cargado(
    sb: Client, dsl: str, r_fut: str, caja_by_id: dict[str, dict[str, Any]]
) -> tuple[dict[str, float], int, int]:
    """Totales VES/USD/USDT con montos de cuenta (patch_030 o monto_moneda); conteos ok / sin monto."""
    tot: dict[str, float] = defaultdict(float)
    n_ok = 0
    n_sin = 0
    sel_ext = "categoria_gasto,moneda,monto_moneda,monto_bs,monto_usdt,monto_usd_caja,caja_id"
    sel_leg = "categoria_gasto,moneda,monto_moneda,caja_id"
    try:
        try:
            mr = (
                sb.table("movimientos_caja")
                .select(sel_ext)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .eq("tipo", "Egreso")
                .execute()
            )
        except Exception:
            mr = (
                sb.table("movimientos_caja")
                .select(sel_leg)
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .eq("tipo", "Egreso")
                .execute()
            )
        for row in mr.data or []:
            cg = row.get("categoria_gasto")
            if cg is None or not str(cg).strip():
                continue
            bk, amt = _gasto_op_bucket_solo_monto_cargado(row, caja_by_id)
            if bk is None or amt is None:
                n_sin += 1
                continue
            n_ok += 1
            tot[bk] += amt
    except Exception:
        pass
    return dict(tot), n_ok, n_sin


def _fmt_linea_gastos_solo_cargados(tot: dict[str, float]) -> str:
    parts: list[str] = []
    for bk, pref in (("VES", "Bs"), ("USD", "US$"), ("USDT", "USDT")):
        v = float(tot.get(bk) or 0)
        if abs(v) >= 0.005:
            parts.append(f"{pref} {v:,.2f}")
    return " · ".join(parts) if parts else ""


def _tarjeta_gasto_cat_solo_cargados(cat: str, sub: dict[str, float]) -> None:
    lab = cat if len(cat) <= 44 else cat[:41] + "…"
    main = _fmt_linea_gastos_solo_cargados(sub)
    if not main:
        _dash_mercado_card(lab, "—", foot="Sin monto en moneda guardado en BD")
    else:
        _dash_mercado_card(lab, main, foot="Suma en moneda de cuenta (Bs / USDT / US$ en BD)")


def _fmt_dash_bucket_label(bk: str) -> str:
    return {
        "VES": "Bs",
        "USD": "US$",
        "USDT": "USDT",
        "VES_MOTOR_USD": "Cuenta Bs (VES)",
        "USDT_MOTOR_USD": "Cuenta USDT",
        "USD_equiv": "Equiv. USD",
    }.get(bk, bk)


def _markdown_lineas_flujo_caja(d: dict[str, float]) -> str:
    if not d:
        return "*Sin movimientos.*"
    lines: list[str] = []
    order = ("VES", "USD", "USDT", "VES_MOTOR_USD", "USDT_MOTOR_USD", "USD_equiv")
    for bk in order:
        v = float(d.get(bk) or 0)
        if abs(v) >= 0.005:
            lab = _fmt_dash_bucket_label(bk)
            if bk in ("VES_MOTOR_USD", "USDT_MOTOR_USD"):
                lines.append(f"- **{lab}:** equiv. US$ {_round_money_2(v):,.2f}")
            else:
                lines.append(f"- **{lab}:** {_round_money_2(v):,.2f}")
    for bk, v in sorted(d.items()):
        if bk in order:
            continue
        vf = float(v or 0)
        if abs(vf) >= 0.005:
            lines.append(f"- **{bk}:** {_round_money_2(vf):,.2f}")
    return "\n".join(lines) if lines else "*Sin movimientos.*"


def _gastos_op_totales_por_moneda(by_cat_mm: dict[str, dict[str, float]]) -> dict[str, float]:
    acc: dict[str, float] = defaultdict(float)
    for sub in by_cat_mm.values():
        for bk, v in sub.items():
            acc[str(bk)] += float(v or 0)
    return dict(acc)


def _fmt_multimon_bucket_line(
    d: dict[str, float], *, legacy_suffix: str = "registros viejos"
) -> tuple[str, str | None]:
    """Línea compacta: moneda nativa o moneda de cuenta (VES/USDT/USD) + cifra; pie solo si hay mezcla."""
    parts_vis: list[str] = []
    for bk in ("VES", "USD", "USDT"):
        v = float(d.get(bk) or 0)
        if abs(v) >= 0.005:
            parts_vis.append(f"{_fmt_dash_bucket_label(bk)} {_round_money_2(v):,.2f}")
    v_vm = float(d.get("VES_MOTOR_USD") or 0)
    if abs(v_vm) >= 0.005:
        parts_vis.append(f"VES · equiv. US$ {_round_money_2(v_vm):,.2f}")
    v_ut = float(d.get("USDT_MOTOR_USD") or 0)
    if abs(v_ut) >= 0.005:
        parts_vis.append(f"USDT · equiv. US$ {_round_money_2(v_ut):,.2f}")
    ueq = float(d.get("USD_equiv") or 0)
    if parts_vis:
        foot: str | None = None
        if abs(ueq) >= 0.005:
            foot = f"+ US$ {_round_money_2(ueq):,.2f} ({legacy_suffix})"
        return " · ".join(parts_vis), foot
    main = f"US$ {_round_money_2(ueq):,.2f}" if abs(ueq) >= 0.005 else "—"
    return main, None


def _tarjeta_gasto_cat_multimon(cat: str, sub: dict[str, float]) -> None:
    main, foot = _fmt_multimon_bucket_line(sub, legacy_suffix="registros viejos")
    lab = cat if len(cat) <= 44 else cat[:41] + "…"
    _dash_mercado_card(lab, main, foot=foot or "Gasto operativo")


def render_dashboard_resumen_cuentas_flujo_y_detalle(
    sb: Client,
    d_a: date,
    r_fut: str,
    *,
    bitacora_rows: list[dict[str, Any]] | None = None,
    t: dict[str, Any] | None = None,
) -> None:
    """Tarjetas: saldos por cuenta, bitácora origen→destino, ingresos/egresos, gastos, compras."""
    dsl = f"{d_a.isoformat()}T00:00:00"
    _br = list(bitacora_rows or [])
    _cmap_dash = _caja_map_por_id(sb)

    st.markdown("##### Dónde está el dinero (cuentas activas · saldo hoy)")
    st.caption(
        "Equivalente **USD** por cuenta (**caja / banco / wallet**), ordenado de mayor a menor. "
        "Si en el período registraste **bitácora con movimientos de caja** (con **Aplicar movimientos en caja**), los **Bs** **salieron** del **origen VES** "
        "y el **equivalente** quedó en la **cuenta destino** (**USD** o **USDT**). "
        "Lo que sigue figurando en **VES** es saldo que **no** moviste en esa operación (u otras entradas posteriores)."
    )
    cajas = _cajas_fetch_rows(sb, solo_activas=True)
    with_saldo = [c for c in cajas if float(c.get("saldo_actual_usd") or 0) >= 0.005]
    with_saldo.sort(key=lambda x: -float(x.get("saldo_actual_usd") or 0))
    if not with_saldo:
        st.info("No hay saldos positivos en **cajas activas** (o todo está en cero). Revisá **Cajas y bancos**.")
    else:
        for i in range(0, len(with_saldo), 3):
            chunk = with_saldo[i : i + 3]
            cols = st.columns(3)
            for j in range(3):
                with cols[j]:
                    if j < len(chunk):
                        c = chunk[j]
                        et = _caja_etiqueta_lista(c)
                        if len(et) > 44:
                            et = et[:41] + "…"
                        mon = str(c.get("moneda_cuenta") or "—").strip().upper() or "—"
                        buck = _dash_liquidity_bucket(
                            tipo=str(c.get("tipo") or ""),
                            nombre=str(c.get("nombre") or ""),
                            entidad=str(c.get("entidad") or ""),
                        )
                        _dash_mercado_card(
                            et,
                            f"US$ {_round_money_2(c.get('saldo_actual_usd')):,.2f}",
                            foot=f"{mon} · {buck}",
                        )

    st.markdown("##### Cambios Bs → USD / USDT (bitácora del período)")
    st.caption(
        "Cada fila enlaza **de qué cuenta salieron los bolívares** y **en qué cuenta quedó** el valor en moneda estable. "
        "Con **patch_024** y al guardar con **Aplicar movimientos en caja**, el sistema genera **egreso en VES** e **ingreso en la caja destino**; "
        "los saldos de arriba ya deberían reflejarlo. Si solo anotaste bitácora, los saldos **no** cambian."
    )
    if not _br:
        st.info("No hay registros de **bitácora** en el período **Desde/Hasta**.")
    else:
        _bf = _bitacora_filas_con_cajas(sb, _br, t)
        if _bf:
            st.dataframe(pd.DataFrame(_bf), use_container_width=True, hide_index=True)

    st.markdown("##### Qué entró y qué salió de cajas (en el período)")
    st.caption(
        "Suma de **movimientos de caja** entre **Desde** y **Hasta**. "
        "Cuando el movimiento guarda **moneda de cuenta** (`monto_bs` / `monto_usdt` / `monto_usd_caja` con **`patch_030`**, o `monto_moneda` + moneda), se muestra en **Bs / US$ / USDT**. "
        "Si **no** hay nativo en BD, solo se muestra el **equiv. USD** ya guardado en `monto_usd` al registrar (**el panel no recalcula** Bs ni USDT con tasas del día). "
        "**Entró:** cobros, **ingreso por bitácora** (USD/USDT destino), etc. "
        "**Salió:** compras al contado, **gastos operativos**, **egreso VES por bitácora**, pagos CXP, traspasos, etc. "
        "**No** mostramos **ingresos − egresos** en una sola cifra: en cambios y traspasos el mismo valor figura como egreso en una cuenta e ingreso en otra."
    )
    sum_in = sum_out = 0.0
    try:
        m_all = (
            sb.table("movimientos_caja")
            .select("tipo,monto_usd")
            .gte("created_at", dsl)
            .lte("created_at", r_fut)
            .execute()
        )
        rows_m = m_all.data or []
        sum_in = sum(float(r.get("monto_usd") or 0) for r in rows_m if r.get("tipo") == "Ingreso")
        sum_out = sum(float(r.get("monto_usd") or 0) for r in rows_m if r.get("tipo") == "Egreso")
    except Exception:
        pass
    ing_m, egr_m = _flow_ingreso_egreso_por_moneda(sb, dsl, r_fut, _cmap_dash)
    fx1, fx2 = st.columns(2)
    with fx1:
        st.markdown("**Ingresos a caja**")
        st.markdown(_markdown_lineas_flujo_caja(ing_m), unsafe_allow_html=False)
        st.caption(f"Suma equiv. USD (motor de saldos): **US$ {_round_money_2(sum_in):,.2f}**")
    with fx2:
        st.markdown("**Egresos de caja**")
        st.markdown(_markdown_lineas_flujo_caja(egr_m), unsafe_allow_html=False)
        st.caption(f"Suma equiv. USD (motor de saldos): **US$ {_round_money_2(sum_out):,.2f}**")

    st.markdown("##### Qué se gastó (gastos operativos por categoría)")
    st.caption(
        "Solo egresos con **categoría** (**`patch_025`**). "
        "Aquí se **suman solo los montos guardados en BD** (columnas **`monto_bs` / `monto_usdt` / `monto_usd_caja`** o `monto_moneda`). "
        "**No** se convierte con tasas: si no hay monto en moneda de cuenta, la tarjeta muestra **—** hasta que corrijas el movimiento. "
        "Otros egresos (compras al contado, etc.) no entran."
    )
    by_cat_mm = _gastos_op_por_categoria_multimoneda(sb, dsl, r_fut, _cmap_dash)
    by_cat_solo = _gastos_op_por_categoria_solo_cargado(sb, dsl, r_fut, _cmap_dash)
    by_cat_usd_sort: dict[str, float] = {}
    try:
        mr_s = (
            sb.table("movimientos_caja")
            .select("monto_usd,categoria_gasto")
            .gte("created_at", dsl)
            .lte("created_at", r_fut)
            .eq("tipo", "Egreso")
            .execute()
        )
        for row in mr_s.data or []:
            cg = row.get("categoria_gasto")
            if cg is not None and str(cg).strip():
                k = str(cg).strip()
                by_cat_usd_sort[k] = by_cat_usd_sort.get(k, 0.0) + float(row.get("monto_usd") or 0)
    except Exception:
        by_cat_usd_sort = {}
    if not by_cat_mm:
        st.info(
            "No hay **gastos operativos categorizados** en el período. "
            "Registrálos en **Gastos operativos** o ejecutá **`supabase/patch_025_gastos_operativos.sql`** si falta la columna."
        )
    else:
        items = sorted(by_cat_mm.items(), key=lambda x: -by_cat_usd_sort.get(x[0], 0.0))[:9]
        for i in range(0, len(items), 3):
            chunk = items[i : i + 3]
            cols = st.columns(3)
            for j in range(3):
                with cols[j]:
                    if j < len(chunk):
                        cat, _ = chunk[j]
                        _tarjeta_gasto_cat_solo_cargados(cat, by_cat_solo.get(cat, {}))

    st.markdown("##### Qué se compró (compras a proveedores en el período)")
    st.caption("Suma en **USD** de **compras** registradas en el rango, agrupada por **proveedor** (mercancía / inventario).")
    by_prov: dict[str, float] = {}
    try:
        cp = (
            sb.table("compras")
            .select("proveedor,total_usd")
            .gte("fecha", str(d_a))
            .lte("fecha", r_fut)
            .execute()
        )
        for row in cp.data or []:
            pv = str(row.get("proveedor") or "").strip() or "Sin nombre"
            by_prov[pv] = by_prov.get(pv, 0.0) + float(row.get("total_usd") or 0)
    except Exception:
        by_prov = {}
    if not by_prov:
        st.info("No hay **compras a proveedores** en el período.")
    else:
        items_p = sorted(by_prov.items(), key=lambda x: -x[1])[:9]
        for i in range(0, len(items_p), 3):
            chunk = items_p[i : i + 3]
            cols = st.columns(3)
            for j in range(3):
                with cols[j]:
                    if j < len(chunk):
                        pr, amt = chunk[j]
                        lab = pr if len(pr) <= 44 else pr[:41] + "…"
                        _dash_mercado_card(lab, f"US$ {_round_money_2(amt):,.2f}", foot="Compra mercancía")


def _pdf_resumen_ejecutivo_bytes(
    sb: Client,
    t: dict[str, Any] | None,
    d_a: date,
    d_b: date,
    r_fut: str,
    *,
    ventas_usd: float,
    margen_bruto_usd: float,
    unidades_stock: float,
    n_sku: int,
    liquidez_total_usd: float,
    compras_periodo_usd: float,
    gastos_op_periodo_usd: float,
    total_salidas_usd: float,
    bitacora_rows: list[dict[str, Any]],
) -> bytes:
    """PDF imprimible del resumen ejecutivo (A4, márgenes ~18 mm)."""
    from copy import deepcopy
    from xml.sax.saxutils import escape as xml_esc

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    dsl = f"{d_a.isoformat()}T00:00:00"
    tz_v = ZoneInfo("America/Caracas")
    gen_txt = datetime.now(tz_v).strftime("%d/%m/%Y %H:%M")

    cajas = _cajas_fetch_rows(sb, solo_activas=True)
    with_saldo = sorted(
        [c for c in cajas if float(c.get("saldo_actual_usd") or 0) >= 0.005],
        key=lambda x: -float(x.get("saldo_actual_usd") or 0),
    )
    rows_cuentas: list[list[str]] = []
    for c in with_saldo:
        et = _caja_etiqueta_lista(c)
        if len(et) > 52:
            et = et[:49] + "…"
        mon = str(c.get("moneda_cuenta") or "—").strip().upper() or "—"
        buck = _dash_liquidity_bucket(
            tipo=str(c.get("tipo") or ""),
            nombre=str(c.get("nombre") or ""),
            entidad=str(c.get("entidad") or ""),
        )
        rows_cuentas.append(
            [
                xml_esc(et),
                xml_esc(mon),
                xml_esc(buck),
                f"{_round_money_2(c.get('saldo_actual_usd')):,.2f}",
            ]
        )

    cmap_pdf = _caja_map_por_id(sb)
    ing_m_pdf, egr_m_pdf = _flow_ingreso_egreso_por_moneda(sb, dsl, r_fut, cmap_pdf)
    by_cat_solo_pdf = _gastos_op_por_categoria_solo_cargado(sb, dsl, r_fut, cmap_pdf)
    go_pdf_tot, _, _ = _gastos_op_totales_solo_cargado(sb, dsl, r_fut, cmap_pdf)
    _gop_pdf_txt = _fmt_linea_gastos_solo_cargados(go_pdf_tot) or "—"

    sum_in = sum_out = 0.0
    try:
        m_all = (
            sb.table("movimientos_caja")
            .select("tipo,monto_usd")
            .gte("created_at", dsl)
            .lte("created_at", r_fut)
            .execute()
        )
        for r in m_all.data or []:
            if r.get("tipo") == "Ingreso":
                sum_in += float(r.get("monto_usd") or 0)
            elif r.get("tipo") == "Egreso":
                sum_out += float(r.get("monto_usd") or 0)
    except Exception:
        pass

    def _fmt_pdf_flow_lines(d: dict[str, float]) -> str:
        parts: list[str] = []
        for bk in ("VES", "USD", "USDT", "VES_MOTOR_USD", "USDT_MOTOR_USD", "USD_equiv"):
            v = float(d.get(bk) or 0)
            if abs(v) >= 0.005:
                parts.append(f"{_fmt_dash_bucket_label(bk)} {_round_money_2(v):,.2f}")
        return " · ".join(parts) if parts else "—"

    by_cat: dict[str, float] = {}
    try:
        mr = (
            sb.table("movimientos_caja")
            .select("monto_usd,categoria_gasto")
            .gte("created_at", dsl)
            .lte("created_at", r_fut)
            .eq("tipo", "Egreso")
            .execute()
        )
        for row in mr.data or []:
            cg = row.get("categoria_gasto")
            if cg is not None and str(cg).strip():
                k = str(cg).strip()
                by_cat[k] = by_cat.get(k, 0.0) + float(row.get("monto_usd") or 0)
    except Exception:
        pass
    rows_gasto: list[list[str]] = []
    for cat, usdtot in sorted(by_cat.items(), key=lambda x: -x[1])[:20]:
        sub_solo = by_cat_solo_pdf.get(cat, {})
        det = _fmt_linea_gastos_solo_cargados(sub_solo)
        if not det:
            det = "— (sin monto en moneda en BD)"
        rows_gasto.append([xml_esc(cat), xml_esc(det)])

    by_prov: dict[str, float] = {}
    try:
        cp = (
            sb.table("compras")
            .select("proveedor,total_usd")
            .gte("fecha", str(d_a))
            .lte("fecha", r_fut)
            .execute()
        )
        for row in cp.data or []:
            pv = str(row.get("proveedor") or "").strip() or "Sin nombre"
            by_prov[pv] = by_prov.get(pv, 0.0) + float(row.get("total_usd") or 0)
    except Exception:
        pass
    rows_prov = [[xml_esc(a), f"{_round_money_2(v):,.2f}"] for a, v in sorted(by_prov.items(), key=lambda x: -x[1])[:20]]

    live = get_live_exchange_rates()
    ves = live.get("ves_bs_por_usd")
    p2p = live.get("usdt_x_ves_p2p") or live.get("p2p_bs_por_usdt_aprox")
    linea_mercado = "Sin datos web (revisá conexión)."
    if live.get("ok") and ves is not None:
        linea_mercado = f"Bs/USD web ref.: {float(ves):,.2f}"
        if p2p is not None:
            linea_mercado += f" · Bs/USDT: {float(p2p):,.4f}"

    sum_bs_bt = sum(float(r.get("monto_ves") or 0) for r in bitacora_rows)
    sum_usd_bt = sum(float(r.get("monto_usd_obtenido") or 0) for r in bitacora_rows)

    buf = BytesIO()
    lm = rm = 18 * mm
    tm = 14 * mm
    bm = 18 * mm
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=lm, rightMargin=rm, topMargin=tm, bottomMargin=bm)
    styles = getSampleStyleSheet()
    story: list[Any] = []
    tw = float(A4[0] - lm - rm)

    if BRAND_LOGO_PATH.is_file():
        try:
            ir = ImageReader(str(BRAND_LOGO_PATH))
            iw, ih = ir.getSize()
            if iw > 0 and ih > 0:
                target_w = 40 * mm
                target_h = target_w * (float(ih) / float(iw))
                max_h = 20 * mm
                if target_h > max_h:
                    target_h = max_h
                    target_w = target_h * (float(iw) / float(ih))
                lg = RLImage(str(BRAND_LOGO_PATH), width=target_w, height=target_h)
                lg.hAlign = "CENTER"
                story.append(lg)
                story.append(Spacer(1, 2 * mm))
        except Exception:
            pass

    tit = deepcopy(styles["Title"])
    tit.fontSize = 15
    tit.leading = 17
    tit.textColor = colors.HexColor("#1a1f2e")
    tit.alignment = TA_CENTER
    story.append(Paragraph(xml_esc("Resumen ejecutivo"), tit))
    meta = deepcopy(styles["Normal"])
    meta.fontSize = 9
    meta.alignment = TA_CENTER
    story.append(
        Paragraph(
            xml_esc(f"Movi Motor's Importadora · Período: {d_a.isoformat()} → {d_b.isoformat()} · Generado: {gen_txt} (Caracas)"),
            meta,
        )
    )
    story.append(Paragraph(xml_esc(linea_mercado), meta))
    story.append(Spacer(1, 3 * mm))

    h_style = ParagraphStyle(name="rh", parent=styles["Heading2"], fontSize=11, textColor=colors.HexColor("#1a1f2e"), spaceAfter=4)
    # Celdas con texto largo en PDF: usar Paragraph para que haga wrap (evita solapamiento entre columnas).
    cell_bit_l = ParagraphStyle(
        name="cbit_l",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
    )
    cell_bit_r = ParagraphStyle(
        name="cbit_r",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        alignment=TA_RIGHT,
    )

    def _tbl(data: list[list[str]], hdr: list[str], wcols: list[float]) -> Table:
        h_esc = [[xml_esc(x) for x in hdr]]
        body = h_esc + data
        t = Table(body, colWidths=wcols, repeatRows=1)
        stl: list[tuple[Any, ...]] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eaf0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1a1f2e")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cfd4dc")),
        ]
        if len(body) > 1:
            stl.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fb")]))
        t.setStyle(TableStyle(stl))
        return t

    story.append(Paragraph(xml_esc("Indicadores del período"), h_style))
    kpi_data = [
        ["Ventas totales (USD)", f"{ventas_usd:,.2f}"],
        ["Margen bruto (USD)", f"{margen_bruto_usd:,.2f}"],
        ["Unidades en stock", f"{unidades_stock:,.0f}"],
        ["SKU activos", str(int(n_sku))],
        ["Liquidez total (USD equiv.)", f"{liquidez_total_usd:,.2f}"],
        ["Compras a proveedores (USD)", f"{compras_periodo_usd:,.2f}"],
        ["Gastos operativos (suma montos en moneda en BD)", _gop_pdf_txt],
        ["Total salidas motor (compras USD + Σ gastos monto_usd)", f"{total_salidas_usd:,.2f}"],
    ]
    story.append(
        _tbl([[xml_esc(a), xml_esc(str(b))] for a, b in kpi_data], ["Concepto", "Valor"], [tw * 0.62, tw * 0.38])
    )
    story.append(Spacer(1, 4 * mm))

    if t:
        story.append(
            Paragraph(
                xml_esc(
                    f"Tasas en BD (documentos): Bs/USD {float(t['tasa_bs']):,.2f} · USDT/USD {float(t['tasa_usdt']):,.6f}"
                ),
                meta,
            )
        )
        story.append(Spacer(1, 2 * mm))

    story.append(Paragraph(xml_esc("Dónde está el dinero (cuentas activas)"), h_style))
    if not rows_cuentas:
        story.append(Paragraph(xml_esc("Sin saldos positivos en cajas activas."), meta))
    else:
        story.append(
            _tbl(
                rows_cuentas,
                ["Cuenta", "Moneda", "Origen / tipo", "USD equiv."],
                [tw * 0.36, tw * 0.1, tw * 0.3, tw * 0.24],
            )
        )
    story.append(Spacer(1, 2 * mm))
    story.append(
        Paragraph(
            xml_esc(
                "Saldos VES: es lo que sigue en bolívares en esa cuenta. Si guardaste la bitácora con movimientos de caja, "
                "los Bs de la tabla siguiente ya egresaron del origen y el equivalente ingresó en la cuenta destino (USD o USDT)."
            ),
            meta,
        )
    )
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(xml_esc("Bitácora: salida de bolívares e ingreso en moneda estable"), h_style))
    bf_pdf = _bitacora_filas_con_cajas(sb, list(bitacora_rows or []), t)
    if not bf_pdf:
        story.append(Paragraph(xml_esc("Sin operaciones de bitácora en el período."), meta))
    else:
        wcols_bt = [tw * 0.09, tw * 0.21, tw * 0.21, tw * 0.07, tw * 0.11, tw * 0.11, tw * 0.20]
        hdr_bt = ["Fecha", "Origen VES", "Destino", "Mon.", "Bs", "USD eq.", "Detalle"]
        rows_bt_pdf: list[list[Any]] = [[xml_esc(x) for x in hdr_bt]]
        for r in bf_pdf:
            det = str(r.get("Qué representa") or "")
            rows_bt_pdf.append(
                [
                    Paragraph(xml_esc(str(r.get("Fecha") or "")), cell_bit_l),
                    Paragraph(xml_esc(str(r.get("Salió de (bolívares)") or "")), cell_bit_l),
                    Paragraph(xml_esc(str(r.get("Llegó a (destino)") or "")), cell_bit_l),
                    Paragraph(xml_esc(str(r.get("Moneda destino") or "")), cell_bit_l),
                    Paragraph(xml_esc(f"{float(r['Bs usados']):,.2f}"), cell_bit_r),
                    Paragraph(xml_esc(f"{float(r['USD equivalente']):,.2f}"), cell_bit_r),
                    Paragraph(xml_esc(det), cell_bit_l),
                ]
            )
        t_bt = Table(rows_bt_pdf, colWidths=wcols_bt, repeatRows=1)
        stl_bt: list[tuple[Any, ...]] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eaf0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1a1f2e")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cfd4dc")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        if len(rows_bt_pdf) > 1:
            stl_bt.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fb")]))
        t_bt.setStyle(TableStyle(stl_bt))
        story.append(t_bt)
        story.append(
            Paragraph(
                xml_esc(
                    f"Totales bitácora (período): {len(bf_pdf)} operación(es) · Bs usados: {sum_bs_bt:,.2f} · USD equiv. pactado: {_round_money_2(sum_usd_bt):,.2f}"
                ),
                meta,
            )
        )
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(xml_esc("Movimientos de caja en el período"), h_style))
    story.append(
        Paragraph(
            xml_esc(
                "Incluye cobros, pagos y, si aplica, el par egreso VES + ingreso USD/USDT generado al guardar la bitácora con movimientos. "
                "No se incluye una fila de neto global (ingresos − egresos): en traspasos y bitácora el mismo valor sale de una caja y entra en otra."
            ),
            meta,
        )
    )
    story.append(Spacer(1, 1 * mm))
    mov_rows = [
        ["Ingresos (por moneda registrada)", _fmt_pdf_flow_lines(ing_m_pdf)],
        ["Ingresos (suma equiv. USD motor)", f"{_round_money_2(sum_in):,.2f}"],
        ["Egresos (por moneda registrada)", _fmt_pdf_flow_lines(egr_m_pdf)],
        ["Egresos (suma equiv. USD motor)", f"{_round_money_2(sum_out):,.2f}"],
    ]
    story.append(_tbl([[xml_esc(a), b] for a, b in mov_rows], ["Concepto", "Detalle"], [tw * 0.42, tw * 0.58]))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(xml_esc("Gastos operativos por categoría"), h_style))
    if not rows_gasto:
        story.append(Paragraph(xml_esc("Sin gastos con categoría en el período (ver módulo Gastos operativos / patch 025)."), meta))
    else:
        story.append(_tbl(rows_gasto, ["Categoría", "Importes"], [tw * 0.38, tw * 0.62]))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(xml_esc("Compras por proveedor"), h_style))
    if not rows_prov:
        story.append(Paragraph(xml_esc("Sin compras en el período."), meta))
    else:
        story.append(_tbl(rows_prov, ["Proveedor", "USD"], [tw * 0.62, tw * 0.38]))
    story.append(Spacer(1, 4 * mm))

    foot = deepcopy(styles["Normal"])
    foot.fontSize = 7.5
    foot.textColor = colors.HexColor("#666666")
    story.append(Spacer(1, 5 * mm))
    story.append(
        Paragraph(
            xml_esc(
                "Referencia de mercado web con caché ~2 min. Gastos por categoría requieren categoria_gasto en movimientos. "
                "Cifras orientativas; no sustituyen asesoría contable."
            ),
            foot,
        )
    )

    doc.build(story)
    return buf.getvalue()


def _dash_semaforo(*, stock: float, minimo: float, vendido_periodo: float) -> str:
    """Rojo / amarillo / verde para priorizar liquidación o reposición."""
    if stock <= minimo:
        return "🔴 Bajo mínimo"
    if stock <= minimo * 1.25:
        return "🟡 Cerca del mínimo"
    if vendido_periodo < 0.001 and stock > minimo:
        return "🟡 Baja rotación"
    return "🟢 OK"


def _ingreso_cobro_native_y_equiv(r: dict[str, Any]) -> tuple[str, float, float]:
    """(moneda bucket, monto en esa moneda, monto_usd equiv.). Prioriza monto_bs / monto_usdt / monto_usd_caja."""
    mu = float(r.get("monto_usd") or 0)
    mon_d = str(r.get("moneda") or "USD").strip().upper()

    def _pos(col: str) -> float | None:
        v = r.get(col)
        if v is None or str(v).strip() == "":
            return None
        try:
            x = float(v)
        except (TypeError, ValueError):
            return None
        return x if x > 0 else None

    b = _pos("monto_bs")
    if b is not None:
        return "VES", b, mu
    ut = _pos("monto_usdt")
    if ut is not None:
        return "USDT", ut, mu
    uc = _pos("monto_usd_caja")
    if uc is not None:
        if mon_d == "ZELLE":
            return "ZELLE", uc, mu
        if mon_d in ("VES", "USDT"):
            return mon_d, uc, mu
        return "USD", uc, mu

    mm = r.get("monto_moneda")
    native = mu
    if mm is not None and str(mm).strip() != "":
        try:
            native = float(mm)
        except (TypeError, ValueError):
            native = mu
    return mon_d, native, mu


def _dashboard_resumen_cobros_por_moneda(sb: Client, *, d_a: date, r_fut: str) -> None:
    """Ingresos a caja en el período: totales VES / USD / USDT y detalle por cuenta."""
    dsl = d_a.isoformat()
    sel_ext = (
        "caja_id, monto_usd, moneda, monto_moneda, monto_bs, monto_usdt, monto_usd_caja, "
        "concepto, nota_operacion, created_at"
    )
    sel_mid = "caja_id, monto_usd, moneda, monto_moneda, concepto, nota_operacion, created_at"
    sel_leg = "caja_id, monto_usd, moneda, monto_moneda, concepto, created_at"
    try:
        try:
            mh = (
                sb.table("movimientos_caja")
                .select(sel_ext)
                .eq("tipo", "Ingreso")
                .gte("created_at", dsl)
                .lte("created_at", r_fut)
                .execute()
            )
        except Exception:
            try:
                mh = (
                    sb.table("movimientos_caja")
                    .select(sel_mid)
                    .eq("tipo", "Ingreso")
                    .gte("created_at", dsl)
                    .lte("created_at", r_fut)
                    .execute()
                )
            except Exception:
                mh = (
                    sb.table("movimientos_caja")
                    .select(sel_leg)
                    .eq("tipo", "Ingreso")
                    .gte("created_at", dsl)
                    .lte("created_at", r_fut)
                    .execute()
                )
    except Exception:
        st.caption(
            "Para ver cobros en **Bs / USD / USDT** por caja, ejecutá en Supabase "
            "`supabase/patch_008_movimientos_moneda_cobros.sql`."
        )
        return

    rows = mh.data or []
    if not rows:
        st.caption("Sin ingresos de caja en el período.")
        return

    caj_map = {str(c["id"]): _caja_etiqueta_lista(c) for c in _cajas_fetch_rows(sb, solo_activas=False)}

    recs: list[dict[str, Any]] = []
    tot_ves = tot_usdt = tot_usd_cash = tot_zelle = 0.0
    sum_equiv_usd = 0.0
    for r in rows:
        mon, native, mu = _ingreso_cobro_native_y_equiv(r)
        sum_equiv_usd += mu
        if mon == "VES":
            tot_ves += native
        elif mon == "USDT":
            tot_usdt += native
        elif mon == "ZELLE":
            tot_zelle += native
        elif mon == "USD":
            tot_usd_cash += native
        else:
            tot_usd_cash += native
        recs.append(
            {
                "Fecha": str(r.get("created_at", ""))[:19],
                "Caja": caj_map.get(str(r.get("caja_id")), str(r.get("caja_id"))),
                "Moneda": mon,
                "Monto moneda": _round_money_2(native),
                "Equiv. USD": _round_money_2(mu),
                "Concepto": (r.get("concepto") or "")[:80],
                "Nota tesorería": str(r.get("nota_operacion") or "")[:100],
            }
        )

    m1, m2, m3, m4, m5 = st.columns(5)
    with m1:
        st.metric("Cobrado VES (Bs)", f"{_round_money_2(tot_ves):,.2f}")
    with m2:
        st.metric("Cobrado USDT", f"{_round_money_2(tot_usdt):,.2f}")
    with m3:
        st.metric("Cobrado USD", f"US$ {_round_money_2(tot_usd_cash):,.2f}")
    with m4:
        st.metric("Zelle (USD)", f"US$ {_round_money_2(tot_zelle):,.2f}")
    with m5:
        st.metric("Ingresos (equiv. USD)", f"US$ {_round_money_2(sum_equiv_usd):,.2f}")

    st.caption(
        "Equiv. USD es lo que suma al **saldo de cada caja** (VES y USDT convertidos con la tasa de la venta / cobro)."
    )
    st.dataframe(
        pd.DataFrame(recs).sort_values("Fecha", ascending=False),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Monto moneda": st.column_config.NumberColumn(format="%.2f"),
            "Equiv. USD": st.column_config.NumberColumn(format="%.2f"),
        },
    )


def _cambios_tesoreria_en_rango(sb: Client, d_a: date, r_fut: str) -> list[dict[str, Any]]:
    """Filas de `cambios_tesoreria` (bitácora Bs→USD/USDT) en [d_a 00:00, r_fut]."""
    try:
        q = (
            sb.table("cambios_tesoreria")
            .select("*")
            .gte("fecha", f"{d_a.isoformat()}T00:00:00")
            .lte("fecha", r_fut)
            .order("fecha", desc=True)
            .execute()
        )
        return q.data or []
    except Exception:
        return []


def _dashboard_seccion_cambios_tesoreria(
    sb: Client,
    *,
    t: dict[str, Any] | None,
    d_a: date,
    d_b: date,
    r_fut: str,
    rows_raw: list[dict[str, Any]] | None = None,
    key_suffix: str = "",
) -> None:
    """Bitácora Bs→USD/USDT: precio pactado + comparación opcional vs BCV/mercado."""
    erp_uid = str(st.session_state.get("erp_uid") or "").strip()
    st.markdown('<div class="dash-bento">', unsafe_allow_html=True)
    st.markdown("##### Bitácora: Bs → USD / USDT")
    st.caption(
        "Registrás **cuántos Bs** usaste en el cambio y **qué equivalente en USD** tomás como referencia en el sistema. "
        "Si **marcás** «Aplicar movimientos en caja», el ERP registra **egreso de bolívares** en la cuenta **VES** origen e **ingreso** del equivalente en la cuenta destino: "
        "esos Bs **dejan de figurar** en bolívares y el valor pasa a **USD**, **USDT** o **efectivo en dólares** según la caja destino (**Zelle** suele ser una caja en **USD**). "
        "La comparación con otra tasa (BCV/mercado) sigue siendo opcional."
    )

    if rows_raw is not None:
        raw_rows = list(rows_raw)
    else:
        try:
            cambios_q = (
                sb.table("cambios_tesoreria")
                .select("*")
                .gte("fecha", f"{d_a.isoformat()}T00:00:00")
                .lte("fecha", r_fut)
                .order("fecha", desc=True)
                .execute()
            )
            raw_rows = cambios_q.data or []
        except Exception as ex:
            st.info(
                f"Ejecutá en Supabase **`supabase/patch_018_cambios_tesoreria.sql`** y luego **`supabase/patch_019_cambios_tesoreria_tasa_compra.sql`**. Detalle: {ex}"
            )
            st.markdown("</div>", unsafe_allow_html=True)
            return

    rows_caj = _cajas_fetch_rows(sb, solo_activas=False)
    id_to_et = {str(r["id"]): _caja_etiqueta_lista(r) for r in rows_caj}
    ves_ids = [str(c["id"]) for c in rows_caj if str(c.get("moneda_cuenta") or "").strip().upper() == "VES"]
    stab_ids = [str(c["id"]) for c in rows_caj if str(c.get("moneda_cuenta") or "").strip().upper() in ("USD", "USDT")]

    id_to_moneda = {str(c["id"]): str(c.get("moneda_cuenta") or "").strip().upper() or "—" for c in rows_caj}
    t_usdt_ref = float(_nf(t.get("tasa_usdt")) or 1.0) if t else 1.0
    if t_usdt_ref <= 0:
        t_usdt_ref = 1.0

    recs: list[dict[str, Any]] = []
    diffs_comp: list[float] = []
    for r in raw_rows:
        m_ves = float(r.get("monto_ves") or 0)
        m_usd = float(r.get("monto_usd_obtenido") or 0)
        tc_raw = r.get("tasa_compra_bs_por_usd")
        t_compra = float(tc_raw) if tc_raw is not None else ((m_ves / m_usd) if m_usd > 0 else 0.0)
        tr_raw = r.get("tasa_referencia_bs_por_usd")
        t_comp = float(tr_raw) if tr_raw is not None and float(tr_raw) > 0 else None
        usd_a_comp = (m_ves / t_comp) if t_comp and t_comp > 0 else None
        diff_usd = (m_usd - usd_a_comp) if usd_a_comp is not None else None
        if diff_usd is not None:
            diffs_comp.append(float(diff_usd))
        oid = r.get("caja_origen_id")
        did = r.get("caja_destino_id")
        _md = id_to_moneda.get(str(did), "—") if did else "—"
        _usdt_ref_dia = m_usd * t_usdt_ref
        recs.append(
            {
                "Fecha": r.get("fecha"),
                "Origen": id_to_et.get(str(oid), "—") if oid else "—",
                "Destino": id_to_et.get(str(did), "—") if did else "—",
                "Moneda dest.": _md,
                "Bs usados": _round_money_2(m_ves),
                "Equiv. USD (pactado)": _round_money_2(m_usd),
                "≈ USDT (ref. día)": _round_money_2(_usdt_ref_dia),
                "Precio compra (Bs/USD)": round(t_compra, 4) if t_compra else None,
                "Comparar con (Bs/USD)": round(t_comp, 4) if t_comp else None,
                "USD a esa comparación": _round_money_2(usd_a_comp) if usd_a_comp is not None else None,
                "Diff vs comparación (USD)": _round_money_2(diff_usd) if diff_usd is not None else None,
                "Nota": (r.get("nota") or "")[:120],
            }
        )

    tot_bs_cambios = sum(float(r.get("Bs usados") or 0) for r in recs)
    tot_usd_pact = sum(float(r.get("Equiv. USD (pactado)") or 0) for r in recs)
    tot_usdt_ref = sum(float(r.get("≈ USDT (ref. día)") or 0) for r in recs)

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("Operaciones en bitácora", f"{len(recs)}")
    with m2:
        st.metric("Bs pasados a cambio (suma)", f"{tot_bs_cambios:,.2f}")
    with m3:
        st.metric("Equiv. USD obtenido (suma pactada)", f"{_round_money_2(tot_usd_pact):,.2f}")
    with m4:
        st.metric(
            "≈ USDT ref. día (suma)",
            f"{_round_money_2(tot_usdt_ref):,.2f}",
            help="Equiv. USD × tasa USDT/USD del día operativo (seguimiento; no es saldo on-chain).",
        )

    m5, m6, m7 = st.columns(3)
    with m5:
        tot_diff = sum(diffs_comp) if diffs_comp else None
        st.metric(
            "Suma diff vs comparación (USD)",
            f"{_round_money_2(tot_diff):+,.2f}" if tot_diff is not None else "—",
            help="Solo suma filas donde cargaste *Comparar con*. Positivo = mejor que esa tasa.",
        )
    with m6:
        st.metric(
            "Saldo equiv. VES en cajas (ahora)",
            f"US$ {sum(float(c.get('saldo_actual_usd') or 0) for c in rows_caj if c.get('activo') and str(c.get('moneda_cuenta') or '').strip().upper() == 'VES'):,.2f}",
            help="Solo el **remanente** en cuentas **VES** después de ventas, egresos y cambios. Lo que ya pasaste a bitácora **con movimientos de caja** **ya salió** de este saldo.",
        )
    with m7:
        st.caption(
            "**Seguimiento:** en la tabla, los Bs de cada fila son los que **destinaste al cambio**; con movimientos aplicados, ese monto **ya no está** en la caja en bolívares "
            "y el equivalente quedó en la cuenta **destino** (USD, USDT o dólares físicos según la caja)."
        )

    if recs:
        df_c = pd.DataFrame(recs)
        df_c["Fecha"] = pd.to_datetime(df_c["Fecha"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M")
        st.dataframe(
            df_c,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Bs usados": st.column_config.NumberColumn(format="%.2f"),
                "Equiv. USD (pactado)": st.column_config.NumberColumn(format="%.2f"),
                "≈ USDT (ref. día)": st.column_config.NumberColumn(format="%.2f"),
                "Precio compra (Bs/USD)": st.column_config.NumberColumn(format="%.2f"),
                "Comparar con (Bs/USD)": st.column_config.NumberColumn(format="%.2f"),
                "USD a esa comparación": st.column_config.NumberColumn(format="%.2f"),
                "Diff vs comparación (USD)": st.column_config.NumberColumn(format="%.2f"),
            },
        )
    else:
        st.caption("No hay registros de cambio en el rango de fechas del panel.")

    def_tasa = 0.0
    if t:
        def_tasa = (
            float(_nf(t.get("bcv_bs_por_usd")) or 0)
            or float(_nf(t.get("tasa_bs")) or 0)
            or float(_nf(t.get("paralelo_bs_por_usd")) or 0)
        )
    if def_tasa <= 0:
        liv = get_live_exchange_rates()
        def_tasa = float(_nf(liv.get("ves_bs_por_usd")) or 0)
    if def_tasa <= 0:
        def_tasa = 1.0

    if not erp_uid:
        st.warning("Sesión sin usuario ERP; no se puede registrar cambio.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    _open_cambio = bool(st.session_state.pop("dash_open_cambio_tesoreria", False))
    _ks = (key_suffix or "").strip()
    if _ks:
        _ks = "_" + _ks
    with st.expander("Registrar cambio (bitácora)", expanded=_open_cambio):
        with st.form(
            f"f_cambio_tesoreria_dash{_ks}_{int(st.session_state.get('dash_cambio_teso_form_nonce', 0))}"
        ):
            opt_none = "__none__"
            opt_o = [opt_none] + ves_ids
            opt_d = [opt_none] + stab_ids

            def _fmt_caja(cid: str) -> str:
                if cid == opt_none:
                    return "Sin especificar"
                return id_to_et.get(cid, cid)

            so = st.selectbox("Caja origen (VES)", options=opt_o, format_func=_fmt_caja, key=f"dash_ct_orig{_ks}")
            sd = st.selectbox("Caja destino (USD/USDT)", options=opt_d, format_func=_fmt_caja, key=f"dash_ct_dest{_ks}")

            _row_o = next((c for c in rows_caj if str(c.get("id")) == str(so)), None) if so != opt_none else None
            _saldo_o_usd = float(_nf((_row_o or {}).get("saldo_actual_usd")) or 0.0)
            _saldo_o_bs_est = _saldo_o_usd * float(def_tasa)
            if _row_o is not None:
                st.caption(
                    f"Saldo disponible en origen: **US$ {_round_money_2(_saldo_o_usd):,.2f}** "
                    f"(≈ **{_round_money_2(_saldo_o_bs_est):,.2f} Bs** usando {def_tasa:,.4f} Bs/USD como referencia)."
                )

                # Prefill: sugerir convertir el 100% del saldo disponible.
                if f"dash_ct_musd{_ks}" not in st.session_state:
                    st.session_state[f"dash_ct_musd{_ks}"] = max(0.0001, float(_saldo_o_usd))
                if f"dash_ct_mves{_ks}" not in st.session_state:
                    st.session_state[f"dash_ct_mves{_ks}"] = max(0.0001, float(_saldo_o_bs_est))

            m_ves_in = st.number_input(
                "Monto en bolívares (Bs) que usaste en el cambio",
                min_value=0.0001,
                format="%.4f",
                key=f"dash_ct_mves{_ks}",
            )
            m_usd_in = st.number_input(
                "Equivalente obtenido (USD de referencia en el sistema)",
                min_value=0.0001,
                format="%.4f",
                key=f"dash_ct_musd{_ks}",
                help="Si compraste **USDT**, cargá el equivalente en **USD** que uses para valorar (o el monto en USDT si lo tratás 1:1 con USD en esta bitácora).",
            )
            implied = (float(m_ves_in) / float(m_usd_in)) if float(m_usd_in) > 1e-12 else 0.0
            if implied > 0:
                st.caption(f"Precio **implícito** Bs/USD según montos: **{implied:,.4f}** (podés usarlo abajo si es el pactado).")
            t_compra_in = st.number_input(
                "A qué precio compraste (Bs por 1 USD)",
                min_value=0.00000001,
                value=float(implied) if implied > 0 else 1.0,
                format="%.6f",
                key=f"dash_ct_tcompra{_ks}",
                help="Lo que acordaste con ese cambista: cuántos Bs te cobraron por cada dólar.",
            )
            t_comp_in = st.number_input(
                "Comparar con (Bs por 1 USD, opcional — 0 = no comparar)",
                min_value=0.0,
                value=0.0,
                format="%.6f",
                key=f"dash_ct_tcomp{_ks}",
                help="Ej.: BCV u otra cotización. Si es 0, no se calcula diff vs ‘mercado’.",
            )
            if float(t_comp_in) <= 0 and def_tasa > 0:
                st.caption(f"Sugerido para comparar (BCV/ref. guardada): **{def_tasa:,.4f}** — copiá si querés medir vs eso.")
            aplicar_mov = st.checkbox(
                "Aplicar movimientos en caja (egreso Bs en origen + ingreso en cuenta destino)",
                value=True,
                key=f"dash_ct_aplicar_mov{_ks}",
                help="**Marcado (recomendado):** se registran egreso en la cuenta **VES** e ingreso en **USD/USDT** según la caja destino, y se actualizan saldos. "
                "**Desmarcado:** solo queda la anotación en la bitácora (sin movimientos de caja). Si ya tenías registros «solo anotación», ejecutá de nuevo con esta opción marcada y cajas origen/destino.",
            )
            nota_in = st.text_input("Nota (opcional)", key=f"dash_ct_nota{_ks}")
            if st.form_submit_button("Guardar registro"):
                try:
                    if float(m_usd_in) > 1e-12:
                        usd_impl = float(m_ves_in) / float(t_compra_in)
                        if abs(usd_impl - float(m_usd_in)) > max(0.02, float(m_usd_in) * 0.002):
                            st.warning(
                                f"El precio compra ({t_compra_in:,.4f}) no cuadra del todo con Bs/USD ingresados "
                                f"(implicaría ~{usd_impl:,.4f} USD). Revisá montos o la tasa pactada."
                            )
                    if aplicar_mov and (so == opt_none or sd == opt_none):
                        st.error("Para mover dinero entre cuentas elegí **caja origen (VES)** y **caja destino (USD/USDT)**.")
                    elif aplicar_mov and _row_o is not None and float(m_usd_in) - _saldo_o_usd > 0.00001:
                        st.error(
                            "Saldo insuficiente en la caja origen: el **equivalente USD** a egresar supera el saldo disponible. "
                            f"Disponible: **US$ {_round_money_2(_saldo_o_usd):,.2f}** · Intentás egresar: **US$ {_round_money_2(float(m_usd_in)) :,.2f}**."
                        )
                    else:
                        payload_rpc: dict[str, Any] = {
                            "p_usuario_id": erp_uid,
                            "p_caja_origen_id": None if so == opt_none else str(so),
                            "p_caja_destino_id": None if sd == opt_none else str(sd),
                            "p_monto_ves": float(m_ves_in),
                            "p_monto_usd_obtenido": float(m_usd_in),
                            "p_tasa_compra_bs_por_usd": float(t_compra_in),
                            "p_aplicar_movimientos": aplicar_mov,
                        }
                        if float(t_comp_in) > 0:
                            payload_rpc["p_tasa_comparacion_bs_por_usd"] = float(t_comp_in)
                        nn = (nota_in or "").strip()
                        if nn:
                            payload_rpc["p_nota"] = nn
                        _dm_dest = ""
                        if aplicar_mov:
                            _dm_dest = next(
                                (
                                    str(c.get("moneda_cuenta") or "").strip().upper()
                                    for c in rows_caj
                                    if str(c.get("id")) == str(sd)
                                ),
                                "",
                            )
                            if _dm_dest == "USDT":
                                _t_ut = float(_nf(t.get("tasa_usdt")) or 0) if t else 0.0
                                if _t_ut <= 0:
                                    _lt = latest_tasas(sb) or {}
                                    _t_ut = float(_nf(_lt.get("tasa_usdt")) or 0)
                                if _t_ut <= 0:
                                    st.error("Falta **tasa USDT/USD** (cargá tasas del día en **Tasas**).")
                                    st.stop()
                                payload_rpc["p_tasa_usdt"] = _t_ut
                        _resp_ct = sb.rpc("registrar_cambio_tesoreria_erp", payload_rpc).execute()
                        _cambio_id = _rpc_resp_uuid(_resp_ct)
                        if aplicar_mov and _cambio_id:
                            try:
                                _mv_chk = (
                                    sb.table("movimientos_caja")
                                    .select("id")
                                    .eq("referencia", str(_cambio_id).strip())
                                    .execute()
                                )
                                _n_mv = len(_mv_chk.data or [])
                                if _n_mv < 2:
                                    st.error(
                                        "La bitácora se guardó, pero **no se crearon los movimientos de caja** (se esperan egreso VES + ingreso en destino). "
                                        "En Supabase SQL Editor ejecutá **`supabase/patch_024_cambio_tesoreria_movimientos.sql`** (redefine la función `registrar_cambio_tesoreria_erp` para insertar en `movimientos_caja`). "
                                        "Luego podés registrar un nuevo cambio con esta opción marcada."
                                    )
                                    st.stop()
                            except Exception as ex_mv:
                                st.warning(
                                    f"No se pudo comprobar movimientos de caja automáticamente ({ex_mv}). "
                                    "Revisá **Caja** o ejecutá **patch_024** si los saldos no se actualizan."
                                )
                        st.session_state["dash_bitacora_flash"] = {
                            "aplicar_mov": bool(aplicar_mov),
                            "monto_bs": float(m_ves_in),
                            "monto_usd": float(m_usd_in),
                            "dest_moneda": _dm_dest or None,
                            "dest_etiqueta": id_to_et.get(str(sd), "") if sd != opt_none else "",
                        }
                        _movi_bump_form_nonce("dash_cambio_teso_form_nonce")
                        _movi_ss_pop_keys(
                            f"dash_ct_orig{_ks}",
                            f"dash_ct_dest{_ks}",
                            f"dash_ct_mves{_ks}",
                            f"dash_ct_musd{_ks}",
                            f"dash_ct_tcompra{_ks}",
                            f"dash_ct_tcomp{_ks}",
                            f"dash_ct_nota{_ks}",
                            f"dash_ct_aplicar_mov{_ks}",
                        )
                        st.rerun()
                except Exception as ex:
                    st.error(str(ex))

    st.markdown("</div>", unsafe_allow_html=True)


def _dashboard_kpis_periodo(sb: Client, d_a: date, d_b: date) -> dict[str, Any]:
    """Ventas, margen, stock, liquidez, compras, gastos op. y bitácora en [d_a, d_b]."""
    return compute_dashboard_kpis_periodo(
        sb,
        d_a,
        d_b,
        cambios_tesoreria_en_rango=_cambios_tesoreria_en_rango,
        caja_map_por_id=_caja_map_por_id,
        gastos_op_totales_solo_cargado=_gastos_op_totales_solo_cargado,
        fmt_linea_gastos_solo_cargados=_fmt_linea_gastos_solo_cargados,
    )


def panel_resumen_ejecutivo_periodo_ui(
    sb: Client,
    t: dict[str, Any] | None,
    d_a: date,
    d_b: date,
    k: dict[str, Any],
    *,
    pdf_download_key: str,
) -> None:
    """Contenido del resumen ejecutivo (antes en Dashboard): KPIs, PDF, cuentas, gráficos."""
    r_fut = str(k["r_fut"])
    rows_cambios_bitacora = k["rows_cambios_bitacora"]
    ventas_usd = float(k["ventas_usd"])
    ventas_prev = float(k["ventas_prev"])
    margen_usd = float(k["margen_usd"])
    margen_prev = float(k["margen_prev"])
    unidades_stock = float(k["unidades_stock"])
    n_sku = int(k["n_sku"])
    liquidez = float(k["liquidez"])
    compras_period_usd = float(k["compras_period_usd"])
    gastos_op_period_usd = float(k["gastos_op_period_usd"])
    total_salidas_op_usd = float(k["total_salidas_op_usd"])
    go_kpi_main = k["go_kpi_main"]
    go_kpi_sub = k["go_kpi_sub"]

    render_dashboard_mercado_live_tarjetas(t)
    try:
        _pdf_re_sum = _pdf_resumen_ejecutivo_bytes(
            sb,
            t,
            d_a,
            d_b,
            r_fut,
            ventas_usd=ventas_usd,
            margen_bruto_usd=margen_usd,
            unidades_stock=unidades_stock,
            n_sku=int(n_sku),
            liquidez_total_usd=liquidez,
            compras_periodo_usd=compras_period_usd,
            gastos_op_periodo_usd=gastos_op_period_usd,
            total_salidas_usd=total_salidas_op_usd,
            bitacora_rows=list(rows_cambios_bitacora or []),
        )
        _ts_res = datetime.now(ZoneInfo("America/Caracas")).strftime("%Y%m%d_%H%M")
        st.download_button(
            label=f"Descargar PDF — resumen ejecutivo ({_ts_res})",
            data=_pdf_re_sum,
            file_name=f"resumen_ejecutivo_{_ts_res}.pdf",
            mime="application/pdf",
            key=pdf_download_key,
            help="Incluye KPIs, cuentas, movimientos, gastos por categoría, compras por proveedor y bitácora. Listo para imprimir.",
        )
    except Exception as ex_pdf:
        st.caption(f"No se pudo preparar el PDF del resumen ({ex_pdf}). ¿Tenés **reportlab** instalado?")
    st.divider()
    st.markdown("##### Caja y cambios (bitácora)")
    st.caption(
        "Vista gerencial: saldos por cuenta y registro de cambios VES→USD/USDT. "
        "Usá **Registrar cambio ahora** en el aviso de VES del **Dashboard** para abrir el formulario prellenado."
    )
    _dashboard_seccion_cambios_tesoreria(
        sb, t=t, d_a=d_a, d_b=d_b, r_fut=r_fut, rows_raw=rows_cambios_bitacora, key_suffix="rep_re"
    )
    st.divider()
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        _dash_kpi_card(
            "Ventas totales (USD)",
            f"US$ {ventas_usd:,.2f}",
            _dash_trend_pct(ventas_usd, ventas_prev),
            f"Período {d_a} → {d_b}",
        )
    with k2:
        _dash_kpi_card(
            "Margen bruto (USD)",
            f"US$ {margen_usd:,.2f}",
            _dash_trend_pct(margen_usd, margen_prev),
            "Sobre costo de productos vendidos",
        )
    with k3:
        _dash_kpi_card(
            "Unidades en stock",
            f"{unidades_stock:,.0f}",
            None,
            f"{n_sku} SKU activos",
        )
    with k4:
        _dash_kpi_card(
            "Liquidez total",
            f"US$ {liquidez:,.2f}",
            None,
            "Saldos en cajas / bancos / wallets",
        )
    st.divider()
    render_dashboard_resumen_cuentas_flujo_y_detalle(sb, d_a, r_fut, bitacora_rows=rows_cambios_bitacora, t=t)
    st.divider()
    st.markdown("##### Compras, gastos operativos y flujo (mismo período)")
    st.caption(
        "**Gastos operativos:** suma **solo** montos en moneda de cuenta en BD (**`monto_bs` / `monto_usdt` / `monto_usd_caja`** con **`patch_030`**, o `monto_moneda`) — **sin** convertir con tasas. "
        "Si un gasto no tiene ese dato, **no** entra en el total hasta que lo corrijas. **Compras** siguen en USD de documentos. Requiere **`patch_025`**."
    )
    s1, s2, s3 = st.columns(3)
    with s1:
        _dash_kpi_card(
            "Compras a proveedores (USD)",
            f"US$ {compras_period_usd:,.2f}",
            None,
            "Mercancía en el período",
        )
    with s2:
        _dash_kpi_card("Gastos operativos (lo cargado en BD)", go_kpi_main, None, go_kpi_sub)
    with s3:
        _total_kpi_val = f"Compras (docs): US$ {compras_period_usd:,.2f} · Gastos op.: {go_kpi_main}"
        _dash_kpi_card(
            "Compras + gastos op. (periodo)",
            _total_kpi_val,
            None,
            "Compras en USD; gastos operativos = suma por moneda de lo guardado en BD (columnas nativas o `monto_moneda`).",
        )
    st.markdown("##### Seguimiento: Bs → USD / USDT (bitácora de tesorería)")
    if rows_cambios_bitacora:
        _sum_bs = sum(float(r.get("monto_ves") or 0) for r in rows_cambios_bitacora)
        _sum_usd = sum(float(r.get("monto_usd_obtenido") or 0) for r in rows_cambios_bitacora)
        _t_ut_d = float(_nf(t.get("tasa_usdt")) or 1.0) if t else 1.0
        if _t_ut_d <= 0:
            _t_ut_d = 1.0
        _sum_usdt_ref = _sum_usd * _t_ut_d
        st.success(
            f"En este período registraste **{len(rows_cambios_bitacora)}** cambio(s): **{_sum_bs:,.2f} Bs** "
            f"que **destinaste al cambio** a moneda más estable (equiv. **US$ {_round_money_2(_sum_usd):,.2f}** al pactado; **≈ {_round_money_2(_sum_usdt_ref):,.2f} USDT** ref. "
            f"con tasa USDT/USD **{_t_ut_d:,.6f}** del día). "
            "Si esos registros llevan **movimientos de caja**, esos bolívares **ya no están** en la cuenta VES: el valor quedó en **USD**, **USDT** o dólares físicos según la caja destino. "
            "Detalle y tabla en **Caja, cobros y tasas** (Dashboard)."
        )
        b1, b2, b3, b4 = st.columns(4)
        with b1:
            st.metric("Operaciones (bitácora)", len(rows_cambios_bitacora))
        with b2:
            st.metric("Bs en cambios (suma)", f"{_sum_bs:,.2f}")
        with b3:
            st.metric("Equiv. USD pactado (suma)", f"{_round_money_2(_sum_usd):,.2f}")
        with b4:
            st.metric("≈ USDT ref. día (suma)", f"{_round_money_2(_sum_usdt_ref):,.2f}")
        st.caption(
            "**¿Por qué abajo no ves el cambio en el gráfico de barras?** Los bloques **Origen de la liquidez** y **Compras por categoría** "
            "no leen la bitácora: el primero muestra **saldos actuales** (los Bs ya convertidos **bajan** en VES y **suben** en USD/USDT si aplicaste movimientos); el segundo, **compras a proveedores**. "
            "El gráfico de abajo resume solo los **equiv. USD** anotados en **bitácora** en el período."
        )
        _df_bt = pd.DataFrame(
            [
                {
                    "Operación": f"#{i + 1} · {str(r.get('fecha') or '')[:16]}",
                    "Equiv. USD (pactado)": _round_money_2(r.get("monto_usd_obtenido")),
                    "Bs usados": _round_money_2(r.get("monto_ves")),
                }
                for i, r in enumerate(rows_cambios_bitacora)
            ]
        )
        fig_bt = px.bar(
            _df_bt,
            x="Operación",
            y="Equiv. USD (pactado)",
            hover_data={"Bs usados": ":,.2f"},
            labels={"Equiv. USD (pactado)": "Equiv. USD (bitácora)", "Operación": ""},
        )
        fig_bt.update_traces(texttemplate="%{y:,.2f} USD", textposition="outside")
        _plotly_apply_dash_theme(fig_bt, title="Tus cambios Bs → moneda estable (equiv. USD en bitácora)")
        fig_bt.update_layout(height=300, showlegend=False)
        st.plotly_chart(fig_bt, use_container_width=True)
    else:
        st.info(
            "No hay **registros de bitácora** (Bs → USD / USDT / efectivo dólar) en el rango **Desde/Hasta**. "
            "Si pasaste Bs a **USDT**, **Zelle** (caja USD) o **dólares en efectivo**, registrá el cambio en **Dashboard → Caja, cobros y tasas → Registrar cambio (bitácora)** "
            "para que quede claro que esos bolívares **salieron** de VES y el valor quedó en la cuenta destino (con movimientos de caja si corresponde)."
        )
    row2a, row2b = st.columns([1.15, 1])
    with row2a:
        st.markdown('<div class="dash-bento">', unsafe_allow_html=True)
        st.markdown("**Origen de la liquidez** (USD por tipo de caja)")
        _caj_liq = _cajas_fetch_rows(sb, solo_activas=True)
        caj_df = pd.DataFrame(
            [
                {
                    "nombre": r.get("nombre"),
                    "tipo": r.get("tipo"),
                    "saldo_actual_usd": r.get("saldo_actual_usd"),
                    "entidad": r.get("entidad") or "",
                }
                for r in _caj_liq
            ]
        )
        if caj_df.empty:
            st.caption("No hay cajas activas.")
        else:
            caj_df["origen"] = caj_df.apply(
                lambda r: _dash_liquidity_bucket(
                    tipo=str(r.get("tipo", "")),
                    nombre=str(r.get("nombre", "")),
                    entidad=str(r.get("entidad", "") or ""),
                ),
                axis=1,
            )
            caj_df["saldo_actual_usd"] = pd.to_numeric(caj_df["saldo_actual_usd"], errors="coerce").fillna(0)
            agg_l = caj_df.groupby("origen", as_index=False)["saldo_actual_usd"].sum()
            order = ["Caja fuerte", "Bancos nacionales", "Binance (crypto)", "Otros"]
            agg_l["origen"] = pd.Categorical(agg_l["origen"], categories=order, ordered=True)
            agg_l = agg_l.sort_values("origen")
            agg_l = agg_l[agg_l["saldo_actual_usd"] > 0]
            if agg_l.empty:
                st.info(
                    "**Sin barras:** no hay **saldo positivo** en cajas activas para graficar (o todo está en cero). "
                    "Si registraste un cambio con **movimientos de caja**, los saldos ya deberían reflejar egreso de Bs e ingreso en USD/USDT."
                )
            else:
                colors = ["#00e5ff", "#ff9100", "#b388ff", "#78909c"]
                fig_liq = go.Figure()
                for idx, (_, row) in enumerate(agg_l.iterrows()):
                    fig_liq.add_trace(
                        go.Bar(
                            name=str(row["origen"]),
                            x=["Liquidez"],
                            y=[float(row["saldo_actual_usd"])],
                            marker_color=colors[idx % len(colors)],
                            text=[f"{float(row['saldo_actual_usd']):,.2f}"],
                            textposition="inside",
                            hovertemplate="%{fullData.name}: %{y:,.2f} USD<extra></extra>",
                        )
                    )
                fig_liq.update_layout(
                    barmode="stack",
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="top", y=-0.2, x=0),
                )
                _plotly_apply_dash_theme(fig_liq, title="Composición por origen")
                fig_liq.update_layout(hoverlabel=dict(bgcolor="#1a1f2e", font_size=12), height=320)
                st.plotly_chart(fig_liq, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
    with row2b:
        st.markdown('<div class="dash-bento">', unsafe_allow_html=True)
        st.markdown("**Compras por categoría** (USD en el período · proxy de gasto / abastecimiento)")
        cids = [
            str(x["id"])
            for x in (
                sb.table("compras")
                .select("id")
                .gte("fecha", str(d_a))
                .lte("fecha", r_fut)
                .execute()
                .data
                or []
            )
        ]
        if not cids:
            st.info(
                "**Sin compras en el rango:** no registraste **compras a proveedores** entre las fechas elegidas. "
                "Eso es independiente de la **bitácora** (cambio Bs→USD/USDT)."
            )
        else:
            cdet = sb.table("compras_detalles").select("producto_id, subtotal_usd").in_("compra_id", cids).execute()
            cdf = pd.DataFrame(cdet.data or [])
            if cdf.empty:
                st.caption("Sin líneas de compra.")
            else:
                plist = sb.table("productos").select("id, categoria_id, categorias(nombre)").execute().data or []
                id_cat: dict[str, str] = {}
                for p in plist:
                    cid = str(p["id"])
                    cat = p.get("categorias")
                    if isinstance(cat, list) and cat:
                        cat = cat[0]
                    if isinstance(cat, dict) and cat.get("nombre"):
                        id_cat[cid] = str(cat["nombre"])
                    else:
                        id_cat[cid] = "Sin categoría"
                cdf["categoria"] = cdf["producto_id"].astype(str).map(lambda x: id_cat.get(x, "Sin categoría"))
                cdf["subtotal_usd"] = pd.to_numeric(cdf["subtotal_usd"], errors="coerce").fillna(0)
                gcat = cdf.groupby("categoria", as_index=False)["subtotal_usd"].sum()
                fig_d = px.pie(
                    gcat,
                    names="categoria",
                    values="subtotal_usd",
                    hole=0.52,
                    color_discrete_sequence=px.colors.sequential.Teal_r,
                )
                fig_d.update_traces(
                    textposition="inside", textinfo="percent+label", hovertemplate="%{label}<br>%{value:,.2f} USD<extra></extra>"
                )
                _plotly_apply_dash_theme(fig_d, title="Distribución (donut)")
                st.plotly_chart(fig_d, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)


def module_dashboard(sb: Client, t: dict[str, Any] | None) -> None:
    """Panel ejecutivo estilo Bento (dark + acentos cian/naranja). Streamlit + Plotly."""
    h1, h2, h3 = st.columns([2.2, 2.2, 1.6])
    with h1:
        st.markdown('<p class="dash-header-title">Panel Movi Motors</p>', unsafe_allow_html=True)
        st.markdown(
            '<p class="dash-header-sub">ERP automotriz · multimoneda · vista ejecutiva</p>',
            unsafe_allow_html=True,
        )
    with h2:
        d_a = st.date_input("Desde", value=date.today() - timedelta(days=30), key="dash_d0")
        d_b = st.date_input("Hasta", value=date.today(), key="dash_d1")
    with h3:
        q_search = st.text_input(
            "Buscar", placeholder="Producto, código…", key="dash_global_search", label_visibility="visible"
        )

    if d_b < d_a:
        st.error("La fecha *Hasta* debe ser ≥ *Desde*.")
        st.stop()

    _bit_flash = st.session_state.pop("dash_bitacora_flash", None)
    if isinstance(_bit_flash, dict):
        if _bit_flash.get("aplicar_mov"):
            _dm_f = str(_bit_flash.get("dest_moneda") or "").strip().upper()
            if _dm_f == "USDT":
                _dest_txt = "El equivalente quedó en **USDT** (moneda estable en la caja destino)."
            elif _dm_f == "USD":
                _dest_txt = (
                    "El equivalente quedó en **USD** en la cuenta destino (banco, efectivo en dólares o **Zelle**, según cómo tengas nombrada esa caja)."
                )
            else:
                _dest_txt = "El equivalente quedó en la **cuenta destino** en moneda estable."
            st.success(
                f"**Bitácora guardada con movimientos en caja:** salieron **{_bit_flash.get('monto_bs', 0):,.2f} Bs** del saldo en bolívares; "
                f"ingresó **~US$ {float(_bit_flash.get('monto_usd') or 0):,.2f}** de equivalente en destino. {_dest_txt} "
                "Ese monto en **VES** ya no debería figurar como bolívares en caja; si algo no cuadra, revisá **Caja, cobros y tasas**."
            )
        else:
            st.info(
                "**Bitácora guardada solo como anotación:** no se movieron saldos entre cajas. "
                "Para que los Bs **salgan** de la cuenta VES y el equivalente **entre** a USD/USDT/efectivo dólar, "
                "marcá **Aplicar movimientos en caja**, elegí origen y destino, y guardá de nuevo."
            )

    if st.session_state.get("dash_open_cambio_tesoreria"):
        st.info("Abriendo formulario de cambio (bitácora)…")

    _caj_ves_alert = _cajas_fetch_rows(sb, solo_activas=True)
    sum_ves_equiv_usd = sum(
        float(c.get("saldo_actual_usd") or 0)
        for c in _caj_ves_alert
        if str(c.get("moneda_cuenta") or "").strip().upper() == "VES"
    )
    # Aviso solo si el remanente VES (equiv. USD) supera este monto (evita ruido por centavos / ~US$ 1).
    _ves_remanente_alerta_min_usd = 20.0
    if sum_ves_equiv_usd > _ves_remanente_alerta_min_usd:
        _ves_cajas = [
            c
            for c in _caj_ves_alert
            if str(c.get("moneda_cuenta") or "").strip().upper() == "VES" and float(c.get("saldo_actual_usd") or 0) > 0.00001
        ]
        _ves_cajas.sort(key=lambda x: -float(x.get("saldo_actual_usd") or 0))
        _ves_top = _ves_cajas[0] if _ves_cajas else None

        st.warning(
            f"**Todavía hay equivalente en cuentas en bolívares (VES):** suman ~**US$ {sum_ves_equiv_usd:,.2f}** en el sistema — es lo que **aún está** en caja/banco en Bs. "
            "Si **registraste un cambio con movimientos de caja**, lo que cambiaste **ya salió** de ahí: el valor pasó a la cuenta destino en **USD**, **USDT** o **efectivo en dólares** "
            "(**Zelle** suele registrarse en una caja en **USD**). Este aviso solo refleja **lo que sigue** en VES, no lo ya convertido."
        )
        c1, c2 = st.columns([1.4, 2.6])
        with c1:
            if st.button("Registrar cambio ahora", use_container_width=True, key="dash_cta_registrar_cambio"):
                _tasa_pref = 1.0
                if t:
                    _tasa_pref = float(_nf(t.get("bcv_bs_por_usd")) or 0) or float(_nf(t.get("tasa_bs")) or 0) or float(
                        _nf(t.get("paralelo_bs_por_usd")) or 0
                    ) or 1.0
                if _ves_top is not None:
                    # Prefill para ambas instancias del formulario (Resumen y Caja).
                    for _ks in ("_res", "_caja"):
                        st.session_state[f"dash_ct_orig{_ks}"] = str(_ves_top.get("id"))
                        st.session_state[f"dash_ct_musd{_ks}"] = float(_ves_top.get("saldo_actual_usd") or 0.0)
                        st.session_state[f"dash_ct_mves{_ks}"] = float(_ves_top.get("saldo_actual_usd") or 0.0) * float(
                            _tasa_pref
                        )
                    # Nota: no conocemos Bs exactos; sugerimos Bs ref. usando la tasa del panel.
                st.session_state["dash_open_cambio_tesoreria"] = True
                st.session_state["dash_cta_ack"] = (st.session_state.get("dash_cta_ack") or 0) + 1
                st.rerun()
        with c2:
            st.caption(
                "Recomendación operativa: convertí el remanente en VES a una cuenta estable (USD/USDT/Zelle) y registralo en la bitácora para que los saldos del panel reflejen la realidad."
            )

    k_dash = _dashboard_kpis_periodo(sb, d_a, d_b)
    r_fut = k_dash["r_fut"]
    dsl_mov = k_dash["dsl_mov"]
    rows_cambios_bitacora = k_dash["rows_cambios_bitacora"]
    ventas_usd = k_dash["ventas_usd"]
    ventas_prev = k_dash["ventas_prev"]
    margen_usd = k_dash["margen_usd"]
    margen_prev = k_dash["margen_prev"]
    vids = k_dash["vids"]
    unidades_stock = k_dash["unidades_stock"]
    n_sku = k_dash["n_sku"]
    liquidez = k_dash["liquidez"]
    compras_period_usd = k_dash["compras_period_usd"]
    gastos_op_period_usd = k_dash["gastos_op_period_usd"]
    total_salidas_op_usd = k_dash["total_salidas_op_usd"]
    n_gastos_op_movs = k_dash["n_gastos_op_movs"]
    go_tot_carg = k_dash["go_tot_carg"]
    n_go_moneda = k_dash["n_go_moneda"]
    n_go_sin_moneda = k_dash["n_go_sin_moneda"]
    go_kpi_main = k_dash["go_kpi_main"]
    go_kpi_sub = k_dash["go_kpi_sub"]

    with st.expander("Información del panel", expanded=False, key="modinfo_exp_dashboard"):
        st.markdown(
            "**Cómo recorrer el dashboard:** 1) Elegí **Desde / Hasta** arriba. 2) Pestaña **Mercado en vivo** → cotizaciones. "
            "El **resumen ejecutivo completo** (KPIs, PDF, cuentas, gráficos) está en **Reportes → Resumen ejecutivo**. "
            "3) **Inventario y stock** → semáforo y valor (el **Buscar** de arriba filtra la tabla). 4) **Caja, cobros y tasas** → flujo, cobros por moneda, bitácora, tasas y últimos movimientos. "
            "**Usuarios** del ERP están en **Mantenimiento** (superusuario)."
        )
    tab_d_mercado, tab_d_inv, tab_d_caja = st.tabs(
        [
            "Mercado en vivo",
            "Inventario y stock",
            "Caja, cobros y tasas",
        ]
    )

    with tab_d_mercado:
        render_dashboard_mercado_live_tarjetas(t)
        st.info(
            "**Resumen ejecutivo** (KPIs, PDF imprimible, cuentas, flujo y gráficos del período): abrí el módulo **Reportes** "
            "y la pestaña **Resumen ejecutivo**. Allí elegís **Desde / Hasta** propios del reporte."
        )

    pinv = (
        sb.table("productos")
        .select("id, codigo, descripcion, stock_actual, stock_minimo, costo_usd, precio_v_usd, categorias(nombre)")
        .eq("activo", True)
        .execute()
        .data
        or []
    )

    def _dash_inv_sem_prio(s: str) -> int:
        if str(s).startswith("🔴"):
            return 0
        if str(s).startswith("🟡"):
            return 1
        return 2

    ventas_qty_dash: dict[str, float] = {}
    if vids:
        dq_dash = (
            sb.table("ventas_detalles")
            .select("producto_id, cantidad")
            .in_("venta_id", vids)
            .execute()
            .data
            or []
        )
        for r in dq_dash:
            pid = str(r["producto_id"])
            ventas_qty_dash[pid] = ventas_qty_dash.get(pid, 0) + float(r.get("cantidad") or 0)
    rows_inv_dash: list[dict[str, Any]] = []
    for p in pinv:
        pid = str(p["id"])
        st_a = float(p.get("stock_actual") or 0)
        st_m = float(p.get("stock_minimo") or 0)
        if st_a <= 0:
            continue
        vq = ventas_qty_dash.get(pid, 0.0)
        cat = p.get("categorias")
        if isinstance(cat, list) and cat:
            cat = cat[0]
        cat_n = cat.get("nombre") if isinstance(cat, dict) else "—"
        rows_inv_dash.append(
            {
                "Semáforo": _dash_semaforo(stock=st_a, minimo=st_m, vendido_periodo=vq),
                "Producto": str(p.get("descripcion", ""))[:80],
                "Código": str(p.get("codigo") or "—"),
                "Categoría": str(cat_n or "—")[:40],
                "Stock": st_a,
                "Mín.": st_m,
                "Vendido (período)": vq,
                "Valor inv. USD": round(st_a * float(p.get("costo_usd") or 0), 2),
            }
        )
    dfi_inv = pd.DataFrame(rows_inv_dash)
    if not dfi_inv.empty:
        dfi_inv["_prio"] = dfi_inv["Semáforo"].map(_dash_inv_sem_prio)
        dfi_inv = dfi_inv.sort_values(["_prio", "Valor inv. USD"], ascending=[True, False]).drop(columns=["_prio"])

    with tab_d_inv:
        st.caption(
            "El campo **Buscar** del encabezado filtra la tabla de la izquierda. La columna *Vendido (período)* usa las mismas fechas **Desde/Hasta** que arriba."
        )
        row3a, row3b = st.columns([1.2, 1])
        with row3a:
            st.markdown('<div class="dash-bento">', unsafe_allow_html=True)
            st.markdown("**Productos con baja rotación** · semáforo de inventario")
            st.caption("🟢 OK · 🟡 revisar · 🔴 bajo mínimo")
            if dfi_inv.empty:
                st.info("No hay stock positivo para analizar.")
            else:
                dfi_show = dfi_inv.copy()
                if q_search and q_search.strip():
                    qs = q_search.strip().lower()
                    mask = dfi_show["Producto"].str.lower().str.contains(qs, na=False) | dfi_show[
                        "Código"
                    ].str.lower().str.contains(qs, na=False)
                    dfi_show = dfi_show[mask]
                if dfi_show.empty:
                    st.info("Ningún producto coincide con la búsqueda. Probá otro texto o vaciá el buscador.")
                else:
                    st.dataframe(dfi_show, use_container_width=True, hide_index=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with row3b:
            st.markdown('<div class="dash-bento">', unsafe_allow_html=True)
            st.markdown("**Valor de inventario por categoría** (costo × stock)")
            if not pinv:
                st.caption("Sin productos.")
            else:
                rows_cat: list[dict[str, Any]] = []
                for p in pinv:
                    cat = p.get("categorias")
                    if isinstance(cat, list) and cat:
                        cat = cat[0]
                    cn = str(cat.get("nombre")) if isinstance(cat, dict) and cat.get("nombre") else "Sin categoría"
                    rows_cat.append(
                        {
                            "categoria": cn,
                            "valor_usd": float(p.get("stock_actual") or 0) * float(p.get("costo_usd") or 0),
                        }
                    )
                dfc2 = pd.DataFrame(rows_cat).groupby("categoria", as_index=False)["valor_usd"].sum()
                dfc2 = dfc2[dfc2["valor_usd"] > 0]
                if dfc2.empty:
                    st.caption("Sin valor de inventario (costos en cero o sin stock).")
                else:
                    fig_iv = px.bar(
                        dfc2.sort_values("valor_usd", ascending=True),
                        x="valor_usd",
                        y="categoria",
                        orientation="h",
                        labels={"valor_usd": "USD", "categoria": ""},
                    )
                    fig_iv.update_traces(marker_color="#00e5ff", hovertemplate="%{y}: %{x:,.2f} USD<extra></extra>")
                    _plotly_apply_dash_theme(fig_iv, title="Inventario por categoría")
                    st.plotly_chart(fig_iv, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

    with tab_d_caja:
        st.caption("Movimientos del período, cobros por moneda, comparación ventas/compras y **tasas** (expandibles abajo).")
        st.divider()
        st.markdown("##### Flujo de caja y ventas (detalle)")
        dsl = d_a.isoformat()
        mh = (
            sb.table("movimientos_caja")
            .select("created_at, tipo, monto_usd")
            .gte("created_at", dsl)
            .execute()
        )
        dfm = pd.DataFrame(mh.data or [])
        if dfm.empty:
            st.caption("Sin movimientos de caja en el rango seleccionado.")
        else:
            ts = pd.to_datetime(dfm["created_at"], errors="coerce")
            dfm["dia"] = ts.dt.strftime("%Y-%m-%d")
            dfm["monto_usd"] = pd.to_numeric(dfm["monto_usd"], errors="coerce").fillna(0)
            ing = dfm[dfm["tipo"] == "Ingreso"].groupby("dia", as_index=False)["monto_usd"].sum().rename(columns={"monto_usd": "Ingreso"})
            egr = dfm[dfm["tipo"] == "Egreso"].groupby("dia", as_index=False)["monto_usd"].sum().rename(columns={"monto_usd": "Egreso"})
            merged_ie = pd.merge(
                pd.DataFrame({"dia": sorted(dfm["dia"].unique())}),
                ing,
                on="dia",
                how="left",
            )
            merged_ie = pd.merge(merged_ie, egr, on="dia", how="left").fillna(0)
            merged_ie["Ingreso"] = pd.to_numeric(merged_ie["Ingreso"], errors="coerce").fillna(0).round(2)
            merged_ie["Egreso"] = pd.to_numeric(merged_ie["Egreso"], errors="coerce").fillna(0).round(2)
            fig_ie = px.bar(
                merged_ie,
                x="dia",
                y=["Ingreso", "Egreso"],
                barmode="group",
                labels={"value": "USD", "dia": "Día", "variable": ""},
            )
            fig_ie.update_traces(
                marker_line_width=0,
                hovertemplate="<b>%{data.name}</b><br>%{x}<br>%{y:,.2f} USD<extra></extra>",
            )
            _plotly_apply_dash_theme(fig_ie, title="Ingresos y egresos por día")
            st.plotly_chart(fig_ie, use_container_width=True)

        _dashboard_seccion_cambios_tesoreria(
            sb, t=t, d_a=d_a, d_b=d_b, r_fut=r_fut, rows_raw=rows_cambios_bitacora, key_suffix="caja"
        )

        st.markdown("##### Resumen: qué entró en Bs, USD y USDT (y en qué cuenta)")
        _dashboard_resumen_cobros_por_moneda(sb, d_a=d_a, r_fut=r_fut)

        vrows = sb.table("ventas").select("fecha, total_usd").gte("fecha", str(d_a)).lte("fecha", r_fut).execute()
        crows = sb.table("compras").select("fecha, total_usd").gte("fecha", str(d_a)).lte("fecha", r_fut).execute()
        dfv = pd.DataFrame(vrows.data or [])
        dfc_v = pd.DataFrame(crows.data or [])
        if dfv.empty and dfc_v.empty:
            st.caption("Sin ventas ni compras en el rango.")
        else:
            vsum = (
                dfv.assign(
                    dia=pd.to_datetime(dfv["fecha"], errors="coerce").dt.strftime("%Y-%m-%d"),
                    total_usd=pd.to_numeric(dfv["total_usd"], errors="coerce").fillna(0),
                )
                .groupby("dia", as_index=False)["total_usd"]
                .sum()
                .rename(columns={"total_usd": "Ventas USD"})
                if not dfv.empty
                else pd.DataFrame(columns=["dia", "Ventas USD"])
            )
            csum = (
                dfc_v.assign(
                    dia=pd.to_datetime(dfc_v["fecha"], errors="coerce").dt.strftime("%Y-%m-%d"),
                    total_usd=pd.to_numeric(dfc_v["total_usd"], errors="coerce").fillna(0),
                )
                .groupby("dia", as_index=False)["total_usd"]
                .sum()
                .rename(columns={"total_usd": "Compras USD"})
                if not dfc_v.empty
                else pd.DataFrame(columns=["dia", "Compras USD"])
            )
            dias = sorted(
                set(vsum["dia"].tolist() if not vsum.empty else []) | set(csum["dia"].tolist() if not csum.empty else [])
            )
            out_vc = pd.DataFrame({"dia": dias})
            out_vc = out_vc.merge(vsum, on="dia", how="left").merge(csum, on="dia", how="left").fillna(0)
            for _col in ("Ventas USD", "Compras USD"):
                if _col in out_vc.columns:
                    out_vc[_col] = pd.to_numeric(out_vc[_col], errors="coerce").fillna(0.0)

            out_vc_long = out_vc.melt(
                id_vars=["dia"],
                value_vars=[c for c in ("Ventas USD", "Compras USD") if c in out_vc.columns],
                var_name="serie",
                value_name="value",
            )
            fig_vc = px.line(
                out_vc_long,
                x="dia",
                y="value",
                color="serie",
                markers=True,
                labels={"value": "USD", "dia": "Día", "serie": ""},
            )
            fig_vc.update_traces(line=dict(width=2))
            _plotly_apply_dash_theme(fig_vc, title="Ventas vs compras (USD)")
            st.plotly_chart(fig_vc, use_container_width=True)

        with st.expander("Tasas en vivo y tabla guardada (BCV · ref. mercado / P2P Binance)", expanded=False):
            render_tasas_tiempo_real(key_suffix="dash", t_guardado=t)
            if t:
                st.caption(f"Registro tasas **{t.get('fecha', '—')}**")
                render_tabla_tasas_ui(build_tasas_tabla_detalle(t))
            else:
                st.warning("Sin tasas del día en base de datos.")

        rol_dash = str(st.session_state.get("erp_rol", ""))
        if role_can(rol_dash, "tasas"):
            with st.expander("Cargar / editar tasas (BCV, ref. P2P Binance Bs/USD, USDT P2P)", expanded=False):
                module_tasas(sb, embedded=True)

        st.divider()
        st.markdown("##### Últimos movimientos de caja")
        try:
            mov = (
                sb.table("movimientos_caja")
                .select("created_at, tipo, monto_usd, concepto, caja_id, moneda, nota_operacion")
                .order("created_at", desc=True)
                .limit(15)
                .execute()
            )
        except Exception:
            mov = (
                sb.table("movimientos_caja")
                .select("created_at, tipo, monto_usd, concepto, caja_id")
                .order("created_at", desc=True)
                .limit(15)
                .execute()
            )
        if mov.data:
            _mc_cajas = _cajas_fetch_rows(sb, solo_activas=False)
            _mc_map = {str(c["id"]): _caja_etiqueta_lista(c) for c in _mc_cajas}
            df_mc = pd.DataFrame(mov.data)
            df_mc["caja"] = df_mc["caja_id"].map(lambda x: _mc_map.get(str(x), str(x)[:8] + "…"))
            df_mc = df_mc.drop(columns=["caja_id"], errors="ignore")
            cols = ["created_at", "tipo", "monto_usd", "moneda", "caja", "concepto", "nota_operacion"]
            df_mc = df_mc[[c for c in cols if c in df_mc.columns]]
            if "monto_usd" in df_mc.columns:
                df_mc["monto_usd"] = pd.to_numeric(df_mc["monto_usd"], errors="coerce").fillna(0).round(2)
            st.dataframe(
                df_mc,
                use_container_width=True,
                hide_index=True,
                column_config=(
                    {"monto_usd": st.column_config.NumberColumn(format="%.2f")}
                    if "monto_usd" in df_mc.columns
                    else {}
                ),
            )
        else:
            st.caption("Sin movimientos.")


def module_tasas(sb: Client, *, embedded: bool = False) -> None:
    """Delega en `movi.modules.tasas`; dependencias siguen en app.py para evitar ciclos."""
    render_module_tasas(
        sb,
        embedded=embedded,
        deps=TasasModuleDeps(
            modulo_titulo_info=_modulo_titulo_info,
            auto_tasa_abs_min_bs=AUTO_TASA_ABS_MIN_BS,
            auto_tasa_sync_rel_min=AUTO_TASA_SYNC_REL_MIN,
            auto_tasa_sync_min_seconds=AUTO_TASA_SYNC_MIN_SECONDS,
            latest_tasas=latest_tasas,
            nf=_nf,
            render_tasas_tiempo_real=render_tasas_tiempo_real,
            get_live_exchange_rates=get_live_exchange_rates,
            render_tabla_tasas_ui=render_tabla_tasas_ui,
            build_tasas_tabla_detalle=build_tasas_tabla_detalle,
            infer_tasa_bs_oper_index=_infer_tasa_bs_oper_index,
            refresh_productos_bs_equiv_note=_refresh_productos_bs_equiv_note,
            movi_bump_form_nonce=_movi_bump_form_nonce,
        ),
    )


def module_inventario(sb: Client, erp_uid: str, t: dict[str, Any] | None) -> None:
    _modulo_titulo_info(
        "Inventario",
        key="inventario",
        ayuda_md=(
            f"**Categorías:** en la base hay las categorías disponibles; respaldos y listados en PDF/Excel están en **Mantenimiento** y **Reportes**.\n\n"
            "**Buscar y modificar:** el **stock** baja con las **ventas** y sube con las **compras** (y cargas nuevas). "
            "Buscá por código, **OEM**, descripción, **marca del repuesto** o **marcas de carro**. "
            "La **tabla masiva** quedó **al final** del módulo (después de alta y CSV)."
        ),
    )
    if not t:
        st.warning("Registre tasas en **Dashboard** (expander *Cargar / editar tasas en base de datos*) para ver equivalentes.")

    try:
        cats = sb.table("categorias").select("id,nombre").order("nombre").execute()
        cats_list = cats.data or []
    except Exception as ex:
        cats_list = []
        st.error(f"No se pudieron leer **categorías** desde la base: {ex}")

    cat_opts = {c["nombre"]: c["id"] for c in cats_list if c.get("nombre")}
    _id_to_nombre_cat, _nombre_a_id_cat, _cat_select_opts = _categoria_maps_from_rows(cats_list)
    _marcas_veh_catalogo = _fetch_marcas_vehiculo_catalogo(sb)

    df = _inv_enrich_compat_columns(_normalize_productos_inventario_df(_fetch_productos_inventario_df(sb)))
    if not df.empty:
        if "categoria_id" in df.columns:
            df["categoria"] = df["categoria_id"].apply(
                lambda x: _id_to_nombre_cat.get(str(x).strip(), "")
                if x is not None and not (isinstance(x, float) and pd.isna(x)) and str(x).strip()
                else ""
            )
        else:
            df["categoria"] = ""
    else:
        df["categoria"] = pd.Series(dtype=object)

    st.caption(f"**{len(cats_list)}** categoría(s) en la base.")

    st.markdown("##### Buscar y modificar productos")
    _fc1, _fc2 = st.columns([2, 1])
    with _fc1:
        _inv_q = st.text_input(
            "Buscar (código, OEM, descripción, marca repuesto, marcas de carro…)",
            value="",
            key="inv_prod_filter",
            placeholder="Ej: filtro, 04465, Toyota, Bosch…",
        )
    with _fc2:
        _nombres_cat_ord = sorted(cat_opts.keys(), key=str.casefold)
        _opts_filtro_cat = ["Todas las categorías"] + _nombres_cat_ord + ["(Sin categoría)"]
        _sel_una_cat = st.selectbox(
            "Solo categoría",
            options=_opts_filtro_cat,
            index=0,
            key="inv_filtro_una_categoria",
            help="Filtrá por una categoría. Combinado con la búsqueda de texto.",
        )

    df_view = df
    if not df.empty:
        if _inv_q.strip():
            _q = _inv_q.strip()
            _mask = df_view.apply(lambda r: _inv_row_matches_query(r, _q), axis=1)
            df_view = df_view.loc[_mask]
        if _sel_una_cat == "(Sin categoría)":
            if "categoria_id" in df_view.columns:
                _m_sin_id = df_view["categoria_id"].isna() | (df_view["categoria_id"].astype(str).str.strip() == "")
            else:
                _m_sin_id = pd.Series(True, index=df_view.index)
            _m_sin_nom = df_view["categoria"].fillna("").astype(str).str.strip() == ""
            df_view = df_view.loc[_m_sin_id | _m_sin_nom]
        elif _sel_una_cat != "Todas las categorías":
            _cid_f = cat_opts.get(_sel_una_cat)
            if _cid_f is not None and "categoria_id" in df_view.columns:
                df_view = df_view.loc[df_view["categoria_id"].astype(str) == str(_cid_f)]
            else:
                df_view = df_view.loc[df_view["categoria"].fillna("").astype(str).str.strip() == _sel_una_cat]

    if df.empty:
        st.info(
            "No hay **productos** en la base. Más abajo podés **crear categorías y productos** o importar un **CSV**."
        )
    elif df_view.empty:
        st.info("No hay productos que coincidan con la búsqueda o la categoría elegida.")
    else:
        _sa = df["stock_actual"].map(_inv_stock_int)
        _sm = df["stock_minimo"].map(_inv_stock_int)
        crit_all = df.loc[_sa <= _sm]
        if len(crit_all):
            st.error(f"Alertas de stock crítico: {len(crit_all)} ítems")
            st.dataframe(crit_all, use_container_width=True, hide_index=True)

    if not df.empty and df_view.empty:
        crit_all = df.loc[df["stock_actual"].map(_inv_stock_int) <= df["stock_minimo"].map(_inv_stock_int)]
        if len(crit_all):
            st.error(f"Alertas de stock crítico: {len(crit_all)} ítems")
            st.dataframe(crit_all, use_container_width=True, hide_index=True)

    st.divider()
    with st.expander(
        "Productos — Editar · Nuevo · Kit/compuesto · Eliminar · Carga/descarga inventario",
        expanded=True,
    ):
        _t_edit, _t_prod, _t_kit, _t_del, _t_mov = st.tabs(
            [
                "Editar Productos",
                "Nuevo Producto",
                "Kit / compuesto",
                "Eliminar Producto",
                "Carga /Descarga de inventario",
            ]
        )
        with _t_edit:
            st.caption(
                "Usa los mismos **filtros de búsqueda y categoría** de la sección *Buscar y modificar productos* arriba."
            )
            if df.empty:
                st.info(
                    "No hay **productos** en la base. Usá **Nuevo Producto** (y *Nueva categoría* dentro de esa pestaña) o importá un **CSV** más abajo."
                )
            elif df_view.empty:
                st.info(
                    "No hay productos que coincidan con la búsqueda o la categoría elegida. "
                    "Vacía el filtro o cambiá **Solo categoría** para ver ítems y editarlos desde acá."
                )
            else:
                _max_ficha = 200
                if len(df_view) > _max_ficha:
                    st.caption(
                        f"Hay **{len(df_view)}** productos en pantalla. Afiná la búsqueda o la categoría "
                        f"(menos de {_max_ficha}) para usar esta ficha."
                    )
                else:
                    _labels: dict[str, str] = {}
                    for _, _r in df_view.iterrows():
                        _cod = _export_cell_txt(_r.get("codigo")) or "—"
                        _desc = _export_cell_txt(_r.get("descripcion")) or "—"
                        _lid = str(_r.get("id") or "").strip()
                        if not _lid:
                            continue
                        _lab = f"{_cod} · {_desc[:48]}" + ("" if len(_desc) <= 48 else "…")
                        if _lab in _labels:
                            _lab = f"{_lab} [{_lid[:8]}]"
                        _labels[_lab] = _lid
                    _pick_labs = sorted(_labels.keys(), key=str.casefold)
                    _sel_lab = st.selectbox(
                        "Elegí el producto a editar",
                        options=["—"] + _pick_labs,
                        index=0,
                        key="inv_ficha_producto_pick",
                        help="El listado respeta búsqueda y categoría. Cambiá acá de ítem sin guardar el formulario.",
                    )
                    if _sel_lab != "—" and _sel_lab in _labels:
                        _pid = _labels[_sel_lab]
                        _row = df_view[df_view["id"].astype(str) == _pid]
                        if len(_row) != 1:
                            st.error("No se encontró el producto.")
                        else:
                            _rw = _row.iloc[0]
                            _d_comp = _inv_compat_as_dict(_rw.get("compatibilidad"))
                            _marcas_act: list[str] = []
                            for _m in _d_comp.get("marcas_vehiculo") or _d_comp.get("marcas") or []:
                                _s = str(_m).strip()
                                if _s:
                                    _marcas_act.append(_s)
                            _en_cat = [m for m in _marcas_act if m in _marcas_veh_catalogo]
                            _extras_list = [m for m in _marcas_act if m not in _marcas_veh_catalogo]
                            _extras_str = ", ".join(_extras_list)
                            _anos_ini = _inv_compat_anos_str(_d_comp)
                            _cat_nm = _export_cell_txt(_rw.get("categoria"))
                            try:
                                _cat_ix = _cat_select_opts.index(_cat_nm) if _cat_nm in _cat_select_opts else 0
                            except ValueError:
                                _cat_ix = 0
                            _cond_ini = _rw.get("condicion")
                            if _cond_ini not in ("Nuevo", "Usado"):
                                _cond_ini = "Nuevo"
                            st.caption(
                                f"**Producto** · ID interno `{_pid}`. "
                                "Ubicación y foto van **fuera** del botón *Guardar cambios* para que el archivo no se pierda al enviar."
                            )
                            fu1, fu2 = st.columns(2)
                            _fubi = fu1.text_input(
                                "Ubicación en almacén",
                                value=_export_cell_txt(_rw.get("ubicacion")),
                                key=f"inv_ficha_ubi_{_pid}",
                            )
                            with fu2:
                                st.markdown("**Imagen del producto**")
                                if _catalogo_storage_portada_enabled():
                                    _ficha_img_file = st.file_uploader(
                                        "Subir foto (JPG/PNG/WebP)",
                                        type=["jpg", "jpeg", "png", "webp"],
                                        accept_multiple_files=False,
                                        key=f"inv_ficha_img_file_{_pid}",
                                    )
                                    _fimg = st.text_input(
                                        "o pegar URL",
                                        value=_export_cell_txt(_rw.get("imagen_url")),
                                        key=f"inv_ficha_img_{_pid}",
                                        help="Si subís una foto arriba, al guardar se usa esa y queda como portada en catálogo/Storage.",
                                    )
                                else:
                                    _ficha_img_file = None
                                    _fimg = st.text_input(
                                        "URL de imagen (opcional)",
                                        value=_export_cell_txt(_rw.get("imagen_url")),
                                        key=f"inv_ficha_img_{_pid}",
                                        help="Subida de archivos desactivada (`[catalogo] enabled = false`). Solo enlace externo; no usa Storage.",
                                    )
                            with st.form(
                                f"inv_ficha_prod_form_{_pid}_{int(st.session_state.get(f'inv_ficha_form_nonce_{_pid}', 0))}"
                            ):
                                st.caption("Completá los datos y pulsá **Guardar cambios del producto** al final.")
                                fa, fb = st.columns(2)
                                _fcod = fa.text_input(
                                    "Código interno",
                                    value=_export_cell_txt(_rw.get("codigo")),
                                    key=f"inv_ficha_cod_{_pid}",
                                )
                                _fsku = fb.text_input(
                                    "OEM / código parte",
                                    value=_export_cell_txt(_rw.get("sku_oem")),
                                    key=f"inv_ficha_sku_{_pid}",
                                )
                                _fdesc = st.text_input(
                                    "Descripción",
                                    value=_export_cell_txt(_rw.get("descripcion")) or "Sin descripción",
                                    max_chars=500,
                                    key=f"inv_ficha_desc_{_pid}",
                                )
                                fm1, fm2 = st.columns(2)
                                _fmprod = fm1.text_input(
                                    "Marca del repuesto (ej. Bosch)",
                                    value=_export_cell_txt(_rw.get("marca_producto")),
                                    key=f"inv_ficha_mprod_{_pid}",
                                )
                                _fcond = fm2.selectbox(
                                    "Condición",
                                    ["Nuevo", "Usado"],
                                    index=0 if _cond_ini == "Nuevo" else 1,
                                    key=f"inv_ficha_cond_{_pid}",
                                )
                                st.markdown("**Compatibilidad (marcas de carro)**")
                                if _marcas_veh_catalogo:
                                    _fpick_mv = st.multiselect(
                                        "Del catálogo en base",
                                        options=_marcas_veh_catalogo,
                                        default=_en_cat,
                                        key=f"inv_ficha_mv_{_pid}",
                                    )
                                else:
                                    st.caption("Sin catálogo: ejecutá **patch_012** o escribí las marcas solo en texto.")
                                    _fpick_mv = []
                                fc1, fc2 = st.columns(2)
                                _fextra_mv = fc1.text_input(
                                    "Otras marcas (coma)",
                                    value=_extras_str,
                                    key=f"inv_ficha_mvx_{_pid}",
                                )
                                _fanos = fc2.text_input(
                                    "Años / rango",
                                    value=_anos_ini,
                                    key=f"inv_ficha_anos_{_pid}",
                                )
                                st.markdown("**Stock y precios (USD)**")
                                fs1, fs2 = st.columns(2)
                                _ns = fs1.number_input(
                                    "Stock actual (unidades)",
                                    min_value=0,
                                    value=_inv_stock_int(_rw.get("stock_actual")),
                                    step=1,
                                    format="%d",
                                    key=f"inv_ficha_st_{_pid}",
                                )
                                _nsmin = fs2.number_input(
                                    "Stock mínimo (alerta)",
                                    min_value=0,
                                    value=_inv_stock_int(_rw.get("stock_minimo")),
                                    step=1,
                                    format="%d",
                                    key=f"inv_ficha_smin_{_pid}",
                                )
                                fp1, fp2 = st.columns(2)
                                _nco = fp1.number_input(
                                    "Costo USD",
                                    min_value=0.0,
                                    value=float(_rw.get("costo_usd") or 0),
                                    step=0.01,
                                    format="%.2f",
                                    key=f"inv_ficha_co_{_pid}",
                                )
                                _npv = fp2.number_input(
                                    "Precio venta USD (precio_v_usd)",
                                    min_value=0.0,
                                    value=float(_rw.get("precio_v_usd") or 0),
                                    step=0.01,
                                    format="%.2f",
                                    key=f"inv_ficha_pv_{_pid}",
                                )
                                if float(_nco) > 0:
                                    st.caption(
                                        f"Margen bruto: **{((float(_npv) - float(_nco)) / float(_nco) * 100):.1f}%** "
                                        f"· Diferencia USD: **{float(_npv) - float(_nco):.2f}**"
                                    )
                                st.markdown("**Estado y categoría**")
                                fx1, fx2 = st.columns(2)
                                _factivo = fx1.checkbox(
                                    "Producto activo",
                                    value=bool(_rw.get("activo", True)),
                                    key=f"inv_ficha_act_{_pid}",
                                )
                                _fsel_cat = fx2.selectbox(
                                    "Categoría",
                                    options=_cat_select_opts,
                                    index=_cat_ix,
                                    key=f"inv_ficha_cat_{_pid}",
                                )
                                _srl_ini_f = "\n".join(_inv_compat_seriales_motor_list(_d_comp))
                                st.markdown("**Números de serie (opcional)**")
                                st.caption(
                                    "Un valor por **unidad en stock** (motores, alternadores, etc.). "
                                    "Si cargás seriales y ejecutaste **patch_023**, al vender se eligen y se guardan en la factura."
                                )
                                _f_srl = st.text_area(
                                    "Seriales — uno por línea o separados por coma",
                                    value=_srl_ini_f,
                                    height=120,
                                    key=f"inv_ficha_srl_{_pid}",
                                )
                                if st.form_submit_button("Guardar cambios del producto"):
                                    _merged_mv = _inv_merge_marcas_catalogo_texto(_fpick_mv, _fextra_mv)
                                    _compat_f = _inv_build_compat_dict(_merged_mv, _fanos)
                                    _srl_parsed = _inv_parse_seriales_motor_texto(str(_f_srl))
                                    _compat_f = _inv_compat_merge_seriales(_compat_f, _srl_parsed)
                                    if _srl_parsed and int(_ns) != len(_srl_parsed):
                                        st.warning(
                                            f"Hay **{len(_srl_parsed)}** serie(s) cargadas y **stock {_ns}** unidades: "
                                            "conviene que coincidan si controlás cada unidad."
                                        )
                                    _cid_f = (
                                        _nombre_a_id_cat.get(str(_fsel_cat).strip())
                                        if str(_fsel_cat or "").strip()
                                        else None
                                    )
                                    _upd_f: dict[str, Any] = {
                                        "codigo": str(_fcod).strip() or None,
                                        "descripcion": str(_fdesc).strip() or "Sin descripción",
                                        "stock_actual": int(_ns),
                                        "stock_minimo": int(_nsmin),
                                        "costo_usd": float(_nco),
                                        "precio_v_usd": float(_npv),
                                        "activo": bool(_factivo),
                                        "categoria_id": _cid_f,
                                        "condicion": _fcond if _fcond in ("Nuevo", "Usado") else "Nuevo",
                                        "compatibilidad": _compat_f,
                                    }
                                    if "sku_oem" in df_view.columns:
                                        _upd_f["sku_oem"] = str(_fsku).strip() or None
                                    if "marca_producto" in df_view.columns:
                                        _upd_f["marca_producto"] = str(_fmprod).strip() or None
                                    if "ubicacion" in df_view.columns:
                                        _upd_f["ubicacion"] = str(_fubi).strip() or None
                                    if "imagen_url" in df_view.columns:
                                        _upd_f["imagen_url"] = str(_fimg).strip() or None
                                    if (
                                        _catalogo_storage_portada_enabled()
                                        and _ficha_img_file is not None
                                        and "imagen_url" in df_view.columns
                                    ):
                                        try:
                                            bucket = _catalogo_bucket_name()
                                            data = _ficha_img_file.getvalue()
                                            if data:
                                                sb.table("producto_fotos").update({"is_primary": False}).eq(
                                                    "producto_id", str(_pid)
                                                ).execute()
                                                obj = _catalogo_upload_producto_foto(
                                                    sb,
                                                    bucket=bucket,
                                                    producto_id=str(_pid),
                                                    filename=str(
                                                        getattr(_ficha_img_file, "name", "") or "foto"
                                                    ),
                                                    content_type=str(
                                                        getattr(_ficha_img_file, "type", "")
                                                        or "application/octet-stream"
                                                    ),
                                                    data=data,
                                                )
                                                _row_pf_ficha: dict[str, Any] = {
                                                    "producto_id": str(_pid),
                                                    "storage_path": obj,
                                                    "is_primary": True,
                                                }
                                                _cb_fc = _erp_user_uuid_or_none(erp_uid)
                                                if _cb_fc:
                                                    _row_pf_ficha["created_by"] = _cb_fc
                                                sb.table("producto_fotos").insert(_row_pf_ficha).execute()
                                                _catalogo_sync_primary_foto(
                                                    sb, bucket=bucket, producto_id=str(_pid), storage_path=obj
                                                )
                                                _upd_f["imagen_url"] = _storage_public_object_url(bucket, obj)
                                        except Exception as ex_img:
                                            st.warning(
                                                f"No se pudo subir la **foto al Storage**; **el producto sí se actualizó** en la base "
                                                f"(código, stock, texto de URL, etc.—solo falló el archivo). Detalle: {ex_img}"
                                                + _movi_foto_upload_bucket_hint(bucket, ex_img)
                                            )
                                    try:
                                        sb.table("productos").update(_upd_f).eq("id", _pid).execute()
                                        st.success("Producto actualizado.")
                                        _movi_reset_inv_ficha_product_keys(_pid)
                                        _movi_bump_form_nonce(f"inv_ficha_form_nonce_{_pid}")
                                        st.rerun()
                                    except Exception as ex:
                                        st.error(
                                            f"{ex} · Revisá que esté aplicado **patch_011** (columnas repuestos) en Supabase."
                                        )
        with _t_prod:
            with st.expander("Nueva categoría", expanded=False):
                st.caption("Creá la categoría acá si aún no existe; después elegila en el formulario de abajo.")
                with st.form(f"f_cat_{int(st.session_state.get('inv_cat_form_nonce', 0))}"):
                    cn = st.text_input("Nombre categoría", key="inv_alta_cat_nombre")
                    submitted_cat = st.form_submit_button("Crear categoría")
                    if submitted_cat:
                        if not cn.strip():
                            st.error("Escribí un nombre para la categoría.")
                        else:
                            try:
                                sb.table("categorias").insert({"nombre": cn.strip()}).execute()
                                st.success("Categoría guardada en la base.")
                                _movi_ss_pop_keys("inv_alta_cat_nombre")
                                _movi_bump_form_nonce("inv_cat_form_nonce")
                                st.rerun()
                            except Exception as ex:
                                st.error(
                                    f"No se pudo guardar. Si el nombre ya existe, elegí otro (las categorías son únicas). Detalle: {ex}"
                                )
            st.caption(
                "**Ubicación"
                + (" y foto** van" if _catalogo_storage_portada_enabled() else "** va")
                + " fuera del formulario de guardado"
                + (
                    ": así Streamlit no pierde el archivo al pulsar *Guardar producto* (selector de archivos fuera de `st.form`)."
                    if _catalogo_storage_portada_enabled()
                    else "."
                )
            )
            _ua1, _ua2 = st.columns(2)
            ubic = _ua1.text_input("Ubicación en almacén", key="inv_alta_ubic")
            with _ua2:
                if _catalogo_storage_portada_enabled():
                    st.markdown("**Foto del producto (opcional)**")
                    img_file = st.file_uploader(
                        "Subir foto (JPG/PNG/WebP)",
                        type=["jpg", "jpeg", "png", "webp"],
                        accept_multiple_files=False,
                        key="inv_alta_img_file",
                    )
                    img_url = st.text_input(
                        "o pegar URL (opcional)",
                        key="inv_alta_img",
                        help="Si subís foto arriba, al guardar se usa esa y se llena `imagen_url` automáticamente.",
                    )
                else:
                    img_file = None
                    st.markdown("**Imagen (solo enlace, opcional)**")
                    img_url = st.text_input(
                        "URL de imagen (opcional)",
                        key="inv_alta_img",
                        help="Subida a Storage desactivada (`[catalogo] enabled = false`).",
                    )
            with st.form(f"f_prod_{int(st.session_state.get('inv_prod_form_nonce', 0))}"):
                desc = st.text_input("Descripción", max_chars=500, key="inv_alta_prod_desc")
                cx, mx = st.columns(2)
                cname = cx.selectbox(
                    "Categoría",
                    options=[""] + sorted(cat_opts.keys(), key=str.casefold),
                    key="inv_alta_prod_cat",
                    help="Para código automático es obligatoria. Primeras 3 letras/números del nombre → tramo del código.",
                )
                marca_rep = mx.text_input(
                    "Marca del repuesto (ej. Bosch, Denso)",
                    key="inv_alta_marca_prod",
                    help="Primeras 3 letras/números → tramo del código. Vacío = **GEN** (ej. FIL-GEN-0001).",
                )
                cond_alta = st.selectbox("Condición", ["Nuevo", "Usado"], index=0, key="inv_alta_cond")
                st.markdown("**Código interno**")
                cod_auto = st.checkbox(
                    "Generar automático: **categoría + marca + número** (ej. `FIL-BOS-0001`)",
                    value=True,
                    key="inv_alta_cod_auto",
                )
                codigo_manual = ""
                if cod_auto:
                    if cname:
                        _prev_cod = _siguiente_codigo_interno_producto(sb, cname, marca_rep)
                        st.info(
                            f"Vista previa: **`{_prev_cod}`** · Al guardar se vuelve a calcular el siguiente libre "
                            "(por si otro usuario cargó algo al mismo tiempo)."
                        )
                    else:
                        st.warning("Elegí una **categoría** para poder generar el código automático.")
                else:
                    codigo_manual = st.text_input(
                        "Código manual",
                        key="inv_alta_prod_codigo",
                        placeholder="Ej. ARR-REN-01",
                    )
                a1, a2 = st.columns(2)
                sku_oem = a1.text_input("Código OEM / parte / fabricante", key="inv_alta_prod_sku_oem")
                a2.caption("El **código interno** es el tuyo para buscar en inventario; el OEM es el de la caja/fábrica.")
                if _marcas_veh_catalogo:
                    st.multiselect(
                        "Marcas de carro (catálogo en base)",
                        options=_marcas_veh_catalogo,
                        default=[],
                        key="inv_alta_marcas_pick",
                        help="Listado cargado con **patch_012_marcas_vehiculo.sql**. Podés sumar más en el campo de texto.",
                    )
                else:
                    st.caption(
                        "No hay catálogo de marcas en la base. Ejecutá **supabase/patch_012_marcas_vehiculo.sql** "
                        "en Supabase y recargá la app."
                    )
                    st.session_state["inv_alta_marcas_pick"] = []
                c1, c2 = st.columns(2)
                marcas_auto = c1.text_input(
                    "Otras marcas de carro (texto, coma)",
                    key="inv_alta_marcas_veh",
                    placeholder="Ej. otra que no esté en la lista…",
                    help="Se unen con las que marcaste arriba en el catálogo.",
                )
                anos_auto = c2.text_input("Años / rango (opcional)", key="inv_alta_anos", placeholder="2010-2015")
                stock = st.number_input(
                    "Stock actual (unidades)",
                    min_value=0,
                    value=0,
                    step=1,
                    format="%d",
                    key="inv_alta_stock",
                )
                smin = st.number_input(
                    "Stock mínimo (alerta)",
                    min_value=0,
                    value=0,
                    step=1,
                    format="%d",
                    key="inv_alta_smin",
                )
                st.markdown("**Números de serie (opcional)**")
                st.caption(
                    "Un valor por cada unidad en **stock** (ej. motor, pieza serializada). "
                    "Vacío = sin tracking por serie. Con **patch_023**, al vender se validan y registran en la factura."
                )
                _alta_srl = st.text_area(
                    "Seriales — uno por línea o separados por coma",
                    height=100,
                    key="inv_alta_seriales",
                    placeholder="Ej. SN-ABC123 (una línea por unidad si cargás varias)",
                )
                costo = st.number_input("Costo USD", min_value=0.0, value=0.0, format="%.2f", key="inv_alta_costo")
                pv = st.number_input(
                    "Precio venta USD (precio_v_usd)",
                    min_value=0.0,
                    value=0.0,
                    format="%.2f",
                    key="inv_alta_pv",
                )
                if float(costo) > 0:
                    st.caption(
                        f"Margen bruto sobre costo: **{((float(pv) - float(costo)) / float(costo) * 100):.1f}%** "
                        f"· Diferencia USD: **{float(pv) - float(costo):.2f}**"
                    )
                cid = cat_opts.get(cname) if cname else None
                if st.form_submit_button("Guardar producto"):
                    try:
                        _cb_foto = _erp_user_uuid_or_none(erp_uid)
                        _pick_mv = list(st.session_state.get("inv_alta_marcas_pick") or [])
                        _merged_mv = _inv_merge_marcas_catalogo_texto(_pick_mv, marcas_auto)
                        _compat_ins = _inv_build_compat_dict(_merged_mv, anos_auto)
                        _srl_alta = _inv_parse_seriales_motor_texto(str(_alta_srl))
                        _compat_ins = _inv_compat_merge_seriales(_compat_ins, _srl_alta)
                        if _srl_alta and int(stock) != len(_srl_alta):
                            st.warning(
                                f"Seriales: **{len(_srl_alta)}** vs stock **{int(stock)}** — revisá que coincidan."
                            )
                        if cod_auto:
                            if not cname:
                                st.error("Para código automático elegí una **categoría**.")
                            else:
                                codigo_final: str | None = _siguiente_codigo_interno_producto(sb, cname, marca_rep)
                                _insert_ok = False
                                _last_ex: Exception | None = None
                                for _ in range(10):
                                    try:
                                        ins = _movi_productos_insert_execute(
                                            sb,
                                            {
                                                "codigo": codigo_final,
                                                "sku_oem": sku_oem.strip() or None,
                                                "descripcion": desc.strip() or "Sin descripción",
                                                "marca_producto": marca_rep.strip() or None,
                                                "condicion": cond_alta,
                                                "ubicacion": ubic.strip() or None,
                                                "compatibilidad": _compat_ins,
                                                "imagen_url": img_url.strip() or None,
                                                "stock_actual": int(stock),
                                                "stock_minimo": int(smin),
                                                "costo_usd": float(costo),
                                                "precio_v_usd": float(pv),
                                                "categoria_id": cid,
                                                "activo": True,
                                            },
                                        )
                                        new_id = _inv_resolve_producto_id_after_insert(
                                            sb, ins_data=ins.data, codigo=codigo_final
                                        )
                                        if not new_id:
                                            raise RuntimeError(_inv_alta_producto_id_missing_help())
                                        if _catalogo_storage_portada_enabled() and img_file is not None and new_id:
                                            try:
                                                bucket = _catalogo_bucket_name()
                                                data = img_file.getvalue()
                                                if data:
                                                    obj = _catalogo_upload_producto_foto(
                                                        sb,
                                                        bucket=bucket,
                                                        producto_id=new_id,
                                                        filename=str(getattr(img_file, "name", "") or "foto"),
                                                        content_type=str(getattr(img_file, "type", "") or "application/octet-stream"),
                                                        data=data,
                                                    )
                                                    row_pf: dict[str, Any] = {
                                                        "producto_id": new_id,
                                                        "storage_path": obj,
                                                        "is_primary": True,
                                                    }
                                                    if _cb_foto:
                                                        row_pf["created_by"] = _cb_foto
                                                    sb.table("producto_fotos").insert(row_pf).execute()
                                                    _catalogo_sync_primary_foto(
                                                        sb, bucket=bucket, producto_id=new_id, storage_path=obj
                                                    )
                                            except Exception as _ex_foto:
                                                st.warning(
                                                    f"Producto guardado, pero la foto no se registró: {_ex_foto}. "
                                                    "Revisá Storage (bucket), **patch_021** y `GRANT` sobre `producto_fotos`."
                                                    + _movi_foto_upload_bucket_hint(bucket, _ex_foto)
                                                )
                                        _insert_ok = True
                                        break
                                    except Exception as ex_i:
                                        _last_ex = ex_i
                                        es = str(ex_i).lower()
                                        if "duplicate" in es or "unique" in es or "23505" in es:
                                            codigo_final = _siguiente_codigo_interno_producto(sb, cname, marca_rep)
                                        else:
                                            raise
                                if _insert_ok:
                                    st.success(f"Producto guardado con código **{codigo_final}**.")
                                    _movi_reset_producto_alta_fields()
                                    _movi_bump_form_nonce("inv_prod_form_nonce")
                                    st.rerun()
                                else:
                                    st.error(
                                        str(_last_ex)
                                        if _last_ex
                                        else "No se pudo asignar código único. Reintentá o usá código manual."
                                    )
                        else:
                            _cm = (codigo_manual or "").strip()
                            if not _cm:
                                st.error("Ingresá un **código manual** o activá el generador automático.")
                            else:
                                insm = _movi_productos_insert_execute(
                                    sb,
                                    {
                                        "codigo": _cm or None,
                                        "sku_oem": sku_oem.strip() or None,
                                        "descripcion": desc.strip() or "Sin descripción",
                                        "marca_producto": marca_rep.strip() or None,
                                        "condicion": cond_alta,
                                        "ubicacion": ubic.strip() or None,
                                        "compatibilidad": _compat_ins,
                                        "imagen_url": img_url.strip() or None,
                                        "stock_actual": int(stock),
                                        "stock_minimo": int(smin),
                                        "costo_usd": float(costo),
                                        "precio_v_usd": float(pv),
                                        "categoria_id": cid,
                                        "activo": True,
                                    },
                                )
                                new_id = _inv_resolve_producto_id_after_insert(
                                    sb, ins_data=insm.data, codigo=_cm
                                )
                                if not new_id:
                                    raise RuntimeError(_inv_alta_producto_id_missing_help())
                                if _catalogo_storage_portada_enabled() and img_file is not None and new_id:
                                    try:
                                        bucket = _catalogo_bucket_name()
                                        data = img_file.getvalue()
                                        if data:
                                            obj = _catalogo_upload_producto_foto(
                                                sb,
                                                bucket=bucket,
                                                producto_id=new_id,
                                                filename=str(getattr(img_file, "name", "") or "foto"),
                                                content_type=str(getattr(img_file, "type", "") or "application/octet-stream"),
                                                data=data,
                                            )
                                            row_pfm: dict[str, Any] = {
                                                "producto_id": new_id,
                                                "storage_path": obj,
                                                "is_primary": True,
                                            }
                                            if _cb_foto:
                                                row_pfm["created_by"] = _cb_foto
                                            sb.table("producto_fotos").insert(row_pfm).execute()
                                            _catalogo_sync_primary_foto(
                                                sb, bucket=bucket, producto_id=new_id, storage_path=obj
                                            )
                                    except Exception as _ex_foto_m:
                                        st.warning(
                                            f"Producto guardado, pero la foto no se registró: {_ex_foto_m}. "
                                            "Revisá Storage (bucket), **patch_021** y `GRANT` sobre `producto_fotos`."
                                            + _movi_foto_upload_bucket_hint(bucket, _ex_foto_m)
                                        )
                                st.success("Producto guardado en la base.")
                                _movi_reset_producto_alta_fields()
                                _movi_bump_form_nonce("inv_prod_form_nonce")
                                st.rerun()
                    except Exception as ex:
                        _em = str(ex)
                        st.error(
                            f"No se pudo guardar el producto: {_em}. "
                            "Si falta alguna columna, ejecutá **patch_011_productos_repuestos.sql** en Supabase."
                            + (
                                " · Si el error dice **SyncQueryRequestBuilder** y **select**, hacé **git pull**, "
                                "reiniciá Streamlit y confirmá que corrés este `app.py` del repo (no una copia vieja)."
                                if "select" in _em.lower() and "syncquery" in _em.lower()
                                else ""
                            )
                        )

        with _t_kit:
            st.caption(
                "Creá el ítem cabecera en **Nuevo Producto** (ej. **Conversión Corsa**) con su **precio de venta**. "
                "Acá definís los repuestos por cada **1 kit** vendido; en **Ventas** el detalle muestra el kit y el stock baja de los **componentes**. "
                "Requiere **supabase/patch_014_productos_kit.sql** en Supabase."
            )
            _pxk = None
            try:
                _pxk = (
                    sb.table("productos")
                    .select("id,descripcion,stock_actual,es_compuesto,activo")
                    .eq("activo", True)
                    .order("descripcion")
                    .limit(3000)
                    .execute()
                )
            except Exception as _exk:
                _msgk = str(_exk).lower()
                if "es_compuesto" in _msgk or "column" in _msgk:
                    try:
                        _pxk = (
                            sb.table("productos")
                            .select("id,descripcion,stock_actual,activo")
                            .eq("activo", True)
                            .order("descripcion")
                            .limit(3000)
                            .execute()
                        )
                        for _zr in _pxk.data or []:
                            _zr["es_compuesto"] = False
                    except Exception as _exk2:
                        st.error(str(_exk2))
                else:
                    st.error(str(_exk))
            _plk = list(_pxk.data or []) if _pxk else []
            if not _plk:
                st.info("No hay productos activos.")
            else:

                def _kit_parse_comp_opt(opt: str) -> str:
                    if not opt or " ‖ " not in opt:
                        return ""
                    return str(opt.split(" ‖ ", 1)[-1]).strip()

                _lkp: dict[str, str] = {}
                for _p in _plk:
                    _i = str(_p.get("id") or "").strip()
                    if not _i:
                        continue
                    _d = (_export_cell_txt(_p.get("descripcion")) or "—")[:56]
                    _suf = " (kit)" if _p.get("es_compuesto") else ""
                    _lkp[f"{_d}{_suf}"] = _i
                _k_keys = sorted(_lkp.keys(), key=str.casefold)
                _selk_lab = st.selectbox("Producto kit (cabecera)", options=_k_keys, key="inv_kit_parent_lab")
                _kid = _lkp[str(_selk_lab)]
                _comp_pool = [p for p in _plk if str(p.get("id")) != _kid and not p.get("es_compuesto")]
                if not _comp_pool:
                    _comp_pool = [p for p in _plk if str(p.get("id")) != _kid]
                _opt_empty = ""
                _kit_opt_vals = [_opt_empty] + [
                    f"{(_export_cell_txt(p.get('descripcion')) or '—')[:55]} ‖ {p['id']}"
                    for p in sorted(_comp_pool, key=lambda x: str(x.get("descripcion") or "").casefold())
                ]
                _kit_map = _fetch_kit_items_by_kit(sb)
                _existing = _kit_map.get(_kid, [])
                if _existing:
                    _df_rows = []
                    for it in _existing:
                        _cid = str(it["componente_producto_id"])
                        _match = next((v for v in _kit_opt_vals if v.endswith(f" ‖ {_cid}")), _opt_empty)
                        _df_rows.append(
                            {"componente": _match, "cantidad_por_kit": float(it["cantidad"])}
                        )
                else:
                    _df_rows = [{"componente": _opt_empty, "cantidad_por_kit": 1.0}]
                while len(_df_rows) < 3:
                    _df_rows.append({"componente": _opt_empty, "cantidad_por_kit": 1.0})
                _dfk = pd.DataFrame(_df_rows)
                _rev_map = st.session_state.setdefault("_inv_kit_rev", {})
                _ed_key = f"inv_kit_ed_{_kid}_{_rev_map.get(_kid, 0)}"
                _kc0, _kc1 = st.columns([1, 2])
                with _kc0:
                    _kchk = st.checkbox(
                        "Este producto es un kit (compuesto)",
                        value=bool(next((p.get("es_compuesto") for p in _plk if str(p.get("id")) == _kid), False)),
                        key=f"inv_kit_flag_{_kid}",
                    )
                with _kc1:
                    _stk_by = {str(p["id"]): _inv_stock_int(p.get("stock_actual")) for p in _plk}
                    _avail = _kit_cantidad_armable(_stk_by, _kit_map.get(_kid, []))
                    st.caption(
                        f"Según componentes en almacén: podés armar aprox. **{_avail}** kit(s). "
                        "El **stock del ítem cabecera** no se usa al vender; valida la RPC con los repuestos."
                    )
                _ded = st.data_editor(
                    _dfk,
                    key=_ed_key,
                    num_rows="dynamic",
                    column_config={
                        "componente": st.column_config.SelectboxColumn(
                            "Componente",
                            options=_kit_opt_vals,
                            required=False,
                        ),
                        "cantidad_por_kit": st.column_config.NumberColumn(
                            "Cant. por 1 kit",
                            min_value=0.001,
                            format="%.3f",
                            required=True,
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )
                if st.button("Guardar definición del kit", key=f"inv_kit_save_{_kid}"):
                    if not _kchk:
                        try:
                            sb.table("productos").update({"es_compuesto": False}).eq("id", _kid).execute()
                            sb.table("productos_kit_items").delete().eq("kit_producto_id", _kid).execute()
                            _rev_map[_kid] = int(_rev_map.get(_kid, 0)) + 1
                            st.success("Kit desactivado; se quitaron los componentes de la definición.")
                            st.rerun()
                        except Exception as _ek:
                            st.error(f"{_ek} · ¿Ejecutaste **patch_014_productos_kit.sql**?")
                    else:
                        _ins_rows: list[dict[str, Any]] = []
                        for _, __r in _ded.iterrows():
                            _cid = _kit_parse_comp_opt(str(__r.get("componente") or ""))
                            try:
                                _cq = float(__r.get("cantidad_por_kit") or 0)
                            except (TypeError, ValueError):
                                _cq = 0.0
                            if not _cid or _cq <= 0:
                                continue
                            if _cid == _kid:
                                st.error("Un kit no puede incluirse a sí mismo como componente.")
                                _ins_rows = []
                                break
                            _ins_rows.append(
                                {
                                    "kit_producto_id": _kid,
                                    "componente_producto_id": _cid,
                                    "cantidad": _cq,
                                }
                            )
                        if not _ins_rows:
                            st.error("Definí al menos un componente con cantidad > 0.")
                        else:
                            try:
                                sb.table("productos").update({"es_compuesto": True}).eq("id", _kid).execute()
                                sb.table("productos_kit_items").delete().eq("kit_producto_id", _kid).execute()
                                sb.table("productos_kit_items").insert(_ins_rows).execute()
                                _rev_map[_kid] = int(_rev_map.get(_kid, 0)) + 1
                                st.success("Kit guardado. Ya podés venderlo en **Ventas / CXC**.")
                                st.rerun()
                            except Exception as _ek:
                                st.error(f"{_ek} · ¿Ejecutaste **patch_014_productos_kit.sql**?")

        with _t_del:
            st.caption(
                "Es la misma tabla **`productos`**: un **kit** es un producto con **compuesto** marcado; "
                "no se lee `productos_kit_items` acá. Solo se **elimina** de la base si el ítem tiene **stock 0** "
                "y sin ventas/compras que lo referencien; si no podés borrarlo, bajá stock en **Carga/Descarga** o "
                "desactivalo (**Activo** = no) en la grilla."
            )
            try:
                _rows_z, _ord_hint_z = _inv_fetch_productos_para_dropdown_eliminar(sb)
            except Exception as exz:
                _rows_z, _ord_hint_z = [], ""
                st.error(str(exz))
            if _ord_hint_z:
                st.caption(_ord_hint_z)
            _hide_kits_del = st.checkbox(
                "Ocultar kits (solo ítems simples en la lista)",
                value=False,
                key="inv_del_hide_kits",
                help="Los kits siguen siendo filas en productos; ocultalos si buscás un repuesto suelto.",
            )
            if _hide_kits_del:
                _rows_z = [r for r in _rows_z if not r.get("es_compuesto")]
            if not _rows_z:
                st.info("No hay productos que mostrar (o ninguno coincide con el filtro de kits).")
            else:
                _lab_z: dict[str, str] = {}
                for _rz in _rows_z:
                    _iz = str(_rz.get("id") or "").strip()
                    if not _iz:
                        continue
                    _cz = _export_cell_txt(_rz.get("codigo")) or "—"
                    _dz = (_export_cell_txt(_rz.get("descripcion")) or "")[:56]
                    _stz = _inv_stock_int(_rz.get("stock_actual"))
                    _ktz = " · kit" if _rz.get("es_compuesto") else ""
                    _base_z = f"{_cz} · {_dz} · stock {_stz}{_ktz}"
                    _lab_one = f"{_base_z} [{_iz[:8]}]"
                    if _lab_one in _lab_z:
                        _lab_one = f"{_base_z} [{_iz}]"
                    _lab_z[_lab_one] = _iz
                _keys_z = sorted(_lab_z.keys(), key=str.casefold)
                with st.form(f"inv_form_del_prod_{int(st.session_state.get('inv_del_prod_form_nonce', 0))}"):
                    _sel_z = st.selectbox("Producto a eliminar", options=_keys_z, key="inv_del_prod_sel")
                    _cf_z = st.text_input('Confirmación: escribí **ELIMINAR**', key="inv_del_prod_conf")
                    if st.form_submit_button("Eliminar definitivamente"):
                        _pid_z = _lab_z.get(str(_sel_z))
                        if not _pid_z:
                            st.error("Elegí un producto.")
                        else:
                            _ok_z, _msg_z = _inv_eliminar_producto_stock_cero(sb, _pid_z, _cf_z)
                            if _ok_z:
                                st.success(_msg_z)
                                _movi_ss_pop_keys("inv_del_prod_sel", "inv_del_prod_conf")
                                _movi_bump_form_nonce("inv_del_prod_form_nonce")
                                st.rerun()
                            else:
                                st.error(_msg_z)

        with _t_mov:
            st.caption(
                "**Entrada** suma stock (hallazgo, ajuste de inventario). **Salida** resta (merma, rotura). "
                "Se guarda historial en **movimientos_inventario** si ejecutaste **patch_013_movimientos_inventario.sql**."
            )
            try:
                _pm = (
                    sb.table("productos")
                    .select("id,codigo,descripcion,stock_actual,es_compuesto")
                    .eq("activo", True)
                    .order("descripcion")
                    .limit(2000)
                    .execute()
                )
                _rows_m = _pm.data or []
            except Exception:
                try:
                    _pm = (
                        sb.table("productos")
                        .select("id,codigo,descripcion,stock_actual")
                        .eq("activo", True)
                        .order("descripcion")
                        .limit(2000)
                        .execute()
                    )
                    _rows_m = _pm.data or []
                except Exception as exm:
                    _rows_m = []
                    st.error(str(exm))
            if not _rows_m:
                st.warning("No hay productos activos.")
            else:
                _lab_m: dict[str, str] = {}
                for _rm in _rows_m:
                    _im = str(_rm.get("id") or "").strip()
                    if not _im:
                        continue
                    if _rm.get("es_compuesto"):
                        continue
                    _cm = _export_cell_txt(_rm.get("codigo")) or "—"
                    _dm = (_export_cell_txt(_rm.get("descripcion")) or "")[:48]
                    _stm = _inv_stock_int(_rm.get("stock_actual"))
                    _lab_m[f"{_cm} · {_dm} · stock {_stm}"] = _im
                _keys_m = sorted(_lab_m.keys(), key=str.casefold)
                if not _keys_m:
                    st.info(
                        "No hay ítems simples para carga/descarga manual: los **kits** no se listan acá "
                        "(el stock del kit lo movés vendiendo el kit o ajustando cada **componente**)."
                    )
                else:
                    with st.form(f"inv_form_mov_stock_{int(st.session_state.get('inv_mov_stock_form_nonce', 0))}"):
                        _sel_m = st.selectbox("Producto", options=_keys_m, key="inv_mov_prod_sel")
                        _tipo_m = st.radio("Movimiento", ["Entrada", "Salida"], horizontal=True, key="inv_mov_tipo")
                        _cant_m = st.number_input(
                            "Cantidad (unidades)",
                            min_value=1,
                            value=1,
                            step=1,
                            format="%d",
                            key="inv_mov_cant",
                        )
                        _mot_m = st.text_area(
                            "Motivo (obligatorio)",
                            key="inv_mov_mot",
                            placeholder="Ej. Inventario físico, merma, hallazgo en bodega…",
                        )
                        if st.form_submit_button("Aplicar movimiento"):
                            _pid_m = _lab_m.get(str(_sel_m))
                            if not _pid_m:
                                st.error("Elegí un producto.")
                            elif not (_mot_m or "").strip():
                                st.error("El **motivo** es obligatorio.")
                            else:
                                _ok_m, _msg_m = _inv_aplicar_movimiento_stock(
                                    sb,
                                    erp_uid,
                                    _pid_m,
                                    str(_tipo_m),
                                    int(_cant_m),
                                    str(_mot_m),
                                )
                                if _ok_m:
                                    st.success(_msg_m)
                                    _movi_ss_pop_keys("inv_mov_prod_sel", "inv_mov_tipo", "inv_mov_cant", "inv_mov_mot")
                                    _movi_bump_form_nonce("inv_mov_stock_form_nonce")
                                    st.rerun()
                                else:
                                    st.error(_msg_m)
                try:
                    _hist = (
                        sb.table("movimientos_inventario")
                        .select("created_at,tipo,cantidad,motivo,stock_antes,stock_despues,producto_id")
                        .order("created_at", desc=True)
                        .limit(25)
                        .execute()
                    )
                    if _hist.data:
                        st.markdown("##### Últimos movimientos")
                        st.dataframe(pd.DataFrame(_hist.data), use_container_width=True, hide_index=True)
                except Exception:
                    pass

    _ts_inv_md = _backup_file_timestamp()
    with st.expander("Inventario activo en Markdown (.md) — para asistente / IA", expanded=False):
        st.caption(
            "Mismo contenido que en **Reportes → Inventario**: tabla en Markdown con **solo productos activos**, "
            "stock, **precio en USD** y texto de **disponibilidad** por ítem."
        )
        try:
            _md_inv = inventario_activos_markdown_ia(sb, t)
            st.download_button(
                label=f"Descargar — inventario_ia_{_ts_inv_md}.md",
                data=_md_inv.encode("utf-8"),
                file_name=f"inventario_ia_{_ts_inv_md}.md",
                mime="text/markdown",
                key="inv_dl_md_ia",
                use_container_width=True,
            )
        except Exception as e:
            st.error(str(e))

    st.caption(
        "CSV: obligatorias **codigo** (en import no hay autogenerado; ponelos vos), **descripcion**, **stock_actual**, "
        "**stock_minimo**, **costo_usd**, **precio_v_usd**. Opcional: **categoria**, **sku_oem**, **marca_producto**, "
        "**condicion**, **ubicacion**, **marcas_vehiculo**, **años**, **imagen_url**."
    )
    up = st.file_uploader("CSV", type=["csv"], key="inv_csv_upload")
    if up is not None:
        df_csv = pd.read_csv(up)
        required = {
            "codigo",
            "descripcion",
            "stock_actual",
            "stock_minimo",
            "costo_usd",
            "precio_v_usd",
        }
        if not required.issubset(set(df_csv.columns.str.lower())):
            st.error(f"Faltan columnas. Requeridas: {required}")
        else:
            df_csv.columns = [c.lower() for c in df_csv.columns]
            if st.button("Insertar filas", key="inv_csv_insert"):
                rows_csv = df_csv.to_dict(orient="records")
                batch_ins: list[dict[str, Any]] = []
                err_cat: list[str] = []
                for i, row in enumerate(rows_csv):
                    cid_csv: str | None = None
                    if "categoria" in df_csv.columns and row.get("categoria") is not None:
                        raw_c = str(row["categoria"]).strip()
                        if raw_c and raw_c.lower() not in ("nan", "none"):
                            cid_csv = _resolve_categoria_id_por_nombre(raw_c, _nombre_a_id_cat)
                            if cid_csv is None:
                                err_cat.append(f"fila {i + 1}: categoría «{raw_c}» no encontrada")
                    _mv = ""
                    if "marcas_vehiculo" in df_csv.columns and row.get("marcas_vehiculo") is not None:
                        _mv = str(row["marcas_vehiculo"])
                    elif "marcas vehiculo" in df_csv.columns and row.get("marcas vehiculo") is not None:
                        _mv = str(row["marcas vehiculo"])
                    _an = ""
                    if "años" in df_csv.columns and row.get("años") is not None:
                        _an = str(row["años"])
                    elif "anos" in df_csv.columns and row.get("anos") is not None:
                        _an = str(row["anos"])
                    _compat_csv = _inv_build_compat_dict(_mv, _an)
                    _cond_csv = "Nuevo"
                    if "condicion" in df_csv.columns and row.get("condicion") is not None:
                        cs = str(row["condicion"]).strip()
                        if cs in ("Nuevo", "Usado"):
                            _cond_csv = cs
                    _row_ins: dict[str, Any] = {
                        "codigo": str(row["codigo"]).strip() or None,
                        "descripcion": str(row["descripcion"]),
                        "stock_actual": _inv_stock_int(row["stock_actual"]),
                        "stock_minimo": _inv_stock_int(row["stock_minimo"]),
                        "costo_usd": float(row["costo_usd"]),
                        "precio_v_usd": float(row["precio_v_usd"]),
                        "categoria_id": cid_csv,
                        "activo": True,
                        "condicion": _cond_csv,
                        "compatibilidad": _compat_csv,
                    }
                    if "sku_oem" in df_csv.columns and row.get("sku_oem") is not None:
                        s = str(row["sku_oem"]).strip()
                        _row_ins["sku_oem"] = s or None
                    if "marca_producto" in df_csv.columns and row.get("marca_producto") is not None:
                        s = str(row["marca_producto"]).strip()
                        _row_ins["marca_producto"] = s or None
                    if "ubicacion" in df_csv.columns and row.get("ubicacion") is not None:
                        s = str(row["ubicacion"]).strip()
                        _row_ins["ubicacion"] = s or None
                    if "imagen_url" in df_csv.columns and row.get("imagen_url") is not None:
                        s = str(row["imagen_url"]).strip()
                        _row_ins["imagen_url"] = s or None
                    batch_ins.append(_row_ins)
                if err_cat:
                    st.error("Revisá el CSV:\n- " + "\n- ".join(err_cat[:12]))
                    if len(err_cat) > 12:
                        st.caption(f"… y {len(err_cat) - 12} errores más.")
                else:
                    _insert_rows_batched(sb, "productos", batch_ins)
                    st.success(f"Insertados {len(batch_ins)} productos.")
                    st.rerun()

    st.divider()
    st.markdown("##### Tabla de productos (edición masiva — al final del módulo)")
    st.caption(
        "Mismos **filtros de búsqueda y categoría** de arriba. Podés editar **varias filas** y guardar una vez. "
        "**OEM** = código de parte; **vehículos** = marcas de carro separadas por coma. "
        "**precio_v_bs_ref** / **costo_bs_ref**: solo referencia en Bs (solo lectura aquí)."
    )
    if df.empty:
        st.info("Cuando cargues productos en la base, la tabla aparecerá acá.")
    elif df_view.empty:
        st.info(
            "No hay filas con el filtro actual. Cambiá la búsqueda o la categoría para ver la tabla, "
            "o usá la pestaña **Editar Productos** en el expander *Productos* más arriba."
        )
    else:
        _inv_skip_cols = {"id", "categoria_id", "compatibilidad"}
        _inv_editor_order = [
            "codigo",
            "sku_oem",
            "descripcion",
            "marca_producto",
            "condicion",
            "vehiculos_compat",
            "años_compat",
            "ubicacion",
            "imagen_url",
            "stock_actual",
            "stock_minimo",
            "costo_usd",
            "precio_v_usd",
            "precio_v_bs_ref",
            "costo_bs_ref",
            "activo",
            "categoria",
        ]
        _editor_cols = [c for c in _inv_editor_order if c in df_view.columns]
        for c in df_view.columns:
            if c not in _editor_cols and c not in _inv_skip_cols:
                _editor_cols.append(c)
        _ed_df = df_view[_editor_cols].copy()
        _disabled_cols = ["id"]
        if "precio_v_bs_ref" in _ed_df.columns:
            _disabled_cols.extend(["precio_v_bs_ref", "costo_bs_ref"])
        _inv_col_cfg_tbl: dict[str, Any] = {
            "categoria": st.column_config.SelectboxColumn(
                "Categoría",
                options=_cat_select_opts,
            ),
            "condicion": st.column_config.SelectboxColumn(
                "Condición",
                options=["Nuevo", "Usado"],
                required=True,
            ),
            "stock_actual": st.column_config.NumberColumn(
                "Stock",
                min_value=0,
                step=1,
                format="%d",
            ),
            "stock_minimo": st.column_config.NumberColumn(
                "Stock mín.",
                min_value=0,
                step=1,
                format="%d",
            ),
        }
        edited = st.data_editor(
            _ed_df,
            num_rows="fixed",
            disabled=_disabled_cols,
            column_config=_inv_col_cfg_tbl,
            use_container_width=True,
            key="editor_prod",
        )
        if st.button("Guardar cambios de inventario", key="inv_btn_guardar_tabla"):
            for _, row in edited.iterrows():
                _cv = row.get("categoria")
                if _cv is None or (isinstance(_cv, float) and pd.isna(_cv)):
                    _cv = ""
                _cv = str(_cv).strip()
                _cid_up = _nombre_a_id_cat.get(_cv) if _cv else None
                _cond = row.get("condicion")
                if _cond not in ("Nuevo", "Usado"):
                    _cond = "Nuevo"
                _sku = row.get("sku_oem")
                _sku = None if _sku is None or (isinstance(_sku, float) and pd.isna(_sku)) else str(_sku).strip() or None
                _mprod = row.get("marca_producto")
                _mprod = (
                    None
                    if _mprod is None or (isinstance(_mprod, float) and pd.isna(_mprod))
                    else str(_mprod).strip() or None
                )
                _ubi = row.get("ubicacion")
                _ubi = None if _ubi is None or (isinstance(_ubi, float) and pd.isna(_ubi)) else str(_ubi).strip() or None
                _img = row.get("imagen_url")
                _img = None if _img is None or (isinstance(_img, float) and pd.isna(_img)) else str(_img).strip() or None
                _compat = _inv_build_compat_dict(
                    str(row.get("vehiculos_compat") or ""),
                    str(row.get("años_compat") or ""),
                )
                _upd: dict[str, Any] = {
                    "codigo": None
                    if row.get("codigo") is None
                    or (isinstance(row.get("codigo"), float) and pd.isna(row.get("codigo")))
                    else (str(row.get("codigo")).strip() or None),
                    "descripcion": str(row.get("descripcion") or ""),
                    "stock_actual": _inv_stock_int(row.get("stock_actual")),
                    "stock_minimo": _inv_stock_int(row.get("stock_minimo")),
                    "costo_usd": float(row.get("costo_usd", 0)),
                    "precio_v_usd": float(row.get("precio_v_usd", 0)),
                    "activo": bool(row.get("activo", True)),
                    "categoria_id": _cid_up,
                    "condicion": _cond,
                    "compatibilidad": _compat,
                }
                if "sku_oem" in df_view.columns:
                    _upd["sku_oem"] = _sku
                if "marca_producto" in df_view.columns:
                    _upd["marca_producto"] = _mprod
                if "ubicacion" in df_view.columns:
                    _upd["ubicacion"] = _ubi
                if "imagen_url" in df_view.columns:
                    _upd["imagen_url"] = _img
                try:
                    sb.table("productos").update(_upd).eq("id", str(row["id"])).execute()
                except Exception as ex:
                    st.error(
                        f"Error al guardar fila id={row.get('id')}: {ex}. "
                        "¿Ejecutaste **supabase/patch_011_productos_repuestos.sql** en Supabase?"
                    )
                    st.stop()
            st.success("Productos actualizados.")
            st.rerun()


def module_ventas(sb: Client, erp_uid: str, t: dict[str, Any] | None) -> None:
    _modulo_titulo_info(
        "Ventas y CXC",
        key="ventas",
        ayuda_md=(
            "Flujo: **cliente y condición** → **productos** (el stock baja al registrar) → **totales** → **cobros**. "
            "Montos en **USD** (1:1)."
        ),
    )
    if not t:
        st.stop()

    _okb = st.session_state.get("venta_ok_banner")
    if isinstance(_okb, dict) and _okb.get("id"):
        _n = _okb.get("numero")
        _tid = str(_okb.get("id") or "")
        _tot = float(_okb.get("total_usd") or 0)
        _cli = str(_okb.get("cliente") or "—")
        _fp = str(_okb.get("forma_pago") or "—")
        st.success(
            f"**Venta registrada** · Nº interno **{_n or '—'}** · Total **US$ {_tot:,.2f}** · "
            f"Cliente: {_cli} · Forma: `{_fp}`"
        )
        st.caption(f"ID en base: `{_tid}` — también podés verla en **Reportes → Ventas**.")
        if (_okb.get("extra") or "").strip():
            st.caption(str(_okb.get("extra")).strip())
        st.info(
            "**Para no duplicar:** no vuelvas a pulsar **Registrar venta** si ya ves este aviso. "
            "Si te equivocaste, **Mantenimiento → Anular venta** (superusuario)."
        )
        if st.button("Ocultar aviso y seguir", key="venta_ok_banner_dismiss"):
            del st.session_state["venta_ok_banner"]
            st.rerun()
        st.divider()

    t_usdt = float(t["tasa_usdt"])

    try:
        prods = (
            sb.table("productos")
            .select(
                "id,descripcion,precio_v_usd,stock_actual,es_compuesto,categoria_id,compatibilidad,categorias(nombre)"
            )
            .eq("activo", True)
            .order("descripcion")
            .execute()
        )
    except Exception:
        try:
            prods = (
                sb.table("productos")
                .select("id,descripcion,precio_v_usd,stock_actual,es_compuesto,compatibilidad")
                .eq("activo", True)
                .order("descripcion")
                .execute()
            )
        except Exception:
            prods = (
                sb.table("productos")
                .select("id,descripcion,precio_v_usd,stock_actual,compatibilidad")
                .eq("activo", True)
                .order("descripcion")
                .execute()
            )
    plist = prods.data or []
    for _p in plist:
        _p.setdefault("es_compuesto", False)

    def _venta_cat_nombre(p: dict[str, Any]) -> str:
        c = p.get("categorias")
        if isinstance(c, dict):
            return str(c.get("nombre") or "").strip()
        return ""

    def _venta_pide_seriales_motor(p: dict[str, Any]) -> bool:
        if _inv_categoria_sugiere_seriales_motor(_venta_cat_nombre(p)):
            return True
        if _inv_compat_seriales_motor_list(_inv_compat_as_dict(p.get("compatibilidad"))):
            return True
        return False

    def _venta_validar_seriales_motor_lineas(lines: list[dict[str, Any]]) -> str | None:
        for ln in lines:
            pid = str(ln.get("producto_id") or "")
            p = next((x for x in plist if str(x.get("id")) == pid), None)
            if not p or not _venta_pide_seriales_motor(p):
                continue
            n = int(ln.get("cantidad") or 0)
            srl = [str(s).strip() for s in (ln.get("seriales") or []) if str(s).strip()]
            if len(srl) != n:
                return (
                    f"**Seriales:** en esta línea hacen falta tantos números como unidades vendidas. "
                    f"Producto «{str(p.get('descripcion') or '')[:48]}»: necesitás **{n}** valor(es)."
                )
            if len(set(srl)) != len(srl):
                return "**Seriales:** no repetir el mismo número en una misma línea de venta."
            pool = _inv_compat_seriales_motor_list(_inv_compat_as_dict(p.get("compatibilidad")))
            for sv in srl:
                if not _venta_serial_en_pool_motor(sv, pool):
                    return (
                        f"**Serial `{sv}`:** no está cargado en **Inventario** para «{str(p.get('descripcion') or '')[:48]}». "
                        "Abrí **Inventario → editar ese producto** y agregá ese número en **Números de serie** (debe coincidir con lo que vendés). "
                        "Después volvé a registrar la venta."
                    )
        return None

    if not plist:
        st.warning("No hay productos activos.")
        st.stop()

    _kit_items_v = _fetch_kit_items_by_kit(sb)
    _stock_v = {str(p["id"]): _inv_stock_int(p.get("stock_actual")) for p in plist}

    def _venta_prod_label(p: dict[str, Any]) -> str:
        pid = str(p["id"])
        if p.get("es_compuesto") and _kit_items_v.get(pid):
            k = _kit_cantidad_armable(_stock_v, _kit_items_v[pid])
            return f"{p['descripcion']} · kit (~{k} armables)"
        return f"{p['descripcion']} (stock {p['stock_actual']})"

    id_to_label = {str(p["id"]): _venta_prod_label(p) for p in plist}
    id_to_price = {str(p["id"]): float(p["precio_v_usd"]) for p in plist}

    caja_rows_act = _cajas_fetch_rows(sb, solo_activas=True)
    caja_ids, caja_fmt = _caja_select_options(caja_rows_act) if caja_rows_act else ([], lambda x: str(x))

    if "venta_lines" not in st.session_state:
        st.session_state["venta_lines"] = [
            {"producto_id": str(plist[0]["id"]), "cantidad": 1, "precio_unitario_usd": id_to_price[str(plist[0]["id"])]}
        ]

    st.session_state.setdefault("venta_n_cobros", 1)

    st.markdown("### 1. Cliente y condición de venta")
    c_cli, c_not = st.columns([1, 1])
    with c_cli:
        cliente = st.text_input("Cliente", key="venta_cli", autocomplete="off")
    with c_not:
        notas = st.text_area(
            "Notas (opcional)",
            key="venta_notas",
            help="Apartado, entrega en taller, teléfono, etc.",
            height=100,
        )
    c_f1, c_f2 = st.columns([1, 1])
    with c_f1:
        forma = st.selectbox(
            "Forma de pago",
            ["contado", "credito"],
            key="venta_forma",
            help="**Crédito:** el cliente se lleva la mercancía y debe el saldo hasta la fecha límite. "
            "Podés marcar abono el mismo día (apartado con seña). **Contado:** debe pagar todo en el acto.",
        )
    with c_f2:
        fv = st.date_input(
            "Fecha límite para saldar (solo venta a crédito)",
            value=date.today() + timedelta(days=30),
            key="venta_fv",
            disabled=(forma != "credito"),
            help="Solo aplica si la forma de pago es **crédito** (CXC).",
        )

    st.markdown("### 2. Productos e inventario")
    st.caption(
        "Al cambiar producto o cantidad, **seriales** y subtotales se actualizan al instante. "
        "El **stock** se descuenta al pulsar **Registrar venta**."
    )
    _ba, _bb = st.columns(2)
    with _ba:
        if st.button("➕ Añadir línea de producto", key="venta_btn_add_line"):
            st.session_state["venta_lines"].append(
                {
                    "producto_id": str(plist[0]["id"]),
                    "cantidad": 1,
                    "precio_unitario_usd": id_to_price[str(plist[0]["id"])],
                }
            )
            st.rerun()
    with _bb:
        if st.button(
            "Limpiar pantalla (nueva venta)",
            key="venta_btn_clear",
            help="Borra cliente, notas, líneas y cobros en pantalla sin guardar en la base.",
        ):
            _movi_reset_venta_session_nueva(plist, id_to_price)
            st.rerun()

    st.caption("Líneas (montos en USD)")
    new_lines: list[dict[str, Any]] = []
    for i, line in enumerate(st.session_state["venta_lines"]):
        c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
        pid = c1.selectbox(
            f"Producto {i+1}",
            options=list(id_to_label.keys()),
            format_func=lambda x: id_to_label[x],
            key=f"vp_{i}",
            index=list(id_to_label.keys()).index(line["producto_id"]) if line["producto_id"] in id_to_label else 0,
        )
        qty = c2.number_input(
            "Cant.",
            min_value=1,
            value=_line_qty_int(line.get("cantidad"), default=1),
            step=1,
            format="%d",
            key=f"vq_{i}",
        )
        pu = c3.number_input("P.U. USD", min_value=0.0, value=float(id_to_price.get(pid, 0)), format="%.2f", key=f"vpu_{i}")
        row_line: dict[str, Any] = {"producto_id": pid, "cantidad": int(qty), "precio_unitario_usd": float(pu)}
        _p_sel = next((x for x in plist if str(x.get("id")) == str(pid)), None)
        if _p_sel and _venta_pide_seriales_motor(_p_sel):
            _pool_s = _inv_compat_seriales_motor_list(_inv_compat_as_dict(_p_sel.get("compatibilidad")))
            _nq = int(qty)
            if _pool_s and len(_pool_s) >= _nq:
                st.caption(
                    f"**Seriales (línea {i + 1}):** elegí **exactamente {_nq}** de la lista (cargados en **Inventario**)."
                )
                _sel_ser = st.multiselect(
                    f"Seriales en stock — línea {i + 1}",
                    options=_pool_s,
                    default=[],
                    key=f"vsrl_ms_{i}_{pid}",
                    max_selections=_nq,
                    help=f"Seleccioná {_nq} número(s) distintos. Solo aparecen los que están en la ficha del producto.",
                )
                if len(_sel_ser) == _nq and len(set(_sel_ser)) == len(_sel_ser):
                    row_line["seriales"] = list(_sel_ser)
                elif _sel_ser and len(_sel_ser) != _nq:
                    st.caption(f"Elegí **{_nq}** seriales (marcados: {len(_sel_ser)}).")
            elif _pool_s and len(_pool_s) < _nq:
                st.error(
                    f"**Seriales (línea {i + 1}):** en inventario hay **{len(_pool_s)}** serial(es) cargado(s) y la cantidad es **{_nq}**. "
                    "Bajá la cantidad o cargá más seriales en **Inventario**."
                )
            else:
                st.warning(
                    f"**Seriales (línea {i + 1}):** no hay seriales cargados en la ficha del producto. "
                    "Cargalos en **Inventario → editar producto** o escribí abajo (solo si no podés usar la lista)."
                )
                _srl_v = st.text_area(
                    f"Seriales línea {i + 1} (texto libre)",
                    height=70,
                    key=f"vsrl_{i}",
                    placeholder="Uno por línea o separados por coma",
                    label_visibility="collapsed",
                )
                _srl_list = _inv_parse_seriales_motor_texto(str(_srl_v))
                if _srl_list:
                    row_line["seriales"] = _srl_list
        new_lines.append(row_line)

    est_total = round(sum(float(l["cantidad"]) * float(l["precio_unitario_usd"]) for l in new_lines), 2)

    st.markdown("### 3. Tasas y referencia en bolívares")
    doc_tasa = st.radio(
        "Referencia Bs/USD (para equivalentes y cobros en bolívares):",
        options=DOC_TASA_BS_OPTS,
        index=_infer_tasa_bs_oper_index(t),
        horizontal=True,
        key="venta_doc_tasa_bs",
        help="Sugiere la tasa BCV o P2P. Si cobraste Bs a otra tasa, ajustala abajo.",
    )

    try:
        t_bs_sugerida = _tasa_bs_para_documento(t, usar_bcv=(doc_tasa == DOC_TASA_BS_OPTS[0]))
    except Exception:
        t_bs_sugerida = 0.0

    t_bs_doc_live = st.number_input(
        "¿A qué tasa recibiste los bolívares? (Bs por 1 USD)",
        min_value=0.0,
        value=float(t_bs_sugerida or 0.0),
        format="%.2f",
        key="venta_tasa_bs_override",
        help=(
            "Tasa real para cobrar en VES y cuadrar montos. Se guarda en la venta."
        ),
    )

    st.markdown("### 4. Totales de esta venta")
    m1, m2 = st.columns(2)
    with m1:
        st.metric("Total venta (USD)", f"{est_total:,.2f}")
    with m2:
        st.caption("Equivalentes según tasas del operador y la tasa de arriba.")

    try:
        _tb_bcv_v = _tasa_bs_para_documento(t, usar_bcv=True)
        _tb_p2p_v = _tasa_bs_para_documento(t, usar_bcv=False)
        _tb_doc_v = float(t_bs_doc_live) if float(t_bs_doc_live or 0) > 0 else None
    except ValueError:
        _tb_bcv_v = _tb_p2p_v = _tb_doc_v = None

    if _tb_bcv_v is not None and _tb_p2p_v is not None:
        _bs_bcv = est_total * _tb_bcv_v
        _bs_p2p = est_total * _tb_p2p_v
        st.info(
            f"**Equivalente en bolívares:** **BCV** {_bs_bcv:,.2f} Bs (@ {_tb_bcv_v:,.2f}) · "
            f"**P2P** {_bs_p2p:,.2f} Bs (@ {_tb_p2p_v:,.2f})."
        )
        if _tb_doc_v is not None:
            _bs_doc = est_total * _tb_doc_v
            st.markdown(
                f"Con la **tasa aplicada** ({doc_tasa}): **{_bs_doc:,.2f} Bs** "
                f"(US$ {est_total:,.2f} × {_tb_doc_v:,.2f} Bs/USD)."
            )
    elif _tb_doc_v is not None:
        st.caption(
            f"Equivalente Bs del total (tasa **{doc_tasa}**): **{est_total * _tb_doc_v:,.2f} Bs** "
            f"(@ {_tb_doc_v:,.2f} Bs/USD)."
        )

    st.markdown("### 5. Cobros y registro")
    _bp1, _bp2, _bp3 = st.columns([1, 1, 2])
    with _bp1:
        if st.button(
            "➕ Otra forma de pago",
            key="venta_btn_mas_cobro",
            help="Hasta 10 filas: una caja y un monto por medio (Zelle, efectivo, Bs, USDT…).",
        ):
            st.session_state["venta_n_cobros"] = min(10, int(st.session_state.get("venta_n_cobros", 1)) + 1)
            st.rerun()
    with _bp2:
        if st.button(
            "↺ Una sola fila de cobro",
            key="venta_btn_una_cobro",
            help="Vuelve a un solo medio de pago.",
        ):
            st.session_state["venta_n_cobros"] = 1
            st.rerun()
    with _bp3:
        with st.expander("Ayuda: pago mixto y cuadre"):
            st.caption(
                "Podés combinar **varias cajas y monedas** en la misma venta. La **suma en USD equivalente** "
                "de todas las filas debe coincidir con el total (±0,05). **Zelle** cuenta como USD 1:1."
            )

    with st.form(f"f_venta_{int(st.session_state.get('venta_form_nonce', 0))}"):
        cobros_pl: list[dict[str, Any]] = []
        if forma == "contado":
            st.markdown("**Cobro al contado — una o varias cajas / medios**")
            st.caption(
                f"Total venta **US$ {est_total:,.2f}**. Usá **Otra forma de pago** si necesitás más filas. "
                "La suma en **USD equivalente** debe cuadrar con el total (±0,05). En *Nota de tesorería* podés detallar referencia."
            )
            if not caja_ids:
                st.error("No hay cajas activas. Cree una en el módulo Cajas.")
            n_cob = int(st.session_state.get("venta_n_cobros", 1))
            for i in range(n_cob if caja_ids else 0):
                r1, r2, r3 = st.columns([2, 1, 1])
                cid = r1.selectbox(
                    f"Caja cobro {i + 1}",
                    options=caja_ids,
                    format_func=caja_fmt,
                    key=f"vcb_ck_{i}",
                )
                mon = r2.selectbox(
                    "Forma / moneda",
                    options=list(COBRO_MONEDAS),
                    format_func=_fmt_moneda_cobro,
                    key=f"vcb_mon_{i}",
                    help="Monto en la moneda elegida. Zelle = USD vía Zelle.",
                )
                if n_cob == 1 and i == 0 and mon in ("USD", "ZELLE"):
                    default_m = float(est_total)
                elif n_cob == 1 and i == 0 and mon == "VES" and _tb_doc_v is not None and float(_tb_doc_v) > 0:
                    default_m = round(float(est_total) * float(_tb_doc_v), 2)
                else:
                    default_m = 0.0
                _fmt_am = "%.2f" if mon in ("USD", "ZELLE", "VES") else "%.4f"
                _monto_help = (
                    f"En **VES**, ingresá bolívares cobrados. Referencia total: **{est_total * float(_tb_doc_v):,.2f} Bs** "
                    f"con tasa **{doc_tasa}** ({float(_tb_doc_v):,.2f} Bs/USD × US$ {est_total:,.2f})."
                    if mon == "VES" and _tb_doc_v is not None
                    else "Monto en la moneda elegida."
                )
                mval = r3.number_input(
                    f"Monto ({mon})",
                    min_value=0.0,
                    value=default_m,
                    format=_fmt_am,
                    key=f"vcb_mv_{i}",
                    help=_monto_help,
                )
                nota_cob = st.text_input(
                    "Nota de tesorería (opcional)",
                    placeholder="Ej.: Cliente pagó en Bs; cambié a USD en caja fuerte el … / Zelle recibido de …",
                    key=f"vcb_no_{i}",
                )
                _row_c: dict[str, Any] = {"caja_id": str(cid), "moneda": mon, "monto": float(mval)}
                if (nota_cob or "").strip():
                    _row_c["nota_operacion"] = (nota_cob or "").strip()
                cobros_pl.append(_row_c)

            if cobros_pl and _tb_doc_v is not None:
                _sum_eq_v = round(
                    sum(
                        _monto_nativo_a_usd(r["moneda"], float(r["monto"]), float(_tb_doc_v), t_usdt)
                        for r in cobros_pl
                    ),
                    2,
                )
                _dif_v = round(_sum_eq_v - est_total, 2)
                st.markdown(
                    f"**Resumen de cobros:** equivalente **US$ {_sum_eq_v:,.2f}** · Total venta **US$ {est_total:,.2f}** · "
                    f"Diferencia **US$ {_dif_v:+,.2f}**"
                )
                if abs(_dif_v) <= 0.05:
                    st.success("Los cobros cuadran con el total de la venta.")
                else:
                    st.warning(
                        "Ajustá los montos o agregá más filas hasta que la diferencia quede en **±0,05 US$** (redondeos de Bs/USDT)."
                    )

        abono_hoy = False
        if forma == "credito":
            abono_hoy = st.checkbox(
                "El cliente deja **seña o abono hoy** (apartado: entra dinero a caja y el resto queda por cobrar)",
                value=False,
                key="venta_abono_credito",
            )
            if abono_hoy:
                st.markdown("**Dinero que entra hoy (seña o abono) — una o varias cuentas**")
                st.caption(
                    f"Total **US$ {est_total:,.2f}**. El abono en USD equivalente debe ser **menor** que el total; el resto queda en **CXC**. "
                    "Si cobra todo hoy, usá forma **contado**."
                )
                if not caja_ids:
                    st.error("No hay cuentas activas en Cajas. Creá una para registrar el abono.")
                n_ab = int(st.session_state.get("venta_n_cobros", 1))
                for i in range(n_ab if caja_ids else 0):
                    r1, r2, r3 = st.columns([2, 1, 1])
                    cid = r1.selectbox(
                        f"Cuenta del abono {i + 1}",
                        options=caja_ids,
                        format_func=caja_fmt,
                        key=f"vca_ck_{i}",
                    )
                    mon = r2.selectbox(
                        "Forma / moneda",
                        options=list(COBRO_MONEDAS),
                        format_func=_fmt_moneda_cobro,
                        key=f"vca_mon_{i}",
                    )
                    _fmt_aa = "%.2f" if mon in ("USD", "ZELLE") else "%.4f"
                    mval = r3.number_input(
                        f"Monto ({mon})",
                        min_value=0.0,
                        value=0.0,
                        format=_fmt_aa,
                        key=f"vca_mv_{i}",
                    )
                    nota_ab = st.text_input(
                        "Nota de tesorería (opcional)",
                        placeholder="Ej.: Seña en Bs, liquidación luego a Zelle…",
                        key=f"vca_no_{i}",
                    )
                    _row_a: dict[str, Any] = {"caja_id": str(cid), "moneda": mon, "monto": float(mval)}
                    if (nota_ab or "").strip():
                        _row_a["nota_operacion"] = (nota_ab or "").strip()
                    cobros_pl.append(_row_a)

                if cobros_pl and _tb_doc_v is not None:
                    _sum_ab = round(
                        sum(
                            _monto_nativo_a_usd(r["moneda"], float(r["monto"]), float(_tb_doc_v), t_usdt)
                            for r in cobros_pl
                        ),
                        2,
                    )
                    _pend_ab = round(est_total - _sum_ab, 2)
                    st.markdown(
                        f"**Resumen del abono:** equivalente **US$ {_sum_ab:,.2f}** · Total venta **US$ {est_total:,.2f}** · "
                        f"Saldo a financiar ~**US$ {_pend_ab:,.2f}**"
                    )
                    if _sum_ab <= 0:
                        st.warning("Cargá montos en las filas o desmarcá el abono si no entra dinero hoy.")
                    elif _sum_ab >= est_total - 0.05:
                        st.warning("El abono cubre casi todo el total; para eso usá forma de pago **contado**.")
                    else:
                        st.success("El abono es menor al total; el resto quedará en cuentas por cobrar.")

        if st.form_submit_button("Registrar venta (atómica)"):
            _err_srl_v = _venta_validar_seriales_motor_lineas(new_lines)
            if _err_srl_v:
                st.error(_err_srl_v)
            else:
                _sig_v = _venta_firma_registro(cliente, est_total, new_lines)
                _now_v = time.time()
                if (
                    st.session_state.get("venta_last_success_sig") == _sig_v
                    and (_now_v - float(st.session_state.get("venta_last_success_ts", 0))) < 35
                ):
                    st.error(
                        "Acabás de registrar una venta con los mismos datos. Revisá el aviso verde arriba o **Reportes → Ventas**. "
                        "No vuelvas a registrar para no duplicar."
                    )
                    st.stop()
                t_bs_doc = float(t_bs_doc_live or 0)
                if t_bs_doc <= 0:
                    st.error("Tasa Bs/USD inválida para esta venta. Ajustala arriba (por ejemplo la tasa P2P real).")
                else:
                    payload: dict[str, Any] = {
                        "p_usuario_id": erp_uid,
                        "p_cliente": cliente,
                        "p_forma_pago": forma,
                        "p_caja_id": None,
                        "p_tasa_bs": t_bs_doc,
                        "p_tasa_usdt": t_usdt,
                        "p_fecha_vencimiento": str(fv) if forma == "credito" else None,
                        "p_notas": notas,
                        "p_lineas": new_lines,
                    }
                    if forma == "contado":
                        if not caja_ids:
                            st.error("Sin cajas.")
                        else:
                            p_cobros = [
                                {"caja_id": str(row["caja_id"]), "moneda": row["moneda"], "monto": row["monto"]}
                                for row in cobros_pl
                            ]
                            sum_eq = round(
                                sum(
                                    _monto_nativo_a_usd(r["moneda"], r["monto"], t_bs_doc, t_usdt) for r in p_cobros
                                ),
                                2,
                            )
                            if any(r["monto"] <= 0 for r in p_cobros):
                                st.error("Cada línea de cobro debe tener monto > 0.")
                            elif abs(sum_eq - est_total) > 0.05:
                                st.error(
                                    f"Los cobros equivalen a ~**US$ {sum_eq:,.2f}**; el total de la venta es **US$ {est_total:,.2f}**."
                                )
                            else:
                                payload["p_caja_id"] = str(p_cobros[0]["caja_id"])
                                payload["p_cobros"] = p_cobros
                                try:
                                    _resp_v = sb.rpc("crear_venta_erp", payload).execute()
                                    _vid = _rpc_resp_uuid(_resp_v)
                                    if not _vid:
                                        st.error(
                                            "La venta pudo haberse guardado igual. Revisá **Reportes → Ventas** y no registres de nuevo."
                                        )
                                    else:
                                        st.session_state["venta_last_success_sig"] = _sig_v
                                        st.session_state["venta_last_success_ts"] = time.time()
                                        _venta_set_ok_banner_desde_id(
                                            sb,
                                            venta_id=_vid,
                                            cliente=cliente,
                                            est_total=est_total,
                                            forma=forma,
                                        )
                                        _movi_reset_venta_session_nueva(plist, id_to_price)
                                        st.rerun()
                                except Exception as e:
                                    err = _error_msg_from_supabase_exc(e)
                                    if "p_cobros" in err or "could not find" in err.lower():
                                        st.error(
                                            f"{err} · Si falta el parámetro en BD, ejecutá `supabase/patch_008_movimientos_moneda_cobros.sql`."
                                        )
                                    elif "Serial" in err and "no está en el inventario" in err:
                                        st.error(
                                            f"**{err}** · Cargá ese número en **Inventario** (ficha del producto → números de serie) "
                                            "y volvé a registrar la venta."
                                        )
                                    else:
                                        st.error(f"No se pudo registrar: {err}")
                    else:
                        if abono_hoy:
                            if not caja_ids:
                                st.error("Sin cuentas activas para el abono.")
                            elif not cobros_pl:
                                st.error("Marcaste abono hoy pero no hay líneas de cobro.")
                            else:
                                p_cobros = [
                                    {"caja_id": str(row["caja_id"]), "moneda": row["moneda"], "monto": row["monto"]}
                                    for row in cobros_pl
                                ]
                                sum_eq = round(
                                    sum(
                                        _monto_nativo_a_usd(r["moneda"], r["monto"], t_bs_doc, t_usdt)
                                        for r in p_cobros
                                    ),
                                    2,
                                )
                                if any(r["monto"] <= 0 for r in p_cobros):
                                    st.error("Cada línea del abono debe tener monto mayor a cero.")
                                elif sum_eq >= est_total - 0.05:
                                    st.error(
                                        "El abono cubre casi todo el total. Para eso usá **contado**. "
                                        "El abono en crédito debe dejar **algo pendiente** por cobrar."
                                    )
                                elif sum_eq <= 0.05:
                                    st.error(
                                        "El abono en dólares equivalente es casi cero. Quitá el tilde de abono o cargá el monto."
                                    )
                                else:
                                    payload["p_cobros"] = p_cobros
                                    try:
                                        _resp_v = sb.rpc("crear_venta_erp", payload).execute()
                                        _vid = _rpc_resp_uuid(_resp_v)
                                        if not _vid:
                                            st.error(
                                                "La venta pudo haberse guardado igual. Revisá **Reportes → Ventas** y no registres de nuevo."
                                            )
                                        else:
                                            st.session_state["venta_last_success_sig"] = _sig_v
                                            st.session_state["venta_last_success_ts"] = time.time()
                                            _venta_set_ok_banner_desde_id(
                                                sb,
                                                venta_id=_vid,
                                                cliente=cliente,
                                                est_total=est_total,
                                                forma=forma,
                                                extra_caption=(
                                                    f"Crédito con abono: abono ~US$ {sum_eq:,.2f}; "
                                                    f"pendiente ~US$ {est_total - sum_eq:,.2f} en cuentas por cobrar."
                                                ),
                                            )
                                            _movi_reset_venta_session_nueva(plist, id_to_price)
                                            st.rerun()
                                    except Exception as e:
                                        err = _error_msg_from_supabase_exc(e)
                                        if "abono" in err.lower() or "cobros" in err.lower():
                                            st.error(
                                                f"{err} · Si la base aún no acepta abono en crédito, ejecutá "
                                                "`supabase/patch_016_venta_credito_abono_inicial.sql` en Supabase."
                                            )
                                        elif "Serial" in err and "no está en el inventario" in err:
                                            st.error(
                                                f"**{err}** · Cargá ese número en **Inventario** (ficha del producto → números de serie) "
                                                "y volvé a registrar la venta."
                                            )
                                        else:
                                            st.error(f"No se pudo registrar: {err}")
                        else:
                            try:
                                _resp_v = sb.rpc("crear_venta_erp", payload).execute()
                                _vid = _rpc_resp_uuid(_resp_v)
                                if not _vid:
                                    st.error(
                                        "La venta pudo haberse guardado igual. Revisá **Reportes → Ventas** y no registres de nuevo."
                                    )
                                else:
                                    st.session_state["venta_last_success_sig"] = _sig_v
                                    st.session_state["venta_last_success_ts"] = time.time()
                                    _venta_set_ok_banner_desde_id(
                                        sb,
                                        venta_id=_vid,
                                        cliente=cliente,
                                        est_total=est_total,
                                        forma=forma,
                                        extra_caption="Crédito: todo pendiente de cobro (CXC).",
                                    )
                                    _movi_reset_venta_session_nueva(plist, id_to_price)
                                    st.rerun()
                            except Exception as e:
                                err = _error_msg_from_supabase_exc(e)
                                if "Serial" in err and "no está en el inventario" in err:
                                    st.error(
                                        f"**{err}** · Cargá ese número en **Inventario** (ficha del producto → números de serie) "
                                        "y volvé a registrar la venta."
                                    )
                                else:
                                    st.error(f"No se pudo registrar: {err}")

    def _venta_fmt_hora_vz(fecha_val: Any) -> str:
        if fecha_val is None:
            return "—"
        try:
            s = str(fecha_val).strip()
            if s.endswith("Z"):
                s = s[:-1] + "+00:00"
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(ZoneInfo("America/Caracas")).strftime("%H:%M")
        except Exception:
            return str(fecha_val)[:16]

    st.divider()
    st.markdown("### Ventas de hoy (cuadre de caja)")
    _tz_vz = ZoneInfo("America/Caracas")
    _now_vz = datetime.now(_tz_vz)
    _start = _now_vz.replace(hour=0, minute=0, second=0, microsecond=0)
    _end = _start + timedelta(days=1)
    try:
        v_hoy = (
            sb.table("ventas")
            .select("numero,cliente,fecha,total_usd,forma_pago,caja_id")
            .gte("fecha", _start.isoformat())
            .lt("fecha", _end.isoformat())
            .order("fecha", desc=True)
            .execute()
        )
    except Exception as ex:
        v_hoy = None
        st.warning(f"No se pudo cargar las ventas del día: {ex}")

    if v_hoy is not None:
        rows_v = v_hoy.data or []
        st.caption(
            f"**{_now_vz.strftime('%d/%m/%Y')}** (America/Caracas). Sumás para cruzar con efectivo, Zelle y movimientos por caja."
        )
        if not rows_v:
            st.info("Aún no hay ventas registradas hoy.")
        else:
            sum_all = sum(float(r.get("total_usd") or 0) for r in rows_v)
            sum_c = sum(float(r.get("total_usd") or 0) for r in rows_v if r.get("forma_pago") == "contado")
            sum_cr = sum(float(r.get("total_usd") or 0) for r in rows_v if r.get("forma_pago") == "credito")
            k1, k2, k3, k4 = st.columns(4)
            with k1:
                st.metric("Ventas hoy (cantidad)", len(rows_v))
            with k2:
                st.metric("Total USD (todas)", f"{sum_all:,.2f}")
            with k3:
                st.metric("Contado (USD)", f"{sum_c:,.2f}")
            with k4:
                st.metric("Crédito (USD)", f"{sum_cr:,.2f}")

            dfvh: list[dict[str, Any]] = []
            for r in rows_v:
                cid = r.get("caja_id")
                caja_txt = caja_fmt(str(cid)) if cid and caja_ids else "—"
                dfvh.append(
                    {
                        "Hora (VZ)": _venta_fmt_hora_vz(r.get("fecha")),
                        "Nº": int(r.get("numero") or 0),
                        "Cliente": str(r.get("cliente") or "")[:44],
                        "Total USD": float(r.get("total_usd") or 0),
                        "Forma": str(r.get("forma_pago") or "—"),
                        "Caja ref.": caja_txt,
                    }
                )
            st.dataframe(pd.DataFrame(dfvh), use_container_width=True, hide_index=True)

    st.divider()
    st.caption(
        "**Cobrar lo que falta:** cada fila es una venta a crédito (con o sin abono inicial). "
        "Podés cobrar el total pendiente o **solo una parte**; el sistema marca *Parcial* hasta que se liquide."
    )
    cxc = (
        sb.table("cuentas_por_cobrar")
        .select("id, venta_id, monto_pendiente_usd, fecha_vencimiento, estado")
        .in_("estado", ["Pendiente", "Parcial"])
        .execute()
    )
    if cxc.data:
        df = pd.DataFrame(cxc.data)
        st.dataframe(df, use_container_width=True, hide_index=True)
        row_id = st.selectbox("ID CXC", options=[str(x["id"]) for x in cxc.data])
        _cxc_row = next(x for x in cxc.data if str(x["id"]) == row_id)
        pend_usd = float(_cxc_row["monto_pendiente_usd"])
        _vid = str(_cxc_row["venta_id"])
        _tvr = sb.table("ventas").select("tasa_bs,tasa_usdt").eq("id", _vid).limit(1).execute()
        _t_bs_v = float((_tvr.data or [{}])[0].get("tasa_bs") or t["tasa_bs"])
        _t_ut_v = float((_tvr.data or [{}])[0].get("tasa_usdt") or t["tasa_usdt"])
        if not caja_ids:
            st.warning("No hay cajas activas para registrar el cobro.")
        caja_cobro = (
            st.selectbox("Cuenta donde entra el cobro", options=caja_ids, format_func=caja_fmt, key="cxc_caja")
            if caja_ids
            else None
        )
        cxc_mon = st.selectbox(
            "Cómo paga el cliente (moneda o medio)",
            options=list(COBRO_MONEDAS),
            format_func=_fmt_moneda_cobro,
            key="cxc_mon",
            help="El sistema convierte a USD con las **tasas de la venta** original.",
        )
        _def_nat = pend_usd if cxc_mon in ("USD", "ZELLE") else 0.0
        monto_nativo = st.number_input(
            f"Monto en {cxc_mon}",
            min_value=0.01,
            value=float(_def_nat),
            format="%.2f" if cxc_mon in ("USD", "ZELLE") else "%.4f",
            key="cxc_monto_nat",
        )
        sum_eq_cxc = round(_monto_nativo_a_usd(cxc_mon, float(monto_nativo), _t_bs_v, _t_ut_v), 2)
        st.caption(
            f"Equivale a **~ US$ {sum_eq_cxc:,.2f}** en el sistema. Pendiente en esta cuenta: **US$ {pend_usd:,.2f}**."
        )
        nota_cxc = st.text_input(
            "Nota de tesorería (opcional)",
            placeholder="Ej.: Bs recibidos en Banesco; cambiados a efectivo USD / recibí Zelle de …",
            key="cxc_nota_op",
        )
        if sum_eq_cxc > pend_usd + 0.05:
            st.warning("El equivalente en USD es mayor al pendiente; ajustá el monto o la moneda.")
        if st.button("Registrar cobro CXC", disabled=not caja_ids):
            if sum_eq_cxc > pend_usd + 0.05:
                st.error("No se registró: el cobro supera lo pendiente.")
            else:
                try:
                    _pc0: dict[str, Any] = {
                        "caja_id": str(caja_cobro),
                        "moneda": cxc_mon,
                        "monto": float(monto_nativo),
                    }
                    if (nota_cxc or "").strip():
                        _pc0["nota_operacion"] = (nota_cxc or "").strip()
                    sb.rpc(
                        "cobrar_cxc_erp",
                        {
                            "p_usuario_id": erp_uid,
                            "p_cxc_id": row_id,
                            "p_caja_id": str(caja_cobro),
                            "p_monto_usd": float(sum_eq_cxc),
                            "p_cobros": [_pc0],
                        },
                    ).execute()
                    st.success("Cobro registrado.")
                    _movi_ss_pop_keys("cxc_caja", "cxc_mon", "cxc_monto_nat", "cxc_nota_op")
                    st.rerun()
                except Exception as e:
                    err = str(e)
                    if "nota_operacion" in err or "column" in err.lower():
                        st.error(
                            f"{err} · Ejecutá en Supabase `supabase/patch_017_movimientos_nota_zelle.sql` "
                            "y recargá el esquema."
                        )
                    else:
                        st.error(err)
    else:
        st.info("No hay cuentas por cobrar pendientes.")


def _compra_parse_lineas_csv(
    raw: bytes,
    id_to_label: dict[str, str],
) -> tuple[list[dict[str, Any]] | None, str | None]:
    """CSV con columnas: producto_id, cantidad, costo_unitario_usd
    o: descripcion, cantidad, costo_unitario_usd (descripción exacta como en inventario)."""
    try:
        df = pd.read_csv(BytesIO(raw))
    except Exception as ex:
        return None, f"No se pudo leer el CSV: {ex}"
    df.columns = [str(c).strip().lower() for c in df.columns]
    has_pid = "producto_id" in df.columns
    has_desc = "descripcion" in df.columns
    if not has_pid and not has_desc:
        return None, "El CSV debe incluir columna **producto_id** o **descripcion** (y cantidad, costo_unitario_usd)."
    need = {"cantidad", "costo_unitario_usd"}
    if not need.issubset(df.columns):
        return None, f"Faltan columnas obligatorias: {', '.join(sorted(need))}."
    desc_to_id = {v.strip().lower(): k for k, v in id_to_label.items()}
    out: list[dict[str, Any]] = []

    def _resolve_pid(row: Any, ix: int) -> tuple[str | None, str | None]:
        raw_id = row.get("producto_id") if has_pid else None
        if raw_id is not None and str(raw_id).strip() and str(raw_id).strip().lower() != "nan":
            pid = str(raw_id).strip()
            if pid not in id_to_label:
                return None, f"Fila {ix}: producto_id `{pid}` no existe o no está activo."
            return pid, None
        if has_desc:
            d = str(row.get("descripcion") or "").strip().lower()
            pid = desc_to_id.get(d)
            if not pid:
                return None, f"Fila {ix}: descripción «{row.get('descripcion')}» no coincide con un producto activo."
            return pid, None
        return None, f"Fila {ix}: indicá **producto_id** o **descripcion**."

    for _ix, row in df.iterrows():
        row_no = int(_ix) + 2  # encabezado = 1
        pid, err = _resolve_pid(row, row_no)
        if err or not pid:
            return None, err
        try:
            q = int(float(row["cantidad"]))
        except (TypeError, ValueError):
            return None, f"Fila {row_no}: cantidad inválida."
        if q < 1:
            return None, f"Fila {row_no}: cantidad debe ser ≥ 1."
        try:
            cu = float(row["costo_unitario_usd"])
        except (TypeError, ValueError):
            return None, f"Fila {row_no}: costo_unitario_usd inválido."
        if cu < 0:
            return None, f"Fila {row_no}: costo no puede ser negativo."
        out.append({"producto_id": pid, "cantidad": q, "costo_unitario_usd": cu})
    if not out:
        return None, "El archivo no tiene filas de datos."
    return out, None


def module_compras(sb: Client, erp_uid: str, t: dict[str, Any] | None) -> None:
    _modulo_titulo_info(
        "Compras y CXP",
        key="compras",
        ayuda_md=(
            "Podés cargar **varias líneas** (botón abajo o **CSV**). Montos en **USD** (1:1). "
            "Para equivalente en **bolívares**, elegí **BCV** o **P2P Binance**.\n\n"
            "**Líneas de la compra:** **Añadir línea** agrega una fila más. Para facturas con muchos ítems, usá **importar CSV** "
            "(exportá desde Excel con las columnas indicadas en el expander de importación)."
        ),
    )
    if not t:
        st.stop()

    t_usdt = float(t["tasa_usdt"])

    prods = (
        sb.table("productos")
        .select("id,descripcion,costo_usd")
        .eq("activo", True)
        .order("descripcion")
        .execute()
    )
    plist = prods.data or []
    if not plist:
        st.warning("No hay productos.")
        st.stop()

    id_to_label = {str(p["id"]): p["descripcion"] for p in plist}
    id_to_cost = {str(p["id"]): float(p["costo_usd"]) for p in plist}

    caja_rows_act_c = _cajas_fetch_rows(sb, solo_activas=True)
    caja_ids_c, caja_fmt_c = _caja_select_options(caja_rows_act_c) if caja_rows_act_c else ([], lambda x: str(x))

    if "compra_lines" not in st.session_state:
        st.session_state["compra_lines"] = [
            {"producto_id": str(plist[0]["id"]), "cantidad": 1, "costo_unitario_usd": id_to_cost[str(plist[0]["id"])]}
        ]

    st.markdown("#### Datos del documento")
    prov = st.text_input("Proveedor", key="compra_prov", autocomplete="off")
    c_fp, c_caja = st.columns([1, 1])
    with c_fp:
        forma = st.selectbox("Forma de pago compra", ["contado", "credito"], key="forma_compra")
    with c_caja:
        caja_id_compra = (
            st.selectbox("Caja (solo contado)", options=caja_ids_c, format_func=caja_fmt_c, key="caja_compra")
            if caja_ids_c
            else None
        )
    fv = st.date_input("Vencimiento (crédito compra)", value=date.today() + timedelta(days=30), key="fv_compra")
    notas = st.text_area("Notas compra", key="compra_notas")

    doc_tasa_c = st.radio(
        "Tasa Bs/USD para esta compra:",
        options=DOC_TASA_BS_OPTS,
        index=_infer_tasa_bs_oper_index(t),
        horizontal=True,
        key="compra_doc_tasa_bs",
        help="Montos de línea en USD sin cambio; la tasa queda en el registro de compra.",
    )

    st.markdown("#### Líneas de la compra")
    _ba, _bb, _bc = st.columns([1, 1, 2])
    with _ba:
        if st.button("➕ Añadir línea", key="compra_btn_add_line"):
            st.session_state["compra_lines"].append(
                {
                    "producto_id": str(plist[0]["id"]),
                    "cantidad": 1,
                    "costo_unitario_usd": id_to_cost[str(plist[0]["id"])],
                }
            )
            st.rerun()
    with _bb:
        if len(st.session_state["compra_lines"]) > 1 and st.button(
            "➖ Quitar última línea", key="compra_btn_drop_line"
        ):
            st.session_state["compra_lines"].pop()
            st.rerun()

    with st.expander("Importar líneas desde CSV (facturas largas)", expanded=False):
        st.markdown(
            "Columnas requeridas: **`cantidad`**, **`costo_unitario_usd`**, y además **`producto_id`** "
            "(UUID) **o** **`descripcion`** (texto **idéntico** al producto en inventario, sin importar mayúsculas)."
        )
        sample = (
            "descripcion,cantidad,costo_unitario_usd\n"
            f'"{plist[0]["descripcion"]}",2,10.50\n'
        )
        st.download_button(
            "Descargar ejemplo CSV",
            data=sample.encode("utf-8"),
            file_name="ejemplo_compra_lineas.csv",
            mime="text/csv",
            key="compra_dl_sample_csv",
        )
        up = st.file_uploader("Archivo .csv", type=["csv"], key="compra_csv_up")
        if st.button("Importar líneas del archivo", key="compra_csv_apply"):
            if up is None:
                st.error("Elegí un archivo CSV primero.")
            else:
                raw = up.getvalue()
                parsed, err = _compra_parse_lineas_csv(raw, id_to_label)
                if err:
                    st.error(err)
                elif parsed:
                    st.session_state["compra_lines"] = parsed
                    _movi_reset_compra_form_fields()
                    _movi_bump_form_nonce("compra_form_nonce")
                    st.success(f"Se cargaron **{len(parsed)}** líneas. Revisalas abajo y pulsá **Registrar compra**.")
                    st.rerun()

    new_lines: list[dict[str, Any]] = []
    for i, line in enumerate(st.session_state["compra_lines"]):
        c1, c2, c3 = st.columns([3, 1, 1])
        pid = c1.selectbox(
            f"Producto {i+1}",
            options=list(id_to_label.keys()),
            format_func=lambda x: id_to_label[x],
            key=f"cp_{i}",
            index=list(id_to_label.keys()).index(line["producto_id"]) if line["producto_id"] in id_to_label else 0,
        )
        qty = c2.number_input(
            "Cant.",
            min_value=1,
            value=_line_qty_int(line.get("cantidad"), default=1),
            step=1,
            format="%d",
            key=f"cq_{i}",
        )
        cu = c3.number_input(
            "Costo u. USD", min_value=0.0, value=float(line.get("costo_unitario_usd", id_to_cost.get(pid, 0))), format="%.2f", key=f"ccu_{i}"
        )
        new_lines.append({"producto_id": pid, "cantidad": int(qty), "costo_unitario_usd": float(cu)})

    n_lin = len(new_lines)
    tot_usd = sum(float(x["cantidad"]) * float(x["costo_unitario_usd"]) for x in new_lines)
    st.metric("Líneas en esta compra", f"{n_lin} · Total USD {tot_usd:,.2f}")

    with st.form(f"f_compra_{int(st.session_state.get('compra_form_nonce', 0))}"):
        st.caption("Un solo **registro atómico** en base: stock, costos, caja o CXP.")
        if st.form_submit_button("Registrar compra (atómica)"):
            try:
                t_bs_doc = _tasa_bs_para_documento(t, usar_bcv=(doc_tasa_c == DOC_TASA_BS_OPTS[0]))
            except ValueError as e:
                st.error(str(e))
            else:
                payload = {
                    "p_usuario_id": erp_uid,
                    "p_proveedor": prov,
                    "p_forma_pago": forma,
                    "p_caja_id": str(caja_id_compra) if forma == "contado" and caja_id_compra else None,
                    "p_tasa_bs": t_bs_doc,
                    "p_tasa_usdt": t_usdt,
                    "p_fecha_vencimiento": str(fv) if forma == "credito" else None,
                    "p_notas": notas,
                    "p_lineas": new_lines,
                }
                if forma == "contado" and not caja_id_compra:
                    st.error("Elegí una caja activa para compra al contado (o creá una en Cajas).")
                else:
                    try:
                        sb.rpc("crear_compra_erp", payload).execute()
                        st.success("Compra registrada.")
                        st.session_state["compra_lines"] = [
                            {
                                "producto_id": str(plist[0]["id"]),
                                "cantidad": 1,
                                "costo_unitario_usd": id_to_cost[str(plist[0]["id"])],
                            }
                        ]
                        _movi_reset_compra_form_fields()
                        _movi_bump_form_nonce("compra_form_nonce")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo registrar: {e}")


def _movi_fetch_egresos_caja_recientes(sb: Client, *, limit: int = 50) -> tuple[list[dict[str, Any]], bool]:
    """Egresos recientes; intenta id/venta_id/compra_id/moneda (patch_008/028) y categoria_gasto (patch_025)."""
    attempts: list[tuple[str, bool]] = [
        (
            "created_at,caja_id,monto_usd,concepto,referencia,categoria_gasto,nota_operacion,moneda,monto_moneda,monto_bs,monto_usdt,monto_usd_caja,id,venta_id,compra_id",
            True,
        ),
        (
            "created_at,caja_id,monto_usd,concepto,referencia,categoria_gasto,nota_operacion,moneda,monto_moneda,id,venta_id,compra_id",
            True,
        ),
        (
            "created_at,caja_id,monto_usd,concepto,referencia,categoria_gasto,nota_operacion,id,venta_id,compra_id",
            True,
        ),
        ("created_at,caja_id,monto_usd,concepto,referencia,categoria_gasto,nota_operacion", True),
        ("created_at,caja_id,monto_usd,concepto,referencia,nota_operacion", False),
    ]
    for sel, tiene_cat in attempts:
        try:
            r = (
                sb.table("movimientos_caja")
                .select(sel)
                .eq("tipo", "Egreso")
                .order("created_at", desc=True)
                .limit(limit)
                .execute()
            )
            rows = list(r.data or [])
            for row in rows:
                for k in (
                    "moneda",
                    "monto_moneda",
                    "monto_bs",
                    "monto_usdt",
                    "monto_usd_caja",
                    "id",
                    "venta_id",
                    "compra_id",
                    "categoria_gasto",
                ):
                    row.setdefault(k, None)
            return rows, tiene_cat
        except Exception:
            continue
    return [], False


def _gasto_op_fmt_monto_tabla(r: dict[str, Any], cmap: dict[str, dict[str, Any]]) -> str:
    musd = float(r.get("monto_usd") or 0)
    ex_bk, ex_amt = _movimiento_monto_explicito_columnas(r)
    if ex_bk == "VES" and ex_amt is not None:
        return f"Bs {_round_money_2(ex_amt):,.2f} (≈ US$ {_round_money_2(musd):,.2f})"
    if ex_bk == "USDT" and ex_amt is not None:
        return f"USDT {_round_money_2(ex_amt):,.2f} (≈ US$ {_round_money_2(musd):,.2f})"
    if ex_bk == "USD" and ex_amt is not None:
        return f"US$ {_round_money_2(ex_amt):,.2f} (equiv. motor {_round_money_2(musd):,.2f})"
    mm = r.get("monto_moneda")
    rmon = str(r.get("moneda") or "").strip().upper()
    if mm is not None and str(mm).strip() != "":
        try:
            mv = float(mm)
        except (TypeError, ValueError):
            mv = None
        if mv is not None and mv > 0:
            lab = {"VES": "Bs", "USD": "US$", "USDT": "USDT"}.get(rmon, rmon or "?")
            return f"{lab} {_round_money_2(mv):,.2f} (≈ US$ {_round_money_2(musd):,.2f})"
    cid = str(r.get("caja_id") or "")
    cj = cmap.get(cid) or {}
    cmon_raw = str(cj.get("moneda_cuenta") or "").strip().upper()
    cmon = "VES" if cmon_raw in ("VES", "BS") else cmon_raw
    if cmon == "VES":
        return f"VES · equiv. US$ {_round_money_2(musd):,.2f}"
    if cmon == "USDT":
        return f"USDT · equiv. US$ {_round_money_2(musd):,.2f}"
    return f"US$ {_round_money_2(musd):,.2f}"


def module_gastos_operativos(sb: Client, erp_uid: str, t: dict[str, Any] | None) -> None:
    _modulo_titulo_info(
        "Gastos operativos",
        key="gastos_op",
        ayuda_md=(
            "Salidas de efectivo que **no** son compra de mercancía para inventario (eso va en **Compras / CXP**). "
            "El **monto** se ingresa en la **moneda de la cuenta** (Bs, US$ o USDT); el sistema guarda también el **equivalente USD** para el saldo. "
            "Podés **corregir** egresos **manuales** (sin venta/compra ligada) en el expander. "
            "Ejecutá **`patch_025`** (categoría), **`patch_028`** (moneda nativa + corrección) y **`patch_030`** (columnas Bs / USDT / USD de cuenta en movimientos) en Supabase cuando corresponda."
        ),
    )

    caja_rows = _cajas_fetch_rows(sb, solo_activas=True)
    caja_ids, caja_fmt = _caja_select_options(caja_rows) if caja_rows else ([], lambda x: str(x))
    if not caja_ids:
        st.warning("No hay cajas activas. Creá una en **Cajas y bancos**.")
        st.stop()

    cmap_full = {str(c["id"]): c for c in _cajas_fetch_rows(sb, solo_activas=False)}
    cmap_lab = {str(c["id"]): _caja_etiqueta_lista(c) for c in caja_rows}

    cat_opts = GASTO_OPERATIVO_CATEGORIAS + (GASTO_OPERATIVO_OTRO,)

    with st.form(f"f_gasto_op_{int(st.session_state.get('gasto_op_form_nonce', 0))}"):
        cat = st.selectbox("Categoría", options=cat_opts, key="gasto_op_cat")
        cat_custom = ""
        if cat == GASTO_OPERATIVO_OTRO:
            cat_custom = st.text_input("Nombre de la categoría", key="gasto_op_cat_custom", placeholder="Ej.: Suscripción software")
        desc = st.text_input("Descripción / detalle", key="gasto_op_desc", placeholder="Ej.: Alquiler enero local principal")
        cid = st.selectbox("Caja de donde sale el dinero", options=caja_ids, format_func=caja_fmt, key="gasto_op_caja")
        acc = cmap_full.get(str(cid), {}) or {}
        _mc_raw = acc.get("moneda_cuenta")
        if _mc_raw is None or not str(_mc_raw).strip():
            st.warning(
                "Esta caja **no tiene moneda de cuenta** definida en **Cajas y bancos**. "
                "El formulario tratará la cuenta como **USD** y **no** guardará el monto en **bolívares** en el movimiento. "
                "Editá la caja y elegí **VES**, **USD** o **USDT**."
            )
        mon_cuenta = str(_mc_raw or "USD").strip().upper()
        if mon_cuenta not in ("VES", "USD", "USDT"):
            st.warning(
                f"Moneda de cuenta **{mon_cuenta}** no estándar; usá **VES**, **USD** o **USDT** en la ficha de la caja para cobros y gastos coherentes."
            )
        monto_usd_calc = 0.01
        p_mon: str | None = None
        p_mm: float | None = None
        if mon_cuenta == "VES":
            m_bs = st.number_input("Monto en bolívares (Bs)", min_value=0.01, format="%.2f", key="gasto_op_m_bs")
            try:
                _tb0 = float(_tasa_bs_para_documento(t, usar_bcv=False)) if t else 0.0
            except (ValueError, TypeError):
                _tb0 = 0.0
            tasa_bs = st.number_input(
                "Tasa Bs por 1 USD (para equivalente en sistema)",
                min_value=0.01,
                value=float(_tb0) if _tb0 > 0 else 300.0,
                format="%.4f",
                key="gasto_op_tasa_bs",
                help="Solo para convertir a USD interno; el gasto se muestra en Bs en reportes.",
            )
            monto_usd_calc = float(m_bs) / float(tasa_bs) if tasa_bs > 0 else 0.0
            p_mon, p_mm = "VES", float(m_bs)
            st.caption(f"Equiv. en sistema: **US$ {_round_money_2(monto_usd_calc):,.2f}** (Bs ÷ tasa).")
        elif mon_cuenta == "USDT":
            m_ut = st.number_input("Monto en USDT", min_value=0.0001, format="%.4f", key="gasto_op_m_ut")
            t_ut = float(_nf(t.get("tasa_usdt")) or 1.0) if t else 1.0
            if t_ut <= 0:
                t_ut = 1.0
            monto_usd_calc = float(m_ut) / t_ut
            p_mon, p_mm = "USDT", float(m_ut)
            st.caption(
                f"Equiv. en sistema: **US$ {_round_money_2(monto_usd_calc):,.2f}** (USDT ÷ tasa USDT/USD **{t_ut:,.6f}** del día en BD)."
            )
        else:
            m_us = st.number_input("Monto en USD", min_value=0.01, format="%.2f", key="gasto_op_m_usd")
            monto_usd_calc = float(m_us)
            p_mon, p_mm = "USD", float(m_us)
        ref = st.text_input("Referencia (Nº factura, recibo…)", key="gasto_op_ref")
        nota = st.text_input(
            "Nota de tesorería (opcional)",
            key="gasto_op_nota",
            placeholder="Ej.: Pago Banesco / efectivo",
        )
        submitted = st.form_submit_button("Registrar gasto operativo")
        if submitted:
            cat_final = (cat_custom or "").strip() if cat == GASTO_OPERATIVO_OTRO else cat
            if cat == GASTO_OPERATIVO_OTRO and not cat_final:
                st.error("Completá el nombre de la categoría.")
            elif not (desc or "").strip():
                st.error("La descripción es obligatoria.")
            elif monto_usd_calc <= 0:
                st.error("El equivalente USD calculado debe ser mayor que cero (revisá monto y tasa).")
            else:
                payload: dict[str, Any] = {
                    "p_usuario_id": erp_uid,
                    "p_caja_id": str(cid),
                    "p_tipo": "Egreso",
                    "p_monto_usd": float(monto_usd_calc),
                    "p_concepto": (desc or "").strip(),
                    "p_referencia": (ref or "").strip(),
                }
                if (nota or "").strip():
                    payload["p_nota_operacion"] = (nota or "").strip()
                payload_cat = cat_final if cat_final else None
                if payload_cat:
                    payload["p_categoria_gasto"] = payload_cat
                if p_mon and p_mm is not None and p_mm > 0:
                    payload["p_moneda"] = p_mon
                    payload["p_monto_moneda"] = float(p_mm)

                def _try_reg(pl: dict[str, Any]) -> Exception | None:
                    try:
                        sb.rpc("registrar_movimiento_caja_erp", pl).execute()
                        return None
                    except Exception as ex:
                        return ex

                err = _try_reg(payload)
                if err is not None:
                    err_s = _error_msg_from_supabase_exc(err)
                    low = err_s.lower()
                    if "p_moneda" in low or "p_monto_moneda" in low:
                        payload.pop("p_moneda", None)
                        payload.pop("p_monto_moneda", None)
                        err = _try_reg(payload)
                        if err is None:
                            st.warning(
                                "Gasto guardado **sin** moneda nativa en BD. Ejecutá **`supabase/patch_028_mov_caja_moneda_nativa_y_correccion.sql`** "
                                "o revisá que la firma del RPC en Supabase incluya `p_moneda` y `p_monto_moneda`."
                            )
                    if err is not None:
                        err_s = _error_msg_from_supabase_exc(err)
                        low = err_s.lower()
                        if payload_cat and any(
                            x in low for x in ("categoria_gasto", "p_categoria", "could not find", "42883", "42725")
                        ):
                            payload.pop("p_categoria_gasto", None)
                            payload["p_concepto"] = f"[{cat_final}] {(desc or '').strip()}"
                            err = _try_reg(payload)
                            if err is None:
                                st.warning(
                                    "Se guardó **sin** columna de categoría. Ejecutá **`patch_025`** en Supabase."
                                )
                        if err is not None:
                            st.error(_error_msg_from_supabase_exc(err))
                if err is None:
                    st.success("Gasto registrado.")
                    _movi_ss_pop_keys(
                        "gasto_op_cat",
                        "gasto_op_cat_custom",
                        "gasto_op_desc",
                        "gasto_op_m_bs",
                        "gasto_op_tasa_bs",
                        "gasto_op_m_ut",
                        "gasto_op_m_usd",
                        "gasto_op_caja",
                        "gasto_op_ref",
                        "gasto_op_nota",
                    )
                    _movi_bump_form_nonce("gasto_op_form_nonce")
                    st.rerun()

    rows_eg, tiene_cat = _movi_fetch_egresos_caja_recientes(sb, limit=200)
    editable = [
        r
        for r in rows_eg
        if not r.get("venta_id") and not r.get("compra_id") and r.get("id")
    ]
    with st.expander("✏️ Corregir un egreso manual (monto o texto)", expanded=False):
        _flash_corr = st.session_state.pop("flash_gasto_op_correccion", None)
        if isinstance(_flash_corr, dict) and _flash_corr.get("ok"):
            st.success(str(_flash_corr.get("msg") or "Cambios guardados."))
        elif isinstance(_flash_corr, dict) and not _flash_corr.get("ok"):
            st.error(str(_flash_corr.get("msg") or "No se pudo guardar."))
        st.caption(
            "Solo movimientos **sin** venta ni compra ligada. "
            "En cuentas **VES** / **USDT** cargás el **monto en la moneda del banco**; el ERP calcula el **equiv. USD** (`monto_usd`) que usa el **motor de saldos** (igual que al dar de alta un gasto). "
            "En cuenta **USD** el monto es directamente en dólares. Requiere **`patch_028`** en Supabase; si falla *schema cache*, ejecutá **`patch_029_corregir_movimiento_fn_recreate.sql`**."
        )
        if not editable:
            st.info("No hay egresos editables en los últimos registros, o falta columna **id** en la consulta.")
        else:
            id_opts = [str(r["id"]) for r in editable]

            def _fmt_ed(mid: str) -> str:
                rr = next(x for x in editable if str(x.get("id")) == mid)
                ts = str(rr.get("created_at") or "")[:19].replace("T", " ")
                co = (rr.get("concepto") or "")[:42]
                return f"{ts} — {co}"

            pick = st.selectbox("Elegí el movimiento", options=id_opts, format_func=_fmt_ed, key="gasto_op_edit_pick")
            er = next(x for x in editable if str(x.get("id")) == pick)
            _fk = str(pick).replace("-", "")[:32]
            ecaja = cmap_full.get(str(er.get("caja_id") or ""), {})
            emon_raw = str(ecaja.get("moneda_cuenta") or "USD").strip().upper()
            emon = "VES" if emon_raw in ("VES", "BS") else emon_raw
            _musd_er = float(er.get("monto_usd") or 0.01)
            _mm_er = er.get("monto_moneda")
            _mm_f = float(_mm_er) if _mm_er is not None and str(_mm_er).strip() != "" else 0.0
            with st.form(f"f_gasto_op_corregir_{_fk}"):
                st.markdown(f"**Cuenta:** {emon} · {_caja_etiqueta_lista(ecaja) if ecaja else '—'}")
                if emon == "VES" and _mm_f <= 0:
                    st.warning(
                        "Este movimiento **no tiene Bs guardados** en BD (`monto_moneda` vacío). "
                        "Cargá abajo el **monto real en bolívares** que salió del banco y la **tasa** con la que querés valorar en el sistema."
                    )

                nu: float
                p_mon_e: str | None = None
                p_mm_e: float | None = None
                upd_nat: bool

                if emon == "VES":
                    try:
                        _tb0 = float(_tasa_bs_para_documento(t, usar_bcv=False)) if t else 0.0
                    except (ValueError, TypeError):
                        _tb0 = float(_nf(t.get("tasa_bs")) or 0) if t else 0.0
                    if _tb0 <= 0:
                        _tb0 = 300.0
                    _sug_bs = _mm_f if _mm_f > 0 else max(0.01, _musd_er * _tb0)
                    m_bs_in = st.number_input(
                        "Monto en bolívares (Bs) — lo que salió de la cuenta",
                        min_value=0.01,
                        value=float(_round_money_2(_sug_bs)),
                        format="%.2f",
                        key=f"gasto_op_edit_bs_{_fk}",
                    )
                    tasa_bs_in = st.number_input(
                        "Tasa Bs por 1 USD (para calcular el equiv. interno `monto_usd`)",
                        min_value=0.01,
                        value=float(_tb0),
                        format="%.4f",
                        key=f"gasto_op_edit_tasa_bs_{_fk}",
                        help="Mismo criterio que al registrar un gasto en Bs: Bs ÷ esta tasa = valor que guarda el motor de saldos en USD.",
                    )
                    if float(tasa_bs_in) <= 0:
                        nu = _musd_er
                    else:
                        nu = float(m_bs_in) / float(tasa_bs_in)
                    st.caption(
                        f"En base se guardará: **moneda** = VES, **monto_moneda** = {_round_money_2(m_bs_in):,.2f} Bs, "
                        f"**monto_usd** = US$ {_round_money_2(nu):,.2f} (equiv. para saldos)."
                    )
                    upd_nat = True
                    p_mon_e, p_mm_e = "VES", float(m_bs_in)
                elif emon == "USDT":
                    t_ut0 = float(_nf(t.get("tasa_usdt")) or 1.0) if t else 1.0
                    if t_ut0 <= 0:
                        t_ut0 = 1.0
                    _sug_ut = _mm_f if _mm_f > 0 else max(0.0001, _musd_er * t_ut0)
                    m_ut_in = st.number_input(
                        "Monto en USDT",
                        min_value=0.0001,
                        value=float(_round_money_2(_sug_ut)),
                        format="%.4f",
                        key=f"gasto_op_edit_ut_{_fk}",
                    )
                    t_ut_in = st.number_input(
                        "Tasa USDT por 1 USD",
                        min_value=0.0000001,
                        value=float(t_ut0),
                        format="%.6f",
                        key=f"gasto_op_edit_tasa_ut_{_fk}",
                    )
                    if float(t_ut_in) <= 0:
                        nu = _musd_er
                    else:
                        nu = float(m_ut_in) / float(t_ut_in)
                    st.caption(
                        f"En base: **moneda** = USDT, **monto_moneda** = {_round_money_2(m_ut_in):,.4f} USDT, "
                        f"**monto_usd** = US$ {_round_money_2(nu):,.2f}."
                    )
                    upd_nat = True
                    p_mon_e, p_mm_e = "USDT", float(m_ut_in)
                else:
                    nu = st.number_input(
                        "Monto en USD (cuenta en dólares)",
                        min_value=0.01,
                        value=float(_round_money_2(_musd_er)),
                        format="%.2f",
                        key=f"gasto_op_edit_usd_{_fk}",
                    )
                    upd_nat = st.checkbox(
                        "Guardar también en columnas moneda / monto_moneda (USD)",
                        value=True,
                        key=f"gasto_op_edit_upd_nat_{_fk}",
                    )
                    if upd_nat:
                        p_mon_e, p_mm_e = "USD", float(nu)
                nc = st.text_input("Concepto", value=str(er.get("concepto") or ""), key=f"gasto_op_edit_conc_{_fk}")
                nr = st.text_input("Referencia", value=str(er.get("referencia") or ""), key=f"gasto_op_edit_ref_{_fk}")
                ncat = st.text_input(
                    "Categoría (vacío = quitar categoría)",
                    value=str(er.get("categoria_gasto") or ""),
                    key=f"gasto_op_edit_cat_{_fk}",
                )
                if st.form_submit_button("Guardar corrección"):
                    if not (erp_uid or "").strip():
                        st.error("Sesión sin usuario ERP; no se puede guardar.")
                    elif not (nc or "").strip():
                        st.error("El **concepto** no puede quedar vacío (lo exige la base de datos).")
                    elif nu <= 0:
                        st.error("El equivalente USD calculado debe ser mayor que cero (revisá monto y tasa).")
                    else:
                        _rpc_mon = (p_mon_e or None) if upd_nat else None
                        _rpc_mm = float(p_mm_e) if upd_nat and p_mm_e is not None else None
                        pl_e: dict[str, Any] = {
                            "p_usuario_id": erp_uid,
                            "p_movimiento_id": pick,
                            "p_monto_usd": float(nu),
                            "p_concepto": (nc or "").strip(),
                            "p_referencia": (nr or "").strip(),
                            "p_categoria_gasto": (ncat or "").strip() or None,
                            "p_actualizar_moneda_nativa": bool(upd_nat),
                            "p_moneda": _rpc_mon,
                            "p_monto_moneda": _rpc_mm,
                        }
                        try:
                            sb.rpc("corregir_movimiento_caja_manual_erp", pl_e).execute()
                            st.session_state["flash_gasto_op_correccion"] = {
                                "ok": True,
                                "msg": "Movimiento actualizado y guardado en base de datos.",
                            }
                            st.rerun()
                        except Exception as ex:
                            _em = _error_msg_from_supabase_exc(ex)
                            st.session_state["flash_gasto_op_correccion"] = {
                                "ok": False,
                                "msg": f"{_em} · Si falta el RPC: **`patch_029_corregir_movimiento_fn_recreate.sql`** en Supabase y **Reload schema** (API).",
                            }
                            st.rerun()

    if not tiene_cat:
        st.caption(
            "Para ver la columna **Categoría** en la tabla de abajo, aplicá **`patch_025_gastos_operativos.sql`** en Supabase."
        )
    if rows_eg:
        disp: list[dict[str, Any]] = []
        for r in rows_eg:
            cid_s = str(r.get("caja_id") or "")
            row_d: dict[str, Any] = {
                "Fecha": str(r.get("created_at") or "")[:19].replace("T", " "),
                "Caja": cmap_lab.get(cid_s, cid_s[:8] + "…"),
                "Monto": _gasto_op_fmt_monto_tabla(r, cmap_full),
                "Concepto": r.get("concepto") or "",
            }
            if tiene_cat:
                row_d["Categoría"] = r.get("categoria_gasto") or "—"
            if r.get("referencia"):
                row_d["Ref."] = r.get("referencia")
            disp.append(row_d)
        st.markdown("##### Últimos egresos de caja (incluye compras al contado y otros pagos)")
        st.dataframe(pd.DataFrame(disp), use_container_width=True, hide_index=True)
        st.caption(
            "Si necesitás filtrar **solo** gastos operativos, usá la columna **Categoría** tras el patch, o revisá **Reportes → movimientos de caja**."
        )
    else:
        st.info("Aún no hay egresos registrados en caja.")


def module_cajas(sb: Client, erp_uid: str) -> None:
    _modulo_titulo_info(
        "Cajas y bancos",
        key="cajas",
        ayuda_md=(
            "Cada fila es una **cuenta concreta**: banco o entidad (Banesco, Bancamiga…), alias interno, moneda de la cuenta (VES/USD/USDT), número y titular. "
            "**Saldo en cuenta** en **Bs** en cuentas VES, **USD** o **USDT** según la moneda; el valor interno en BD sigue siendo equiv. USD para el motor. "
            "Los **gastos operativos** (alquiler, servicios, nómina…) podés registrarlos en el módulo **Gastos operativos**."
        ),
    )
    rows = sb.table("cajas_bancos").select("*").order("nombre").execute()
    if rows.data:
        df_c = pd.DataFrame(rows.data)
        _t_cajas = latest_tasas(sb) or {}
        _sc: list[str] = []
        _eq: list[str] = []
        for _, _r in df_c.iterrows():
            _a, _b = _caja_saldo_cuenta_y_equiv(_r.get("moneda_cuenta"), _r.get("saldo_actual_usd"), _t_cajas)
            _sc.append(_a)
            _eq.append(_b)
        df_c["Saldo en cuenta"] = _sc
        df_c["Equiv. USD (sistema)"] = _eq
        pref = [
            "entidad",
            "nombre",
            "tipo",
            "moneda_cuenta",
            "numero_cuenta",
            "titular",
            "Saldo en cuenta",
            "Equiv. USD (sistema)",
            "activo",
        ]
        cols_show = [c for c in pref if c in df_c.columns] + [
            c for c in df_c.columns if c not in pref and c != "saldo_actual_usd"
        ]
        st.dataframe(df_c[cols_show], use_container_width=True, hide_index=True)

    with st.expander("Nueva caja / cuenta"):
        st.caption("Si falla al guardar, ejecutá en Supabase `supabase/patch_015_cajas_detalle.sql`.")
        with st.form(f"f_caja_{int(st.session_state.get('caja_alta_form_nonce', 0))}"):
            entidad = st.text_input("Banco / entidad (ej. Banesco, Bancamiga)", placeholder="Opcional si es efectivo")
            nombre = st.text_input("Nombre o alias en el ERP", help="Ej. Corriente USD proveedores")
            tipo = st.selectbox("Tipo", ["Banco", "Wallet", "Efectivo"])
            moneda_cuenta = st.selectbox("Moneda de la cuenta", ["USD", "VES", "USDT"], index=0)
            numero_cuenta = st.text_input("Número de cuenta (últimos dígitos o completo)", placeholder="Opcional")
            titular = st.text_input("Titular", placeholder="Opcional")
            if st.form_submit_button("Crear"):
                if not nombre or not str(nombre).strip():
                    st.error("El nombre es obligatorio.")
                else:
                    try:
                        sb.table("cajas_bancos").insert(
                            {
                                "nombre": nombre.strip(),
                                "tipo": tipo,
                                "saldo_actual_usd": 0,
                                "entidad": entidad.strip() or None,
                                "numero_cuenta": numero_cuenta.strip() or None,
                                "titular": titular.strip() or None,
                                "moneda_cuenta": moneda_cuenta,
                            }
                        ).execute()
                        st.success("Caja creada.")
                        _movi_bump_form_nonce("caja_alta_form_nonce")
                        st.rerun()
                    except Exception as e:
                        st.error(
                            f"{e} · Si falta columna en BD, aplicá el parche **patch_015_cajas_detalle** en Supabase."
                        )

    caja_rows_mov = _cajas_fetch_rows(sb, solo_activas=True)
    caja_ids_mov, caja_fmt_mov = _caja_select_options(caja_rows_mov) if caja_rows_mov else ([], lambda x: str(x))
    if not caja_ids_mov:
        st.info("No hay cajas activas: creá al menos una para movimientos manuales.")
        st.stop()

    st.caption("Movimiento manual (ajuste de caja)")
    with st.form(f"f_mov_{int(st.session_state.get('caja_mov_form_nonce', 0))}"):
        cid_mov = st.selectbox("Caja", options=caja_ids_mov, format_func=caja_fmt_mov)
        tipo_m = st.selectbox("Tipo movimiento", ["Ingreso", "Egreso"])
        monto = st.number_input("Monto USD", min_value=0.01, format="%.2f")
        concepto = st.text_input("Concepto")
        ref = st.text_input("Referencia")
        nota_mov = st.text_input(
            "Nota de tesorería (opcional)",
            placeholder="Ej.: Cambio Bs→USD, traspaso entre cuentas…",
            key="caja_mov_nota_op",
        )
        if st.form_submit_button("Registrar"):
            try:
                _rpc_mov: dict[str, Any] = {
                    "p_usuario_id": erp_uid,
                    "p_caja_id": str(cid_mov),
                    "p_tipo": tipo_m,
                    "p_monto_usd": float(monto),
                    "p_concepto": concepto,
                    "p_referencia": ref,
                }
                if (nota_mov or "").strip():
                    _rpc_mov["p_nota_operacion"] = (nota_mov or "").strip()
                sb.rpc("registrar_movimiento_caja_erp", _rpc_mov).execute()
                st.success("Movimiento registrado.")
                _movi_ss_pop_keys("caja_mov_nota_op")
                _movi_bump_form_nonce("caja_mov_form_nonce")
                st.rerun()
            except Exception as e:
                st.error(str(e))


def panel_reportes_inventario_export(sb: Client, t: dict[str, Any] | None) -> None:
    _ts_md = _backup_file_timestamp()
    with st.expander("Markdown (.md) para asistente / IA — inventario activo", expanded=False):
        st.caption(
            "Archivo **.md** con productos **activos**: stock, **precio de venta solo en USD**, categoría y compatibilidad. "
            "Sirve para subirlo a una IA que consulte disponibilidad y precios."
        )
        try:
            _md_ia = inventario_activos_markdown_ia(sb, t)
            st.download_button(
                label=f"Descargar — inventario_ia_{_ts_md}.md",
                data=_md_ia.encode("utf-8"),
                file_name=f"inventario_ia_{_ts_md}.md",
                mime="text/markdown",
                key="rep_inv_dl_md_ia",
                use_container_width=True,
            )
        except Exception as e:
            st.error(str(e))
    if not t:
        st.warning("Registrá tasas en el **Dashboard** para que el reporte muestre referencias en Bs.")
        return
    st.caption(
        "Podés filtrar por categoría, costo o precio y bajar el listado en **Excel**, **PDF** o una página **HTML** para imprimir. "
        "Si el botón de Excel no aparece, avisá a quien instaló el programa (hace falta una librería llamada openpyxl; el PDF usa reportlab)."
    )
    try:
        cats_list_rp = (sb.table("categorias").select("id,nombre").order("nombre").execute().data or [])
    except Exception:
        cats_list_rp = []
    _id_rp, _, _ = _categoria_maps_from_rows(cats_list_rp)
    df = _inv_enrich_compat_columns(_normalize_productos_inventario_df(_fetch_productos_inventario_df(sb)))
    if not df.empty:
        if "categoria_id" in df.columns:
            df["categoria"] = df["categoria_id"].apply(
                lambda x: _id_rp.get(str(x).strip(), "")
                if x is not None and not (isinstance(x, float) and pd.isna(x)) and str(x).strip()
                else ""
            )
        else:
            df["categoria"] = ""
    else:
        df["categoria"] = pd.Series(dtype=object)

    _from_db = sorted(
        {str(c.get("nombre") or "").strip() for c in cats_list_rp if str(c.get("nombre") or "").strip()},
        key=str.casefold,
    )
    _from_prod = (
        sorted(
            {str(x) for x in df["categoria"].map(_inv_cat_display).tolist() if str(x).strip()},
            key=str.casefold,
        )
        if not df.empty and "categoria" in df.columns
        else []
    )
    _lab_cat = sorted(set(_from_db) | set(_from_prod), key=str.casefold)
    if not _lab_cat:
        _lab_cat = ["(Sin categoría)"]
    _pick_cat = st.multiselect(
        "Categorías a incluir en el reporte",
        options=_lab_cat,
        default=_lab_cat,
        key="rep_inv_print_cats",
    )
    _use_cats = _pick_cat if _pick_cat else _lab_cat
    _inv_col_mode = st.selectbox(
        "Qué columnas incluir (PDF, HTML y Excel)",
        options=["interno", "lista_cliente", "analisis_precios", "personalizado"],
        format_func=lambda x: {
            "interno": "Completo — núcleo + detalles opcionales (marca, ubicación…)",
            "lista_cliente": "Lista de precios (cliente) — núcleo + detalles opcionales",
            "analisis_precios": "Análisis de precios — núcleo + detalles opcionales",
            "personalizado": "Personalizado — elijo columnas en el multiselect",
        }[x],
        key="rep_inv_col_mode",
        help="En los tres primeros modos, **descripción** usa más ancho si no marcás columnas de detalle. En personalizado, todo sale del multiselect.",
    )
    _col_keys_f: frozenset[str] | None
    if _inv_col_mode == "personalizado":
        _opts_k = list(INV_REP_COL_KEYS_PERSONALIZADO)
        _cust = st.multiselect(
            "Columnas del reporte",
            options=_opts_k,
            default=_opts_k,
            format_func=lambda k: INV_REP_COL_META_DICT[k],
            key="rep_inv_col_custom",
        )
        _col_keys_f = frozenset(_cust) if _cust else None
    else:
        _col_keys_f = INV_REP_PRESET_COLS[_inv_col_mode]
    st.markdown("**Precios en el reporte** (activá solo lo que necesites)")
    _m1, _m2, _m3 = st.columns(3)
    with _m1:
        _show_usd = st.checkbox("USD (costo y precio)", value=True, key="rep_inv_cur_usd")
    with _m2:
        _show_bs = st.checkbox("Bs (ref., según tasa del sistema)", value=False, key="rep_inv_cur_bs")
    with _m3:
        _show_usdt = st.checkbox("USDT (ref., USD × tasa USDT)", value=False, key="rep_inv_cur_usdt")
    _k_tpl = _inv_rep_merge_template_keys(_col_keys_f)
    _k_ext = _inv_rep_extend_currency_columns(_k_tpl, show_bs=_show_bs, show_usdt=_show_usdt)
    _col_keys_export = _inv_rep_apply_currency_prefs(
        _k_ext, show_usd=_show_usd, show_bs=_show_bs, show_usdt=_show_usdt
    )
    _det_marca = _det_cond = _det_veh = _det_anos = _det_smin = _det_ubi = False
    if _inv_col_mode != "personalizado":
        st.markdown("**Columnas de detalle (opcional)** — desmarcadas = más espacio horizontal para **descripción**")
        _d1, _d2, _d3 = st.columns(3)
        with _d1:
            _det_marca = st.checkbox("Marca del repuesto", value=False, key="rep_inv_det_marca")
            _det_cond = st.checkbox("Condición", value=False, key="rep_inv_det_cond")
        with _d2:
            _det_veh = st.checkbox("Marcas carro (compatibilidad)", value=False, key="rep_inv_det_veh")
            _det_anos = st.checkbox("Años (compatibilidad)", value=False, key="rep_inv_det_anos")
        with _d3:
            _det_smin = st.checkbox("Stock mínimo", value=False, key="rep_inv_det_smin")
            _det_ubi = st.checkbox("Ubicación", value=False, key="rep_inv_det_ubi")
        _col_keys_export = _inv_rep_extend_detail_columns(
            _col_keys_export,
            marca=_det_marca,
            cond=_det_cond,
            veh=_det_veh,
            anos=_det_anos,
            stock_min=_det_smin,
            ubi=_det_ubi,
        )
    else:
        st.caption(
            "Modo **personalizado**: activá marca, condición, años, ubicación, etc. desde el **multiselect** de columnas."
        )
    _ic1, _ic2 = st.columns(2)
    with _ic1:
        _solo_act = st.checkbox("Solo productos activos", value=True, key="rep_inv_print_act")
    with _ic2:
        _agrup_cat = st.checkbox("Agrupar por categoría en el impreso", value=True, key="rep_inv_print_grp")
    _oc1, _oc2 = st.columns(2)
    with _oc1:
        st.markdown("**Costo USD**")
        _cmin = st.number_input("Desde (0 = sin mínimo)", min_value=0.0, value=0.0, step=0.01, key="rep_inv_cmin")
        _cmax = st.number_input("Hasta (0 = sin máximo)", min_value=0.0, value=0.0, step=0.01, key="rep_inv_cmax")
    with _oc2:
        st.markdown("**Precio venta USD**")
        _pmin = st.number_input("Desde (0 = sin mínimo)", min_value=0.0, value=0.0, step=0.01, key="rep_inv_pmin")
        _pmax = st.number_input("Hasta (0 = sin máximo)", min_value=0.0, value=0.0, step=0.01, key="rep_inv_pmax")
    _orden_lbl = st.selectbox(
        "Ordenar por",
        options=[
            "Descripción (A-Z)",
            "Código (A-Z)",
            "Costo USD: menor → mayor",
            "Costo USD: mayor → menor",
            "Precio venta USD: menor → mayor",
            "Precio venta USD: mayor → menor",
        ],
        key="rep_inv_sort",
    )
    _orden_map = {
        "Descripción (A-Z)": "descripcion",
        "Código (A-Z)": "codigo",
        "Costo USD: menor → mayor": "costo_asc",
        "Costo USD: mayor → menor": "costo_desc",
        "Precio venta USD: menor → mayor": "precio_asc",
        "Precio venta USD: mayor → menor": "precio_desc",
    }
    _orden_key = _orden_map[_orden_lbl]

    _df_p = _df_inventario_filtrado_impresion(
        df,
        categorias_sel=_use_cats,
        costo_min=float(_cmin),
        costo_max=float(_cmax),
        precio_min=float(_pmin),
        precio_max=float(_pmax),
        solo_activos=bool(_solo_act),
    )
    _parts_sub: list[str] = []
    if _pick_cat and len(_pick_cat) < len(_lab_cat):
        _parts_sub.append(f"categorías: {len(_pick_cat)} seleccionadas")
    if _cmin > 0 or _cmax > 0:
        _parts_sub.append(f"costo USD {_cmin if _cmin > 0 else '…'} — {_cmax if _cmax > 0 else '…'}")
    if _pmin > 0 or _pmax > 0:
        _parts_sub.append(f"precio USD {_pmin if _pmin > 0 else '…'} — {_pmax if _pmax > 0 else '…'}")
    _parts_sub.append(_orden_lbl)
    if _agrup_cat:
        _parts_sub.append("agrupado por categoría")
    _parts_sub.append(
        {
            "interno": "plantilla: completo (núcleo)",
            "lista_cliente": "plantilla: lista cliente",
            "analisis_precios": "plantilla: análisis precios",
            "personalizado": "plantilla: personalizado",
        }[_inv_col_mode]
    )
    if _inv_col_mode != "personalizado":
        _det_bits: list[str] = []
        if _det_marca:
            _det_bits.append("marca")
        if _det_cond:
            _det_bits.append("cond.")
        if _det_veh:
            _det_bits.append("marcas carro")
        if _det_anos:
            _det_bits.append("años")
        if _det_smin:
            _det_bits.append("st.mín")
        if _det_ubi:
            _det_bits.append("ubic.")
        if _det_bits:
            _parts_sub.append("detalle: " + ", ".join(_det_bits))
        else:
            _parts_sub.append("detalle: solo núcleo (descripción amplia)")
    _mon_lbl: list[str] = []
    if _show_usd:
        _mon_lbl.append("USD")
    if _show_bs:
        _mon_lbl.append("Bs ref.")
    if _show_usdt:
        _mon_lbl.append("USDT ref.")
    if _mon_lbl:
        _parts_sub.append("precios: " + " · ".join(_mon_lbl))
    _sub_f = " · ".join(_parts_sub)

    if _df_p.empty:
        st.warning("No hay productos con esos filtros. Igual podés descargar archivos con solo encabezados.")
        _df_out = _df_p
    else:
        _df_out = _df_inventario_orden_impresion(_df_p, _orden_key, agrupar_categoria=bool(_agrup_cat))
    _html_inv = _html_inventario_listado(
        _df_out,
        t,
        agrupar_categoria=bool(_agrup_cat),
        subtitulo_filtros=_sub_f,
        column_keys=_col_keys_export,
    )
    _df_flat = _df_inventario_export_flat(_df_out, t, column_keys=_col_keys_export)
    _ts_p = _backup_file_timestamp()
    _bx, _bp = st.columns(2)
    with _bx:
        try:
            _xlsx_b = _xlsx_inventario_bytes(_df_flat)
        except ImportError:
            st.caption("Instalá **openpyxl** para Excel.")
        else:
            st.download_button(
                label=f"Excel — inventario_{_ts_p}.xlsx",
                data=_xlsx_b,
                file_name=f"inventario_{_ts_p}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="rep_inv_dl_xlsx",
                use_container_width=True,
            )
    with _bp:
        try:
            _pdf_b = _pdf_inventario_bytes(
                _df_out,
                t,
                agrupar_categoria=bool(_agrup_cat),
                subtitulo_filtros=_sub_f,
                column_keys=_col_keys_export,
            )
        except ImportError:
            st.caption("Instalá **reportlab** para PDF.")
        else:
            st.download_button(
                label=f"PDF — inventario_{_ts_p}.pdf",
                data=_pdf_b,
                file_name=f"inventario_{_ts_p}.pdf",
                mime="application/pdf",
                key="rep_inv_dl_pdf",
                use_container_width=True,
            )
    st.download_button(
        label=f"HTML — inventario_{_ts_p}.html",
        data=_html_inv.encode("utf-8"),
        file_name=f"inventario_{_ts_p}.html",
        mime="text/html",
        key="rep_inv_dl_html",
    )
    st.divider()
    st.markdown("##### Toma de inventario físico")
    st.caption(
        "PDF con **código**, **categoría**, **descripción**, **stock del sistema** y columna **Conteo físico** en blanco "
        "para completar al imprimir. Usa los **mismos filtros y orden** que el listado de arriba."
    )
    try:
        _pdf_tf_b = _pdf_toma_inventario_fisico_bytes(_df_out, subtitulo_filtros=_sub_f)
    except ImportError:
        st.caption("Instalá **reportlab** para PDF.")
    else:
        st.download_button(
            label=f"PDF — toma_fisica_{_ts_p}.pdf",
            data=_pdf_tf_b,
            file_name=f"toma_fisica_inventario_{_ts_p}.pdf",
            mime="application/pdf",
            key="rep_inv_dl_pdf_toma_fisica",
            use_container_width=True,
        )
    st.caption("Vista previa (Ctrl+P desde el recuadro o abrí el HTML descargado).")
    components.html(_html_inv, height=560, scrolling=True)


def _rep_parse_fecha_venc(x: Any) -> date | None:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None


def _rep_texto_plazo_vencimiento(fv: Any) -> str:
    """Explicación en palabras según la fecha límite de pago o cobro."""
    d = _rep_parse_fecha_venc(x=fv)
    if d is None:
        return "Sin fecha cargada"
    hoy = date.today()
    dias = (hoy - d).days
    if dias < 0:
        faltan = -dias
        if faltan == 1:
            return "Todavía a tiempo — vence mañana"
        return f"Todavía a tiempo — faltan {faltan} días"
    if dias == 0:
        return "La fecha límite es hoy"
    if dias <= 30:
        return f"Vencido hace {dias} día(s) (menos de un mes)"
    if dias <= 60:
        return f"Vencido hace {dias} días (entre 1 y 2 meses)"
    if dias <= 90:
        return f"Vencido hace {dias} días (entre 2 y 3 meses)"
    return f"Muy atrasado — más de 3 meses ({dias} días)"


def _rep_bucket_antiguedad(fv: Any) -> str:
    """Grupo corto para totales (quién debe / a quién debemos)."""
    d = _rep_parse_fecha_venc(x=fv)
    if d is None:
        return "Sin fecha"
    hoy = date.today()
    dias = (hoy - d).days
    if dias < 0:
        return "Aún no vence"
    if dias == 0:
        return "Vence hoy"
    if dias <= 30:
        return "Vencido 1–30 días"
    if dias <= 60:
        return "Vencido 31–60 días"
    if dias <= 90:
        return "Vencido 61–90 días"
    return "Vencido más de 90 días"


def _rep_movimientos_caja_filtrados(
    sb: Client, *, desde: date, hasta: date, caja_id: str | None, tipo_mov: str | None
) -> list[dict[str, Any]]:
    ds = desde.isoformat()
    hf = f"{hasta.isoformat()}T23:59:59"
    sel_ext = (
        "created_at, tipo, monto_usd, moneda, monto_moneda, monto_bs, monto_usdt, monto_usd_caja, "
        "concepto, referencia, nota_operacion, caja_id, usuario_id"
    )
    sel_full = (
        "created_at, tipo, monto_usd, moneda, monto_moneda, concepto, referencia, nota_operacion, caja_id, usuario_id"
    )
    sel_leg = "created_at, tipo, monto_usd, concepto, referencia, caja_id, usuario_id"
    try:
        try:
            q = (
                sb.table("movimientos_caja")
                .select(sel_ext)
                .gte("created_at", ds)
                .lte("created_at", hf)
                .order("created_at", desc=True)
            )
            r = q.execute()
        except Exception:
            q = (
                sb.table("movimientos_caja")
                .select(sel_full)
                .gte("created_at", ds)
                .lte("created_at", hf)
                .order("created_at", desc=True)
            )
            r = q.execute()
    except Exception:
        q = (
            sb.table("movimientos_caja")
            .select(sel_leg)
            .gte("created_at", ds)
            .lte("created_at", hf)
            .order("created_at", desc=True)
        )
        r = q.execute()
        for row in r.data or []:
            row.setdefault("moneda", None)
            row.setdefault("monto_moneda", None)
            row.setdefault("monto_bs", None)
            row.setdefault("monto_usdt", None)
            row.setdefault("monto_usd_caja", None)
            row.setdefault("nota_operacion", None)
    rows = r.data or []
    if caja_id:
        rows = [x for x in rows if str(x.get("caja_id")) == caja_id]
    if tipo_mov in ("Ingreso", "Egreso"):
        rows = [x for x in rows if str(x.get("tipo") or "") == tipo_mov]
    return rows


def panel_reportes_catalogo_fotos(sb: Client, erp_uid: str) -> None:
    st.markdown("#### Catálogo y etiquetas imprimibles")
    storage_on = _catalogo_storage_portada_enabled()
    st.info(
        "**Catálogo (lo principal):** más abajo armás una **página HTML** con código, descripción y precio para **imprimir** o mandar listados y etiquetas. "
        "**Fotos de portada** (subir archivos a la nube, galería) son **opcionales** y no cambian stock ni precios. "
        "Si no las querés, en `secrets` → `[catalogo]` poné **`storage_fotos = false`** (o `enabled = false`): se ocultan subida y galería, pero **el HTML del catálogo sigue**. "
        "Una imagen en el HTML puede salir igual si cargás una **URL** a mano en **Inventario**."
    )
    if storage_on:
        st.caption(
            "Elegí un producto para subir fotos, galería o generar el HTML. La **portada** es la imagen que el sistema prioriza en ese HTML."
        )
    else:
        st.caption(
            "**Subida y galería apagadas** en configuración. Usá el producto de abajo para el modo *uno* y el bloque HTML; imágenes solo si hay URL en la ficha del producto."
        )

    try:
        prows = (
            sb.table("productos")
            .select("id,codigo,descripcion,imagen_url,activo")
            .order("descripcion")
            .limit(3000)
            .execute()
            .data
            or []
        )
    except Exception as ex:
        st.error(f"No se pudieron leer productos: {ex}")
        return
    if not prows:
        st.warning("No hay productos en inventario.")
        return

    labels: list[str] = []
    id_by_label: dict[str, str] = {}
    for p in prows:
        pid = str(p.get("id") or "").strip()
        if not pid:
            continue
        cod = (p.get("codigo") or "") or ""
        desc = (p.get("descripcion") or "") or ""
        lab = f"{desc[:70]} · {cod}".strip(" ·")
        labels.append(lab)
        id_by_label[lab] = pid
    labels = sorted(set(labels), key=str.casefold)
    sel = st.selectbox(
        "Producto",
        options=labels,
        key="cat_prod_sel",
        help="Sirve para el modo *uno* del HTML y, si hay subida de fotos, para ese producto. Escribí para filtrar.",
    )
    pid = id_by_label.get(sel, "")
    if not pid:
        st.error("Producto inválido.")
        return

    head = next((x for x in prows if str(x.get("id") or "") == pid), {})
    cur_img = str(head.get("imagen_url") or "").strip()

    if storage_on:
        bucket = _catalogo_bucket_name()
        st.caption(
            f"Almacenamiento de fotos: bucket **{bucket}** (`secrets` → `[catalogo] bucket = ...`)."
        )
        c0, c1 = st.columns([1, 1])
        with c0:
            st.markdown("#### Foto de portada (vista previa)")
            st.caption("Imagen que muestra el sistema hoy para este producto (listados y catálogo HTML).")
            if cur_img:
                st.image(cur_img, use_container_width=True)
            else:
                st.caption("Sin URL todavía. Subí una acá o en **Inventario**, o elegí una en la galería.")
        with c1:
            st.markdown("#### Subir fotos")
            up = st.file_uploader(
                "Elegí una o varias imágenes",
                type=["jpg", "jpeg", "png", "webp"],
                accept_multiple_files=True,
                key=f"cat_upl_{pid}",
                help="JPG, PNG o WebP. Podés subir varias a la vez; en la galería elegís cuál será la foto de portada.",
            )
            make_primary_first = st.checkbox(
                "La primera foto que suba pasa a ser la de portada automáticamente",
                value=True,
                key=f"cat_make_primary_{pid}",
                help="Si lo desmarcás, las nuevas fotos solo se guardan en la galería; después tocá **Poner como portada** en la que quieras.",
            )
            if up and st.button("Subir", key=f"cat_do_upload_{pid}", use_container_width=True):
                try:
                    any_uploaded = False
                    inserted_paths: list[str] = []
                    for f in up:
                        data = f.getvalue()
                        if not data:
                            continue
                        obj_path = _catalogo_upload_producto_foto(
                            sb,
                            bucket=bucket,
                            producto_id=pid,
                            filename=str(getattr(f, "name", "") or "foto"),
                            content_type=str(getattr(f, "type", "") or "application/octet-stream"),
                            data=data,
                        )
                        _row_pf_cat: dict[str, Any] = {
                            "producto_id": pid,
                            "storage_path": obj_path,
                            "is_primary": False,
                        }
                        _cb_cat = _erp_user_uuid_or_none(erp_uid)
                        if _cb_cat:
                            _row_pf_cat["created_by"] = _cb_cat
                        sb.table("producto_fotos").insert(_row_pf_cat).execute()
                        inserted_paths.append(obj_path)
                        any_uploaded = True
                    if not any_uploaded:
                        st.warning("No se subió nada (archivos vacíos).")
                        return

                    if make_primary_first and inserted_paths:
                        fotos_now = _catalogo_fetch_fotos(sb, pid)
                        match = next((r for r in fotos_now if str(r.get("storage_path") or "") == inserted_paths[0]), None)
                        if match:
                            _catalogo_set_primary(sb, producto_id=pid, foto_id=str(match["id"]))
                            pub = _storage_public_object_url(bucket, inserted_paths[0])
                            sb.table("productos").update({"imagen_url": pub}).eq("id", pid).execute()

                    st.success("Fotos subidas.")
                    st.rerun()
                except Exception as ex:
                    st.error(
                        f"No se pudo subir. Verificá que exista el bucket **{bucket}** en Supabase Storage (ideal público) "
                        f"y que tu key tenga permisos. Detalle: {ex}"
                    )

        st.divider()
        st.markdown("#### Galería")
        st.caption(
            "Todas las fotos del producto en Storage. **Solo una** es la de portada (la que se prioriza en el HTML). "
            "El resto son adicionales."
        )
        try:
            fotos = _catalogo_fetch_fotos(sb, pid)
        except Exception as ex:
            st.error(
                f"No se pueden leer fotos desde `producto_fotos`. Ejecutá `supabase/patch_021_catalogo_fotos_productos.sql`. "
                f"Detalle: {ex}"
            )
            return
        if not fotos:
            st.info("Este producto aún no tiene fotos en la galería.")
            fotos = []

        ncol = 3
        for i in range(0, len(fotos), ncol):
            grp = fotos[i : i + ncol]
            cs = st.columns(ncol)
            for c, r in zip(cs, grp):
                fid = str(r.get("id") or "")
                path = str(r.get("storage_path") or "")
                is_p = _catalogo_row_is_primary(r)
                url = _storage_public_object_url(bucket, path) if path else ""
                with c:
                    if is_p:
                        st.markdown("**Foto de portada**")
                        st.caption("Esta es la que se muestra en listados y catálogo.")
                    else:
                        st.caption(" ")
                    if url:
                        st.image(url, use_container_width=True)
                    else:
                        st.warning("Sin URL.")
                    b1, b2 = st.columns(2)
                    with b1:
                        if st.button(
                            "Poner como portada",
                            key=f"cat_primary_{fid}",
                            disabled=is_p,
                            use_container_width=True,
                            help="Esta imagen pasará a ser la que se ve en inventario y en el HTML imprimible.",
                        ):
                            try:
                                _catalogo_set_primary(sb, producto_id=pid, foto_id=fid)
                                if path:
                                    sb.table("productos").update({"imagen_url": _storage_public_object_url(bucket, path)}).eq(
                                        "id", pid
                                    ).execute()
                                st.rerun()
                            except Exception as ex:
                                st.error(str(ex))
                    with b2:
                        if st.button("Eliminar", key=f"cat_del_{fid}", use_container_width=True):
                            try:
                                _catalogo_delete_foto(sb, bucket=bucket, foto_row=r)
                                after = _catalogo_fetch_fotos(sb, pid)
                                prim = next((x for x in after if bool(x.get("is_primary"))), None)
                                if not prim and after:
                                    _catalogo_set_primary(sb, producto_id=pid, foto_id=str(after[0]["id"]))
                                    sp = str(after[0].get("storage_path") or "")
                                    if sp:
                                        sb.table("productos").update(
                                            {"imagen_url": _storage_public_object_url(bucket, sp)}
                                        ).eq("id", pid).execute()
                                elif not after:
                                    sb.table("productos").update({"imagen_url": None}).eq("id", pid).execute()
                                st.rerun()
                            except Exception as ex:
                                st.error(str(ex))
    else:
        st.markdown("#### Imagen (solo si hay URL en el producto)")
        st.caption(
            "No hay subida a Storage desde acá. Si en **Inventario** cargaste una **URL** en el producto, se previsualiza abajo y puede salir en el HTML."
        )
        if cur_img:
            st.image(cur_img, use_container_width=True)
        else:
            st.caption("Este producto no tiene URL de imagen en la base.")

    st.divider()
    st.markdown("#### Catálogo / etiquetas imprimibles (HTML)")
    st.caption(
        "Elegí **una ficha**, **varios ítems** o un **listado** con tope. Imprimí con **Ctrl+P** o descargá el HTML."
    )
    modo_imp = st.radio(
        "Qué incluir en la página imprimible (HTML)",
        options=["uno", "varios", "listado"],
        format_func=lambda v: {
            "uno": "Solo el producto elegido arriba",
            "varios": "Varios productos (elegís en una lista)",
            "listado": "Listado con tope y filtros de precio",
        }[v],
        horizontal=True,
        key="cat_print_mode",
        help="Genera una vista para imprimir (Ctrl+P) o descargar como archivo HTML.",
    )

    only_active = st.checkbox(
        "Solo productos activos",
        value=True,
        key="cat_print_only_active",
        help="Si está marcado, no entran productos dados de baja en el inventario.",
    )
    precio_min = st.number_input(
        "Precio de venta USD mínimo (0 = sin mínimo)",
        min_value=0.0,
        value=0.0,
        step=1.0,
        key="cat_print_pmin",
        help="Filtra por precio de venta en dólares. 0 deja pasar cualquier precio por abajo.",
    )
    precio_max = st.number_input(
        "Precio de venta USD máximo (0 = sin máximo)",
        min_value=0.0,
        value=0.0,
        step=1.0,
        key="cat_print_pmax",
        help="0 significa sin tope superior.",
    )

    listado_limit = 300
    if modo_imp == "listado":
        listado_limit = int(
            st.number_input(
                "Máximo de filas en el listado",
                min_value=50,
                max_value=3000,
                value=300,
                step=50,
                key="cat_print_limit",
                help="Límite de seguridad para no generar páginas enormes de golpe.",
            )
        )

    fetch_cap = listado_limit if modo_imp == "listado" else 3000

    try:
        q = (
            sb.table("productos")
            .select("id,codigo,sku_oem,descripcion,precio_v_usd,imagen_url,activo")
            .order("descripcion")
        )
        if only_active:
            q = q.eq("activo", True)
        r = q.limit(int(fetch_cap)).execute()
        items_pool = list(r.data or [])
    except Exception as ex:
        st.error(f"No se pudieron leer productos para el catálogo: {ex}")
        items_pool = []

    if storage_on:
        try:
            bucket_html = _catalogo_bucket_name()
            for it in items_pool:
                pid2 = str(it.get("id") or "").strip()
                if not pid2:
                    continue
                sp = _catalogo_primary_path_for_producto(sb, pid2)
                if sp:
                    it["imagen_url"] = _storage_public_object_url(bucket_html, sp)
        except Exception:
            pass

    def _in_rango_precio(x: dict[str, Any]) -> bool:
        p = _nf(x.get("precio_v_usd"))
        if p is None:
            return False
        if precio_min and float(p) < float(precio_min):
            return False
        if precio_max and float(precio_max) > 0 and float(p) > float(precio_max):
            return False
        return True

    items_filtered = [x for x in items_pool if _in_rango_precio(x)]

    items_out: list[dict[str, Any]] = []
    titulo_cat = "Catálogo — Movi Motors"
    subt_extra = ""

    if modo_imp == "uno":
        items_out = [x for x in items_filtered if str(x.get("id") or "") == pid]
        if items_out:
            c0 = str(items_out[0].get("codigo") or "").strip()
            d0 = str(items_out[0].get("descripcion") or "").strip()[:40]
            titulo_cat = f"Etiqueta / ficha — {c0 or d0 or 'producto'}"
        else:
            st.warning(
                "El producto del selector no entra con los filtros (activo / precio). "
                "Probá desmarcar *Solo productos activos* o ajustar precio mínimo/máximo."
            )
    elif modo_imp == "varios":
        pick_labels: dict[str, str] = {}
        for p in sorted(items_filtered, key=lambda x: str(x.get("descripcion") or "").casefold()):
            i = str(p.get("id") or "").strip()
            if not i:
                continue
            lab = f"{str(p.get('descripcion') or '')[:52]} · {p.get('codigo') or ''}".strip(" ·")
            if lab in pick_labels:
                lab = f"{lab} [{i[:8]}]"
            pick_labels[lab] = i
        plabs = sorted(pick_labels.keys(), key=str.casefold)
        chosen = st.multiselect(
            "Productos en la hoja imprimible",
            options=plabs,
            default=[],
            key="cat_print_multisel",
            help="Solo aparecen productos que pasan los filtros de activo y precio. Escribí para acortar la lista.",
        )
        if not chosen:
            st.info("Elegí uno o más productos en la lista para generar la vista imprimible.")
        cids = {pick_labels[L] for L in chosen}
        items_out = [x for x in items_filtered if str(x.get("id") or "") in cids]
        subt_extra = f"{len(items_out)} ítem(s) seleccionados"
    else:
        items_out = items_filtered[: int(listado_limit)]
        subt_extra = f"hasta {len(items_out)} ítems (orden alfabético)"

    sub = []
    if only_active:
        sub.append("solo activos")
    if precio_min:
        sub.append(f"precio ≥ {precio_min:g}")
    if precio_max:
        sub.append(f"precio ≤ {precio_max:g}")
    if subt_extra:
        sub.append(subt_extra)
    subt = " · ".join(sub)

    if items_out:
        html_cat = _html_catalogo_imprimible(items_out, titulo=titulo_cat, subtitulo=subt)
        ts = _backup_file_timestamp()
        st.download_button(
            label=f"Descargar HTML — catalogo_{ts}.html",
            data=html_cat.encode("utf-8"),
            file_name=f"catalogo_{ts}.html",
            mime="text/html",
            key="cat_dl_html",
            use_container_width=True,
        )
        components.html(html_cat, height=560, scrolling=True)
    elif modo_imp == "listado" and not items_filtered:
        st.warning("No hay productos que cumplan activo + rango de precio.")


def module_reportes(sb: Client, erp_uid: str, t: dict[str, Any] | None, rol: str) -> None:
    can_fin = role_can(rol, "reportes")
    can_cat = role_can(rol, "catalogo")
    if not can_fin and not can_cat:
        st.error("Tu rol no tiene acceso a reportes ni al catálogo.")
        return

    have_t = bool(t) if can_fin else True
    t_bs = float(t["tasa_bs"]) if t else 0.0
    t_usdt = float(t["tasa_usdt"]) if t else 0.0

    if can_fin:
        _rep_info_parts = [
            "**Cómo usar reportes:** 1) Elegí la **pestaña** del tema. 2) Ajustá **fechas** y filtros en esa pestaña. "
            "3) Revisá la **tabla o gráfico** principal. 4) Si necesitás profundizar, abrí **Más detalle** al final de la pestaña. "
            "5) **Descargá** Excel o CSV para otra PC o WhatsApp.",
            "Los totales en **USD** son los del sistema; las columnas en **bolívares** son referencia según la tasa del día (cuando esté cargada).",
        ]
        if have_t:
            _rep_info_parts.append(
                f"Referencia: 1 USD equivale a **Bs** {int(round(t_bs)):,d} · **USDT** {int(round(t_usdt)):,d}"
            )
        if can_cat:
            _rep_info_parts.append(
                "**Catálogo y etiquetas:** página HTML para imprimir listados y fichas. "
                "La subida de fotos a la nube es opcional y se puede apagar en `secrets` → `[catalogo]` → `storage_fotos`."
            )
        _modulo_titulo_info("Reportes", key="reportes", ayuda_md="\n\n".join(_rep_info_parts))
        if not have_t:
            st.warning(
                "Aún no hay **tasas del día** cargadas en el Dashboard: los reportes en bolívares no se muestran hasta que las registres. "
                "La pestaña **Catálogo y etiquetas** funciona igual."
            )
    else:
        _modulo_titulo_info(
            "Reportes",
            key="reportes_cat",
            ayuda_md=(
                "**Catálogo y etiquetas:** página HTML para imprimir listados y fichas. "
                "La subida de fotos a la nube es opcional y se puede apagar en `secrets` → `[catalogo]` → `storage_fotos`."
            ),
        )

    if can_fin:
        tab_re, tab_inv, tab_caja, tab_ven, tab_comp, tab_cartera, tab_cat = st.tabs(
            [
                "Resumen ejecutivo",
                "Inventario",
                "Caja",
                "Ventas",
                "Compras",
                "Cartera",
                "Catálogo",
            ]
        )
    else:
        tab_cat = st.tabs(["Catálogo"])[0]

    if can_fin:
        with tab_re:
            st.markdown("#### Resumen ejecutivo")
            st.caption(
                "KPIs, PDF imprimible, cuentas, flujo multimoneda y gráficos del período. "
                "Las fechas de abajo son **solo** para este reporte (independientes del Dashboard)."
            )
            rc1, rc2 = st.columns(2)
            with rc1:
                d_re_a = st.date_input(
                    "Desde",
                    value=date.today() - timedelta(days=30),
                    key="rep_res_ej_desde",
                )
            with rc2:
                d_re_b = st.date_input("Hasta", value=date.today(), key="rep_res_ej_hasta")
            if d_re_b < d_re_a:
                st.error("La fecha *Hasta* debe ser ≥ *Desde*.")
            else:
                k_rep = _dashboard_kpis_periodo(sb, d_re_a, d_re_b)
                panel_resumen_ejecutivo_periodo_ui(
                    sb,
                    t,
                    d_re_a,
                    d_re_b,
                    k_rep,
                    pdf_download_key="rep_resumen_ejecutivo_pdf_dl",
                )

        with tab_inv:
            st.markdown("#### Listado de repuestos y productos")
            st.caption(
                "**Paso 1:** usá los filtros del panel de abajo. **Paso 2:** revisá la vista previa. **Paso 3:** descargá **Excel**, **PDF** o **HTML** (imprimible desde el navegador)."
            )
            panel_reportes_inventario_export(sb, t)
    
        with tab_caja:
            st.markdown("#### Dinero que entró y salió de cada cuenta")
            st.caption(
                "**Paso 1:** fechas y tipo (todo / solo entradas / solo salidas). **Paso 2:** cuenta o *Todas*. **Paso 3:** revisá la tabla y los totales. **Paso 4:** descargá Excel o CSV."
            )
            c1, c2, c3 = st.columns(3)
            d_caja_a = c1.date_input("Desde", value=date.today() - timedelta(days=30), key="rep_caja_desde")
            d_caja_b = c2.date_input("Hasta", value=date.today(), key="rep_caja_hasta")
            tipo_sel = c3.selectbox(
                "Qué mostrar",
                options=["Todo", "Solo entradas de dinero", "Solo salidas de dinero"],
                key="rep_caja_tipo",
                help="Entrada = cobraste o ingresó dinero a la cuenta. Salida = pagaste o retiraste.",
            )
            tipo_f = None
            if tipo_sel == "Solo entradas de dinero":
                tipo_f = "Ingreso"
            elif tipo_sel == "Solo salidas de dinero":
                tipo_f = "Egreso"
    
            cajas_r = _cajas_fetch_rows(sb, solo_activas=False)
            caja_ids_all, caja_fmt_all = _caja_select_options(cajas_r) if cajas_r else ([], lambda x: str(x))
            cuenta_opciones = ["(Todas las cuentas)"] + caja_ids_all
    
            def _fmt_cuenta_opt(x: str) -> str:
                if x == "(Todas las cuentas)":
                    return "Todas las cuentas"
                return caja_fmt_all(x)
    
            caja_pick = st.selectbox(
                "Cuenta o caja",
                options=cuenta_opciones,
                format_func=_fmt_cuenta_opt,
                key="rep_caja_cuenta",
                help="Elegí una cuenta suelta (por ejemplo un banco en bolívares) o dejá **Todas** para ver todo junto.",
            )
            caja_f = None if caja_pick == "(Todas las cuentas)" else str(caja_pick)
    
            movs = _rep_movimientos_caja_filtrados(sb, desde=d_caja_a, hasta=d_caja_b, caja_id=caja_f, tipo_mov=tipo_f)
            umap = {
                str(u["id"]): (u.get("nombre") or u.get("username") or "")
                for u in (sb.table("erp_users").select("id,nombre,username").execute().data or [])
            }
            cmap = {str(c["id"]): _caja_etiqueta_lista(c) for c in cajas_r}
    
            filas_mc: list[dict[str, Any]] = []
            for m in movs:
                mon = (m.get("moneda") or "USD") or "USD"
                mm = m.get("monto_moneda")
                _musd = _round_money_2(m.get("monto_usd"))
                mon_u = str(mon).upper()
                ex_bk, ex_amt = _movimiento_monto_explicito_columnas(m)
                if ex_bk is not None and ex_amt is not None:
                    mon_orig = "ZELLE" if mon_u == "ZELLE" and ex_bk == "USD" else ex_bk
                    _mm_orig = _round_money_2(ex_amt)
                else:
                    mon_orig = mon_u
                    _mm_orig = _round_money_2(mm) if mm is not None and str(mm).strip() != "" else None
                filas_mc.append(
                    {
                        "Fecha y hora": str(m.get("created_at", ""))[:19],
                        "Cuenta": cmap.get(str(m.get("caja_id")), "—"),
                        "Entrada o salida": "Entrada (cobro / ingreso)" if m.get("tipo") == "Ingreso" else "Salida (pago / egreso)",
                        "Monto en USD (sistema)": _musd,
                        "Moneda original": mon_orig,
                        "Monto en moneda original": _mm_orig,
                        "Concepto": (m.get("concepto") or "")[:120],
                        "Referencia": (m.get("referencia") or "")[:80],
                        "Nota tesorería": str(m.get("nota_operacion") or "")[:200],
                        "Registrado por": umap.get(str(m.get("usuario_id")), "—"),
                    }
                )
            df_mc = pd.DataFrame(filas_mc)
            if df_mc.empty:
                st.info("No hay movimientos en esas fechas y filtros. Probá ampliar el rango o elegir **Todas las cuentas**.")
            else:
                st.dataframe(
                    df_mc,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Monto en USD (sistema)": st.column_config.NumberColumn(format="%.2f"),
                        "Monto en moneda original": st.column_config.NumberColumn(format="%.2f"),
                    },
                )
                tot_in = _round_money_2(
                    df_mc[df_mc["Entrada o salida"].str.startswith("Entrada")]["Monto en USD (sistema)"].sum()
                )
                tot_out = _round_money_2(
                    df_mc[df_mc["Entrada o salida"].str.startswith("Salida")]["Monto en USD (sistema)"].sum()
                )
                m1, m2 = st.columns(2)
                with m1:
                    st.metric("Total entradas (USD en el sistema)", f"{tot_in:,.2f}")
                with m2:
                    st.metric("Total salidas (USD en el sistema)", f"{tot_out:,.2f}")
                st.caption(
                    "Los **USD** son el equivalente que guardó el sistema al momento del movimiento (bolívares y USDT convertidos con la tasa de entonces)."
                )
    
            ts_c = _backup_file_timestamp()
            colx, colc = st.columns(2)
            with colx:
                try:
                    st.download_button(
                        label=f"Descargar Excel — movimientos_caja_{ts_c}.xlsx",
                        data=_reporte_tabla_a_excel(df_mc if not df_mc.empty else pd.DataFrame(), nombre_hoja="Movimientos"),
                        file_name=f"movimientos_caja_{ts_c}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="rep_dl_caja_xlsx",
                        use_container_width=True,
                    )
                except ImportError:
                    st.caption("Para Excel hace falta instalar **openpyxl** (ya viene en los requisitos del programa).")
            with colc:
                st.download_button(
                    label=f"Descargar CSV — movimientos_caja_{ts_c}.csv",
                    data=_reporte_tabla_a_csv(df_mc if not df_mc.empty else pd.DataFrame()),
                    file_name=f"movimientos_caja_{ts_c}.csv",
                    mime="text/csv",
                    key="rep_dl_caja_csv",
                    use_container_width=True,
                )
    
        with tab_ven:
            st.markdown("#### Ventas")
            st.caption("**Paso 1:** fechas. **Paso 2:** revisá el **resumen por venta** y el gráfico. **Paso 3 (opcional):** abrí *Más detalle* para ganancias por producto, cada línea facturada y descargas.")
            d1, d2 = st.columns(2)
            a = d1.date_input("Desde", value=date.today() - timedelta(days=30), key="rep_ven_desde")
            b = d2.date_input("Hasta", value=date.today(), key="rep_ven_hasta")
    
            ventas = (
                sb.table("ventas")
                .select("id, numero, cliente, fecha, total_usd, forma_pago, usuario_id")
                .gte("fecha", str(a))
                .lte("fecha", f"{b}T23:59:59")
                .order("fecha", desc=True)
                .execute()
            )
            umap_v = {
                str(u["id"]): (u.get("nombre") or u.get("username") or "—")
                for u in (sb.table("erp_users").select("id,nombre,username").execute().data or [])
            }
    
            st.markdown("##### Resumen por venta")
            if ventas.data:
                dfv = pd.DataFrame(ventas.data)
                if have_t:
                    dfv["Equiv. aprox. en Bs (según tasa de hoy)"] = (
                        dfv["total_usd"].astype(float) * t_bs
                    ).round(0).astype("Int64")
                dfv["Fecha"] = pd.to_datetime(dfv["fecha"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M")
                dfv["Quién la registró"] = dfv["usuario_id"].map(lambda x: umap_v.get(str(x), "—"))
                _rv = {
                    "numero": "Nº interno",
                    "cliente": "Cliente",
                    "total_usd": "Total USD",
                    "forma_pago": "Forma de pago",
                }
                dfv2 = dfv.rename(columns=_rv)
                _cols_v = ["Nº interno", "Fecha", "Cliente", "Forma de pago", "Total USD"]
                if have_t:
                    _cols_v.append("Equiv. aprox. en Bs (según tasa de hoy)")
                _cols_v.append("Quién la registró")
                dfv_disp = dfv2[_cols_v]
                dfv_disp["Total USD"] = _rep_series_montos_enteros(dfv_disp["Total USD"])
                st.dataframe(dfv_disp, use_container_width=True, hide_index=True)
                dfv["fecha_d"] = pd.to_datetime(dfv["fecha"], errors="coerce").dt.strftime("%Y-%m-%d")
                agg = dfv.groupby("fecha_d", as_index=False)["total_usd"].sum()
                fig = px.bar(
                    agg,
                    x="fecha_d",
                    y="total_usd",
                    labels={"fecha_d": "Día", "total_usd": "Dólares (USD)"},
                    title="Vendido en dólares por día (total del día)",
                )
                fig.update_layout(yaxis=dict(tickformat=",d"))
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No hay ventas registradas entre esas fechas.")
    
            vids = [str(x["id"]) for x in (ventas.data or [])]
            det_rows: list[dict[str, Any]] = []
            if vids:
                det = (
                    sb.table("ventas_detalles")
                    .select("venta_id, producto_id, cantidad, precio_unitario_usd, subtotal_usd")
                    .in_("venta_id", vids)
                    .execute()
                )
                det_rows = det.data or []

            with st.expander("Más detalle: ganancias por producto y cada línea de venta (con descargas)", expanded=False):
                st.markdown("##### Cuánto ganaste aproximado por producto (mismo período)")
                st.caption(
                    "**Ganancia bruta** simple: precio de venta menos costo del producto × cantidades. No incluye gastos fijos."
                )
                if det_rows:
                    pmap = {
                        str(p["id"]): p
                        for p in (sb.table("productos").select("id,descripcion,codigo,costo_usd").execute().data or [])
                    }
                    rows_m = []
                    for row in det_rows:
                        pid = str(row["producto_id"])
                        pr = pmap.get(pid, {})
                        desc = pr.get("descripcion", pid)
                        costo = float(pr.get("costo_usd") or 0)
                        cant = float(row["cantidad"])
                        pu = float(row["precio_unitario_usd"])
                        margin = (pu - costo) * cant
                        rows_m.append({"producto": desc, "utilidad_bruta_usd": margin})
                    dfm = pd.DataFrame(rows_m).groupby("producto", as_index=False)["utilidad_bruta_usd"].sum()
                    dfm = dfm.rename(columns={"producto": "Producto", "utilidad_bruta_usd": "Ganancia bruta USD (aprox.)"})
                    dfm["Ganancia bruta USD (aprox.)"] = _rep_series_montos_enteros(dfm["Ganancia bruta USD (aprox.)"])
                    st.dataframe(dfm, use_container_width=True, hide_index=True)
                    if have_t:
                        st.caption(fmt_tri(float(dfm["Ganancia bruta USD (aprox.)"].sum()), t_bs, t_usdt))
                else:
                    st.info("No hay líneas de venta en ese período.")

                st.markdown("##### Cada artículo en cada venta")
                st.caption("Una fila por línea facturada (para buscar un repuesto o exportar).")
                filas_det: list[dict[str, Any]] = []
                if det_rows and ventas.data:
                    pmap2 = {
                        str(p["id"]): p
                        for p in (sb.table("productos").select("id,descripcion,codigo").execute().data or [])
                    }
                    vhead = {str(v["id"]): v for v in ventas.data}
                    for row in det_rows:
                        vid = str(row.get("venta_id"))
                        vh = vhead.get(vid, {})
                        pid = str(row["producto_id"])
                        pr = pmap2.get(pid, {})
                        filas_det.append(
                            {
                                "Nº venta": vh.get("numero", ""),
                                "Fecha venta": str(vh.get("fecha", ""))[:19],
                                "Cliente": vh.get("cliente", ""),
                                "Forma de pago": vh.get("forma_pago", ""),
                                "Quién registró": umap_v.get(str(vh.get("usuario_id")), "—"),
                                "Código": _export_cell_txt(pr.get("codigo")) or "—",
                                "Descripción": _export_cell_txt(pr.get("descripcion")) or pid,
                                "Cantidad": float(row.get("cantidad") or 0),
                                "Precio unitario USD": int(round(float(row.get("precio_unitario_usd") or 0))),
                                "Subtotal USD": int(round(float(row.get("subtotal_usd") or 0))),
                            }
                        )
                df_det = pd.DataFrame(filas_det)
                if df_det.empty:
                    st.info("No hay detalle para mostrar en esas fechas.")
                else:
                    st.dataframe(df_det, use_container_width=True, hide_index=True)
                ts_v = _backup_file_timestamp()
                vx, vc = st.columns(2)
                with vx:
                    try:
                        st.download_button(
                            label=f"Excel — detalle_ventas_{ts_v}.xlsx",
                            data=_reporte_tabla_a_excel(df_det, nombre_hoja="Ventas detalle"),
                            file_name=f"detalle_ventas_{ts_v}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="rep_dl_ven_det_xlsx",
                            use_container_width=True,
                        )
                    except ImportError:
                        pass
                with vc:
                    st.download_button(
                        label=f"CSV — detalle_ventas_{ts_v}.csv",
                        data=_reporte_tabla_a_csv(df_det),
                        file_name=f"detalle_ventas_{ts_v}.csv",
                        mime="text/csv",
                        key="rep_dl_ven_det_csv",
                        use_container_width=True,
                    )

        with tab_comp:
            st.markdown("#### Compras a proveedores")
            st.caption(
                "**Paso 1:** fechas. **Paso 2:** tabla y gráfico del período. **Paso 3 (opcional):** *Más detalle* por artículo y descargas. **Paso 4:** más abajo, **pendientes de pagar** al proveedor."
            )
            d1c, d2c = st.columns(2)
            ac = d1c.date_input("Desde", value=date.today() - timedelta(days=30), key="rep_comp_desde")
            bc = d2c.date_input("Hasta", value=date.today(), key="rep_comp_hasta")
    
            compras_r = (
                sb.table("compras")
                .select("id, numero, proveedor, fecha, total_usd, forma_pago, usuario_id")
                .gte("fecha", str(ac))
                .lte("fecha", f"{bc}T23:59:59")
                .order("fecha", desc=True)
                .execute()
            )
            umap_c = {
                str(u["id"]): (u.get("nombre") or u.get("username") or "—")
                for u in (sb.table("erp_users").select("id,nombre,username").execute().data or [])
            }
    
            st.markdown("##### Compras en el rango")
            if compras_r.data:
                dfcmp = pd.DataFrame(compras_r.data)
                if have_t:
                    dfcmp["Equiv. aprox. en Bs (tasa de hoy)"] = (
                        dfcmp["total_usd"].astype(float) * t_bs
                    ).round(0).astype("Int64")
                dfcmp["Fecha"] = pd.to_datetime(dfcmp["fecha"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M")
                dfcmp["Quién la registró"] = dfcmp["usuario_id"].map(lambda x: umap_c.get(str(x), "—"))
                _rc = {
                    "numero": "Nº interno",
                    "proveedor": "Proveedor",
                    "total_usd": "Total USD",
                    "forma_pago": "Forma de pago",
                }
                dfc2 = dfcmp.rename(columns=_rc)
                _cols_c = ["Nº interno", "Fecha", "Proveedor", "Forma de pago", "Total USD"]
                if have_t:
                    _cols_c.append("Equiv. aprox. en Bs (tasa de hoy)")
                _cols_c.append("Quién la registró")
                dfc_disp = dfc2[_cols_c]
                dfc_disp["Total USD"] = _rep_series_montos_enteros(dfc_disp["Total USD"])
                st.dataframe(dfc_disp, use_container_width=True, hide_index=True)
                dfcmp["fecha_d"] = pd.to_datetime(dfcmp["fecha"], errors="coerce").dt.strftime("%Y-%m-%d")
                aggc = dfcmp.groupby("fecha_d", as_index=False)["total_usd"].sum()
                figc = px.bar(
                    aggc,
                    x="fecha_d",
                    y="total_usd",
                    labels={"fecha_d": "Día", "total_usd": "Dólares (USD)"},
                    title="Compras en dólares por día",
                )
                figc.update_layout(yaxis=dict(tickformat=",d"))
                st.plotly_chart(figc, use_container_width=True)
            else:
                st.info("No hay compras entre esas fechas.")
    
            cids = [str(x["id"]) for x in (compras_r.data or [])]
            det_c: list[dict[str, Any]] = []
            if cids:
                dc = (
                    sb.table("compras_detalles")
                    .select("compra_id, producto_id, cantidad, costo_unitario_usd, subtotal_usd")
                    .in_("compra_id", cids)
                    .execute()
                )
                det_c = dc.data or []

            with st.expander("Más detalle: cada artículo comprado (con descargas)", expanded=False):
                st.markdown("##### Detalle por artículo comprado")
                filas_cd: list[dict[str, Any]] = []
                if det_c and compras_r.data:
                    pmap_c = {
                        str(p["id"]): p
                        for p in (sb.table("productos").select("id,descripcion,codigo").execute().data or [])
                    }
                    head_c = {str(v["id"]): v for v in compras_r.data}
                    for row in det_c:
                        cid = str(row.get("compra_id"))
                        ch = head_c.get(cid, {})
                        pid = str(row["producto_id"])
                        pr = pmap_c.get(pid, {})
                        filas_cd.append(
                            {
                                "Nº compra": ch.get("numero", ""),
                                "Fecha": str(ch.get("fecha", ""))[:19],
                                "Proveedor": ch.get("proveedor", ""),
                                "Código": _export_cell_txt(pr.get("codigo")) or "—",
                                "Descripción": _export_cell_txt(pr.get("descripcion")) or pid,
                                "Cantidad": float(row.get("cantidad") or 0),
                                "Costo unit. USD": int(round(float(row.get("costo_unitario_usd") or 0))),
                                "Subtotal USD": int(round(float(row.get("subtotal_usd") or 0))),
                            }
                        )
                df_cd = pd.DataFrame(filas_cd)
                if df_cd.empty:
                    st.info("No hay líneas de compra en ese período.")
                else:
                    st.dataframe(df_cd, use_container_width=True, hide_index=True)
                ts_cp = _backup_file_timestamp()
                cpx, cpc = st.columns(2)
                with cpx:
                    try:
                        st.download_button(
                            label=f"Excel — detalle_compras_{ts_cp}.xlsx",
                            data=_reporte_tabla_a_excel(df_cd, nombre_hoja="Compras detalle"),
                            file_name=f"detalle_compras_{ts_cp}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="rep_dl_comp_det_xlsx",
                            use_container_width=True,
                        )
                    except ImportError:
                        pass
                with cpc:
                    st.download_button(
                        label=f"CSV — detalle_compras_{ts_cp}.csv",
                        data=_reporte_tabla_a_csv(df_cd),
                        file_name=f"detalle_compras_{ts_cp}.csv",
                        mime="text/csv",
                        key="rep_dl_comp_det_csv",
                        use_container_width=True,
                    )

            st.divider()
            st.markdown("##### Facturas o deudas pendientes de pagar al proveedor")
            cxp = sb.table("cuentas_por_pagar").select("*").execute()
            if cxp.data:
                dfp = pd.DataFrame(cxp.data)
                st.dataframe(dfp, use_container_width=True, hide_index=True)
                pend_p = dfp[dfp["estado"].isin(["Pendiente", "Parcial"])]["monto_pendiente_usd"].astype(float).sum()
                st.metric("Total aún por pagar (USD)", f"{int(round(pend_p)):,d}")
                if have_t:
                    st.caption(fmt_tri(pend_p, t_bs, t_usdt))
            else:
                st.info("No hay cuentas por pagar cargadas.")
    
        with tab_cartera:
            st.markdown("#### Clientes que deben y proveedores a los que debes")
            st.caption(
                "**Paso 1:** revisá totales y **resumen por plazo**. **Paso 2:** abrí *Listado completo* si necesitás todas las filas. **Paso 3:** descargá Excel/CSV abajo. "
                "*¿Qué tan al día está?* indica si la fecha límite ya pasó."
            )
    
            ventas_all = {str(v["id"]): v for v in (sb.table("ventas").select("id, numero, cliente").execute().data or [])}
            compras_all = {str(c["id"]): c for c in (sb.table("compras").select("id, numero, proveedor").execute().data or [])}
    
            df_dl_cob = pd.DataFrame()
            df_dl_pag = pd.DataFrame()
    
            cxc = sb.table("cuentas_por_cobrar").select("*").execute()
            if cxc.data:
                st.markdown("##### Te deben (clientes)")
                rows_cxc: list[dict[str, Any]] = []
                for r in cxc.data:
                    vid = str(r.get("venta_id") or "")
                    vh = ventas_all.get(vid, {})
                    rows_cxc.append(
                        {
                            "Cliente": vh.get("cliente", "—"),
                            "Nº venta": vh.get("numero", "—"),
                            "Estado del crédito": r.get("estado", ""),
                            "Fecha límite de cobro": str(r.get("fecha_vencimiento") or "")[:10],
                            "¿Qué tan al día está?": _rep_texto_plazo_vencimiento(r.get("fecha_vencimiento")),
                            "Grupo (para totales)": _rep_bucket_antiguedad(r.get("fecha_vencimiento")),
                            "Adeudado USD": int(round(float(r.get("monto_pendiente_usd") or 0))),
                        }
                    )
                df_cxc = pd.DataFrame(rows_cxc)
                df_dl_cob = df_cxc
                pend_c = df_cxc[df_cxc["Estado del crédito"].isin(["Pendiente", "Parcial"])]["Adeudado USD"].sum()
                st.metric("Total que te deben (pendiente, USD)", f"{int(round(pend_c)):,d}")
                if have_t:
                    st.caption(fmt_tri(pend_c, t_bs, t_usdt))

                res_c = (
                    df_cxc[df_cxc["Estado del crédito"].isin(["Pendiente", "Parcial"])]
                    .groupby("Grupo (para totales)", as_index=False)["Adeudado USD"]
                    .sum()
                    .sort_values("Adeudado USD", ascending=False)
                )
                res_c["Adeudado USD"] = _rep_series_montos_enteros(res_c["Adeudado USD"])
                st.markdown("**Resumen: te deben — agrupado por plazo**")
                st.dataframe(res_c, use_container_width=True, hide_index=True)
                with st.expander("Listado completo — clientes que te deben (todas las filas)", expanded=False):
                    st.dataframe(df_cxc, use_container_width=True, hide_index=True)
            else:
                st.info("No hay cuentas por cobrar.")
    
            st.divider()
            cxp2 = sb.table("cuentas_por_pagar").select("*").execute()
            if cxp2.data:
                st.markdown("##### Debes a proveedores")
                rows_cxp: list[dict[str, Any]] = []
                for r in cxp2.data:
                    cid = str(r.get("compra_id") or "")
                    ch = compras_all.get(cid, {})
                    rows_cxp.append(
                        {
                            "Proveedor": ch.get("proveedor", "—"),
                            "Nº compra": ch.get("numero", "—"),
                            "Estado": r.get("estado", ""),
                            "Fecha límite de pago": str(r.get("fecha_vencimiento") or "")[:10],
                            "¿Qué tan al día está?": _rep_texto_plazo_vencimiento(r.get("fecha_vencimiento")),
                            "Grupo (para totales)": _rep_bucket_antiguedad(r.get("fecha_vencimiento")),
                            "Debes USD": int(round(float(r.get("monto_pendiente_usd") or 0))),
                        }
                    )
                df_cxp = pd.DataFrame(rows_cxp)
                df_dl_pag = df_cxp
                pend_x = df_cxp[df_cxp["Estado"].isin(["Pendiente", "Parcial"])]["Debes USD"].sum()
                st.metric("Total que debes pagar (pendiente, USD)", f"{int(round(pend_x)):,d}")
                if have_t:
                    st.caption(fmt_tri(pend_x, t_bs, t_usdt))
                res_x = (
                    df_cxp[df_cxp["Estado"].isin(["Pendiente", "Parcial"])]
                    .groupby("Grupo (para totales)", as_index=False)["Debes USD"]
                    .sum()
                    .sort_values("Debes USD", ascending=False)
                )
                res_x["Debes USD"] = _rep_series_montos_enteros(res_x["Debes USD"])
                st.markdown("**Resumen: debes — agrupado por plazo**")
                st.dataframe(res_x, use_container_width=True, hide_index=True)
                with st.expander("Listado completo — proveedores a pagar (todas las filas)", expanded=False):
                    st.dataframe(df_cxp, use_container_width=True, hide_index=True)
            else:
                st.info("No hay cuentas por pagar.")
    
            st.markdown("##### Bajar estos listados a tu computadora")
            ts_car = _backup_file_timestamp()
            _parts_dl = []
            if not df_dl_cob.empty:
                _parts_dl.append(df_dl_cob.assign(**{"Listado": "Clientes que te deben"}))
            if not df_dl_pag.empty:
                _parts_dl.append(df_dl_pag.assign(**{"Listado": "Proveedores a pagar"}))
            df_car_csv = pd.concat(_parts_dl, ignore_index=True) if _parts_dl else pd.DataFrame()
    
            ca1, ca2, ca3 = st.columns(3)
            with ca1:
                try:
                    st.download_button(
                        label=f"Excel — clientes_que_deben_{ts_car}.xlsx",
                        data=_reporte_tabla_a_excel(df_dl_cob, nombre_hoja="Te deben"),
                        file_name=f"clientes_que_deben_{ts_car}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="rep_dl_car_cxc_xlsx",
                        use_container_width=True,
                        disabled=df_dl_cob.empty,
                    )
                except ImportError:
                    st.caption("Instalá **openpyxl** para generar Excel.")
            with ca2:
                try:
                    st.download_button(
                        label=f"Excel — proveedores_a_pagar_{ts_car}.xlsx",
                        data=_reporte_tabla_a_excel(df_dl_pag, nombre_hoja="Debes pagar"),
                        file_name=f"proveedores_a_pagar_{ts_car}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="rep_dl_car_cxp_xlsx",
                        use_container_width=True,
                        disabled=df_dl_pag.empty,
                    )
                except ImportError:
                    pass
            with ca3:
                st.download_button(
                    label=f"CSV — todo_junto_{ts_car}.csv",
                    data=_reporte_tabla_a_csv(df_car_csv),
                    file_name=f"cartera_cobrar_y_pagar_{ts_car}.csv",
                    mime="text/csv",
                    key="rep_dl_car_csv",
                    use_container_width=True,
                    disabled=df_car_csv.empty,
                )

    with tab_cat:
        panel_reportes_catalogo_fotos(sb, erp_uid)

def module_usuarios(sb: Client, *, embedded_in_mantenimiento: bool = False) -> None:
    if embedded_in_mantenimiento:
        _uu1, _uu2 = st.columns([4.5, 1.05], gap="small")
        with _uu1:
            st.markdown("### Usuarios del sistema")
        with _uu2:
            with st.expander("Información", expanded=False, key="modinfo_exp_usuarios_emb"):
                st.markdown(
                    "Solo el **superusuario** puede crear cuentas y definir la contraseña inicial de cada persona."
                )
    else:
        _modulo_titulo_info(
            "Usuarios del sistema",
            key="usuarios",
            ayuda_md="Solo el **superusuario** puede crear cuentas y definir la contraseña inicial de cada persona.",
        )

    r = (
        sb.table("erp_users")
        .select("id,username,nombre,email,rol,activo,created_at")
        .order("nombre")
        .execute()
    )
    rows = r.data or []
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.divider()
    st.markdown("#### Nuevo usuario")
    with st.form(f"f_new_user_{int(st.session_state.get('erp_new_user_form_nonce', 0))}"):
        nu = st.text_input("Usuario (solo letras/números, sin espacios)", key="nu_user")
        nn = st.text_input("Nombre completo", key="nu_nom")
        ne = st.text_input("Correo (opcional)", key="nu_mail")
        nr = st.selectbox(
            "Rol",
            options=["vendedor", "admin", "almacen", "superuser"],
            format_func=lambda x: {
                "vendedor": "Vendedor (ventas y cobros CXC)",
                "admin": "Administrador (compras, cajas, reportes, tasas…)",
                "almacen": "Almacén (inventario; catálogo y etiquetas en Reportes)",
                "superuser": "Superusuario (acceso total)",
            }[x],
            key="nu_rol",
        )
        p1 = st.text_input("Contraseña inicial", type="password", key="nu_p1")
        p2 = st.text_input("Repetir contraseña", type="password", key="nu_p2")
        if st.form_submit_button("Crear usuario"):
            un = (nu or "").strip().lower()
            if not un or not nn.strip():
                st.error("Usuario y nombre son obligatorios.")
            elif not p1 or p1 != p2:
                st.error("Las contraseñas no coinciden o están vacías.")
            elif len(p1) < 4:
                st.error("La contraseña debe tener al menos 4 caracteres.")
            else:
                ex = sb.table("erp_users").select("id").eq("username", un).limit(1).execute()
                if (ex.data or []):
                    st.error("Ese usuario ya existe.")
                else:
                    sb.table("erp_users").insert(
                        {
                            "username": un,
                            "nombre": nn.strip(),
                            "email": ne.strip() or None,
                            "rol": nr,
                            "password_hash": _hash_password(p1),
                            "activo": True,
                        }
                    ).execute()
                    st.success(f"Usuario **{un}** creado. Ya puede iniciar sesión.")
                    _movi_ss_pop_keys("nu_user", "nu_nom", "nu_mail", "nu_rol", "nu_p1", "nu_p2")
                    _movi_bump_form_nonce("erp_new_user_form_nonce")
                    st.rerun()

    if not rows:
        return

    st.divider()
    st.markdown("#### Editar usuario")
    labels = {f"{u['nombre']} (@{u['username']})": u for u in rows}
    pick = st.selectbox("Seleccionar", options=list(labels.keys()))
    u = labels[pick]
    uid = str(u["id"])
    with st.form(f"f_edit_user_{uid}_{int(st.session_state.get(f'erp_edit_user_form_nonce_{uid}', 0))}"):
        act = st.checkbox("Activo", value=bool(u.get("activo", True)), key="ed_act")
        _roles = ["vendedor", "admin", "almacen", "superuser"]
        _ri = _roles.index(u["rol"]) if u["rol"] in _roles else 0
        new_rol = st.selectbox(
            "Rol",
            options=_roles,
            index=_ri,
            format_func=lambda x: {
                "vendedor": "Vendedor",
                "admin": "Administrador",
                "almacen": "Almacén",
                "superuser": "Superusuario",
            }[x],
            key="ed_rol",
        )
        np1 = st.text_input("Nueva contraseña (dejar vacío para no cambiar)", type="password", key="ed_p1")
        np2 = st.text_input("Repetir nueva contraseña", type="password", key="ed_p2")
        if st.form_submit_button("Guardar cambios"):
            if np1 or np2:
                if np1 != np2:
                    st.error("Las contraseñas nuevas no coinciden.")
                elif len(np1) < 4:
                    st.error("La contraseña debe tener al menos 4 caracteres.")
                else:
                    sb.table("erp_users").update(
                        {
                            "activo": act,
                            "rol": new_rol,
                            "password_hash": _hash_password(np1),
                        }
                    ).eq("id", uid).execute()
                    st.success("Usuario actualizado.")
                    _movi_ss_pop_keys("ed_act", "ed_rol", "ed_p1", "ed_p2")
                    _movi_bump_form_nonce(f"erp_edit_user_form_nonce_{uid}")
                    st.rerun()
            else:
                sb.table("erp_users").update({"activo": act, "rol": new_rol}).eq("id", uid).execute()
                st.success("Usuario actualizado.")
                _movi_ss_pop_keys("ed_act", "ed_rol", "ed_p1", "ed_p2")
                _movi_bump_form_nonce(f"erp_edit_user_form_nonce_{uid}")
                st.rerun()

    st.caption(
        "Si eres el único superusuario, evita desactivarte o quedarte sin contraseña."
    )


def render_cambiar_mi_password(sb: Client, erp_uid: str) -> None:
    if st.session_state.pop("pwd_updated_ok", False):
        st.success("Contraseña actualizada correctamente.")
    with st.expander("Cambiar mi contraseña", expanded=False):
        with st.form(f"f_mi_password_{int(st.session_state.get('mi_pwd_form_nonce', 0))}"):
            cur = st.text_input("Contraseña actual", type="password", autocomplete="current-password")
            n1 = st.text_input("Nueva contraseña", type="password", autocomplete="new-password")
            n2 = st.text_input("Confirmar nueva contraseña", type="password", autocomplete="new-password")
            if st.form_submit_button("Guardar nueva contraseña"):
                if not cur or not n1 or not n2:
                    st.error("Completa todos los campos.")
                elif n1 != n2:
                    st.error("La nueva contraseña y la confirmación no coinciden.")
                elif len(n1) < 4:
                    st.error("La nueva contraseña debe tener al menos 4 caracteres.")
                elif n1 == cur:
                    st.error("La nueva contraseña debe ser distinta a la actual.")
                else:
                    r = (
                        sb.table("erp_users")
                        .select("password_hash")
                        .eq("id", erp_uid)
                        .limit(1)
                        .execute()
                    )
                    row = (r.data or [{}])[0]
                    ph = (row.get("password_hash") or "").strip()
                    if not ph or not _password_ok(cur, ph):
                        st.error("La contraseña actual no es correcta.")
                    else:
                        sb.table("erp_users").update({"password_hash": _hash_password(n1)}).eq("id", erp_uid).execute()
                        st.session_state["pwd_updated_ok"] = True
                        _movi_bump_form_nonce("mi_pwd_form_nonce")
                        st.rerun()


def panel_respaldo_inventario_mantenimiento(sb: Client) -> None:
    """Respaldo y restauración JSON solo de categorías + productos (superusuario / Mantenimiento)."""
    st.caption(
        "Incluye **categorías** y **productos**. Para deshacer un cambio grande en el maestro de inventario. "
        "Si hay ventas o compras que usan esos productos, restaurar solo inventario puede fallar: en ese caso usá el **respaldo completo** arriba."
    )
    try:
        inv_payload = build_backup_inventario(sb)
        ts = _backup_file_timestamp()
        st.download_button(
            label=f"Descargar respaldo inventario — movi_inventario_{ts}.json",
            data=_json_backup_bytes(inv_payload),
            file_name=f"movi_inventario_{ts}.json",
            mime="application/json",
            key="mnt_dl_backup_inventario",
        )
        st.download_button(
            label=f"Mismo respaldo (.json.gz) — movi_inventario_{ts}.json.gz",
            data=_json_backup_bytes_compact_gzip(inv_payload),
            file_name=f"movi_inventario_{ts}.json.gz",
            mime="application/gzip",
            key="mnt_dl_backup_inventario_gz",
        )
        st.caption(
            f"**{len(inv_payload.get('categorias') or [])}** categorías · **{len(inv_payload.get('productos') or [])}** productos."
        )
    except Exception as e:
        st.error(f"No se pudo generar el respaldo de inventario: {e}")

    st.divider()
    st.markdown("**Restaurar inventario desde archivo**")
    up_inv = st.file_uploader("JSON o gzip de inventario", type=["json", "gz"], key="mnt_restore_inv_json")
    c_inv = st.text_input("Escribe **RESTAURAR_INVENTARIO** para confirmar", key="mnt_restore_inv_ok")
    if st.button("Restaurar inventario ahora", key="mnt_btn_restore_inv"):
        if up_inv is None:
            st.error("Subí el archivo JSON o .json.gz.")
        elif c_inv.strip() != "RESTAURAR_INVENTARIO":
            st.error("Confirmación incorrecta.")
        else:
            try:
                data_inv = decode_backup_upload_bytes(up_inv.getvalue())
            except Exception as ex:
                st.error(f"Archivo inválido: {ex}")
            else:
                ok_i, msg_i = restore_inventario_desde_json(sb, data_inv)
                if ok_i:
                    st.success(msg_i)
                    st.rerun()
                else:
                    st.error(msg_i)


def panel_anular_venta_compra_mantenimiento(sb: Client, erp_uid: str) -> None:
    st.markdown("#### Anular venta o compra (error de carga)")
    st.warning(
        "Usá esto solo si registraste **mal** una venta o compra. "
        "Se revierten los **movimientos de caja** ligados al documento, se ajusta el **stock** (y el **costo** en compras) y se elimina el registro. "
        "Si **borraste la venta a mano**, el cobro puede haber quedado en caja sin venta: en ese caso usá **Movimientos de caja huérfanos de venta** (más abajo). "
        "Hace falta ejecutar en Supabase **`supabase/patch_020_anular_venta_compra.sql`**."
    )
    t_an1, t_an2 = st.tabs(["Anular venta", "Anular compra"])
    with t_an1:
        try:
            rv = (
                sb.table("ventas")
                .select("id,numero,cliente,fecha,total_usd,forma_pago")
                .order("fecha", desc=True)
                .limit(200)
                .execute()
            )
            vrows = rv.data or []
        except Exception as e:
            st.error(str(e))
            vrows = []
        if not vrows:
            st.info("No hay ventas en la base.")
        else:
            v_opts: dict[str, str] = {}
            for x in vrows:
                lab = (
                    f"Venta #{x.get('numero')} — {str(x.get('cliente') or '')[:36]} — "
                    f"US$ {x.get('total_usd')} — {str(x.get('fecha') or '')[:16]} — {x.get('forma_pago')}"
                )
                v_opts[lab] = str(x["id"])
            pick_v = st.selectbox("Venta a anular", options=list(v_opts.keys()), key="mnt_anul_v_sel")
            conf_v = st.text_input("Escribí **ANULAR_VENTA** para confirmar", key="mnt_anul_v_conf")
            if st.button("Anular venta seleccionada", key="mnt_anul_v_btn"):
                if conf_v.strip() != "ANULAR_VENTA":
                    st.error("Confirmación incorrecta.")
                else:
                    try:
                        sb.rpc(
                            "anular_venta_erp",
                            {"p_usuario_id": erp_uid, "p_venta_id": v_opts[pick_v]},
                        ).execute()
                        st.success("Venta anulada. Revisá caja y stock.")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))
    with t_an2:
        try:
            rc = (
                sb.table("compras")
                .select("id,numero,proveedor,fecha,total_usd,forma_pago")
                .order("fecha", desc=True)
                .limit(200)
                .execute()
            )
            crows = rc.data or []
        except Exception as e:
            st.error(str(e))
            crows = []
        if not crows:
            st.info("No hay compras en la base.")
        else:
            c_opts: dict[str, str] = {}
            for x in crows:
                lab = (
                    f"Compra #{x.get('numero')} — {str(x.get('proveedor') or '')[:36]} — "
                    f"US$ {x.get('total_usd')} — {str(x.get('fecha') or '')[:16]} — {x.get('forma_pago')}"
                )
                c_opts[lab] = str(x["id"])
            pick_c = st.selectbox("Compra a anular", options=list(c_opts.keys()), key="mnt_anul_c_sel")
            conf_c = st.text_input("Escribí **ANULAR_COMPRA** para confirmar", key="mnt_anul_c_conf")
            if st.button("Anular compra seleccionada", key="mnt_anul_c_btn"):
                if conf_c.strip() != "ANULAR_COMPRA":
                    st.error("Confirmación incorrecta.")
                else:
                    try:
                        sb.rpc(
                            "anular_compra_erp",
                            {"p_usuario_id": erp_uid, "p_compra_id": c_opts[pick_c]},
                        ).execute()
                        st.success("Compra anulada. Revisá caja y costos de productos.")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))


def _concepto_es_cobro_venta_huerfano(concepto: object) -> bool:
    c = str(concepto or "").strip()
    return c.startswith("Venta #") or c.startswith("Abono / seña — venta crédito #")


def panel_revertir_movimientos_venta_huerfanos_mantenimiento(sb: Client, erp_uid: str) -> None:
    """Cobros en caja sin venta (p. ej. venta borrada a mano): revierte saldo y borra el movimiento."""
    st.markdown("#### Movimientos de caja huérfanos de venta")
    st.warning(
        "Si **borraste una venta** sin usar **Anular venta**, los **cobros en caja** pueden quedar con la venta en blanco "
        "pero el **saldo de la caja sigue como si el cobro existiera**. Elegí esos movimientos y revertí: se registra el "
        "movimiento opuesto en caja y se elimina el huérfano. En Supabase ejecutá **`supabase/patch_026_revertir_mov_caja_venta_huerfana.sql`**."
    )
    st.caption(
        "Para anular una venta con stock y documento coherentes usá **Anular venta** (arriba), no borres la fila de `ventas` a mano."
    )
    try:
        mh = (
            sb.table("movimientos_caja")
            .select("id,created_at,tipo,monto_usd,moneda,concepto,caja_id")
            .is_("venta_id", "null")
            .is_("compra_id", "null")
            .order("created_at", desc=True)
            .limit(500)
            .execute()
        )
        rows_all = mh.data or []
    except Exception as e:
        st.error(str(e))
        rows_all = []
    rows = [r for r in rows_all if _concepto_es_cobro_venta_huerfano(r.get("concepto"))]
    if not rows:
        st.info("No hay movimientos con concepto de **Venta #…** o **Abono crédito** sin venta/compra asociada (últimos 500 revisados).")
        return
    cajas_map: dict[str, str] = {}
    try:
        cr = sb.table("cajas_bancos").select("id,nombre").execute()
        for c in cr.data or []:
            cajas_map[str(c["id"])] = str(c.get("nombre") or "")[:40]
    except Exception:
        pass
    labels: list[str] = []
    label_to_id: dict[str, str] = {}
    for r in rows:
        cid = str(r.get("caja_id") or "")
        cn = cajas_map.get(cid, cid[:8] + "…")
        ts = str(r.get("created_at") or "")[:19]
        lab = (
            f"{ts} · {r.get('tipo')} · US$ {r.get('monto_usd')} · {str(r.get('moneda') or '')} · {cn} · "
            f"{str(r.get('concepto') or '')[:48]} · id {str(r.get('id'))[:8]}…"
        )
        labels.append(lab)
        label_to_id[lab] = str(r["id"])
    pick = st.multiselect("Movimientos a revertir (huérfanos de venta)", options=labels, key="mnt_huerf_mov_sel")
    conf = st.text_input("Escribí **REVERTIR_HUERFANOS** para confirmar", key="mnt_huerf_mov_conf")
    if st.button("Revertir movimientos seleccionados", key="mnt_huerf_mov_btn"):
        if not pick:
            st.error("Seleccioná al menos un movimiento.")
        elif conf.strip() != "REVERTIR_HUERFANOS":
            st.error("Confirmación incorrecta.")
        else:
            ids = [label_to_id[L] for L in pick]
            try:
                rpc = sb.rpc(
                    "revertir_movimientos_caja_venta_huerfanos_erp",
                    {"p_usuario_id": erp_uid, "p_movimiento_ids": ids},
                ).execute()
                n = _rpc_resp_int(rpc)
                st.success(
                    f"Listo: se revirtieron **{n if n is not None else len(ids)}** movimiento(s). Revisá saldos en Cajas y bancos."
                )
                st.rerun()
            except Exception as e:
                st.error(str(e))


def module_mantenimiento(sb: Client, erp_uid: str) -> None:
    _modulo_titulo_info(
        "Mantenimiento",
        key="mantenimiento",
        ayuda_md=(
            "**Usuarios:** alta y roles (superusuario). **Respaldo:** JSON completo o solo inventario; restauración y herramientas de anulación. "
            "Revisá cada sección antes de operaciones destructivas."
        ),
    )
    module_usuarios(sb, embedded_in_mantenimiento=True)
    st.divider()
    st.markdown("#### Respaldo de seguridad")
    st.caption(
        "Generá un archivo **JSON** con las tablas principales del ERP (ventas, compras, caja, productos, tasas, etc.). "
        "**No** incluye `password_hash` de usuarios: si restaurás en otra base, tendrás que **reasignar contraseñas**."
    )
    try:
        full_payload = build_backup_erp_completo(sb)
        ts = _backup_file_timestamp()
        st.download_button(
            label=f"Descargar respaldo completo — movi_erp_completo_{ts}.json",
            data=_json_backup_bytes(full_payload),
            file_name=f"movi_erp_completo_{ts}.json",
            mime="application/json",
            key="dl_backup_erp_completo",
        )
        st.download_button(
            label=f"Descargar mismo respaldo (liviano .json.gz) — movi_erp_completo_{ts}.json.gz",
            data=_json_backup_bytes_compact_gzip(full_payload),
            file_name=f"movi_erp_completo_{ts}.json.gz",
            mime="application/gzip",
            key="dl_backup_erp_completo_gz",
        )
        st.caption(
            "**Respaldo automático diario** (superusuario): al abrir la app se guarda "
            f"`auto_backups/movi_erp_auto_YYYY-MM-DD.json.gz` (JSON sin indentar + gzip). "
            "Ejecutá `supabase/patch_010_erp_kv.sql` para coordinar el día en la nube. "
            "Opcional: en `secrets.toml`, `[auto_backup]` → `storage_bucket` = bucket privado en Supabase Storage (carpeta `auto/`). "
            "Los archivos locales antiguos se borran pasado `retain_days` (default 14). "
            "Podés **restaurar** subiendo `.json` o `.json.gz`."
        )
        err_part = full_payload.get("meta", {}).get("errores_al_exportar")
        if err_part:
            st.warning("Algunas tablas fallaron al exportar (revisá permisos o columnas en Supabase).")
            st.json(err_part)
    except Exception as e:
        st.error(f"No se pudo generar el respaldo completo: {e}")

    st.divider()
    st.markdown("#### Respaldo y restauración — solo inventario")
    panel_respaldo_inventario_mantenimiento(sb)

    st.divider()
    st.markdown("#### Restaurar todo desde respaldo (1 clic)")
    st.warning(
        "**Sobrescribe** movimientos, ventas, compras, CXC/CXP, productos, categorías, tasas del día y cajas con el contenido del JSON. "
        "**No borra** filas de `erp_users`: actualiza datos básicos y crea usuarios faltantes con clave **Restaurar2025!** (cambiar después). "
        "Hacé un respaldo reciente antes."
    )
    up_full = st.file_uploader(
        "Archivo **movi_erp_completo_*.json** o **.json.gz**", type=["json", "gz"], key="restore_full_json"
    )
    c_full = st.text_input("Confirmación: escribe **RESTAURAR_TODO**", key="restore_full_ok")
    if st.button("Restaurar todo ahora", type="primary", key="btn_restore_full"):
        if up_full is None:
            st.error("Subí el archivo JSON del respaldo completo.")
        elif c_full.strip() != "RESTAURAR_TODO":
            st.error("Escribí exactamente RESTAURAR_TODO para confirmar.")
        else:
            try:
                blob = decode_backup_upload_bytes(up_full.getvalue())
            except Exception as ex:
                st.error(f"Archivo inválido (JSON o gzip+JSON): {ex}")
            else:
                ok_r, msg_r, warns_r = restore_erp_completo_desde_json(sb, blob)
                if ok_r:
                    st.success(msg_r)
                    for w in warns_r:
                        st.warning(w)
                    st.balloons()
                    st.rerun()
                else:
                    st.error(msg_r)
                    for w in warns_r:
                        st.warning(w)

    st.divider()
    panel_anular_venta_compra_mantenimiento(sb, erp_uid)

    st.divider()
    panel_revertir_movimientos_venta_huerfanos_mantenimiento(sb, erp_uid)

    st.divider()
    st.markdown("#### Depuración (peligroso)")
    st.warning(
        "Elimina movimientos, ventas y compras. Reinicia saldos de cajas a 0. **No borra productos ni usuarios.** "
        "**Descargá antes el respaldo completo** de arriba."
    )
    palabra = st.text_input('Escribe ELIMINAR para confirmar')
    if st.button("Ejecutar depuración") and palabra.strip().upper() == "ELIMINAR":
        try:
            try:
                sb.table("cambios_tesoreria").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
            except Exception:
                pass
            sb.table("movimientos_caja").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
            sb.table("ventas").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
            sb.table("compras").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
            sb.table("cajas_bancos").update({"saldo_actual_usd": 0}).neq("id", "00000000-0000-0000-0000-000000000000").execute()
            st.success("Depuración aplicada.")
            st.rerun()
        except Exception as e:
            st.error(str(e))


def main() -> None:
    if not _secrets_ready():
        st.error(
            "Falta configuración en `.streamlit/secrets.toml`. "
            "Copia `.streamlit/secrets.toml.example` y completa la conexión a Supabase."
        )
        st.stop()

    sb = get_supabase()
    cm: Any | None = None
    if _cookie_support():
        try:
            cm = _erp_cookie_manager()
            _try_restore_session_from_cookie(sb, cm)
        except Exception:
            cm = None
    erp = gate_user_login(sb, cm)
    if not erp:
        st.stop()

    rol = str(erp["rol"])
    erp_uid = str(erp["id"])
    username = str(erp.get("username", ""))

    maybe_run_daily_auto_backup(sb, rol)
    _bk_err = st.session_state.pop("_movi_auto_backup_toast_err", None)
    if _bk_err:
        st.toast(f"Respaldo automático no aplicado: {_bk_err}")

    t = latest_tasas(sb)
    t_synced = maybe_auto_sync_tasas_from_web(sb)
    if t_synced is not None:
        t = t_synced
    _auto_msg = st.session_state.pop("_tasas_auto_sync_msg", None)
    if _auto_msg:
        st.toast(_auto_msg)
    _bk_ok = st.session_state.pop("_movi_auto_backup_toast", None)
    if _bk_ok:
        st.toast(_bk_ok)

    with st.sidebar:
        with st.expander("Apariencia", expanded=False):
            render_movi_theme_picker(key_suffix="sb")
        render_brand_logo()
        render_sidebar_welcome(nombre=str(erp.get("nombre", username)), username=username, rol=rol)
        if st.button("Cerrar sesión", key="movi_sidebar_logout", use_container_width=True):
            _logout()
            st.rerun()
        st.caption("Cotizaciones **en vivo**: **Dashboard → Mercado en vivo** · **Resumen ejecutivo** en **Reportes**.")
        render_cambiar_mi_password(sb, erp_uid)
        render_sidebar_calculadora()
        st.markdown('<p class="sb-block-title">Navegación</p>', unsafe_allow_html=True)
        st.caption("Elegí el módulo en la **barra superior** (íconos).")

    opts = movi_nav_options_for_role(rol)
    if not opts:
        st.error("Tu rol no tiene módulos asignados. Pide al superusuario que revise tu cuenta.")
        st.caption("Podés **Cerrar sesión** con el botón de arriba (debajo de tu nombre).")
        st.stop()
    st.session_state.setdefault("movi_mod", opts[0])
    if st.session_state.get("movi_mod") not in opts:
        st.session_state["movi_mod"] = opts[0]
    if len(opts) > 1:
        render_movi_main_module_nav(opts)
    else:
        st.session_state["movi_mod"] = opts[0]
    mod = str(st.session_state["movi_mod"])

    if mod == "Dashboard" and role_can(rol, "dashboard"):
        module_dashboard(sb, t)
    elif mod == "Inventario" and role_can(rol, "inventario"):
        module_inventario(sb, erp_uid, t)
    elif mod == "Ventas / CXC" and role_can(rol, "ventas"):
        module_ventas(sb, erp_uid, t)
    elif mod == "Compras / CXP" and role_can(rol, "compras"):
        module_compras(sb, erp_uid, t)
    elif mod == "Cajas y bancos" and role_can(rol, "cajas"):
        module_cajas(sb, erp_uid)
    elif mod == "Gastos operativos" and role_can(rol, "cajas"):
        module_gastos_operativos(sb, erp_uid, t)
    elif mod == "Reportes" and (role_can(rol, "reportes") or role_can(rol, "catalogo")):
        module_reportes(sb, erp_uid, t, rol)
    elif mod == "Mantenimiento" and rol == "superuser":
        module_mantenimiento(sb, erp_uid)


if __name__ == "__main__":
    main()
